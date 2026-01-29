import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import numpy as np

def check_duplicate_nips(df, column_name='NIP'):
    """Cek NIP duplikat di dataframe dan return baris yang duplikat"""
    # Konversi ke string dan strip whitespace
    df['nip_clean'] = df[column_name].astype(str).str.strip()
    duplicates = df[df.duplicated('nip_clean', keep=False)]
    return duplicates

def check_new_data(df_mentah, df_master):
    """Cek NIP yang ada di data mentah tapi tidak ada di data master"""
    # Bersihkan NIP
    df_mentah['nip_clean'] = df_mentah['NIP'].astype(str).str.strip()
    df_master['nip_clean'] = df_master['NIP'].astype(str).str.strip()
    
    # Cari NIP yang tidak ada di master
    master_nips = set(df_master['nip_clean'].tolist())
    mentah_nips = set(df_mentah['nip_clean'].tolist())
    
    new_nips = mentah_nips - master_nips
    new_data = df_mentah[df_mentah['nip_clean'].isin(new_nips)]
    
    return new_data

def show():
    # Tombol kembali ke Dashboard PPPK
    if st.button("← Kembali ke Dashboard PPPK"):
        st.session_state.current_page = 'dashboard_pppk'
        st.session_state.selected_menu = None
        st.rerun()
    
    st.title("🍽️ Upload Pajak Makan PPPK")
    st.markdown("---")
    
    # ========== PANDUAN PENGGUNAAN ==========
    with st.expander("📚 Panduan Penggunaan - Baca Sebelum Mulai", expanded=False):
        st.markdown("""
        ### **🎯 Cara Menggunakan Fitur Ini:**
        
        1. **Pilih Masa & Tahun Pajak**:
           - Otomatis terdeteksi bulan dan tahun saat ini
           - Pastikan sesuai dengan data yang akan diupload
        
        2. **Siapkan 2 File Excel**:
           - **Data Mentah**: Hasil download sistem per bulan
           - **Data Master**: Database pegawai PPPK
        
        3. **Format Data Mentah**:
           ```
           Header yang HARUS ADA: NIP, NILAI KOTOR, STATUS KAWIN
           
           Header OPSIONAL:
           - PPH: Jika sudah ada nilai PPh langsung
           - PEMOTONG: Jika ingin sistem hitung otomatis
           
           CATATAN PENTING:
           - Minimal salah satu dari PPH atau PEMOTONG harus ada
           - Jika ada PPH, akan digunakan langsung
           - Jika tidak ada PPH, maka PEMOTONG WAJIB ADA
           ```
           
        4. **Format Data Master**:
           ```
           Header wajib: NIP, NIK, ID PENERIMA TKU, STATUS, KODE OBJEK PAJAK, ID TKU
           ```
        
        5. **PERHITUNGAN TARIF OTOMATIS**:
           ```
           RUMUS: Tarif = (PEMOTONG / NILAI KOTOR) × 100
           
           Contoh:
           - Nilai Kotor: 1.000.000
           - Pemotong: 50.000
           - Tarif = (50.000 / 1.000.000) × 100 = 5%
           ```
        
        6. **Validasi Data**:
           - NIP di Data Mentah harus ada di Data Master
           - Status Kawin akan dibandingkan dengan STATUS di Master
           - NPWP akan dibandingkan dengan NIK di Master
        """)
        
        # Hanya menyediakan template data master
        st.download_button(
            label="📥 Template Data Master",
            data=pd.DataFrame({
                'NIP': ['123456789012345678', '987654321098765432', '456789012345678901'],
                'NIK': ['123456789012345', '987654321098765', '456789012345678'],
                'ID PENERIMA TKU': ['ID001', 'ID002', 'ID003'],
                'STATUS': ['K', 'TK', 'K'],
                'KODE OBJEK PAJAK': ['21-100-01', '21-100-01', '21-100-01'],
                'ID TKU': ['TKU001', 'TKU002', 'TKU003']
            }).to_csv(index=False).encode('utf-8'),
            file_name="template_data_master.csv",
            mime="text/csv"
        )
    
    st.markdown("---")
    
    # ========== INFORMASI PERHITUNGAN TARIF & PENGISIAN MANUAL ==========
    st.info("""
    **📊 INFORMASI PERHITUNGAN TARIF BARU:**
    
    **Tarif ditentukan berdasarkan KODE OBJEK PAJAK di Data Master:**
    1. **21-402-02** → Tarif = 5%
    2. **21-402-03** → Tarif = 15%
    3. **21-402-04** → Tarif = 0%
    
    **🎨 KODE WARNA DI HASIL DOWNLOAD:**
    - **Hijau Muda (Soft Green)**: Data hasil sistem (bisa langsung copy)
    - **Oranye (Orange)**: Data yang harus diisi **MANUAL** dari DAFTAR SP2D SATKER
    
    **📋 DATA YANG HARUS DIISI MANUAL (Warna Oranye):**
    1. **Nomor Dok. Referensi** → Ambil dari **Nomor SP2D** di DAFTAR SP2D SATKER
    2. **Tanggal Dok. Referensi** → Ambil dari **Tanggal SP2D** di DAFTAR SP2D SATKER
    3. **Tanggal Pemotongan** → Ambil dari **Tanggal Invoice** di DAFTAR SP2D SATKER
    
    ⚠️ **KODE OBJEK PAJAK wajib ada di Data Master untuk menentukan tarif!**
    """)
    
    # ========== FORM INPUT ==========
    st.subheader("📅 Periode Pajak")
    st.info("⚠️ **Penting**: Pastikan data mentah yang diupload sesuai dengan masa pajak yang dipilih!")
    
    col1, col2 = st.columns(2)
    with col1:
        masa_pajak = st.selectbox(
            "**Masa Pajak (1-12)**",
            options=list(range(1, 13)),
            index=datetime.now().month - 1,
            help="Pilih bulan sesuai data mentah yang akan diupload"
        )
    with col2:
        tahun_sekarang = datetime.now().year
        tahun_pajak = st.selectbox(
            "**Tahun Pajak**",
            options=list(range(2000, tahun_sekarang + 5)),
            index=tahun_sekarang - 2000,
            help="Pilih tahun sesuai data mentah"
        )
    
    st.markdown("---")
    
    # ========== UPLOAD DATA MENTAH ==========
    st.subheader("📤 Upload Data Mentah")
    
    with st.expander("ℹ️ Spesifikasi Data Mentah", expanded=False):
        st.markdown("""
        ### **Header yang HARUS ADA:**
        ```
        1. NIP          → Kunci untuk merge dengan data master (WAJIB)
        2. NILAI KOTOR  → Nilai penghasilan kotor (WAJIB)
        3. STATUS KAWIN → Status pernikahan (K/TK) (WAJIB)
        ```
        
        ### **Catatan Perubahan Logika Tarif:**
        ```
        PERUBAHAN PENTING: Tarif TIDAK lagi dihitung dari PPH atau PEMOTONG!
        Tarif sekarang diambil dari KODE OBJEK PAJAK di Data Master:
        
        - 21-402-02 → Tarif 5%
        - 21-402-03 → Tarif 15%
        - 21-402-04 → Tarif 0%
        ```
        
        ### **📝 KOLOM PPH/PEMOTONG (OPSIONAL):**
        - Kolom PPH dan PEMOTONG sekarang TIDAK WAJIB
        - Jika ada, akan diabaikan dalam perhitungan tarif
        - Tetap dapat diupload jika ada di file
        """)
    
    uploaded_mentah = st.file_uploader(
        "**Pilih file Excel Data Mentah**", 
        type=['xlsx', 'xls'],
        key="upload_mentah_pppk",
        help="Upload file hasil download sistem untuk bulan yang dipilih. Pastikan ada NIP, NILAI KOTOR, dan STATUS KAWIN!"
    )
    
    # ========== UPLOAD DATA MASTER ==========
    st.subheader("📤 Upload Data Master")
    
    with st.expander("ℹ️ Spesifikasi Data Master", expanded=False):
        st.markdown("""
        ### **Header yang wajib ada**:
        ```
        1. NIP               → Untuk merge dengan data mentah
        2. NIK               → Digunakan sebagai NPWP di hasil akhir
        3. ID PENERIMA TKU   → ID penerima penghasilan
        4. STATUS            → Status PTKP (K/ TK/ HB/ etc.)
        5. KODE OBJEK PAJAK  → Kode objek pajak (WAJIB untuk tarif!)
        6. ID TKU            → ID pemotong pajak
        ```
        
        ### **Mapping Kode Objek Pajak ke Tarif**:
        ```
        21-402-02 → 5%   (Tarif 5 persen)
        21-402-03 → 15%  (Tarif 15 persen)
        21-402-04 → 0%   (Tanpa pajak)
        ```
        
        ### **Header tambahan (jika ada)**:
        ```
        7. PNS/PPPK       → Jenis pegawai
        8. NAMA           → Nama lengkap
        9. KDKAWIN        → Kode status kawin
        10. NMREK         → Nama rekening
        11. REKENING      → Nomor rekening
        12. AKTIF/TIDAK   → Status aktif
        ```
        """)
    
    uploaded_master = st.file_uploader(
        "**Pilih file Excel Data Master**", 
        type=['xlsx', 'xls'],
        key="upload_master_pppk",
        help="Upload database master pegawai PPPK. Pastikan ada KODE OBJEK PAJAK!"
    )
    
    st.markdown("---")
    
    # ========== PROSES DATA ==========
    if uploaded_mentah is not None and uploaded_master is not None:
        try:
            # Baca file Excel dengan konversi tipe data yang tepat
            # Untuk NIK, baca sebagai string untuk mencegah munculnya .0
            dtype_master = {
                'NIP': str,
                'NIK': str,
                'ID PENERIMA TKU': str,
                'STATUS': str,
                'KODE OBJEK PAJAK': str,
                'ID TKU': str
            }
            
            df_mentah = pd.read_excel(uploaded_mentah)
            df_master = pd.read_excel(uploaded_master, dtype=dtype_master)
            
            # Normalisasi nama kolom
            df_mentah.columns = df_mentah.columns.astype(str).str.strip().str.upper()
            df_master.columns = df_master.columns.astype(str).str.strip().str.upper()
            
            # Konversi kolom NIK di master menjadi string (jika belum)
            if 'NIK' in df_master.columns:
                df_master['NIK'] = df_master['NIK'].astype(str)
                # Hapus .0 dari NIK jika ada
                df_master['NIK'] = df_master['NIK'].apply(lambda x: x.split('.')[0] if '.' in str(x) else str(x))
                # Hilangkan whitespace
                df_master['NIK'] = df_master['NIK'].str.strip()
            
            st.success("✅ Kedua file berhasil diupload!")
            
            # Tampilkan preview data yang diupload
            col1, col2 = st.columns(2)
            with col1:
                with st.expander(f"📄 Data Mentah ({len(df_mentah)} baris)"):
                    st.write(f"**Kolom yang terdeteksi:**")
                    for col in df_mentah.columns:
                        st.write(f"  • {col}")
                    st.dataframe(df_mentah.head(3))
            
            with col2:
                with st.expander(f"📄 Data Master ({len(df_master)} baris)"):
                    st.write(f"**Kolom yang terdeteksi:**")
                    for col in df_master.columns:
                        st.write(f"  • {col}")
                    st.dataframe(df_master.head(3))
            
            # ========== VALIDASI DATA MENTAH ==========
            st.subheader("🔍 Validasi Data")
            
            # Validasi header wajib
            required_mentah = ['NIP', 'NILAI KOTOR', 'STATUS KAWIN']
            missing_mentah = [col for col in required_mentah if col not in df_mentah.columns]
            
            if missing_mentah:
                st.error(f"❌ **ERROR**: Header berikut tidak ditemukan di Data Mentah:")
                for col in missing_mentah:
                    st.write(f"   - **{col}**")
                st.warning("💡 **Solusi**: Pastikan header ada dan penulisannya benar.")
                st.stop()
            
            # ========== VALIDASI DATA MASTER ==========
            required_master = ['NIP', 'NIK', 'ID PENERIMA TKU', 'STATUS', 'KODE OBJEK PAJAK', 'ID TKU']
            missing_master = [col for col in required_master if col not in df_master.columns]
            
            if missing_master:
                st.error(f"❌ **ERROR**: Header berikut tidak ditemukan di Data Master:")
                for col in missing_master:
                    st.write(f"   - **{col}**")
                st.warning("💡 **Solusi**: Pastikan semua header wajib ada di Data Master")
                st.stop()
            
            # ========== CEK DUPLIKAT NIP ==========
            st.markdown("### **🔍 Cek Duplikat NIP**")
            
            # Cek duplikat di data mentah
            duplicates_mentah = check_duplicate_nips(df_mentah, 'NIP')
            if not duplicates_mentah.empty:
                st.error(f"❌ **Ditemukan {len(duplicates_mentah['NIP'].unique())} NIP duplikat di Data Mentah!**")
                
                # Tampilkan baris duplikat dengan warna merah
                st.warning("**Baris dengan NIP duplikat (ditandai merah):**")
                
                # Buat DataFrame dengan styling untuk duplikat
                def highlight_duplicates(df_original, duplicates_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris duplikat
                    mask = styled_df['NIP'].astype(str).str.strip().isin(
                        duplicates_df['nip_clean'].unique()
                    )
                    
                    # Buat list warna
                    colors = ['background-color: #ffcccc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_duplicates = highlight_duplicates(df_mentah.head(50), duplicates_mentah)
                st.dataframe(styled_duplicates, use_container_width=True)
                
                if len(df_mentah) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_mentah)} baris")
                
                st.error("**PERBAIKI NIP DUPLIKAT DI DATA MENTAH SEBELUM MELANJUTKAN!**")
                st.stop()
            else:
                st.success("✅ Tidak ada NIP duplikat di Data Mentah")
            
            # Cek duplikat di data master
            duplicates_master = check_duplicate_nips(df_master, 'NIP')
            if not duplicates_master.empty:
                st.error(f"❌ **Ditemukan {len(duplicates_master['NIP'].unique())} NIP duplikat di Data Master!**")
                
                # Tampilkan baris duplikat dengan warna merah
                st.warning("**Baris dengan NIP duplikat (ditandai merah):**")
                
                # Buat DataFrame dengan styling untuk duplikat
                def highlight_duplicates_master(df_original, duplicates_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris duplikat
                    mask = styled_df['NIP'].astype(str).str.strip().isin(
                        duplicates_df['nip_clean'].unique()
                    )
                    
                    # Buat list warna
                    colors = ['background-color: #ffcccc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_duplicates_master = highlight_duplicates_master(df_master.head(50), duplicates_master)
                st.dataframe(styled_duplicates_master, use_container_width=True)
                
                if len(df_master) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_master)} baris")
                
                st.error("**PERBAIKI NIP DUPLIKAT DI DATA MASTER SEBELUM MELANJUTKAN!**")
                st.stop()
            else:
                st.success("✅ Tidak ada NIP duplikat di Data Master")
            
            # ========== CEK DATA BARU ==========
            st.markdown("### **🔍 Cek Data Baru**")
            
            # Cek data baru (NIP di mentah tapi tidak di master)
            new_data = check_new_data(df_mentah, df_master)
            
            if not new_data.empty:
                st.warning(f"⚠️ **Ditemukan {len(new_data)} data baru di Data Mentah yang tidak ada di Data Master!**")
                
                # Tampilkan data baru dengan warna hijau
                st.info("**Data baru (ditandai hijau - perlu ditambahkan ke Data Master):**")
                
                def highlight_new_data(df_original, new_data_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris data baru
                    mask = styled_df['NIP'].astype(str).str.strip().isin(
                        new_data_df['nip_clean'].unique()
                    )
                    
                    # Buat list warna
                    colors = ['background-color: #ccffcc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_new_data = highlight_new_data(df_mentah.head(50), new_data)
                st.dataframe(styled_new_data, use_container_width=True)
                
                if len(df_mentah) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_mentah)} baris")
                
                # Tampilkan daftar NIP baru
                st.warning("**Daftar NIP baru yang perlu ditambahkan ke Data Master:**")
                for nip in new_data['NIP'].unique():
                    st.write(f"- {nip}")
                
                # TOMBOL UNTUK MENUJU KE HALAMAN CROSSCHECK PPPK
                st.markdown("---")
                st.error("**DATA BARU HARUS DITAMBAHKAN KE DATA MASTER SEBELUM MELANJUTKAN!**")
                
                # Tambahkan tombol untuk menuju ke halaman croscheck_pppk
                if st.button("➕ Tambahkan Data Baru ke Master (Croscheck PPPK)", type="primary", use_container_width=True):
                    st.session_state.current_page = 'croscheck_pppk'
                    st.session_state.selected_menu = 'croscheck'
                    st.rerun()
                
                st.stop()  # Hentikan proses sampai data baru ditangani
            else:
                st.success("✅ Tidak ditemukan data baru. Semua NIP di Data Mentah ada di Data Master.")
            
            # ========== INFO KODE OBJEK PAJAK ==========
            st.markdown("### **🔢 Informasi Kode Objek Pajak**")
            
            # Hitung distribusi kode objek pajak
            kode_counts = df_master['KODE OBJEK PAJAK'].value_counts()
            
            # Mapping kode objek pajak ke tarif
            mapping_tarif = {
                '21-402-02': 5,
                '21-402-03': 15,
                '21-402-04': 0
            }
            
            # Tampilkan informasi kode objek pajak
            col1, col2, col3, col4 = st.columns(4)
            
            total_kode = len(df_master)
            for kode, jumlah in kode_counts.head(4).items():
                tarif = mapping_tarif.get(str(kode).strip(), 'Tidak dikenal')
                if kode_counts.index.get_loc(kode) < 4:
                    with [col1, col2, col3, col4][kode_counts.index.get_loc(kode)]:
                        st.metric(
                            label=f"Kode: {kode}",
                            value=f"{jumlah} pegawai",
                            delta=f"Tarif: {tarif}%"
                        )
            
            # Periksa kode objek pajak yang tidak dikenal
            kode_unknown = []
            for kode in df_master['KODE OBJEK PAJAK'].unique():
                kode_str = str(kode).strip()
                if kode_str not in mapping_tarif:
                    kode_unknown.append(kode_str)
            
            if kode_unknown:
                st.warning(f"⚠️ **Peringatan**: {len(kode_unknown)} kode objek pajak tidak dikenal:")
                for kode in kode_unknown:
                    st.write(f"   - **{kode}** → Tarif akan dianggap 0%")
            
            # ========== TOMBOL PROSES ==========
            st.markdown("---")
            if st.button("🔄 **PROSES DATA & GENERATE HASIL**", 
                        use_container_width=True, 
                        type="primary",
                        help="Klik untuk memproses data dan menghasilkan file output"):
                
                with st.spinner("🔄 Sedang memproses data..."):
                    try:
                        # ========== PROSES MERGE DATA ==========
                        # Pastikan NIP di mentah juga string
                        df_mentah['NIP'] = df_mentah['NIP'].astype(str).str.strip()
                        df_master['NIP'] = df_master['NIP'].astype(str).str.strip()
                        
                        df_merged = df_mentah.merge(
                            df_master, 
                            on='NIP', 
                            how='inner',  # Menggunakan inner join untuk hanya data yang match
                            suffixes=('_MENTAH', '_MASTER')
                        )
                        
                        # Cek hasil merge
                        if len(df_merged) == 0:
                            st.error("❌ **ERROR**: Tidak ada data yang match antara Data Mentah dan Data Master!")
                            st.warning("Pastikan NIP di kedua file sama dan tidak ada duplikat.")
                            st.stop()
                        
                        st.success(f"✅ Berhasil merge {len(df_merged)} data dari {len(df_mentah)} data mentah")
                        
                        # ========== HITUNG TARIF BERDASARKAN KODE OBJEK PAJAK ==========
                        st.info("📊 **Menghitung tarif berdasarkan KODE OBJEK PAJAK**")
                        
                        # Bersihkan KODE OBJEK PAJAK (hapus whitespace)
                        df_merged['KODE OBJEK PAJAK'] = df_merged['KODE OBJEK PAJAK'].astype(str).str.strip()
                        
                        # Mapping kode objek pajak ke tarif
                        df_merged['TARIF_CALC'] = df_merged['KODE OBJEK PAJAK'].map(mapping_tarif)
                        
                        # Isi NaN dengan 0 (untuk kode yang tidak dikenal)
                        df_merged['TARIF_CALC'] = df_merged['TARIF_CALC'].fillna(0)
                        
                        # Tampilkan distribusi tarif
                        st.success(f"✅ Tarif berhasil dihitung berdasarkan KODE OBJEK PAJAK")
                        
                        # Tampilkan contoh perhitungan
                        with st.expander("📐 Contoh Mapping Kode ke Tarif"):
                            # Ambil 3 data contoh dengan kode yang berbeda
                            sample_codes = []
                            for kode in mapping_tarif.keys():
                                sample = df_merged[df_merged['KODE OBJEK PAJAK'] == kode]
                                if not sample.empty:
                                    sample_codes.append(sample.iloc[0])
                            
                            if sample_codes:
                                st.write("**Contoh Mapping:**")
                                for i, row in enumerate(sample_codes[:3]):
                                    st.write(f"**Contoh {i+1}:**")
                                    st.write(f"  • NIP: {row['NIP']}")
                                    st.write(f"  • Kode Objek Pajak: {row['KODE OBJEK PAJAK']}")
                                    st.write(f"  • Tarif: {row['TARIF_CALC']}%")
                                    st.write("---")
                            
                            # Tampilkan summary tarif
                            tarif_counts = df_merged['TARIF_CALC'].value_counts().sort_index()
                            st.write("**Distribusi Tarif:**")
                            for tarif, jumlah in tarif_counts.items():
                                persentase = (jumlah / len(df_merged)) * 100
                                st.write(f"  • Tarif {tarif}%: {jumlah} pegawai ({persentase:.1f}%)")
                        
                        # ========== BUAT DATA HASIL ==========
                        n_rows = len(df_merged)
                        
                        # Pastikan NIK bersih dari .0
                        nik_clean = []
                        for nik in df_merged['NIK'].fillna('').tolist():
                            if pd.isna(nik):
                                nik_clean.append('')
                            else:
                                # Konversi ke string dan hapus .0 jika ada
                                nik_str = str(nik)
                                if '.' in nik_str:
                                    nik_str = nik_str.split('.')[0]
                                nik_clean.append(nik_str.strip())
                        
                        # Hanya kolom yang diperlukan untuk hasil download
                        data_hasil = {
                            'Masa Pajak': [masa_pajak] * n_rows,
                            'Tahun Pajak': [tahun_pajak] * n_rows,
                            'NPWP': nik_clean,
                            'ID TKU Penerima Penghasilan': df_merged['ID PENERIMA TKU'].fillna('').astype(str).str.strip().tolist(),
                            'Status PTKP': df_merged['STATUS'].fillna('TK').astype(str).str.strip().tolist(),
                            'Fasilitas': ['DTP'] * n_rows,
                            'Kode Objek Pajak': df_merged['KODE OBJEK PAJAK'].fillna('').astype(str).str.strip().tolist(),
                            'Penghasilan': df_merged['NILAI KOTOR'].fillna(0).astype(float).tolist(),
                            'Deemed': [100] * n_rows,
                            'Tarif': df_merged['TARIF_CALC'].fillna(0).astype(float).tolist(),
                            'Jenis Dok. Referensi': ['CommercialInvoice'] * n_rows,
                            'Nomor Dok. Referensi': [''] * n_rows,  # Akan diisi manual dari SP2D
                            'Tanggal Dok. Referensi': [''] * n_rows,  # Akan diisi manual dari SP2D
                            'ID TKU Pemotong': df_merged['ID TKU'].fillna('').astype(str).str.strip().tolist(),
                            'Tanggal Pemotongan': [''] * n_rows  # Akan diisi manual dari Invoice
                        }
                        
                        # Buat DataFrame
                        hasil = pd.DataFrame(data_hasil)
                        
                        # Format Penghasilan tanpa desimal jika angka bulat
                        hasil['Penghasilan'] = hasil['Penghasilan'].apply(lambda x: int(x) if x == int(x) else x)
                        
                        # ========== TAMPILKAN HASIL ==========
                        st.success(f"✅ **Data berhasil diproses!** Total: {len(hasil)} baris")
                        
                        # Info summary
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.info(f"📅 **Masa Pajak:** {masa_pajak}")
                        with col2:
                            st.info(f"📅 **Tahun Pajak:** {tahun_pajak}")
                        with col3:
                            st.info(f"👥 **Total Pegawai:** {len(hasil)}")
                        
                        # Preview hasil dengan warna indikator
                        st.subheader("📊 Preview Hasil (Warna hanya ilustrasi)")
                        
                        # Buat DataFrame untuk preview dengan indikasi kolom
                        preview_df = hasil.head(10).copy()
                        
                        # Tampilkan dengan format
                        st.dataframe(preview_df.style.format({
                            'Penghasilan': 'Rp {:,.0f}',
                            'Tarif': '{:.2f}%'
                        }).apply(
                            lambda x: ['background-color: #e6ffe6' for _ in x], 
                            subset=pd.IndexSlice[:, :'Tarif']
                        ).apply(
                            lambda x: ['background-color: #fff0e6' for _ in x], 
                            subset=pd.IndexSlice[:, 'Jenis Dok. Referensi':'Tanggal Pemotongan']
                        ))
                        
                        # Legenda warna
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            <div style="background-color: #e6ffe6; padding: 10px; border-radius: 5px; border: 1px solid #ccc;">
                            <strong>🟢 Hijau Muda:</strong> Data hasil sistem (siap copy)
                            </div>
                            """, unsafe_allow_html=True)
                        with col2:
                            st.markdown("""
                            <div style="background-color: #fff0e6; padding: 10px; border-radius: 5px; border: 1px solid #ccc;">
                            <strong>🟠 Oranye:</strong> Isi manual dari DAFTAR SP2D SATKER
                            </div>
                            """, unsafe_allow_html=True)
                        
                        # Informasi pengisian manual
                        with st.expander("📝 PETUNJUK PENGISIAN MANUAL (Kolom Oranye)"):
                            st.markdown("""
                            ### **Ambil data dari DAFTAR SP2D SATKER:**
                            
                            1. **Nomor Dok. Referensi** → **Nomor SP2D**
                               - Contoh: `SP2D-001/BPK/2024`
                            
                            2. **Tanggal Dok. Referensi** → **Tanggal SP2D**
                               - Format: DD/MM/YYYY
                               - Contoh: `15/01/2024`
                            
                            3. **Tanggal Pemotongan** → **Tanggal Invoice**
                               - Format: DD/MM/YYYY
                               - Contoh: `10/01/2024`
                            
                            **Langkah:**
                            1. Download DAFTAR SP2D SATKER dari sistem
                            2. Cari SP2D yang sesuai dengan masa pajak **{}/{}**
                            3. Salin data sesuai petunjuk di atas
                            4. Isi ke semua baris dengan data yang sama
                            """.format(masa_pajak, tahun_pajak))
                        
                        # Statistik
                        st.subheader("📈 Statistik Hasil")
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            total_data = len(hasil)
                            st.metric("Total Data", total_data)
                        with col2:
                            total_penghasilan = hasil['Penghasilan'].sum()
                            st.metric("Total Penghasilan", f"Rp {total_penghasilan:,.0f}")
                        with col3:
                            total_tarif = hasil['Tarif'].sum()
                            st.metric("Total Tarif", f"{total_tarif:,.2f}%")
                        with col4:
                            avg_tarif = hasil['Tarif'].mean()
                            st.metric("Rata-rata Tarif", f"{avg_tarif:.2f}%")
                        
                        # Info perhitungan yang digunakan
                        st.info(f"""
                        **📊 Informasi Perhitungan:**
                        - **Sumber Tarif:** KODE OBJEK PAJAK dari Data Master
                        - **Mapping Tarif:**
                          21-402-02 → 5%
                          21-402-03 → 15%
                          21-402-04 → 0%
                        - **Rata-rata Tarif:** {hasil['Tarif'].mean():.2f}%
                        - **Rentang Tarif:** {hasil['Tarif'].min():.2f}% - {hasil['Tarif'].max():.2f}%
                        - **NPWP Clean:** Semua .0 telah dihilangkan dari NIK
                        """)
                        
                        # ========== SIMPAN KE SESSION STATE ==========
                        st.session_state.hasil_pajak_makan_pppk = hasil
                        st.session_state.masa_pajak_pppk_saved = masa_pajak
                        st.session_state.tahun_pajak_pppk_saved = tahun_pajak
                        st.session_state.mode_perhitungan = 'KODE_OBJEK_PAJAK'
                        
                        st.balloons()
                        st.success("🎉 **Data siap untuk didownload dengan format warna!**")
                        
                    except Exception as e:
                        st.error(f"❌ **ERROR**: Terjadi kesalahan saat memproses data")
                        st.error(f"Detail error: {str(e)}")
                        st.info("💡 **Tips**: Periksa kembali format data. Pastikan kolom KODE OBJEK PAJAK ada di Data Master.")
            
            # ========== TOMBOL DOWNLOAD DENGAN FORMAT WARNA ==========
            if 'hasil_pajak_makan_pppk' in st.session_state:
                st.markdown("---")
                st.subheader("💾 Download Hasil dengan Format Warna")
                
                hasil_final = st.session_state.hasil_pajak_makan_pppk
                masa_saved = st.session_state.get('masa_pajak_pppk_saved', masa_pajak)
                tahun_saved = st.session_state.get('tahun_pajak_pppk_saved', tahun_pajak)
                
                # Info file
                st.info(f"📄 **Nama file:** Hasil_Pajak_Makan_PPPK_{masa_saved:02d}_{tahun_saved}.xlsx")
                st.info(f"📊 **Jumlah data:** {len(hasil_final)} baris")
                st.info(f"🔢 **Sumber tarif:** KODE OBJEK PAJAK dari Data Master")
                
                # Buat file Excel dengan format warna menggunakan openpyxl
                output = io.BytesIO()
                
                # Buat workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Pajak Makan PPPK"
                
                # Tulis header
                headers = list(hasil_final.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                
                # Tulis data
                for row_num, row_data in enumerate(hasil_final.values, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        # Format NPWP sebagai teks (mencegah notasi ilmiah untuk NIK panjang)
                        if col_num == headers.index('NPWP') + 1:
                            ws.cell(row=row_num, column=col_num, value=str(cell_value))
                        else:
                            ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Tentukan kolom untuk warna
                green_columns = headers[:headers.index('Jenis Dok. Referensi') + 1]  # Sampai 'Jenis Dok. Referensi'
                orange_columns = ['Nomor Dok. Referensi', 'Tanggal Dok. Referensi', 'Tanggal Pemotongan']
                
                # Cari indeks kolom orange
                orange_indices = []
                for col_name in orange_columns:
                    if col_name in headers:
                        orange_indices.append(headers.index(col_name) + 1)  # +1 karena openpyxl mulai dari 1
                
                # Terapkan warna hijau untuk semua sel data (baris 2 ke atas)
                for row in ws.iter_rows(min_row=2, max_row=len(hasil_final) + 1, 
                                       min_col=1, max_col=len(headers)):
                    for cell in row:
                        # Jika kolom ini termasuk orange columns, beri warna orange
                        if cell.column in orange_indices:
                            cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                        else:
                            # Untuk kolom lainnya, beri warna hijau muda
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                
                # Format angka
                # Format Penghasilan tanpa desimal untuk angka bulat
                for row in ws.iter_rows(min_row=2, max_row=len(hasil_final) + 1):
                    # Kolom Penghasilan (indeks 7 jika mulai dari 0)
                    penghasilan_cell = row[7]
                    if penghasilan_cell.value:
                        if float(penghasilan_cell.value) == int(float(penghasilan_cell.value)):
                            penghasilan_cell.value = int(float(penghasilan_cell.value))
                        penghasilan_cell.number_format = '#,##0'
                
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            cell_value = str(cell.value) if cell.value is not None else ""
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Simpan workbook ke BytesIO
                wb.save(output)
                output.seek(0)
                
                # Tombol download
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="📥 **Download Hasil (Excel dengan Format Warna)**",
                        data=output,
                        file_name=f"Hasil_Pajak_Makan_PPPK_{masa_saved:02d}_{tahun_saved}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        help="File Excel dengan format warna: Hijau untuk data sistem, Oranye untuk data manual"
                    )
                
                # Informasi tambahan
                with st.expander("ℹ️ Informasi Format File yang Didownload"):
                    st.markdown("""
                    ### **🎨 Format Warna di File Excel:**
                    
                    **1. Header (Baris 1):**
                    - Warna: **Hijau Tua**
                    - Font: **Bold**
                    
                    **2. Data Hasil Sistem (Siap Copy):**
                    - Warna: **Hijau Muda** (#E6FFE6)
                    - Data ini sudah diproses oleh sistem
                    - Bisa langsung digunakan/dicopy
                    
                    **3. Data yang Harus Diisi Manual:**
                    - Warna: **Oranye** (#FFD580)
                    - Kolom yang perlu diisi manual:
                      1. **Nomor Dok. Referensi** → dari Nomor SP2D
                      2. **Tanggal Dok. Referensi** → dari Tanggal SP2D
                      3. **Tanggal Pemotongan** → dari Tanggal Invoice
                    
                    ### **📋 Sumber Data Manual:**
                    - **DAFTAR SP2D SATKER** (download dari sistem keuangan)
                    - Pastikan SP2D sesuai dengan masa pajak: **{}/{}**
                    """.format(masa_saved, tahun_saved))
        
        except Exception as e:
            st.error(f"❌ **ERROR**: Terjadi kesalahan saat membaca file")
            st.error(f"Detail: {str(e)}")
            st.info("💡 **Tips**: Pastikan file Excel tidak corrupt. Coba buka file di Excel terlebih dahulu.")
    
    elif uploaded_mentah is not None:
        st.warning("⚠️ **Langkah 2**: Silakan upload juga **Data Master** untuk melanjutkan")
    elif uploaded_master is not None:
        st.warning("⚠️ **Langkah 1**: Silakan upload **Data Mentah** terlebih dahulu")
    
    st.markdown("---")
    st.caption("""
    🔧 **Dukungan Teknis**: 
    - Pastikan KODE OBJEK PAJAK ada di Data Master dengan format yang benar
    - Format kode: 21-402-02, 21-402-03, atau 21-402-04
    - Untuk data manual (kolom oranye): Download DAFTAR SP2D SATKER dari sistem keuangan
    - **NIP tidak boleh duplikat** di kedua file
    - **Data baru** akan ditandai dan harus ditambahkan ke Data Master melalui Croscheck PPPK
    """)