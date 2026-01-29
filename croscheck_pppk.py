import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from io import BytesIO
import sys
import os
import zipfile

def show():
    """Fitur Sistem Master Data PPPK dengan Tracking Bulanan"""
   
    # ===== KONFIGURASI HALAMAN =====
    st.set_page_config(page_title="Sistem Master Data PPPK", layout="wide")
   
    # Tambahkan tombol kembali ke dashboard
    if st.button("← Kembali ke Dashboard"):
        st.session_state.current_page = 'dashboard_pppk'
        st.rerun()
   
    # Breadcrumb navigation
    st.markdown("**Beranda → Dashboard PPPK → Croscheck Data**")
    
    # ========== PANDUAN PENGGUNAAN ==========
    with st.expander("📚 **PANDUAN PENGGUNAAN UNTUK PEMULA**", expanded=False):
        st.markdown("""
        ## **🎯 TUJUAN SISTEM CROSCHECK DATA PPPK**
        
        Sistem ini dirancang untuk membantu Anda dalam:
        1. **Menggabungkan data** dari 3 sumber berbeda menjadi satu Master Data PPPK
        2. **Memvalidasi konsistensi** data antara sistem penggajian dan aplikasi pajak
        3. **Melacak perubahan** data pegawai PPPK dari bulan ke bulan
        4. **Mendeteksi masalah** seperti data duplikat, data tidak lengkap, atau ketidaksesuaian
        
        ## **📋 FILE YANG DIBUTUHKAN**
        
        ### **1. DATA MENTAH PPPK (WAJIB)**
        - **Sumber**: Sistem penggajian PPPK
        - **Fungsi**: Data utama pegawai (NIP, nama, NPWP, golongan, gaji)
        - **Format**: Excel (.xlsx) dengan kolom standar
        
        ### **2. DATA BPMP (WAJIB)**
        - **Sumber**: Aplikasi BPMP (Bukti Potong Pajak)
        - **Fungsi**: Data pajak untuk validasi
        - **Format**: Excel (.xlsx) dengan format tertentu
        
        ### **3. MASTER EXISTING (OPSIONAL)**
        - **Sumber**: Hasil download sistem ini bulan sebelumnya
        - **Fungsi**: Untuk tracking perubahan data
        - **Format**: Excel (.xlsx) hasil sistem ini
        
        ## **🚀 CARA PENGGUNAAN**
        
        **LANGKAH 1: Upload File**
        - Upload ketiga file di kolom yang tersedia
        - File wajib: Data Mentah PPPK dan Data BPMP
        - File opsional: Master Existing (untuk tracking)
        
        **LANGKAH 2: Proses Data**
        - Klik tombol **"🔄 Proses Data"**
        - Sistem akan validasi duplikasi dan konsistensi data
        - Tunggu hingga proses selesai (1-5 menit)
        
        **LANGKAH 3: Analisis Hasil**
        - Gunakan 5 tab untuk melihat hasil berbeda:
          1. **Hasil Master Data Baru** - Download hasil akhir
          2. **Perbandingan** - Lihat perubahan vs bulan lalu
          3. **Validasi Mentah vs BPMP** - Cek kesesuaian data
          4. **Validasi Mentah vs Master** - Cek konsistensi
          5. **Analisis** - Statistik perubahan
        
        ## **⚠️ PENTING: DETEKSI DUPLIKASI**
        
        Sistem akan otomatis mendeteksi dan menampilkan:
        - **NIP yang duplikat** (tidak diperbolehkan)
        - **NPWP yang duplikat** (harus unik)
        - **Data yang sama** di file berbeda
        
        **Harap perbaiki semua duplikasi sebelum melanjutkan!**
        
        ## **🎨 KETERANGAN WARNA**
        
        - **🟢 HIJAU**: Data baru (belum ada di master lama)
        - **🟡 KUNING**: Data sudah ada dan tidak berubah
        - **🟠 ORANGE**: Data ada tapi informasi berubah
        - **🔴 MERAH**: Pegawai tidak aktif (keluar/mutasi)
        
        ## **📥 DOWNLOAD HASIL**
        
        - Hasil akhir bisa di-download dalam format:
          1. **Excel dengan warna** (rekomendasi)
          2. **Excel tanpa warna**
          3. **CSV** (untuk analisis lanjutan)
        
        ## **💡 TIPS PENTING**
        
        1. **Format ID TKU Baru**:
           - ID PENERIMA TKU = NIK + "000000"
           - ID TKU = "0001658723701000000000" (sama untuk semua)
        
        2. **Validasi Sebelum Lanjut**:
           - Periksa semua peringatan duplikasi
           - Pastikan data BPMP match dengan Data Mentah
           - Cek data yang bermasalah di tab Validasi
        
        3. **Simpan untuk Bulan Depan**:
           - Download dan simpan Master Data baru
           - Gunakan sebagai Master Existing bulan berikutnya
        
        ## **❓ BANTUAN TAMBAHAN**
        
        - Jika ada error, periksa format file sesuai template
        - Pastikan kolom wajib ada di setiap file
        - Cek log error yang ditampilkan sistem
        - Hubungi admin jika masalah berlanjut
        """)
    
    # ========== INFORMASI PENTING ==========
    st.info("""
    **📋 INFORMASI PENTING SEBELUM MEMULAI:**
    
    **🎯 TUJUAN MASING-MASING TAB:**
    
    1. **📊 Hasil Master Data Baru PPPK**
       - **Tujuan**: Download hasil akhir Master Data PPPK bulan ini
       - **Fitur**: Tampilan dengan warna status, statistik data, format ID TKU baru
       - **Output**: File Excel/CSV siap pakai
       
    2. **📋 Perbandingan dengan Master Lama**
       - **Tujuan**: Melihat perubahan data vs bulan sebelumnya
       - **Fitur**: Identifikasi data baru, berubah, hilang, sama
       - **Output**: Detail perbedaan per kolom (abaikan kolom bank)
       
    3. **🔄 Validasi Data Mentah vs Data BPMP**
       - **Tujuan**: Memastikan kesesuaian data gaji dengan data pajak
       - **Fitur**: Validasi NIP, NPWP, bulan, tahun, gaji, status kawin
       - **Output**: Data yang perlu diperbaiki dan rekomendasi
       
    4. **⚖️ Validasi Data Mentah vs Data Master**
       - **Tujuan**: Memastikan data bulan ini konsisten dengan referensi master
       - **Fitur**: Validasi Primary Key (NIP), deteksi inkonsistensi
       - **Output**: Data bermasalah dan tindakan perbaikan
       
    5. **📈 Analisis Detail Perubahan**
       - **Tujuan**: Analisis statistik pergerakan pegawai PPPK
       - **Fitur**: Distribusi status, trend perubahan, monitoring
       - **Output**: Laporan analisis dan statistik
       
    **⚠️ PERHATIAN KHUSUS:**
    - **NIP adalah PRIMARY KEY** - tidak boleh duplicate
    - **NPWP harus unik** - tidak boleh duplicate
    - **Data harus match** antara Data Mentah dan BPMP
    - **Perbaikan duplikasi** wajib dilakukan sebelum proses
    - **Simpan Master Data** untuk tracking bulan depan
    """)
   
    # ===== FUNGSI UNTUK FITUR MASTER DATA PPPK =====
    
    # ===== FUNGSI FORMAT NILAI ASLI =====
    def format_nilai_asli(nilai):
        """Mengonversi nilai float yang merupakan integer menjadi string tanpa .0"""
        if nilai is None:
            return ''
        
        # Jika sudah string, kembalikan as is
        if isinstance(nilai, str):
            return nilai.strip()
        
        # Jika float
        if isinstance(nilai, float):
            # Cek apakah ini integer
            if nilai.is_integer():
                # Konversi ke int lalu ke string untuk menghilangkan .0
                return str(int(nilai))
            else:
                # Untuk float non-integer, kembalikan string tanpa trailing zeros
                return str(nilai).rstrip('0').rstrip('.')
        
        # Untuk tipe data lainnya, konversi ke string
        return str(nilai).strip()
    
    def format_angka_panjang(angka_str):
        """Menangani notasi ilmiah menjadi format angka biasa"""
        if not angka_str or pd.isna(angka_str):
            return ''
        
        str_angka = str(angka_str).strip()
        
        # Jika mengandung notasi ilmiah (e+)
        if 'e+' in str_angka.lower():
            try:
                # Konversi dari notasi ilmiah ke float
                num = float(str_angka)
                # Konversi ke int jika tidak ada desimal
                if num.is_integer():
                    return str(int(num))
                else:
                    # Format dengan string tanpa notasi ilmiah
                    return format(num, 'f').rstrip('0').rstrip('.')
            except:
                return str_angka
        
        # Jika mengandung .000000 di akhir
        if str_angka.endswith('.000000'):
            return str_angka.replace('.000000', '')
        
        # Jika mengandung .0 di akhir
        if str_angka.endswith('.0'):
            return str_angka.replace('.0', '')
        
        return str_angka
    # ===== END FUNGSI FORMAT NILAI ASLI =====
   
    # Definisi header untuk setiap file
    HEADERS_MENTAH_PPPK = [
        "kdsatker", "kdanak", "kdsubanak", "bulan", "tahun", "nogaji", "kdjns", "nip", "nmpeg",
        "kdduduk", "kdgol", "npwp", "nmrek", "nm_bank", "rekening", "kdbankspan", "nmbankspan",
        "kdpos", "kdnegara", "kdkppn", "tipesup", "gjpokok", "tjistri", "tjanak", "tjupns",
        "tjstruk", "tjfungs", "tjdaerah", "tjpencil", "tjlain", "tjkompen", "pembul", "tjberas",
        "tjpph", "potpfkbul", "potpfk2", "GajiKotor", "potpfk10", "potpph", "potswrum",
        "potkelbtj", "potlain", "pottabrum", "bersih", "sandi", "kdkawin", "kdjab",
        "thngj", "kdgapok", "bpjs", "bpjs2"
    ]
    
    # Header utama untuk pencarian
    HEADERS_BPMP = [
        "Masa Pajak", "Tahun Pajak", "Status Pegawai", "Posisi", "NPWP/NIK/TIN",
        "Nomor Passport", "Kode Objek Pajak", "Penghasilan Kotor", "Tarif", "ID TKU",
        "Tgl Pemotongan", "TER A", "TER B", "TER C"
    ]
   
    HEADERS_MASTER = [
        "No", "PNS/PPPK", "Nama", "NIK", "ID PENERIMA TKU", "KDGOL", "KODE OBJEK PAJAK",
        "KDKAWIN", "STATUS", "NIP", "nmrek", "nm_bank", "rekening", "kdbankspan",
        "nmbankspan", "kdpos", "ID TKU", "AKTIF/TIDAK", "Keterangan"
    ]
   
    # Mapping KDKAWIN ke STATUS
    KDKAWIN_MAP = {
        "1000": "TK/0", "1001": "TK/1", "1002": "TK/2" , "1100": "K/0",
        "1101": "K/1", "1102": "K/2"
    }
   
    def konversi_kode_objek(kdgol):
        """Konversi kdgol ke kode objek pajak"""
        if pd.isna(kdgol):
            return "-"
        kdgol = str(kdgol).strip()
        if kdgol.startswith("3"):
            return "21-402-02"
        elif kdgol.startswith("4"):
            return "21-402-03"
        elif kdgol.startswith("2"):
            return "21-402-04"
        elif kdgol.startswith("1"):
            return "21-402-04"
        else:
            return "-"
   
    def konversi_status(kdkawin):
        """Konversi kdkawin ke status"""
        if pd.isna(kdkawin):
            return "-"
        kdkawin = str(kdkawin).strip()
        return KDKAWIN_MAP.get(kdkawin, "-")
    
    def read_excel_flexible(uploaded_file, expected_headers, label):
        """Baca Excel dengan pencarian header fleksibel - REVISI: Menggunakan logika dari croscheck_pns.py"""
        if uploaded_file is None:
            return None
        
        try:
            uploaded_file.seek(0)
            wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
            sheet = wb.active
            
            total_rows = sheet.max_row
            total_columns = sheet.max_column
            expected_set = set(h.lower().strip() for h in expected_headers)
            
            st.write(f"📊 {label}: {total_rows} baris, {total_columns} kolom terdeteksi")
            
            # Cari baris header
            best_row = -1
            max_matches = 0
            best_row_values = []
            
            for row_idx in range(1, min(21, total_rows + 1)):
                row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[row_idx]]
                row_values_lower = [v.lower() for v in row_values]
                row_set = set(v for v in row_values_lower if v)
                matches = len(expected_set.intersection(row_set))
                if matches > max_matches:
                    max_matches = matches
                    best_row = row_idx
                    best_row_values = row_values
            
            if best_row > 0 and max_matches >= len(expected_headers) // 2:
                # Handle kolom kosong di kiri
                first_non_empty_col = next((i for i, val in enumerate(best_row_values) if val), 0)
                
                # Ekstrak data mulai dari baris setelah header
                data = []
                for row in sheet.iter_rows(min_row=best_row + 1, max_row=total_rows,
                                          min_col=first_non_empty_col + 1, values_only=True):
                    # Terapkan format_nilai_asli() pada setiap sel
                    formatted_row = []
                    for cell_value in row:
                        if isinstance(cell_value, (int, float)):
                            formatted_row.append(format_nilai_asli(cell_value))
                        else:
                            formatted_row.append(cell_value)
                    data.append(formatted_row)
                
                # Ambil header yang valid (mulai dari kolom pertama yang tidak kosong)
                actual_headers = best_row_values[first_non_empty_col:]
                
                # Handle kolom duplikat dan kosong
                seen = {}
                unique_headers = []
                for i, col in enumerate(actual_headers):
                    if not col or col.strip() == "":
                        col = f"Unnamed_{i}"
                    
                    # Handle duplikat
                    original_col = col
                    counter = 1
                    while col in seen:
                        col = f"{original_col}_{counter}"
                        counter += 1
                    
                    seen[col] = True
                    unique_headers.append(col)
                
                # Batasi jumlah kolom sesuai data
                if len(data) > 0:
                    max_data_cols = len(data[0])
                    unique_headers = unique_headers[:max_data_cols]
                
                df = pd.DataFrame(data, columns=unique_headers)
                
                # Drop kolom yang sepenuhnya kosong
                df = df.dropna(axis=1, how='all')
                
                # Drop baris yang sepenuhnya kosong
                df = df.dropna(how='all').reset_index(drop=True)
                
                st.success(f"✅ {label} berhasil dibaca! Header di baris {best_row}, {len(df)} baris data")
                
                # Bersihkan kolom-kolom penting
                for col in df.columns:
                    if isinstance(col, str):
                        col_lower = col.lower()
                        # Kolom NIP
                        if 'nip' in col_lower:
                            df[col] = df[col].apply(lambda x: format_angka_panjang(format_nilai_asli(x)))
                        # Kolom NPWP
                        elif 'npwp' in col_lower or 'nik' in col_lower or 'tin' in col_lower:
                            df[col] = df[col].apply(lambda x: format_angka_panjang(format_nilai_asli(x)))
                        # Kolom rekening
                        elif 'rekening' in col_lower:
                            df[col] = df[col].apply(lambda x: format_angka_panjang(format_nilai_asli(x)))
                
            else:
                # Fallback: baca seluruh sheet
                data = []
                for row in sheet.iter_rows(min_row=1, max_row=total_rows, values_only=True):
                    # Terapkan format_nilai_asli() pada setiap sel
                    formatted_row = []
                    for cell_value in row:
                        if isinstance(cell_value, (int, float)):
                            formatted_row.append(format_nilai_asli(cell_value))
                        else:
                            formatted_row.append(cell_value)
                    data.append(formatted_row)
                
                if len(data) < 2:
                    st.error(f"❌ {label}: File tidak memiliki cukup data")
                    return None
                
                # Handle header duplikat
                headers = data[0]
                seen = {}
                unique_headers = []
                for i, col in enumerate(headers):
                    col_str = str(col).strip() if col else f"Unnamed_{i}"
                    
                    original_col = col_str
                    counter = 1
                    while col_str in seen:
                        col_str = f"{original_col}_{counter}"
                        counter += 1
                    
                    seen[col_str] = True
                    unique_headers.append(col_str)
                
                df = pd.DataFrame(data[1:], columns=unique_headers)
                df = df.dropna(axis=1, how='all')
                df = df.dropna(how='all').reset_index(drop=True)
                
                # Bersihkan kolom-kolom penting
                for col in df.columns:
                    if isinstance(col, str):
                        col_lower = col.lower()
                        # Kolom NIP
                        if 'nip' in col_lower:
                            df[col] = df[col].apply(lambda x: format_angka_panjang(format_nilai_asli(x)))
                        # Kolom NPWP
                        elif 'npwp' in col_lower or 'nik' in col_lower or 'tin' in col_lower:
                            df[col] = df[col].apply(lambda x: format_angka_panjang(format_nilai_asli(x)))
                        # Kolom rekening
                        elif 'rekening' in col_lower:
                            df[col] = df[col].apply(lambda x: format_angka_panjang(format_nilai_asli(x)))
                
                st.warning(f"⚠️ {label}: Header tidak sepenuhnya cocok, menggunakan baris pertama")
            
            # Tampilkan kolom yang ditemukan
            st.info(f"Kolom ditemukan: {', '.join(df.columns.tolist()[:10])}{'...' if len(df.columns) > 10 else ''}")
            
            return df
        
        except Exception as e:
            st.error(f"❌ Error membaca {label}: {e}")
            import traceback
            st.code(traceback.format_exc())
            return None
    
    def detect_duplicates(df, column_name, label):
        """Deteksi data duplikat dalam kolom tertentu dan tampilkan dengan highlight merah"""
        if df is None or df.empty or column_name not in df.columns:
            return False
        
        # Format nilai untuk deteksi duplikasi
        df[column_name] = df[column_name].apply(lambda x: format_nilai_asli(x))
        
        # Cek duplikasi (termasuk baris pertama)
        duplicates = df[df[column_name].duplicated(keep=False)]
        
        if not duplicates.empty:
            # Hitung jumlah baris duplikat
            duplicate_count = len(duplicates)
            unique_duplicate_values = duplicates[column_name].unique()
            
            # Tampilkan warning
            st.error(f"⚠️ **DETEKSI DUPLIKASI DI {label}**")
            st.error(f"Ditemukan **{duplicate_count} baris** dengan **{column_name} yang sama/duplikat**")
            st.error(f"Nilai {column_name} yang duplikat: {', '.join(map(str, unique_duplicate_values[:10]))}{'...' if len(unique_duplicate_values) > 10 else ''}")
            
            # Tampilkan data duplikat dengan highlight merah
            st.markdown(f"#### 📋 Daftar Data Duplikat ({column_name}):")
            
            # Fungsi untuk highlight baris duplikat
            def highlight_duplicates_row(row):
                is_duplicate = row[column_name] in unique_duplicate_values
                return ['background-color: #FF6B6B' if is_duplicate else '' for _ in row]
            
            # Tampilkan semua baris (tanpa batas)
            st.dataframe(
                df.style.apply(highlight_duplicates_row, axis=1),
                height=400,
                use_container_width=True
            )
            
            # Download button untuk data duplikat
            output_duplicates = BytesIO()
            with pd.ExcelWriter(output_duplicates, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data dengan Duplikat')
                
                workbook = writer.book
                worksheet = writer.sheets['Data dengan Duplikat']
                
                red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                
                # Highlight baris duplikat di Excel
                for idx, row in df.iterrows():
                    if row[column_name] in unique_duplicate_values:
                        excel_row = idx + 2  # +2 karena header di row 1
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=excel_row, column=col).fill = red_fill
            
            output_duplicates.seek(0)
            
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label=f"📥 Download Data {label} dengan Duplikat",
                    data=output_duplicates,
                    file_name=f"duplikat_{column_name}_{label}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_duplicates_{column_name}_{label}"
                )
            with col2:
                if st.button(f"⚠️ PERBAIKI DUPLIKASI {column_name.upper()} TERLEBIH DAHULU", key=f"warning_{column_name}_{label}"):
                    st.warning("Harap perbaiki data duplikat sebelum melanjutkan proses!")
            
            return True
        return False
   
    def fuzzy_match_row(nama, nip, df_master, threshold=80):
        """Cari baris yang cocok menggunakan fuzzy matching"""
        if df_master is None or df_master.empty:
            return None
       
        best_match_idx = None
        best_score = 0
       
        # Cari kolom Nama dan NIP dengan pencarian fleksibel
        nama_col = None
        nip_col = None
       
        for col in df_master.columns:
            col_upper = str(col).upper()
            if 'NAMA' in col_upper and nama_col is None:
                nama_col = col
            if 'NIP' in col.upper() and nip_col is None:
                nip_col = col
       
        if not nama_col and not nip_col:
            return None
       
        for idx, row in df_master.iterrows():
            # Match berdasarkan Nama dan NIP
            nama_master = str(row.get(nama_col, '')).strip() if nama_col else ''
            nip_master = str(row.get(nip_col, '')).strip() if nip_col else ''
           
            nama_score = fuzz.ratio(str(nama).lower(), nama_master.lower()) if nama and nama_master else 0
            nip_score = fuzz.ratio(str(nip).lower(), nip_master.lower()) if nip and nip_master else 0
           
            # Gabungan score (prioritas NIP lebih tinggi)
            if nip_col and nama_col:
                combined_score = (nip_score * 0.7) + (nama_score * 0.3)
            elif nip_col:
                combined_score = nip_score
            elif nama_col:
                combined_score = nama_score
            else:
                combined_score = 0
           
            if combined_score > best_score and combined_score >= threshold:
                best_score = combined_score
                best_match_idx = idx
       
        return best_match_idx
   
    def process_data(df_mentah, df_bpmp, df_master_existing=None):
        """Proses data dari file mentah dan BPMP ke format master"""
       
        if df_mentah is None or df_bpmp is None:
            st.warning("⚠️ Pastikan file Data Mentah dan BPMP sudah di-upload!")
            return None
       
        # ===== DETEKSI DUPLIKASI SEBELUM PROSES =====
        st.markdown("### 🔍 Deteksi Duplikasi Data")
        
        has_duplicates = False
        
        # Deteksi duplikasi NIP di Data Mentah
        if 'nip' in df_mentah.columns:
            if detect_duplicates(df_mentah, 'nip', 'Data Mentah PPPK (NIP)'):
                has_duplicates = True
        
        # Deteksi duplikasi NPWP di Data Mentah
        if 'npwp' in df_mentah.columns:
            if detect_duplicates(df_mentah, 'npwp', 'Data Mentah PPPK (NPWP)'):
                has_duplicates = True
        
        # Deteksi duplikasi di Data BPMP
        # Cari kolom NPWP/NIK/TIN di BPMP
        nik_col_bpmp = None
        for col in df_bpmp.columns:
            if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                nik_col_bpmp = col
                break
        
        if nik_col_bpmp:
            if detect_duplicates(df_bpmp, nik_col_bpmp, 'Data BPMP (NPWP/NIK/TIN)'):
                has_duplicates = True
        
        # Deteksi duplikasi di Master Existing
        if df_master_existing is not None and not df_master_existing.empty:
            # Cari kolom NIP di Master
            nip_col_master = None
            nik_col_master = None
            
            for col in df_master_existing.columns:
                col_upper = str(col).upper()
                if 'NIP' in col_upper:
                    nip_col_master = col
                if 'NIK' in col_upper:
                    nik_col_master = col
            
            if nip_col_master:
                if detect_duplicates(df_master_existing, nip_col_master, 'Master Existing (NIP)'):
                    has_duplicates = True
            
            if nik_col_master:
                if detect_duplicates(df_master_existing, nik_col_master, 'Master Existing (NIK)'):
                    has_duplicates = True
        
        # Jika ada duplikasi, hentikan proses
        if has_duplicates:
            st.error("⛔ **PROSES DIBATALKAN** karena ditemukan data duplikat!")
            st.error("Harap perbaiki data duplikat terlebih dahulu sebelum melanjutkan proses.")
            return None
        
        st.success("✅ Tidak ditemukan data duplikat. Melanjutkan proses...")
        # ===== END DETEKSI DUPLIKASI =====
       
        # Normalisasi kolom - handle duplikat
        df_mentah.columns = [str(col).strip() for col in df_mentah.columns]
        df_bpmp.columns = [str(col).strip() for col in df_bpmp.columns]
       
        # Debug: tampilkan kolom yang tersedia
        st.write("**Kolom Data Mentah PPPK:**", df_mentah.columns.tolist()[:15])
        st.write("**Kolom Data BPMP:**", df_bpmp.columns.tolist())
       
        # Buat DataFrame hasil dengan merge berdasarkan fuzzy matching
        hasil = []
       
        for idx_mentah, row_mentah in df_mentah.iterrows():
            # Ambil data dari file mentah dengan penanganan error - gunakan format_nilai_asli()
            nip = format_nilai_asli(row_mentah.get('nip', '')) if 'nip' in df_mentah.columns else ''
            nama = format_nilai_asli(row_mentah.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else ''
            npwp_mentah = format_nilai_asli(row_mentah.get('npwp', '')) if 'npwp' in df_mentah.columns else ''
           
            # Cari data BPMP yang cocok menggunakan fuzzy matching
            matched_bpmp = None
            best_match_score = 0
           
            for idx_bpmp, row_bpmp in df_bpmp.iterrows():
                # Cek apakah kolom NPWP/NIK/TIN ada
                nik_col = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col = col
                        break
               
                if nik_col:
                    nik_bpmp = format_nilai_asli(row_bpmp.get(nik_col, ''))
                   
                    # Matching berdasarkan NPWP/NIK
                    if npwp_mentah and nik_bpmp and npwp_mentah != '' and nik_bpmp != '':
                        score = fuzz.ratio(npwp_mentah, nik_bpmp)
                        if score > best_match_score and score >= 80:
                            best_match_score = score
                            matched_bpmp = row_bpmp
           
            # Jika tidak ada match berdasarkan NPWP, coba match berdasarkan urutan baris
            if matched_bpmp is None and idx_mentah < len(df_bpmp):
                matched_bpmp = df_bpmp.iloc[idx_mentah]
           
            # Ambil data dari matched BPMP
            posisi = ''
            nik_bpmp = npwp_mentah # default ke NPWP dari mentah
           
            if matched_bpmp is not None:
                # Cari kolom Posisi
                posisi_col = None
                for col in df_bpmp.columns:
                    if 'POSISI' in col.upper():
                        posisi_col = col
                        break
                if posisi_col:
                    posisi = format_nilai_asli(matched_bpmp.get(posisi_col, ''))
               
                # Cari kolom NIK
                nik_col = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col = col
                        break
                if nik_col:
                    nik_from_bpmp = format_nilai_asli(matched_bpmp.get(nik_col, ''))
                    if nik_from_bpmp:
                        nik_bpmp = nik_from_bpmp
           
            # Ambil data lainnya dari mentah
            kdgol = format_nilai_asli(row_mentah.get('kdgol', '')) if 'kdgol' in df_mentah.columns else ''
            kdkawin = format_nilai_asli(row_mentah.get('kdkawin', '')) if 'kdkawin' in df_mentah.columns else ''
           
            # Filter PNS/PPPK
            pns_pppk = 'PPPK' # Default untuk PPPK
           
            # ===== PERUBAHAN: FORMAT ID PENERIMA TKU DAN ID TKU =====
            # Format ID PENERIMA TKU: ambil NIK lalu tambahkan "000000" di akhir
            id_penerima_tku = f"{nik_bpmp}000000" if nik_bpmp and nik_bpmp.strip() != '' else ''
           
            # Format ID TKU: nilai default "0001658723701000000000" untuk semua baris
            id_tku = "0001658723701000000000"
            # ===== END PERUBAHAN =====
           
            row_data = {
                'No': idx_mentah + 1,
                'PNS/PPPK': pns_pppk,
                'Nama': nama,
                'NIK': nik_bpmp,
                'ID PENERIMA TKU': id_penerima_tku, # Menggunakan format baru
                'KDGOL': kdgol,
                'KODE OBJEK PAJAK': konversi_kode_objek(kdgol),
                'KDKAWIN': kdkawin,
                'STATUS': konversi_status(kdkawin),
                'NIP': nip,
                'nmrek': format_nilai_asli(row_mentah.get('nmrek', '')),
                'nm_bank': format_nilai_asli(row_mentah.get('nm_bank', '')),
                'rekening': format_nilai_asli(row_mentah.get('rekening', '')),
                'kdbankspan': format_nilai_asli(row_mentah.get('kdbankspan', '')),
                'nmbankspan': format_nilai_asli(row_mentah.get('nmbankspan', '')),
                'kdpos': format_nilai_asli(row_mentah.get('kdpos', '')),
                'ID TKU': id_tku, # Menggunakan nilai default
                'AKTIF/TIDAK': 'AKTIF',
                'Keterangan': ''
            }
           
            hasil.append(row_data)
       
        df_hasil = pd.DataFrame(hasil)
       
        # Merge dengan master existing jika ada
        if df_master_existing is not None and not df_master_existing.empty:
            # Normalisasi kolom master existing
            df_master_existing.columns = [str(col).strip() for col in df_master_existing.columns]
           
            st.write("**Kolom Master Existing:**", df_master_existing.columns.tolist())
           
            # Tandai data yang sudah ada dengan membandingkan kolom kunci (tidak termasuk kolom bank)
            for idx, row in df_hasil.iterrows():
                match_idx = fuzzy_match_row(row['Nama'], row['NIP'], df_master_existing)
               
                if match_idx is not None:
                    # Cek perbedaan hanya pada kolom kunci, abaikan kolom bank
                    row_master = df_master_existing.iloc[match_idx]
                   
                    # Kolom kunci yang dibandingkan
                    key_columns = ['Nama', 'NIP', 'NIK', 'KDGOL', 'KDKAWIN', 'STATUS', 'KODE OBJEK PAJak', 'PNS/PPPK']
                   
                    is_different = False
                    for col in key_columns:
                        # Cari kolom di master existing
                        master_col = None
                        for c in df_master_existing.columns:
                            if col.upper() in str(c).upper():
                                master_col = c
                                break
                       
                        if master_col:
                            val_master = format_nilai_asli(row_master.get(master_col, ''))
                            val_new = format_nilai_asli(row.get(col, ''))
                           
                            if val_master != val_new:
                                is_different = True
                                break
                   
                    if is_different:
                        # Data ada tapi berbeda pada kolom kunci - ORANGE
                        df_hasil.at[idx, 'Status_Color'] = 'ORANGE'
                        df_hasil.at[idx, 'Keterangan'] = f'Data berubah (kolom kunci berbeda)'
                    else:
                        # Data sama pada kolom kunci - KUNING
                        df_hasil.at[idx, 'Status_Color'] = 'KUNING'
                       
                        # Update keterangan dari master jika ada
                        ket_col = None
                        for col in df_master_existing.columns:
                            if 'KETERANGAN' in str(col).upper():
                                ket_col = col
                                break
                       
                        if ket_col:
                            existing_ket = format_nilai_asli(df_master_existing.at[match_idx, ket_col])
                            if pd.notna(existing_ket) and existing_ket:
                                df_hasil.at[idx, 'Keterangan'] = existing_ket
                else:
                    # Data baru
                    df_hasil.at[idx, 'Status_Color'] = 'HIJAU'
           
            # Cek data lama yang tidak ada di bulan ini (MERAH)
            for idx, row_old in df_master_existing.iterrows():
                # Cari kolom Nama dan NIP di master existing
                nama_col = None
                nip_col = None
               
                for col in df_master_existing.columns:
                    col_upper = str(col).upper()
                    if 'NAMA' in col_upper and nama_col is None:
                        nama_col = col
                    if 'NIP' in col.upper() and nip_col is None:
                        nip_col = col
               
                if not nama_col or not nip_col:
                    continue
               
                nama_old = format_nilai_asli(row_old.get(nama_col, ''))
                nip_old = format_nilai_asli(row_old.get(nip_col, ''))
               
                match_idx = fuzzy_match_row(nama_old, nip_old, df_hasil)
               
                if match_idx is None:
                    # Data tidak ada di bulan ini - tambahkan dengan status TIDAK AKTIF
                    row_old_dict = row_old.to_dict()
                   
                    # ===== PERUBAHAN: UNTUK DATA LAMA YANG TIDAK AKTIF =====
                    # Format ID PENERIMA TKU dari NIK master lama
                    nik_master = format_nilai_asli(row_old_dict.get('NIK', ''))
                    id_penerima_tku_old = f"{nik_master}000000" if nik_master and nik_master.strip() != '' else ''
                   
                    # ID TKU tetap menggunakan nilai default
                    id_tku_old = "0001658723701000000000"
                    # ===== END PERUBAHAN =====
                   
                    # Map kolom dari master existing ke format hasil
                    new_row = {
                        'No': len(df_hasil) + 1,
                        'PNS/PPPK': format_nilai_asli(row_old_dict.get('PNS/PPPK', '')),
                        'Nama': nama_old,
                        'NIK': format_nilai_asli(row_old_dict.get('NIK', '')),
                        'ID PENERIMA TKU': id_penerima_tku_old, # Menggunakan format baru
                        'KDGOL': format_nilai_asli(row_old_dict.get('KDGOL', '')),
                        'KODE OBJEK PAJAK': format_nilai_asli(row_old_dict.get('KODE OBJEK PAJAK', '')),
                        'KDKAWIN': format_nilai_asli(row_old_dict.get('KDKAWIN', '')),
                        'STATUS': format_nilai_asli(row_old_dict.get('STATUS', '')),
                        'NIP': nip_old,
                        'nmrek': format_nilai_asli(row_old_dict.get('nmrek', '')),
                        'nm_bank': format_nilai_asli(row_old_dict.get('nm_bank', '')),
                        'rekening': format_nilai_asli(row_old_dict.get('rekening', '')),
                        'kdbankspan': format_nilai_asli(row_old_dict.get('kdbankspan', '')),
                        'nmbankspan': format_nilai_asli(row_old_dict.get('nmbankspan', '')),
                        'kdpos': format_nilai_asli(row_old_dict.get('kdpos', '')),
                        'ID TKU': id_tku_old, # Menggunakan nilai default
                        'AKTIF/TIDAK': 'TIDAK',
                        'Keterangan': format_nilai_asli(row_old_dict.get('Keterangan', '')),
                        'Status_Color': 'MERAH'
                    }
                   
                    df_hasil = pd.concat([df_hasil, pd.DataFrame([new_row])], ignore_index=True)
        else:
            # Semua data baru
            df_hasil['Status_Color'] = 'HIJAU'
       
        return df_hasil
   
    def create_excel_with_colors(df):
        """Buat Excel dengan warna berdasarkan status"""
        output = BytesIO()
       
        # Hapus kolom helper
        df_export = df.drop(columns=['Status_Color'], errors='ignore')
       
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Format kolom numerik
            for col in df_export.columns:
                if col in ['NIP', 'NIK', 'ID PENERIMA TKU', 'ID TKU', 'rekening']:
                    # Set format teks untuk kolom angka panjang
                    df_export[col] = df_export[col].apply(lambda x: format_nilai_asli(x))
            
            df_export.to_excel(writer, index=False, sheet_name='Master Data')
           
            workbook = writer.book
            worksheet = writer.sheets['Master Data']
           
            # Define fills
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
           
            # Apply colors
            for idx, row in df.iterrows():
                excel_row = idx + 2 # +2 karena header di row 1
                color = row.get('Status_Color', '')
               
                if color == 'KUNING':
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = yellow_fill
                elif color == 'MERAH':
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = red_fill
                elif color == 'ORANGE':
                    # Orange untuk nama yang ada tapi data berbeda
                    worksheet.cell(row=excel_row, column=3).fill = orange_fill # Kolom Nama
                elif color == 'HIJAU':
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = green_fill
       
        output.seek(0)
        return output
   
    # ===== UI UNTUK FITUR MASTER DATA =====
    st.title("🔍 CROSCHECK DATA GAJI PPPK")
    st.markdown("---")
    
    # ===== PROSES 3-STEP =====
    st.info("""
    **📋 PROSES 3-STEP CROSCHECK DATA PPPK:**
    
    1. **📥 UPLOAD FILE** - Upload 3 file yang diperlukan
    2. **🔄 PROSES DATA** - Sistem akan matching dan validasi data
    3. **📊 ANALISIS HASIL** - Lihat hasil di 5 tab berbeda

    **✅ Hasil**: Master Data PPPK baru + laporan validasi lengkap
    """)
   
    col1, col2, col3 = st.columns(3)
    with col1:
        st.subheader("📁 File 1: Data Mentah PPPK")
        uploaded_mentah = st.file_uploader("Upload Data Mentah PPPK", type=["xlsx"], key="mentah_pppk")
        st.caption("File hasil download sistem penggajian PPPK")
        
        if uploaded_mentah:
            st.success(f"✅ {uploaded_mentah.name}")
            st.caption("Klik 'Proses Data' setelah semua file diupload")
    
    with col2:
        st.subheader("📁 File 2: Data BPMP")
        uploaded_bpmp = st.file_uploader("Upload Data BPMP", type=["xlsx"], key="bpmp_pppk")
        st.caption("File dari aplikasi BPMP")
        
        if uploaded_bpmp:
            st.success(f"✅ {uploaded_bpmp.name}")
            st.caption("Pastikan format sesuai template")
    
    with col3:
        st.subheader("📁 File 3: Master Existing")
        uploaded_master = st.file_uploader("Upload Master Lama", type=["xlsx"], key="master_pppk")
        st.caption("Opsional - untuk tracking perubahan")
        
        if uploaded_master:
            st.success(f"✅ {uploaded_master.name}")
            st.caption("Hasil download dari bulan sebelumnya")
        else:
            st.info("ℹ️ Opsional - untuk tracking perubahan")
   
    st.markdown("---")
    
    # Status upload file
    file_status_col1, file_status_col2, file_status_col3 = st.columns(3)
    
    with file_status_col1:
        if uploaded_mentah:
            st.success("✅ Data Mentah PPPK: READY")
        else:
            st.warning("⚠️ Data Mentah PPPK: BELUM DIUPLOAD")
    
    with file_status_col2:
        if uploaded_bpmp:
            st.success("✅ Data BPMP: READY")
        else:
            st.warning("⚠️ Data BPMP: BELUM DIUPLOAD")
    
    with file_status_col3:
        if uploaded_master:
            st.success("✅ Master Existing: READY")
        else:
            st.info("ℹ️ Master Existing: OPSIONAL")
    
    st.markdown("---")
   
    # Baca file dengan metode yang diperbaiki
    df_mentah = read_excel_flexible(uploaded_mentah, HEADERS_MENTAH_PPPK, "Data Mentah PPPK")
    df_bpmp = read_excel_flexible(uploaded_bpmp, HEADERS_BPMP, "Data BPMP")
    df_master_existing = read_excel_flexible(uploaded_master, HEADERS_MASTER, "Master Existing")
    
    # ===== VALIDASI DUPLIKASI DATA =====
    st.subheader("🔍 VALIDASI DUPLIKASI DATA")
    
    duplicate_found = False
    duplicate_info = {}
    
    if df_mentah is not None:
        # Cek NIP duplicate di Data Mentah
        if 'nip' in df_mentah.columns:
            df_mentah['nip_formatted'] = df_mentah['nip'].apply(lambda x: format_nilai_asli(x))
            nip_duplicates = df_mentah[df_mentah['nip_formatted'].duplicated(keep=False)]
            if not nip_duplicates.empty:
                duplicate_found = True
                duplicate_info['NIP Data Mentah'] = {
                    'data': nip_duplicates,
                    'column': 'nip',
                    'count': len(nip_duplicates)
                }
        
        # Cek NPWP duplicate di Data Mentah
        if 'npwp' in df_mentah.columns:
            df_mentah['npwp_formatted'] = df_mentah['npwp'].apply(lambda x: format_nilai_asli(x))
            npwp_duplicates = df_mentah[df_mentah['npwp_formatted'].duplicated(keep=False)]
            if not npwp_duplicates.empty:
                duplicate_found = True
                duplicate_info['NPWP Data Mentah'] = {
                    'data': npwp_duplicates,
                    'column': 'npwp',
                    'count': len(npwp_duplicates)
                }
    
    if df_bpmp is not None:
        # Cek NPWP/NIK/TIN duplicate di Data BPMP
        nik_col_bpmp = None
        for col in df_bpmp.columns:
            if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                nik_col_bpmp = col
                break
        
        if nik_col_bpmp:
            df_bpmp[f'{nik_col_bpmp}_formatted'] = df_bpmp[nik_col_bpmp].apply(lambda x: format_nilai_asli(x))
            nik_bpmp_duplicates = df_bpmp[df_bpmp[f'{nik_col_bpmp}_formatted'].duplicated(keep=False)]
            if not nik_bpmp_duplicates.empty:
                duplicate_found = True
                duplicate_info['NPWP/NIK Data BPMP'] = {
                    'data': nik_bpmp_duplicates,
                    'column': nik_col_bpmp,
                    'count': len(nik_bpmp_duplicates)
                }
    
    if df_master_existing is not None:
        # Cek NIP duplicate di Data Master
        nip_col_master = None
        for col in df_master_existing.columns:
            if 'NIP' in str(col).upper():
                nip_col_master = col
                break
        
        if nip_col_master:
            df_master_existing[f'{nip_col_master}_formatted'] = df_master_existing[nip_col_master].apply(lambda x: format_nilai_asli(x))
            nip_master_duplicates = df_master_existing[df_master_existing[f'{nip_col_master}_formatted'].duplicated(keep=False)]
            if not nip_master_duplicates.empty:
                duplicate_found = True
                duplicate_info['NIP Master Existing'] = {
                    'data': nip_master_duplicates,
                    'column': nip_col_master,
                    'count': len(nip_master_duplicates)
                }
        
        # Cek NIK duplicate di Data Master
        nik_col_master = None
        for col in df_master_existing.columns:
            if 'NIK' in str(col).upper():
                nik_col_master = col
                break
        
        if nik_col_master:
            df_master_existing[f'{nik_col_master}_formatted'] = df_master_existing[nik_col_master].apply(lambda x: format_nilai_asli(x))
            nik_master_duplicates = df_master_existing[df_master_existing[f'{nik_col_master}_formatted'].duplicated(keep=False)]
            if not nik_master_duplicates.empty:
                duplicate_found = True
                duplicate_info['NIK Master Existing'] = {
                    'data': nik_master_duplicates,
                    'column': nik_col_master,
                    'count': len(nik_master_duplicates)
                }
    
    # Tampilkan hasil validasi duplikasi
    if duplicate_found:
        st.error("❌ **DITEMUKAN DUPLIKASI DATA!**")
        
        for key, info in duplicate_info.items():
            with st.expander(f"⚠️ {key} - {info['count']} baris duplicate", expanded=True):
                st.write(f"**Kolom:** {info['column']}")
                st.write(f"**Jumlah baris duplicate:** {info['count']}")
                
                # Tampilkan data duplicate
                st.dataframe(info['data'][[info['column']]].drop_duplicates())
                
                # Download button untuk data duplicate
                output_duplicate = BytesIO()
                info['data'].to_excel(output_duplicate, index=False)
                output_duplicate.seek(0)
                
                st.download_button(
                    label=f"📥 Download Data Duplicate {key}",
                    data=output_duplicate,
                    file_name=f"duplicate_{key.lower().replace(' ', '_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_duplicate_{key}"
                )
        
        st.warning("""
        **⚠️ PERINGATAN:**
        - **NIP** tidak boleh duplicate di Data Mentah maupun Master Existing
        - **NPWP/NIK** tidak boleh duplicate di Data Mentah, BPMP, maupun Master Existing
        - Silakan perbaiki data terlebih dahulu sebelum melanjutkan proses
        """)
        
        # Opsi untuk melanjutkan meskipun ada duplikasi
        continue_with_duplicates = st.checkbox("📌 Lanjutkan proses meskipun ada duplikasi?", value=False)
        if not continue_with_duplicates:
            st.stop()
    else:
        st.success("✅ **TIDAK ADA DUPLIKASI DATA** - Semua data unik dan valid")
    
    st.markdown("---")
    
    # Tampilkan preview data yang di-upload
    preview_col1, preview_col2, preview_col3 = st.columns(3)
    
    with preview_col1:
        if df_mentah is not None:
            with st.expander("👀 Preview Data Mentah PPPK"):
                st.dataframe(df_mentah.head(3))
                st.caption(f"Total: {len(df_mentah)} baris")
    
    with preview_col2:
        if df_bpmp is not None:
            with st.expander("👀 Preview Data BPMP"):
                st.dataframe(df_bpmp.head(3))
                st.caption(f"Total: {len(df_bpmp)} baris")
    
    with preview_col3:
        if df_master_existing is not None:
            with st.expander("👀 Preview Master Existing"):
                st.dataframe(df_master_existing.head(3))
                st.caption(f"Total: {len(df_master_existing)} baris")
   
    # Proses data
    st.markdown("---")
    
    if st.button("🔄 **PROSES DATA**", type="primary", use_container_width=True):
        if df_mentah is None or df_bpmp is None:
            st.error("❌ File Data Mentah PPPK dan Data BPMP wajib diupload!")
        else:
            with st.spinner("Memproses data..."):
                df_hasil = process_data(df_mentah, df_bpmp, df_master_existing)
               
                if df_hasil is not None:
                    st.session_state['df_hasil'] = df_hasil
                    st.session_state['df_master_existing'] = df_master_existing
                    st.session_state['df_mentah'] = df_mentah
                    st.session_state['df_bpmp'] = df_bpmp
                    st.success("✅ Data berhasil diproses!")
   
    # Tampilkan hasil
    if 'df_hasil' in st.session_state:
        st.markdown("---")
        
        # Tab informasi
        st.info("""
        **📊 HASIL PROSES TELAH SIAP!**
        
        **🎯 Gunakan 5 tab di bawah untuk:**
        1. **Master Data Baru** - Download hasil akhir
        2. **Perbandingan** - Lihat perubahan vs master lama
        3. **Validasi Mentah vs BPMP** - Cek kesesuaian data
        4. **Validasi Mentah vs Master** - Cek konsistensi
        5. **Analisis** - Statistik dan trend perubahan
        """)
        
        # Tabs untuk membandingkan
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 Hasil Master Data Baru",
            "📋 Perbandingan dengan Master Lama",
            "🔄 Validasi Data Mentah vs BPMP",
            "⚖️ Validasi Data Mentah vs Master",
            "📈 Analisis Perubahan"
        ])
       
        with tab1:
            st.subheader("📋 Hasil Master Data Baru PPPK")
           
            df_display = st.session_state['df_hasil'].copy()
           
            # Fungsi untuk styling
            def highlight_rows(row):
                color = row.get('Status_Color', '')
                if color == 'KUNING':
                    return ['background-color: #FFFF00'] * len(row)
                elif color == 'MERAH':
                    return ['background-color: #FF6B6B'] * len(row)
                elif color == 'ORANGE':
                    return ['background-color: #FFA500' if i == 2 else '' for i in range(len(row))]
                elif color == 'HIJAU':
                    return ['background-color: #90EE90'] * len(row)
                return [''] * len(row)
           
            # Tampilkan dengan styling
            df_show = df_display.drop(columns=['Status_Color'], errors='ignore')
            st.dataframe(df_show.style.apply(highlight_rows, axis=1), height=400)
           
            # Legend
            st.markdown("""
            **🎨 Keterangan Warna:**
            - 🟢 **Hijau**: Data baru (belum ada di master lama)
            - 🟡 **Kuning**: Data sudah ada di master lama (tidak berubah)
            - 🟠 **Orange**: Nama ada tapi data berubah
            - 🔴 **Merah**: Pegawai tidak aktif (ada di master lama tapi tidak ada di bulan ini)
            """)
           
            # Download button khusus untuk tab 1
            st.markdown("---")
            st.subheader("📥 Download Data Master Baru PPPK")
           
            # Pilihan format download
            download_format = st.radio(
                "Pilih format download:",
                ["Excel dengan warna", "Excel tanpa warna", "CSV"],
                horizontal=True,
                key="tab1_download_pppk"
            )
           
            if download_format == "Excel dengan warna":
                excel_file = create_excel_with_colors(df_display)
                file_name = "master_data_pppk.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif download_format == "Excel tanpa warna":
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Format kolom numerik
                    df_export = df_show.copy()
                    for col in df_export.columns:
                        if col in ['NIP', 'NIK', 'ID PENERIMA TKU', 'ID TKU', 'rekening']:
                            df_export[col] = df_export[col].apply(lambda x: format_nilai_asli(x))
                    df_export.to_excel(writer, index=False, sheet_name='Master Data')
                output.seek(0)
                excel_file = output
                file_name = "master_data_pppk_tanpa_warna.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else: # CSV
                csv_data = df_show.to_csv(index=False)
                excel_file = BytesIO(csv_data.encode())
                file_name = "master_data_pppk.csv"
                mime_type = "text/csv"
           
            st.download_button(
                label=f"📥 Download {download_format}",
                data=excel_file,
                file_name=file_name,
                mime=mime_type,
                key="download_master_baru_pppk"
            )
           
            # Statistik
            st.markdown("---")
            st.subheader("📊 Statistik Master Data Baru PPPK")
           
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
           
            with col_stat1:
                total_aktif = len(df_display[df_display['AKTIF/TIDAK'] == 'AKTIF'])
                st.metric("👥 Total Aktif", total_aktif)
           
            with col_stat2:
                total_tidak_aktif = len(df_display[df_display['AKTIF/TIDAK'] == 'TIDAK'])
                st.metric("❌ Tidak Aktif", total_tidak_aktif)
           
            with col_stat3:
                total_baru = len(df_display[df_display['Status_Color'] == 'HIJAU'])
                st.metric("🆕 Data Baru", total_baru)
           
            with col_stat4:
                total_duplikasi = len(df_display[df_display['Status_Color'] == 'KUNING'])
                st.metric("⚠️ Data Tidak Berubah", total_duplikasi)
       
        with tab2:
            st.subheader("🔍 Perbandingan Master Lama vs Master Baru")
           
            if 'df_master_existing' in st.session_state and st.session_state['df_master_existing'] is not None:
                df_old = st.session_state['df_master_existing']
                df_new = st.session_state['df_hasil'].drop(columns=['Status_Color'], errors='ignore')
               
                # Opsi filter
                st.markdown("### 🎯 Filter Perbandingan")
                col_filter1, col_filter2 = st.columns(2)
               
                with col_filter1:
                    show_option = st.radio(
                        "Tampilkan:",
                        ["Semua Data", "Hanya yang Berbeda", "Hanya yang Sama"],
                        horizontal=True,
                        key="tab2_filter_pppk"
                    )
               
                with col_filter2:
                    highlight_option = st.checkbox("Highlight Perbedaan per Kolom", value=True, key="tab2_highlight_pppk")
               
                st.markdown("---")
               
                # ===== FUNGSI COMPARE ROWS DENGAN IGNORE KETERANGAN DAN KOLOM BANK =====
                def compare_rows(row_old, row_new):
                    """Bandingkan dua row dan return dict perbedaan - IGNORE KETERANGAN DAN KOLOM BANK"""
                    differences = {}
                   
                    # Daftar kolom yang diabaikan dalam perbandingan
                    ignore_columns = {
                        'No', 
                        'Status_Color', 
                        'Keterangan',
                        'nmrek',
                        'nm_bank',
                        'rekening',
                        'kdbankspan',
                        'nmbankspan',
                        'kdpos'
                    }
                   
                    for col in HEADERS_MASTER:
                        if col in ignore_columns:
                            continue
                       
                        val_old = format_nilai_asli(row_old.get(col, '')) if row_old is not None else ''
                        val_new = format_nilai_asli(row_new.get(col, '')) if row_new is not None else ''
                       
                        if val_old != val_new:
                            differences[col] = {
                                'old': val_old,
                                'new': val_new
                            }
                   
                    return differences
                # ===== END PERUBAHAN =====
               
                # Buat dataframe perbandingan
                comparison_data = []
               
                for idx_new, row_new in df_new.iterrows():
                    nama_new = format_nilai_asli(row_new.get('Nama', ''))
                    nip_new = format_nilai_asli(row_new.get('NIP', ''))
                   
                    # Cari matching row di master lama
                    match_idx = fuzzy_match_row(nama_new, nip_new, df_old)
                   
                    if match_idx is not None:
                        row_old = df_old.iloc[match_idx]
                        differences = compare_rows(row_old, row_new)
                       
                        comparison_row = {
                            'Nama': nama_new,
                            'NIP': nip_new,
                            'Status': 'SAMA' if not differences else 'BERBEDA',
                            'Jumlah Perbedaan': len(differences),
                            'Kolom Berbeda': ', '.join(differences.keys()) if differences else '-',
                            'row_new': row_new,
                            'row_old': row_old,
                            'differences': differences
                        }
                    else:
                        comparison_row = {
                            'Nama': nama_new,
                            'NIP': nip_new,
                            'Status': 'BARU',
                            'Jumlah Perbedaan': 0,
                            'Kolom Berbeda': 'Data Baru (tidak ada di master lama)',
                            'row_new': row_new,
                            'row_old': None,
                            'differences': {}
                        }
                   
                    comparison_data.append(comparison_row)
               
                # Tambahkan data yang hilang (ada di master lama tapi tidak di master baru)
                for idx_old, row_old in df_old.iterrows():
                    nama_old = format_nilai_asli(row_old.get('Nama', ''))
                    nip_old = format_nilai_asli(row_old.get('NIP', ''))
                   
                    # Cek apakah kolom ada
                    nama_col = None
                    nip_col = None
                    for col in df_old.columns:
                        col_upper = str(col).upper()
                        if 'NAMA' in col_upper and nama_col is None:
                            nama_col = col
                        if 'NIP' in col.upper() and nip_col is None:
                            nip_col = col
                   
                    if nama_col and nip_col:
                        nama_old = format_nilai_asli(row_old.get(nama_col, ''))
                        nip_old = format_nilai_asli(row_old.get(nip_col, ''))
                       
                        match_idx = fuzzy_match_row(nama_old, nip_old, df_new)
                       
                        if match_idx is None:
                            comparison_row = {
                                'Nama': nama_old,
                                'NIP': nip_old,
                                'Status': 'HILANG',
                                'Jumlah Perbedaan': 0,
                                'Kolom Berbeda': 'Tidak ada di master baru',
                                'row_new': None,
                                'row_old': row_old,
                                'differences': {}
                            }
                            comparison_data.append(comparison_row)
               
                df_comparison = pd.DataFrame(comparison_data)
               
                # Filter berdasarkan pilihan
                if show_option == "Hanya yang Berbeda":
                    df_comparison = df_comparison[df_comparison['Status'].isin(['BERBEDA', 'BARU', 'HILANG'])]
                elif show_option == "Hanya yang Sama":
                    df_comparison = df_comparison[df_comparison['Status'] == 'SAMA']
               
                # Tampilkan ringkasan
                st.markdown("### 📊 Ringkasan Perbandingan")
                col_summary1, col_summary2, col_summary3, col_summary4 = st.columns(4)
               
                with col_summary1:
                    total_sama = len([x for x in comparison_data if x['Status'] == 'SAMA'])
                    st.metric("✅ Data Sama", total_sama)
               
                with col_summary2:
                    total_berbeda = len([x for x in comparison_data if x['Status'] == 'BERBEDA'])
                    st.metric("⚠️ Data Berbeda", total_berbeda)
               
                with col_summary3:
                    total_baru = len([x for x in comparison_data if x['Status'] == 'BARU'])
                    st.metric("🆕 Data Baru", total_baru)
               
                with col_summary4:
                    total_hilang = len([x for x in comparison_data if x['Status'] == 'HILANG'])
                    st.metric("❌ Data Hilang", total_hilang)
               
                st.markdown("---")
               
                # Tampilkan tabel perbandingan ringkas
                st.markdown("### 📋 Tabel Ringkasan Perbandingan")
               
                def color_status(val):
                    if val == 'SAMA':
                        return 'background-color: #90EE90'
                    elif val == 'BERBEDA':
                        return 'background-color: #FFA500'
                    elif val == 'BARU':
                        return 'background-color: #87CEEB'
                    elif val == 'HILANG':
                        return 'background-color: #FF6B6B'
                    return ''
               
                df_summary = df_comparison[['Nama', 'NIP', 'Status', 'Jumlah Perbedaan', 'Kolom Berbeda']].copy()
                st.dataframe(
                    df_summary.style.applymap(color_status, subset=['Status']),
                    height=400
                )
               
                # Download button khusus untuk tab 2
                st.markdown("---")
                st.subheader("📥 Download Data Perbandingan")
               
                # Buat dataframe untuk download
                df_comparison_download = df_comparison[['Nama', 'NIP', 'Status', 'Jumlah Perbedaan', 'Kolom Berbeda']].copy()
               
                # Tambahkan detail perbedaan jika ada
                if highlight_option and 'row_old' in df_comparison.columns and 'row_new' in df_comparison.columns:
                    # Buat dataframe detail
                    detail_data = []
                    for idx, row in df_comparison.iterrows():
                        if row['Status'] == 'BERBEDA' and row['differences']:
                            for col, diff in row['differences'].items():
                                detail_data.append({
                                    'Nama': row['Nama'],
                                    'NIP': row['NIP'],
                                    'Kolom': col,
                                    'Nilai Master Lama': diff['old'],
                                    'Nilai Master Baru': diff['new']
                                })
                   
                    if detail_data:
                        df_detail = pd.DataFrame(detail_data)
                       
                        # Gabungkan dengan summary
                        output_comparison = BytesIO()
                        with pd.ExcelWriter(output_comparison, engine='openpyxl') as writer:
                            df_comparison_download.to_excel(writer, index=False, sheet_name='Ringkasan Perbandingan')
                            df_detail.to_excel(writer, index=False, sheet_name='Detail Perbedaan')
                       
                        file_name = "perbandingan_master_lama_baru_detil_pppk.xlsx"
                    else:
                        output_comparison = BytesIO()
                        with pd.ExcelWriter(output_comparison, engine='openpyxl') as writer:
                            df_comparison_download.to_excel(writer, index=False, sheet_name='Ringkasan Perbandingan')
                       
                        file_name = "perbandingan_master_lama_baru_pppk.xlsx"
                else:
                    output_comparison = BytesIO()
                    with pd.ExcelWriter(output_comparison, engine='openpyxl') as writer:
                        df_comparison_download.to_excel(writer, index=False, sheet_name='Ringkasan Perbandingan')
                   
                    file_name = "perbandingan_master_lama_baru_pppk.xlsx"
               
                output_comparison.seek(0)
               
                st.download_button(
                    label="📥 Download Hasil Perbandingan (Excel)",
                    data=output_comparison,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_perbandingan_pppk"
                )
               
                st.markdown("---")
               
                # Detail perbandingan per pegawai
                if highlight_option and not df_comparison.empty:
                    st.markdown("### 🔍 Detail Perbandingan per Pegawai")
                   
                    # Pilih pegawai untuk detail
                    pegawai_options = [f"{row['Nama']} ({row['NIP']}) - {row['Status']}"
                                      for _, row in df_comparison.iterrows()]
                   
                    if pegawai_options:
                        selected_pegawai = st.selectbox(
                            "Pilih pegawai untuk melihat detail perbandingan:",
                            pegawai_options,
                            key="tab2_select_pegawai_pppk"
                        )
                       
                        selected_idx = pegawai_options.index(selected_pegawai)
                        selected_row = comparison_data[selected_idx]
                       
                        st.markdown(f"#### 👤 {selected_row['Nama']} (NIP: {selected_row['NIP']})")
                        st.markdown(f"**Status:** `{selected_row['Status']}`")
                       
                        if selected_row['Status'] == 'BERBEDA':
                            st.warning(f"⚠️ Ditemukan **{selected_row['Jumlah Perbedaan']}** perbedaan")
                           
                            # Buat tabel perbandingan detail
                            detail_data = []
                           
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan', 'nmrek', 'nm_bank', 'rekening', 'kdbankspan', 'nmbankspan', 'kdpos']:
                                    continue
                               
                                val_old = format_nilai_asli(selected_row['row_old'].get(col, '')) if selected_row['row_old'] is not None else '-'
                                val_new = format_nilai_asli(selected_row['row_new'].get(col, '')) if selected_row['row_new'] is not None else '-'
                               
                                is_different = (col in selected_row['differences'])
                               
                                detail_data.append({
                                    'Kolom': col,
                                    'Master Lama': val_old,
                                    'Master Baru': val_new,
                                    'Status': '❌ BERBEDA' if is_different else '✅ SAMA'
                                })
                           
                            df_detail = pd.DataFrame(detail_data)
                           
                            # Styling untuk highlight perbedaan
                            def highlight_diff(row):
                                if row['Status'] == '❌ BERBEDA':
                                    return ['background-color: #FFE4E1'] * len(row)
                                return [''] * len(row)
                           
                            st.dataframe(
                                df_detail.style.apply(highlight_diff, axis=1),
                                height=500
                            )
                       
                        elif selected_row['Status'] == 'SAMA':
                            st.success("✅ Semua data sama dengan master lama")
                           
                            # Tampilkan data
                            detail_data = []
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan', 'nmrek', 'nm_bank', 'rekening', 'kdbankspan', 'nmbankspan', 'kdpos']:
                                    continue
                                val = format_nilai_asli(selected_row['row_new'].get(col, ''))
                                detail_data.append({
                                    'Kolom': col,
                                    'Nilai': val
                                })
                           
                            st.dataframe(pd.DataFrame(detail_data), height=400)
                       
                        elif selected_row['Status'] == 'BARU':
                            st.info("🆕 Data baru, tidak ada di master lama")
                           
                            detail_data = []
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan', 'nmrek', 'nm_bank', 'rekening', 'kdbankspan', 'nmbankspan', 'kdpos']:
                                    continue
                                val = format_nilai_asli(selected_row['row_new'].get(col, ''))
                                detail_data.append({
                                    'Kolom': col,
                                    'Nilai': val
                                })
                           
                            st.dataframe(pd.DataFrame(detail_data), height=400)
                       
                        elif selected_row['Status'] == 'HILANG':
                            st.error("❌ Data tidak ada di master baru (tidak aktif)")
                           
                            detail_data = []
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan', 'nmrek', 'nm_bank', 'rekening', 'kdbankspan', 'nmbankspan', 'kdpos']:
                                    continue
                                val = format_nilai_asli(selected_row['row_old'].get(col, ''))
                                detail_data.append({
                                    'Kolom': col,
                                    'Nilai': val
                                })
                           
                            st.dataframe(pd.DataFrame(detail_data), height=400)
           
            else:
                st.info("ℹ️ Tidak ada Master Lama yang di-upload untuk dibandingkan")
       
        with tab3:
            st.subheader("🔄 Validasi Data Mentah vs Data BPMP")
            st.markdown("""
            **🎯 TUJUAN VALIDASI:**
            - Memastikan kesesuaian data antara sistem penggajian PPPK (Data Mentah) dan aplikasi pajak (BPMP)
            - Identifikasi perbedaan yang perlu diperbaiki sebelum proses lebih lanjut
            """)
           
            # Ambil dataframe dari session state atau dari file yang diupload
            if 'df_mentah' in st.session_state and 'df_bpmp' in st.session_state:
                df_mentah = st.session_state['df_mentah']
                df_bpmp = st.session_state['df_bpmp']
               
                # Tampilkan jumlah baris
                st.markdown("### 📊 Informasi Jumlah Data")
                col_info1, col_info2, col_info3 = st.columns(3)
               
                with col_info1:
                    st.metric("📄 Total Baris Data Mentah", len(df_mentah))
               
                with col_info2:
                    st.metric("📄 Total Baris Data BPMP", len(df_bpmp))
               
                with col_info3:
                    selisih = len(df_mentah) - len(df_bpmp)
                    st.metric("⚖️ Selisih", selisih,
                             delta=f"Data Mentah {'lebih banyak' if selisih > 0 else 'lebih sedikit' if selisih < 0 else 'sama'}")
               
                if selisih != 0:
                    st.warning(f"⚠️ Jumlah data tidak sama! Data Mentah: {len(df_mentah)}, Data BPMP: {len(df_bpmp)}")
                else:
                    st.success("✅ Jumlah data sama antara Data Mentah dan BPMP")
               
                st.markdown("---")
               
                # ===== CARI KOLOM UNTUK PERBANDINGAN TAMBAHAN =====
                # Cari kolom di Data Mentah untuk perbandingan
                bulan_col_mentah = None
                tahun_col_mentah = None
                gaji_kotor_col_mentah = None
                kdkawin_col_mentah = None
               
                for col in df_mentah.columns:
                    col_lower = str(col).lower()
                    if 'bulan' in col_lower and bulan_col_mentah is None:
                        bulan_col_mentah = col
                    elif 'tahun' in col_lower and tahun_col_mentah is None:
                        tahun_col_mentah = col
                    elif 'gajikotor' in col_lower.replace(' ', '') and gaji_kotor_col_mentah is None:
                        gaji_kotor_col_mentah = col
                    elif 'kdkawin' in col_lower and kdkawin_col_mentah is None:
                        kdkawin_col_mentah = col
               
                # Cari kolom di Data BPMP untuk perbandingan
                masa_pajak_col_bpmp = None
                tahun_pajak_col_bpmp = None
                penghasilan_kotor_col_bpmp = None
                status_col_bpmp = None
               
                for col in df_bpmp.columns:
                    col_lower = str(col).lower()
                    if 'masa' in col_lower and 'pajak' in col_lower and masa_pajak_col_bpmp is None:
                        masa_pajak_col_bpmp = col
                    elif 'tahun' in col_lower and 'pajak' in col_lower and tahun_pajak_col_bpmp is None:
                        tahun_pajak_col_bpmp = col
                    elif 'penghasilan' in col_lower and 'kotor' in col_lower and penghasilan_kotor_col_bpmp is None:
                        penghasilan_kotor_col_bpmp = col
                    elif 'status' in col_lower and 'pegawai' not in col_lower and status_col_bpmp is None:
                        status_col_bpmp = col
                # ===== END CARI KOLOM =====
               
                # Proses validasi
                st.markdown("### 🔍 Hasil Validasi NIP, NPWP, dan Data Lainnya")
               
                validation_data = []
               
                # Cari kolom NPWP/NIK/TIN di BPMP
                nik_col_bpmp = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col_bpmp = col
                        break
               
                # ===== BUAT DICTIONARY UNTUK MAPPING BPMP BERDASARKAN NIK =====
                # Buat mapping dari NIK ke baris BPMP untuk pencarian yang lebih cepat
                bpmp_mapping = {}
                if nik_col_bpmp:
                    for idx_bpmp, row_bpmp in df_bpmp.iterrows():
                        nik_bpmp = format_nilai_asli(row_bpmp.get(nik_col_bpmp, ''))
                        if nik_bpmp:
                            # Simpan row dan data tambahan
                            bpmp_data = {
                                'row': row_bpmp,
                                'masa_pajak': format_nilai_asli(row_bpmp.get(masa_pajak_col_bpmp, '')) if masa_pajak_col_bpmp else '',
                                'tahun_pajak': format_nilai_asli(row_bpmp.get(tahun_pajak_col_bpmp, '')) if tahun_pajak_col_bpmp else '',
                                'penghasilan_kotor': format_nilai_asli(row_bpmp.get(penghasilan_kotor_col_bpmp, '')) if penghasilan_kotor_col_bpmp else '',
                                'status_bpmp': format_nilai_asli(row_bpmp.get(status_col_bpmp, '')) if status_col_bpmp else ''
                            }
                            bpmp_mapping[nik_bpmp] = bpmp_data
                # ===== END MAPPING =====
               
                for idx_mentah, row_mentah in df_mentah.iterrows():
                    nip_mentah = format_nilai_asli(row_mentah.get('nip', '')) if 'nip' in df_mentah.columns else ''
                    npwp_mentah = format_nilai_asli(row_mentah.get('npwp', '')) if 'npwp' in df_mentah.columns else ''
                    nama_mentah = format_nilai_asli(row_mentah.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else ''
                   
                    # ===== AMBIL DATA TAMBAHAN DARI MENTAH =====
                    bulan_mentah = format_nilai_asli(row_mentah.get(bulan_col_mentah, '')) if bulan_col_mentah else ''
                    tahun_mentah = format_nilai_asli(row_mentah.get(tahun_col_mentah, '')) if tahun_col_mentah else ''
                    gaji_kotor_mentah = format_nilai_asli(row_mentah.get(gaji_kotor_col_mentah, '')) if gaji_kotor_col_mentah else ''
                    kdkawin_mentah = format_nilai_asli(row_mentah.get(kdkawin_col_mentah, '')) if kdkawin_col_mentah else ''
                    status_kawin_mentah = konversi_status(kdkawin_mentah) if kdkawin_col_mentah else ''
                    # ===== END DATA TAMBAHAN =====
                   
                    # Cari di BPMP berdasarkan NPWP menggunakan mapping
                    found_in_bpmp = False
                    npwp_bpmp = '-'
                    match_score = 0
                    rekomendasi = ''
                   
                    # ===== INISIALISASI DATA TAMBAHAN BPMP =====
                    masa_pajak_bpmp = ''
                    tahun_pajak_bpmp = ''
                    penghasilan_kotor_bpmp = ''
                    status_bpmp_value = ''
                    status_bulan = 'TIDAK ADA DATA'
                    status_tahun = 'TIDAK ADA DATA'
                    status_gaji = 'TIDAK ADA DATA'
                    status_kawin = 'TIDAK ADA DATA'
                    # ===== END INISIALISASI =====
                   
                    if nik_col_bpmp and npwp_mentah and npwp_mentah in bpmp_mapping:
                        # Exact match ditemukan
                        found_in_bpmp = True
                        npwp_bpmp = npwp_mentah
                        match_score = 100
                       
                        bpmp_data = bpmp_mapping[npwp_mentah]
                        masa_pajak_bpmp = bpmp_data['masa_pajak']
                        tahun_pajak_bpmp = bpmp_data['tahun_pajak']
                        penghasilan_kotor_bpmp = bpmp_data['penghasilan_kotor']
                        status_bpmp_value = bpmp_data['status_bpmp']
                       
                        # ===== BANDINGKAN DATA TAMBAHAN =====
                        # Bandingkan data tambahan
                        if bulan_mentah and masa_pajak_bpmp:
                            try:
                                # Coba konversi ke angka untuk perbandingan
                                bulan_mentah_num = int(bulan_mentah) if bulan_mentah.isdigit() else 0
                                masa_pajak_num = int(masa_pajak_bpmp) if masa_pajak_bpmp.isdigit() else 0
                                status_bulan = 'SESUAI' if bulan_mentah_num == masa_pajak_num else 'TIDAK SESUAI'
                            except:
                                status_bulan = 'TIDAK SESUAI' if bulan_mentah != masa_pajak_bpmp else 'SESUAI'
                       
                        if tahun_mentah and tahun_pajak_bpmp:
                            try:
                                tahun_mentah_num = int(tahun_mentah) if tahun_mentah.isdigit() else 0
                                tahun_pajak_num = int(tahun_pajak_bpmp) if tahun_pajak_bpmp.isdigit() else 0
                                status_tahun = 'SESUAI' if tahun_mentah_num == tahun_pajak_num else 'TIDAK SESUAI'
                            except:
                                status_tahun = 'TIDAK SESUAI' if tahun_mentah != tahun_pajak_bpmp else 'SESUAI'
                       
                        if gaji_kotor_mentah and penghasilan_kotor_bpmp:
                            try:
                                # Bersihkan format angka
                                gaji_mentah_clean = gaji_kotor_mentah.replace('.', '').replace(',', '.')
                                gaji_bpmp_clean = penghasilan_kotor_bpmp.replace('.', '').replace(',', '.')
                               
                                gaji_mentah_num = float(gaji_mentah_clean)
                                gaji_bpmp_num = float(gaji_bpmp_clean)
                               
                                # Toleransi 1 untuk perbedaan pembulatan
                                status_gaji = 'SESUAI' if abs(gaji_mentah_num - gaji_bpmp_num) <= 1 else 'TIDAK SESUAI'
                            except:
                                status_gaji = 'TIDAK SESUAI' if gaji_kotor_mentah != penghasilan_kotor_bpmp else 'SESUAI'
                       
                        # ===== PERBANDINGAN STATUS KAWIN =====
                        if status_kawin_mentah and status_bpmp_value:
                            # Normalisasi nilai untuk perbandingan
                            status_kawin_mentah_clean = status_kawin_mentah.strip().upper()
                            status_bpmp_clean = status_bpmp_value.strip().upper()
                           
                            # Mapping untuk perbandingan fleksibel
                            status_mapping = {
                                'TK/0': ['TK', 'TK/0', 'TK 0', 'TIDAK KAWIN'],
                                'TK/1': ['TK/1', 'TK 1', 'TIDAK KAWIN 1'],
                                'TK/2': ['TK/2', 'TK 2', 'TIDAK KAWIN 2'],
                                'K/0': ['K', 'K/0', 'K 0', 'KAWIN', 'KAWIN 0'],
                                'K/1': ['K/1', 'K 1', 'KAWIN 1'],
                                'K/2': ['K/2', 'K 2', 'KAWIN 2']
                            }
                           
                            # Cek apakah status cocok
                            match_found = False
                            if status_kawin_mentah_clean == status_bpmp_clean:
                                match_found = True
                            else:
                                # Cek mapping fleksibel
                                for key, values in status_mapping.items():
                                    if status_kawin_mentah_clean == key and status_bpmp_clean in values:
                                        match_found = True
                                        break
                                    elif status_bpmp_clean == key and status_kawin_mentah_clean in values:
                                        match_found = True
                                        break
                           
                            status_kawin = 'SESUAI' if match_found else 'TIDAK SESUAI'
                        elif status_kawin_mentah and not status_bpmp_value:
                            status_kawin = 'TIDAK ADA DATA BPMP'
                        elif not status_kawin_mentah and status_bpmp_value:
                            status_kawin = 'TIDAK ADA DATA MENTAH'
                        # ===== END PERBANDINGAN STATUS KAWIN =====
                        # ===== END BANDINGKAN DATA TAMBAHAN =====
                   
                    elif nik_col_bpmp and npwp_mentah:
                        # Coba fuzzy matching jika exact match tidak ditemukan
                        best_score = 0
                        best_nik = None
                       
                        for nik in bpmp_mapping.keys():
                            score = fuzz.ratio(npwp_mentah, nik)
                            if score > best_score and score >= 80:
                                best_score = score
                                best_nik = nik
                       
                        if best_nik:
                            found_in_bpmp = True
                            npwp_bpmp = best_nik
                            match_score = best_score
                           
                            bpmp_data = bpmp_mapping[best_nik]
                            masa_pajak_bpmp = bpmp_data['masa_pajak']
                            tahun_pajak_bpmp = bpmp_data['tahun_pajak']
                            penghasilan_kotor_bpmp = bpmp_data['penghasilan_kotor']
                            status_bpmp_value = bpmp_data['status_bpmp']
                           
                            # Bandingkan data tambahan untuk fuzzy match juga
                            if bulan_mentah and masa_pajak_bpmp:
                                try:
                                    bulan_mentah_num = int(bulan_mentah) if bulan_mentah.isdigit() else 0
                                    masa_pajak_num = int(masa_pajak_bpmp) if masa_pajak_bpmp.isdigit() else 0
                                    status_bulan = 'SESUAI' if bulan_mentah_num == masa_pajak_num else 'TIDAK SESUAI'
                                except:
                                    status_bulan = 'TIDAK SESUAI' if bulan_mentah != masa_pajak_bpmp else 'SESUAI'
                           
                            if tahun_mentah and tahun_pajak_bpmp:
                                try:
                                    tahun_mentah_num = int(tahun_mentah) if tahun_mentah.isdigit() else 0
                                    tahun_pajak_num = int(tahun_pajak_bpmp) if tahun_pajak_bpmp.isdigit() else 0
                                    status_tahun = 'SESUAI' if tahun_mentah_num == tahun_pajak_num else 'TIDAK SESUAI'
                                except:
                                    status_tahun = 'TIDAK SESUAI' if tahun_mentah != tahun_pajak_bpmp else 'SESUAI'
                           
                            if gaji_kotor_mentah and penghasilan_kotor_bpmp:
                                try:
                                    gaji_mentah_clean = gaji_kotor_mentah.replace('.', '').replace(',', '.')
                                    gaji_bpmp_clean = penghasilan_kotor_bpmp.replace('.', '').replace(',', '.')
                                   
                                    gaji_mentah_num = float(gaji_mentah_clean)
                                    gaji_bpmp_num = float(gaji_bpmp_clean)
                                   
                                    status_gaji = 'SESUAI' if abs(gaji_mentah_num - gaji_bpmp_num) <= 1 else 'TIDAK SESUAI'
                                except:
                                    status_gaji = 'TIDAK SESUAI' if gaji_kotor_mentah != penghasilan_kotor_bpmp else 'SESUAI'
                           
                            # ===== PERBANDINGAN STATUS KAWIN UNTUK FUZZY MATCH =====
                            if status_kawin_mentah and status_bpmp_value:
                                # Normalisasi nilai untuk perbandingan
                                status_kawin_mentah_clean = status_kawin_mentah.strip().upper()
                                status_bpmp_clean = status_bpmp_value.strip().upper()
                               
                                # Mapping untuk perbandingan fleksibel
                                status_mapping = {
                                    'TK/0': ['TK', 'TK/0', 'TK 0', 'TIDAK KAWIN'],
                                    'TK/1': ['TK/1', 'TK 1', 'TIDAK KAWIN 1'],
                                    'TK/2': ['TK/2', 'TK 2', 'TIDAK KAWIN 2'],
                                    'K/0': ['K', 'K/0', 'K 0', 'KAWIN', 'KAWIN 0'],
                                    'K/1': ['K/1', 'K 1', 'KAWIN 1'],
                                    'K/2': ['K/2', 'K 2', 'KAWIN 2']
                                }
                               
                                # Cek apakah status cocok
                                match_found = False
                                if status_kawin_mentah_clean == status_bpmp_clean:
                                    match_found = True
                                else:
                                    # Cek mapping fleksibel
                                    for key, values in status_mapping.items():
                                        if status_kawin_mentah_clean == key and status_bpmp_clean in values:
                                            match_found = True
                                            break
                                        elif status_bpmp_clean == key and status_kawin_mentah_clean in values:
                                            match_found = True
                                            break
                               
                                status_kawin = 'SESUAI' if match_found else 'TIDAK SESUAI'
                            elif status_kawin_mentah and not status_bpmp_value:
                                status_kawin = 'TIDAK ADA DATA BPMP'
                            elif not status_kawin_mentah and status_bpmp_value:
                                status_kawin = 'TIDAK ADA DATA MENTAH'
                            # ===== END PERBANDINGAN STATUS KAWIN UNTUK FUZZY MATCH =====
                   
                    # Tentukan status utama
                    if not nip_mentah and not npwp_mentah:
                        status_utama = 'DATA KOSONG'
                        rekomendasi = 'Lengkapi NIP dan NPWP di Data Mentah'
                    elif not nip_mentah:
                        status_utama = 'NIP KOSONG'
                        rekomendasi = 'Lengkapi NIP di Data Mentah'
                    elif not npwp_mentah:
                        status_utama = 'NPWP KOSONG'
                        rekomendasi = 'Lengkapi NPWP di Data Mentah'
                    elif found_in_bpmp:
                        status_utama = 'VALID'
                        rekomendasi = f'Data lengkap dan cocok (Match: {match_score}%)'
                    else:
                        status_utama = 'TIDAK ADA DI BPMP'
                        rekomendasi = f'Tambahkan pegawai "{nama_mentah}" dengan NPWP {npwp_mentah} ke Data BPMP'
                   
                    validation_data.append({
                        'No': idx_mentah + 1,
                        'Nama': nama_mentah,
                        'NIP (Data Mentah)': nip_mentah if nip_mentah else '❌ KOSONG',
                        'NPWP (Data Mentah)': npwp_mentah if npwp_mentah else '❌ KOSONG',
                        'NPWP (Data BPMP)': npwp_bpmp if found_in_bpmp else '❌ TIDAK ADA',
                        'Bulan (Mentah)': bulan_mentah if bulan_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Masa Pajak (BPMP)': masa_pajak_bpmp if found_in_bpmp and masa_pajak_col_bpmp else '❌ TIDAK ADA',
                        'Status Bulan': status_bulan,
                        'Tahun (Mentah)': tahun_mentah if tahun_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Tahun Pajak (BPMP)': tahun_pajak_bpmp if found_in_bpmp and tahun_pajak_col_bpmp else '❌ TIDAK ADA',
                        'Status Tahun': status_tahun,
                        'GajiKotor (Mentah)': gaji_kotor_mentah if gaji_kotor_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Penghasilan Kotor (BPMP)': penghasilan_kotor_bpmp if found_in_bpmp and penghasilan_kotor_col_bpmp else '❌ TIDAK ADA',
                        'Status Gaji Kotor': status_gaji,
                        'KDKAWIN (Mentah)': kdkawin_mentah if kdkawin_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Status Kawin (Mentah)': status_kawin_mentah if kdkawin_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Status (BPMP)': status_bpmp_value if found_in_bpmp and status_col_bpmp else '❌ TIDAK ADA',
                        'Status Perbandingan Kawin': status_kawin,
                        'Status': status_utama,
                        'Rekomendasi': rekomendasi
                    })
               
                df_validation = pd.DataFrame(validation_data)
               
                # Simpan ke session state untuk download
                st.session_state['df_validation_bpmp'] = df_validation
               
                # Statistik validasi utama
                st.markdown("### 📊 Ringkasan Validasi Utama")
                col_val1, col_val2, col_val3, col_val4 = st.columns(4)
               
                with col_val1:
                    total_valid = len(df_validation[df_validation['Status'] == 'VALID'])
                    st.metric("✅ Data Valid", total_valid)
               
                with col_val2:
                    total_tidak_ada = len(df_validation[df_validation['Status'] == 'TIDAK ADA DI BPMP'])
                    st.metric("❌ Tidak Ada di BPMP", total_tidak_ada)
               
                with col_val3:
                    total_kosong = len(df_validation[df_validation['Status'].str.contains('KOSONG')])
                    st.metric("⚠️ Data Kosong", total_kosong)
               
                with col_val4:
                    persentase_valid = round((total_valid / len(df_validation) * 100), 1) if len(df_validation) > 0 else 0
                    st.metric("📈 Persentase Valid", f"{persentase_valid}%")
               
                # ===== STATISTIK VALIDASI TAMBAHAN =====
                st.markdown("### 📊 Ringkasan Validasi Data Tambahan")
                col_val5, col_val6, col_val7, col_val8 = st.columns(4)
               
                with col_val5:
                    total_bulan_sesuai = len(df_validation[df_validation['Status Bulan'] == 'SESUAI'])
                    total_bulan_data = len(df_validation[df_validation['Status Bulan'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_bulan = round((total_bulan_sesuai / total_bulan_data * 100), 1) if total_bulan_data > 0 else 0
                    st.metric("📅 Bulan Sesuai", f"{total_bulan_sesuai}/{total_bulan_data}", f"{persentase_bulan}%")
               
                with col_val6:
                    total_tahun_sesuai = len(df_validation[df_validation['Status Tahun'] == 'SESUAI'])
                    total_tahun_data = len(df_validation[df_validation['Status Tahun'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_tahun = round((total_tahun_sesuai / total_tahun_data * 100), 1) if total_tahun_data > 0 else 0
                    st.metric("📅 Tahun Sesuai", f"{total_tahun_sesuai}/{total_tahun_data}", f"{persentase_tahun}%")
               
                with col_val7:
                    total_gaji_sesuai = len(df_validation[df_validation['Status Gaji Kotor'] == 'SESUAI'])
                    total_gaji_data = len(df_validation[df_validation['Status Gaji Kotor'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_gaji = round((total_gaji_sesuai / total_gaji_data * 100), 1) if total_gaji_data > 0 else 0
                    st.metric("💰 Gaji Sesuai", f"{total_gaji_sesuai}/{total_gaji_data}", f"{persentase_gaji}%")
               
                with col_val8:
                    total_kawin_sesuai = len(df_validation[df_validation['Status Perbandingan Kawin'] == 'SESUAI'])
                    total_kawin_data = len(df_validation[df_validation['Status Perbandingan Kawin'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_kawin = round((total_kawin_sesuai / total_kawin_data * 100), 1) if total_kawin_data > 0 else 0
                    st.metric("💍 Status Kawin Sesuai", f"{total_kawin_sesuai}/{total_kawin_data}", f"{persentase_kawin}%")
                # ===== END STATISTIK VALIDASI TAMBAHAN =====
               
                st.markdown("---")
               
                # Filter tampilan
                col_filter_val1, col_filter_val2 = st.columns(2)
               
                with col_filter_val1:
                    filter_status = st.selectbox(
                        "Filter Status Utama:",
                        ["Semua", "VALID", "TIDAK ADA DI BPMP", "NIP KOSONG", "NPWP KOSONG", "DATA KOSONG"],
                        key="tab3_filter_pppk"
                    )
               
                with col_filter_val2:
                    show_rekomendasi = st.checkbox("Tampilkan Kolom Rekomendasi", value=True, key="tab3_rekomendasi_pppk")
               
                # Apply filter
                if filter_status != "Semua":
                    df_validation_display = df_validation[df_validation['Status'] == filter_status]
                else:
                    df_validation_display = df_validation
               
                # ===== SELALU TAMPILKAN SEMUA DATA PERBANDINGAN =====
                # Tentukan kolom yang akan ditampilkan - SELALU sertakan semua kolom perbandingan
                kolom_tampil = [
                    'No', 'Nama',
                    'NIP (Data Mentah)', 'NPWP (Data Mentah)', 'NPWP (Data BPMP)',
                    'Bulan (Mentah)', 'Masa Pajak (BPMP)', 'Status Bulan',
                    'Tahun (Mentah)', 'Tahun Pajak (BPMP)', 'Status Tahun',
                    'GajiKotor (Mentah)', 'Penghasilan Kotor (BPMP)', 'Status Gaji Kotor',
                    # ===== KOLOM STATUS KAWIN =====
                    'KDKAWIN (Mentah)', 'Status Kawin (Mentah)', 'Status (BPMP)', 'Status Perbandingan Kawin',
                    # ===== END KOLOM STATUS KAWIN =====
                    'Status'
                ]
               
                if show_rekomendasi:
                    kolom_tampil.append('Rekomendasi')
               
                df_validation_display = df_validation_display[kolom_tampil]
                # ===== END TAMPILKAN SEMUA DATA PERBANDINGAN =====
               
                st.markdown("### 📋 Tabel Validasi Lengkap")
                st.info("✅ **Semua data perbandingan (Bulan, Tahun, Gaji, Status Kawin) ditampilkan untuk memudahkan validasi**")
               
                # Styling function
                def color_validation_status(row):
                    colors = [''] * len(row)
                    status = row['Status'] if 'Status' in row.index else ''
                   
                    if status == 'VALID':
                        colors = ['background-color: #90EE90'] * len(row)
                    elif status == 'TIDAK ADA DI BPMP':
                        colors = ['background-color: #FF6B6B'] * len(row)
                    elif 'KOSONG' in status:
                        colors = ['background-color: #FFD700'] * len(row)
                   
                    # Warna untuk status tambahan yang tidak sesuai
                    for i, col_name in enumerate(row.index):
                        if 'Status Bulan' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                        elif 'Status Tahun' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                        elif 'Status Gaji Kotor' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                        elif 'Status Perbandingan Kawin' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                   
                    return colors
               
                # Tampilkan tabel dengan scroll horizontal agar semua kolom terlihat
                st.dataframe(
                    df_validation_display.style.apply(color_validation_status, axis=1),
                    height=500,
                    use_container_width=True
                )
               
                # Download button khusus untuk tab 3
                st.markdown("---")
                st.subheader("📥 Download Hasil Validasi Data Mentah vs BPMP")
               
                # Buat Excel untuk download
                output_validation = BytesIO()
                with pd.ExcelWriter(output_validation, engine='openpyxl') as writer:
                    # Sheet 1: Data lengkap
                    df_validation.to_excel(writer, index=False, sheet_name='Validasi Lengkap')
                   
                    # Sheet 2: Data yang perlu perbaikan
                    df_perlu_perbaikan = df_validation[
                        (df_validation['Status'] == 'TIDAK ADA DI BPMP') |
                        (df_validation['Status'].str.contains('KOSONG'))
                    ]
                    df_perlu_perbaikan.to_excel(writer, index=False, sheet_name='Perlu Perbaikan')
                   
                    # Sheet 3: Data yang tidak sesuai (bulan, tahun, gaji, status kawin)
                    df_tidak_sesuai = df_validation[
                        (df_validation['Status Bulan'] == 'TIDAK SESUAI') |
                        (df_validation['Status Tahun'] == 'TIDAK SESUAI') |
                        (df_validation['Status Gaji Kotor'] == 'TIDAK SESUAI') |
                        (df_validation['Status Perbandingan Kawin'] == 'TIDAK SESUAI')
                    ]
                    df_tidak_sesuai.to_excel(writer, index=False, sheet_name='Data Tidak Sesuai')
                   
                    workbook = writer.book
                   
                    # Warna untuk sheet Perlu Perbaikan
                    if not df_perlu_perbaikan.empty:
                        worksheet = writer.sheets['Perlu Perbaikan']
                        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        yellow_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                       
                        for idx, row in df_perlu_perbaikan.iterrows():
                            excel_row = list(df_perlu_perbaikan.index).index(idx) + 2
                           
                            if row['Status'] == 'TIDAK ADA DI BPMP':
                                for col in range(1, len(df_perlu_perbaikan.columns) + 1):
                                    worksheet.cell(row=excel_row, column=col).fill = red_fill
                            elif 'KOSONG' in row['Status']:
                                for col in range(1, len(df_perlu_perbaikan.columns) + 1):
                                    worksheet.cell(row=excel_row, column=col).fill = yellow_fill
               
                output_validation.seek(0)
               
                st.download_button(
                    label="📥 Download Hasil Validasi (Excel)",
                    data=output_validation,
                    file_name="validasi_mentah_vs_bpmp_pppk.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_validasi_bpmp_pppk"
                )
               
                # Legend
                st.markdown("""
                **🎨 Keterangan Warna:**
                - 🟢 **Hijau**: Data VALID (NIP dan NPWP ada di kedua file)
                - 🔴 **Merah**: Data TIDAK ADA DI BPMP (ada di Data Mentah tapi tidak ada di BPMP)
                - 🟡 **Kuning**: Data KOSONG (NIP atau NPWP kosong di Data Mentah)
                - 🟠 **Salmon**: Data tambahan TIDAK SESUAI (bulan, tahun, gaji, atau status kawin tidak cocok)
               
                **📋 Kolom Perbandingan:**
                - **Bulan (Mentah)** vs **Masa Pajak (BPMP)**
                - **Tahun (Mentah)** vs **Tahun Pajak (BPMP)**
                - **GajiKotor (Mentah)** vs **Penghasilan Kotor (BPMP)**
                - **Status Kawin (Mentah)** vs **Status (BPMP)**
                """)
               
                st.markdown("---")
               
                # Download rekomendasi perbaikan
                if total_tidak_ada > 0 or total_kosong > 0:
                    st.markdown("### 📥 Download Data yang Perlu Diperbaiki")
                   
                    df_perlu_perbaikan = df_validation[
                        (df_validation['Status'] == 'TIDAK ADA DI BPMP') |
                        (df_validation['Status'].str.contains('KOSONG'))
                    ]
                   
                    if not df_perlu_perbaikan.empty:
                        st.info(f"💡 Total **{len(df_perlu_perbaikan)}** data yang perlu diperbaiki")
                else:
                    st.success("🎉 Semua data valid! Tidak ada yang perlu diperbaiki.")
           
            else:
                st.warning("⚠️ Pastikan Data Mentah dan Data BPMP sudah di-upload untuk melakukan validasi")
       
        with tab4:
            st.subheader("⚖️ Validasi Data Mentah vs Data Master")
            st.markdown("""
            **🎯 TUJUAN VALIDASI:**
            - **Data Master sebagai PRIMARY** (Data Benar & Akurat)
            - **Data Mentah sebagai data bulanan** yang harus sesuai dengan Master
            - Identifikasi inkonsistensi antara data bulanan dengan referensi utama
            """)
           
            if 'df_mentah' in st.session_state and 'df_master_existing' in st.session_state and st.session_state['df_master_existing'] is not None:
                df_mentah = st.session_state['df_mentah']
                df_master = st.session_state['df_master_existing']
               
                # VALIDASI PRIMARY KEY: Cek NIP duplikat di Data Mentah
                st.markdown("### 🔑 Validasi Primary Key (NIP)")
               
                if 'nip' in df_mentah.columns:
                    # Format NIP untuk validasi
                    nip_series = df_mentah['nip'].apply(lambda x: format_nilai_asli(x))
                    nip_duplicates = nip_series[nip_series.duplicated(keep=False)]
                   
                    if len(nip_duplicates) > 0:
                        st.error(f"❌ **PERINGATAN: Ditemukan {len(nip_duplicates)} NIP yang duplikat di Data Mentah!**")
                        st.warning("⚠️ NIP adalah Primary Key dan harus UNIQUE. Data dengan NIP duplikat tidak dapat diproses dengan benar.")
                       
                        # Tampilkan NIP yang duplikat
                        duplicate_nips = nip_duplicates.unique()
                        duplicate_data = []
                       
                        for dup_nip in duplicate_nips:
                            if dup_nip and dup_nip != 'nan':
                                dup_rows = df_mentah[nip_series == dup_nip]
                                for idx, row in dup_rows.iterrows():
                                    duplicate_data.append({
                                        'Baris': idx + 2, # +2 karena header dan index mulai dari 0
                                        'NIP (DUPLIKAT)': dup_nip,
                                        'Nama': format_nilai_asli(row.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else '-',
                                        'KDGOL': format_nilai_asli(row.get('kdgol', '')) if 'kdgol' in df_mentah.columns else '-',
                                        'Jumlah Duplikat': len(dup_rows)
                                    })
                       
                        df_duplicates = pd.DataFrame(duplicate_data)
                       
                        st.markdown("#### 📋 Daftar NIP yang Duplikat:")
                        st.dataframe(
                            df_duplicates.style.applymap(lambda x: 'background-color: #FF6B6B'),
                            height=300
                        )
                       
                        # Download duplikat
                        output_duplicates = BytesIO()
                        with pd.ExcelWriter(output_duplicates, engine='openpyxl') as writer:
                            df_duplicates.to_excel(writer, index=False, sheet_name='NIP Duplikat')
                           
                            workbook = writer.book
                            worksheet = writer.sheets['NIP Duplikat']
                           
                            red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                           
                            for row_idx in range(2, len(df_duplicates) + 2):
                                for col_idx in range(1, len(df_duplicates.columns) + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
                       
                        output_duplicates.seek(0)
                       
                        st.download_button(
                            label="📥 Download Daftar NIP Duplikat (Excel)",
                            data=output_duplicates,
                            file_name="nip_duplikat_data_mentah_pppk.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_duplikat_pppk"
                        )
                       
                        st.error("⛔ **PERBAIKI DUPLIKASI NIP TERLEBIH DAHULU** sebelum melanjutkan validasi dengan Master!")
                        st.markdown("---")
                       
                    else:
                        st.success("✅ Validasi Primary Key berhasil: Tidak ada NIP yang duplikat di Data Mentah")
                        st.markdown("---")
                else:
                    st.warning("⚠️ Kolom 'nip' tidak ditemukan di Data Mentah")
                    st.markdown("---")
               
                # Mapping kolom yang akan dibandingkan (tidak termasuk kolom bank)
                # Data Master = Data Mentah
                comparison_mapping = {
                    'Nama': 'nmpeg',
                    'NIP': 'nip',
                    'NIK': 'npwp',
                    'KDGOL': 'kdgol',
                    'KDKAWIN': 'kdkawin',
                    'STATUS': 'kdkawin', # Untuk konversi status
                    'KODE OBJEK PAJAK': 'kdgol' # Untuk konversi kode objek
                }
               
                st.markdown("### 📊 Informasi Jumlah Data")
                col_info1, col_info2, col_info3 = st.columns(3)
               
                with col_info1:
                    st.metric("📄 Total Baris Data Master", len(df_master))
               
                with col_info2:
                    st.metric("📄 Total Baris Data Mentah", len(df_mentah))
               
                with col_info3:
                    selisih = len(df_master) - len(df_mentah)
                    st.metric("⚖️ Selisih", selisih,
                             delta=f"Master {'lebih banyak' if selisih > 0 else 'lebih sedikit' if selisih < 0 else 'sama'}")
               
                st.markdown("---")
               
                # Cari kolom di master dengan pencarian fleksibel
                master_cols = {}
                for master_col in comparison_mapping.keys():
                    for col in df_master.columns:
                        if master_col.upper() in str(col).upper():
                            master_cols[master_col] = col
                            break
               
                st.markdown("### 🔍 Hasil Validasi Data Mentah vs Master")
                st.info(f"**Mapping Kolom:** Nama={comparison_mapping['Nama']}, NIP={comparison_mapping['NIP']}, NIK={comparison_mapping['NIK']}, KDGOL={comparison_mapping['KDGOL']}, KDKAWIN={comparison_mapping['KDKAWIN']}")
               
                validation_master_data = []
               
                for idx_mentah, row_mentah in df_mentah.iterrows():
                    # Ambil data dari mentah dengan format_nilai_asli
                    nip_mentah = format_nilai_asli(row_mentah.get('nip', '')) if 'nip' in df_mentah.columns else ''
                    nmpeg_mentah = format_nilai_asli(row_mentah.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else ''
                    npwp_mentah = format_nilai_asli(row_mentah.get('npwp', '')) if 'npwp' in df_mentah.columns else ''
                    kdgol_mentah = format_nilai_asli(row_mentah.get('kdgol', '')) if 'kdgol' in df_mentah.columns else ''
                    kdkawin_mentah = format_nilai_asli(row_mentah.get('kdkawin', '')) if 'kdkawin' in df_mentah.columns else ''
                   
                    # PRIMARY KEY: Cari matching di master berdasarkan NIP EXACT MATCH
                    match_found = False
                    nama_master = '-'
                    nip_master = '-'
                    nik_master = '-'
                    kdgol_master = '-'
                    kdkawin_master = '-'
                   
                    errors = []
                   
                    # Cari NIP di Master (Primary Key - harus exact match)
                    if 'NIP' in master_cols and nip_mentah:
                        nip_col_master = master_cols['NIP']
                       
                        for idx_master, row_master in df_master.iterrows():
                            nip_master_check = format_nilai_asli(row_master.get(nip_col_master, ''))
                           
                            # PRIMARY KEY: Exact match untuk NIP (tidak pakai fuzzy)
                            if nip_master_check and nip_mentah == nip_master_check:
                                match_found = True
                                nip_master = nip_master_check
                               
                                # Ambil data dari master untuk field lainnya
                                if 'Nama' in master_cols:
                                    nama_master = format_nilai_asli(row_master.get(master_cols['Nama'], ''))
                                if 'NIK' in master_cols:
                                    nik_master = format_nilai_asli(row_master.get(master_cols['NIK'], ''))
                                if 'KDGOL' in master_cols:
                                    kdgol_master = format_nilai_asli(row_master.get(master_cols['KDGOL'], ''))
                                if 'KDKAWIN' in master_cols:
                                    kdkawin_master = format_nilai_asli(row_master.get(master_cols['KDKAWIN'], ''))
                               
                                # Bandingkan HANYA field selain NIP (NIP sudah match sebagai Primary Key)
                                if nama_master and nmpeg_mentah:
                                    if fuzz.ratio(nmpeg_mentah.lower(), nama_master.lower()) < 90:
                                        errors.append('Nama')
                               
                                if nik_master and npwp_mentah:
                                    # Bandingkan NIK master dengan NPWP mentah
                                    if fuzz.ratio(nik_master, npwp_mentah) < 90:
                                        errors.append('NIK/NPWP')
                               
                                if kdgol_master and kdgol_mentah:
                                    if kdgol_master != kdgol_mentah:
                                        errors.append('KDGOL')
                               
                                if kdkawin_master and kdkawin_mentah:
                                    if kdkawin_master != kdkawin_mentah:
                                        errors.append('KDKAWIN')
                               
                                break
                   
                    # Tentukan status berdasarkan Primary Key (NIP)
                    if not nip_mentah or nip_mentah == '':
                        status = 'NIP KOSONG'
                        rekomendasi = 'NIP kosong di Data Mentah - tidak dapat diproses'
                    elif not match_found:
                        # NIP tidak ditemukan di Master
                        status = 'MASTER BELUM LENGKAP'
                        rekomendasi = f'NIP {nip_mentah} tidak ada di Master. Lengkapi Data Master terlebih dahulu.'
                    elif errors:
                        # NIP match, tapi field lain berbeda
                        status = 'TIDAK SESUAI'
                        # Rekomendasi HANYA untuk field yang berbeda (bukan NIP)
                        rekomendasi = f'Perbaiki kolom: {", ".join(errors)} di Data Mentah agar sesuai dengan Master'
                    else:
                        status = 'SESUAI'
                        rekomendasi = 'Semua data sesuai dengan Master'
                   
                    validation_master_data.append({
                        'No': idx_mentah + 1,
                        'NIP (Mentah)': nip_mentah if nip_mentah else '❌ KOSONG',
                        'NIP (Master)': nip_master if match_found else '❌ TIDAK ADA',
                        'Nama (Mentah)': nmpeg_mentah,
                        'Nama (Master)': nama_master,
                        'NPWP (Mentah)': npwp_mentah,
                        'NIK (Master)': nik_master,
                        'KDGOL (Mentah)': kdgol_mentah,
                        'KDGOL (Master)': kdgol_master,
                        'KDKAWIN (Mentah)': kdkawin_mentah,
                        'KDKAWIN (Master)': kdkawin_master,
                        'Status': status,
                        'Kolom Bermasalah': ', '.join(errors) if errors else ('-' if match_found else 'NIP tidak ada di Master'),
                        'Rekomendasi': rekomendasi
                    })
               
                df_validation_master = pd.DataFrame(validation_master_data)
               
                # Simpan ke session state untuk download
                st.session_state['df_validation_master'] = df_validation_master
               
                # Statistik
                st.markdown("### 📊 Ringkasan Validasi")
                col_val1, col_val2, col_val3, col_val4 = st.columns(4)
               
                with col_val1:
                    total_sesuai = len(df_validation_master[df_validation_master['Status'] == 'SESUAI'])
                    st.metric("✅ Data Sesuai", total_sesuai)
               
                with col_val2:
                    total_tidak_sesuai = len(df_validation_master[df_validation_master['Status'] == 'TIDAK SESUAI'])
                    st.metric("⚠️ Tidak Sesuai", total_tidak_sesuai)
               
                with col_val3:
                    total_master_belum = len(df_validation_master[df_validation_master['Status'] == 'MASTER BELUM LENGKAP'])
                    st.metric("❌ Master Belum Lengkap", total_master_belum)
               
                with col_val4:
                    total_nip_kosong = len(df_validation_master[df_validation_master['Status'] == 'NIP KOSONG'])
                    st.metric("⚠️ NIP Kosong", total_nip_kosong)
               
                st.markdown("---")
               
                # Filter
                col_filter1, col_filter2 = st.columns(2)
               
                with col_filter1:
                    filter_status_master = st.selectbox(
                        "Filter Status:",
                        ["Semua", "SESUAI", "TIDAK SESUAI", "MASTER BELUM LENGKAP", "NIP KOSONG"],
                        key="filter_master_pppk"
                    )
               
                with col_filter2:
                    show_rekomendasi_master = st.checkbox("Tampilkan Rekomendasi", value=True, key="show_rek_master_pppk")
               
                # Apply filter
                if filter_status_master != "Semua":
                    df_validation_master_display = df_validation_master[df_validation_master['Status'] == filter_status_master]
                else:
                    df_validation_master_display = df_validation_master
               
                if not show_rekomendasi_master:
                    df_validation_master_display = df_validation_master_display.drop(columns=['Rekomendasi'])
               
                st.markdown("### 📋 Tabel Validasi")
               
                # Styling dengan highlight per kolom
                def highlight_master_differences(row):
                    colors = [''] * len(row)
                    status = row['Status']
                   
                    if status == 'SESUAI':
                        return ['background-color: #90EE90'] * len(row)
                    elif status == 'MASTER BELUM LENGKAP':
                        return ['background-color: #FFD700'] * len(row)
                    elif status == 'NIP KOSONG':
                        return ['background-color: #FFA500'] * len(row)
                    elif status == 'TIDAK SESUAI':
                        # Highlight merah HANYA kolom Data Mentah yang bermasalah (bukan NIP)
                        kolom_bermasalah = str(row.get('Kolom Bermasalah', '')).split(', ')
                       
                        for i, col_name in enumerate(row.index):
                            # NIP tidak di-highlight karena Primary Key (tidak bisa salah)
                            if 'Nama' in kolom_bermasalah and '(Mentah)' in col_name and 'Nama' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif 'NIK/NPWP' in kolom_bermasalah and '(Mentah)' in col_name and 'NPWP' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif 'KDGOL' in kolom_bermasalah and '(Mentah)' in col_name and 'KDGOL' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif 'KDKAWIN' in kolom_bermasalah and '(Mentah)' in col_name and 'KDKAWIN' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif col_name in ['Status', 'Kolom Bermasalah', 'Rekomendasi']:
                                colors[i] = 'background-color: #FFA500'
                   
                    return colors
               
                st.dataframe(
                    df_validation_master_display.style.apply(highlight_master_differences, axis=1),
                    height=500
                )
               
                # Download button khusus untuk tab 4
                st.markdown("---")
                st.subheader("📥 Download Hasil Validasi Data Mentah vs Master")
               
                # Buat Excel untuk download
                output_validation_master = BytesIO()
                with pd.ExcelWriter(output_validation_master, engine='openpyxl') as writer:
                    # Sheet 1: Data lengkap
                    df_validation_master.to_excel(writer, index=False, sheet_name='Validasi Lengkap')
                   
                    # Sheet 2: Data yang perlu diperbaiki
                    df_perlu_perbaikan_master = df_validation_master[
                        (df_validation_master['Status'] == 'TIDAK SESUAI') |
                        (df_validation_master['Status'] == 'MASTER BELUM LENGKAP') |
                        (df_validation_master['Status'] == 'NIP KOSONG')
                    ]
                   
                    if not df_perlu_perbaikan_master.empty:
                        df_perlu_perbaikan_master.to_excel(writer, index=False, sheet_name='Perlu Perbaikan')
                       
                        workbook = writer.book
                        worksheet = writer.sheets['Perlu Perbaikan']
                       
                        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        yellow_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                       
                        for idx, row in df_perlu_perbaikan_master.iterrows():
                            excel_row = list(df_perlu_perbaikan_master.index).index(idx) + 2
                           
                            if row['Status'] == 'TIDAK SESUAI':
                                # Highlight kolom Data Mentah yang bermasalah (bukan NIP)
                                kolom_bermasalah = str(row['Kolom Bermasalah']).split(', ')
                               
                                for col_idx, col_name in enumerate(df_perlu_perbaikan_master.columns, start=1):
                                    if any(kb in col_name for kb in kolom_bermasalah) and '(Mentah)' in col_name and 'NIP' not in col_name:
                                        worksheet.cell(row=excel_row, column=col_idx).fill = red_fill
               
                output_validation_master.seek(0)
               
                st.download_button(
                    label="📥 Download Hasil Validasi (Excel)",
                    data=output_validation_master,
                    file_name="validasi_mentah_vs_master_pppk.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_validasi_master_pppk"
                )
               
                # Legend
                st.markdown("""
                **🎨 Keterangan Warna:**
                - 🟢 **Hijau**: Data SESUAI dengan Master (NIP match dan semua field cocok)
                - 🔴 **Merah**: Highlight pada kolom Data Mentah yang TIDAK SESUAI (NIP match tapi field lain berbeda)
                - 🟠 **Orange**: NIP KOSONG di Data Mentah / Kolom informasi
                - 🟡 **Kuning**: MASTER BELUM LENGKAP (NIP tidak ditemukan di Master)
               
                **📝 Note:** NIP adalah Primary Key - jika NIP tidak match, tidak ada rekomendasi perbaikan karena data tidak dapat dicocokkan.
                """)
               
                st.markdown("---")
               
                # Pisahkan info
                if 'df_perlu_perbaikan_master' in locals() and not df_perlu_perbaikan_master.empty:
                    total_master_belum_lengkap = len(df_perlu_perbaikan_master[df_perlu_perbaikan_master['Status'] == 'MASTER BELUM LENGKAP'])
                    total_tidak_sesuai_perbaikan = len(df_perlu_perbaikan_master[df_perlu_perbaikan_master['Status'] == 'TIDAK SESUAI'])
                    total_nip_kosong_perbaikan = len(df_perlu_perbaikan_master[df_perlu_perbaikan_master['Status'] == 'NIP KOSONG'])
                   
                    st.info(f"""
                    💡 **ANALISIS DATA YANG PERLU DITINDAK LANJUTI:**
                    
                    **Total {len(df_perlu_perbaikan_master)} data yang perlu ditindaklanjuti:**
                    - 🟡 **{total_master_belum_lengkap}** NIP tidak ditemukan di Master (Master belum lengkap)
                    - 🔴 **{total_tidak_sesuai_perbaikan}** field (selain NIP) perlu diperbaiki di Data Mentah
                    - 🟠 **{total_nip_kosong_perbaikan}** NIP kosong di Data Mentah
                    
                    **📌 Catatan:** NIP adalah Primary Key - tidak ada rekomendasi perbaikan untuk NIP yang tidak match.
                    """)
                else:
                    st.success("🎉 Semua data valid! Data Mentah sesuai dengan Master dan Master sudah lengkap.")
           
            else:
                st.warning("⚠️ Pastikan Data Mentah dan Master Existing sudah di-upload untuk melakukan validasi")
       
        with tab5:
            st.subheader("📈 Analisis Detail Perubahan")
           
            df_display = st.session_state['df_hasil'].copy()
           
            if 'df_master_existing' in st.session_state and st.session_state['df_master_existing'] is not None:
               
                # Pegawai yang tidak aktif
                df_tidak_aktif = df_display[df_display['AKTIF/TIDAK'] == 'TIDAK']
                if not df_tidak_aktif.empty:
                    st.markdown("### ❌ Pegawai Tidak Aktif Bulan Ini")
                    st.dataframe(df_tidak_aktif[['No', 'Nama', 'NIP', 'KDGOL', 'Keterangan']].style.apply(
                        lambda x: ['background-color: #FF6B6B'] * len(x), axis=1
                    ))
                   
                    # Statistik tidak aktif
                    col_tidak1, col_tidak2 = st.columns(2)
                   
                    with col_tidak1:
                        st.metric("Total Tidak Aktif", len(df_tidak_aktif))
                   
                    with col_tidak2:
                        persen_tidak = round((len(df_tidak_aktif) / len(df_display) * 100), 1)
                        st.metric("Persentase Tidak Aktif", f"{persen_tidak}%")
                else:
                    st.success("✅ Tidak ada pegawai yang tidak aktif")
               
                st.markdown("---")
               
                # Pegawai baru
                df_baru = df_display[df_display['Status_Color'] == 'HIJAU']
                if not df_baru.empty:
                    st.markdown("### 🆕 Pegawai Baru Bulan Ini")
                    st.dataframe(df_baru[['No', 'Nama', 'NIP', 'KDGOL', 'PNS/PPPK']].style.apply(
                        lambda x: ['background-color: #90EE90'] * len(x), axis=1
                    ))
                   
                    # Statistik baru
                    col_baru1, col_baru2 = st.columns(2)
                   
                    with col_baru1:
                        st.metric("Total Baru", len(df_baru))
                   
                    with col_baru2:
                        persen_baru = round((len(df_baru) / len(df_display) * 100), 1)
                        st.metric("Persentase Baru", f"{persen_baru}%")
                else:
                    st.info("ℹ️ Tidak ada pegawai baru")
               
                st.markdown("---")
               
                # Pegawai dengan data berubah
                df_berubah = df_display[df_display['Status_Color'] == 'ORANGE']
                if not df_berubah.empty:
                    st.markdown("### 🔄 Pegawai dengan Data Berubah")
                    st.dataframe(df_berubah[['No', 'Nama', 'NIP', 'KDGOL', 'Keterangan']].style.apply(
                        lambda x: ['background-color: #FFA500'] * len(x), axis=1
                    ))
                   
                    # Statistik berubah
                    col_berubah1, col_berubah2 = st.columns(2)
                   
                    with col_berubah1:
                        st.metric("Total Berubah", len(df_berubah))
                   
                    with col_berubah2:
                        persen_berubah = round((len(df_berubah) / len(df_display) * 100), 1)
                        st.metric("Persentase Berubah", f"{persen_berubah}%")
                else:
                    st.info("ℹ️ Tidak ada pegawai dengan data berubah")
               
                st.markdown("---")
               
                # Distribusi status
                st.markdown("### 📊 Distribusi Status Pegawai")
                col_chart1, col_chart2 = st.columns(2)
               
                with col_chart1:
                    status_counts = df_display['AKTIF/TIDAK'].value_counts()
                    if not status_counts.empty:
                        st.bar_chart(status_counts)
                        st.caption("Perbandingan Aktif vs Tidak Aktif")
                    else:
                        st.info("Tidak ada data distribusi status")
               
                with col_chart2:
                    if 'Status_Color' in df_display.columns:
                        color_counts = df_display['Status_Color'].value_counts()
                        if not color_counts.empty:
                            color_mapping = {
                                'HIJAU': 'Baru',
                                'KUNING': 'Tidak Berubah',
                                'MERAH': 'Tidak Aktif',
                                'ORANGE': 'Berubah'
                            }
                            color_counts.index = [color_mapping.get(c, c) for c in color_counts.index]
                            st.bar_chart(color_counts)
                            st.caption("Distribusi Perubahan Data")
                        else:
                            st.info("Tidak ada data perubahan")
               
                st.markdown("---")
               
                # Download button khusus untuk tab 5
                st.subheader("📥 Download Analisis Perubahan")
               
                # Buat Excel untuk analisis
                output_analisis = BytesIO()
                with pd.ExcelWriter(output_analisis, engine='openpyxl') as writer:
                    # Sheet 1: Ringkasan Analisis
                    summary_data = {
                        'Kategori': ['Total Data', 'Data Baru', 'Data Tidak Aktif', 'Data Berubah', 'Data Tidak Berubah'],
                        'Jumlah': [
                            len(df_display),
                            len(df_baru) if 'df_baru' in locals() else 0,
                            len(df_tidak_aktif) if 'df_tidak_aktif' in locals() else 0,
                            len(df_berubah) if 'df_berubah' in locals() else 0,
                            len(df_display[df_display['Status_Color'] == 'KUNING']) if 'Status_Color' in df_display.columns else 0
                        ],
                        'Persentase': [
                            '100%',
                            f"{round((len(df_baru) / len(df_display) * 100), 1)}%" if 'df_baru' in locals() else '0%',
                            f"{round((len(df_tidak_aktif) / len(df_display) * 100), 1)}%" if 'df_tidak_aktif' in locals() else '0%',
                            f"{round((len(df_berubah) / len(df_display) * 100), 1)}%" if 'df_berubah' in locals() else '0%',
                            f"{round((len(df_display[df_display['Status_Color'] == 'KUNING']) / len(df_display) * 100), 1)}%" if 'Status_Color' in df_display.columns else '0%'
                        ]
                    }
                    df_summary_analisis = pd.DataFrame(summary_data)
                    df_summary_analisis.to_excel(writer, index=False, sheet_name='Ringkasan Analisis')
                   
                    # Sheet 2: Data Tidak Aktif
                    if 'df_tidak_aktif' in locals() and not df_tidak_aktif.empty:
                        df_tidak_aktif_export = df_tidak_aktif.drop(columns=['Status_Color'], errors='ignore')
                        df_tidak_aktif_export.to_excel(writer, index=False, sheet_name='Pegawai Tidak Aktif')
                   
                    # Sheet 3: Data Baru
                    if 'df_baru' in locals() and not df_baru.empty:
                        df_baru_export = df_baru.drop(columns=['Status_Color'], errors='ignore')
                        df_baru_export.to_excel(writer, index=False, sheet_name='Pegawai Baru')
                   
                    # Sheet 4: Data Berubah
                    if 'df_berubah' in locals() and not df_berubah.empty:
                        df_berubah_export = df_berubah.drop(columns=['Status_Color'], errors='ignore')
                        df_berubah_export.to_excel(writer, index=False, sheet_name='Pegawai Berubah')
                   
                    # Sheet 5: Data Lengkap
                    df_display_export = df_display.drop(columns=['Status_Color'], errors='ignore')
                    df_display_export.to_excel(writer, index=False, sheet_name='Data Lengkap')
               
                output_analisis.seek(0)
               
                st.download_button(
                    label="📥 Download Analisis Perubahan (Excel)",
                    data=output_analisis,
                    file_name="analisis_perubahan_pppk.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_analisis_pppk"
                )
               
                # Ringkasan statistik
                st.markdown("### 📈 Ringkasan Statistik")
                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
               
                with col_stat1:
                    st.metric("Total Pegawai", len(df_display))
               
                with col_stat2:
                    st.metric("Pegawai Baru", len(df_baru) if 'df_baru' in locals() else 0)
               
                with col_stat3:
                    st.metric("Tidak Aktif", len(df_tidak_aktif) if 'df_tidak_aktif' in locals() else 0)
               
                with col_stat4:
                    st.metric("Data Berubah", len(df_berubah) if 'df_berubah' in locals() else 0)
           
            else:
                st.info("ℹ️ Upload Master Lama untuk melihat analisis perubahan detail")
               
                # Tampilkan statistik dasar saja
                st.markdown("### 📊 Statistik Dasar")
                col_stat1, col_stat2 = st.columns(2)
               
                with col_stat1:
                    st.metric("Total Pegawai", len(df_display))
               
                with col_stat2:
                    total_aktif = len(df_display[df_display['AKTIF/TIDAK'] == 'AKTIF'])
                    st.metric("Pegawai Aktif", total_aktif)
               
                # Download button untuk data dasar
                st.markdown("---")
                st.subheader("📥 Download Data Dasar")
               
                output_dasar = BytesIO()
                df_display_export = df_display.drop(columns=['Status_Color'], errors='ignore')
               
                with pd.ExcelWriter(output_dasar, engine='openpyxl') as writer:
                    df_display_export.to_excel(writer, index=False, sheet_name='Data Pegawai')
               
                output_dasar.seek(0)
               
                st.download_button(
                    label="📥 Download Data Pegawai (Excel)",
                    data=output_dasar,
                    file_name="data_pppk_dasar.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_dasar_pppk"
                )

# Untuk running langsung file ini (testing)
if __name__ == "__main__":
    show()