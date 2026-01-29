# croscheck_pns.py
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from io import BytesIO
import sys
import os

def show():
    """Fitur Sistem Master Data Pegawai dengan Tracking Bulanan"""
    
    # ===== KONFIGURASI HALAMAN =====
    st.set_page_config(page_title="Sistem Master Data Pegawai", layout="wide")
    
    # Tambahkan tombol kembali ke dashboard
    if st.button("← Kembali ke Dashboard"):
        st.session_state.current_page = 'dashboard_pns'
        st.rerun()
    
    # Breadcrumb navigation
    st.markdown("**Beranda → Dashboard PNS → Croscheck Data**")
    
    # ===== FUNGSI UNTUK FITUR MASTER DATA PEGAWAI =====
    
    # Definisi header untuk setiap file
    HEADERS_MENTAH = [
        "kdsatker", "kdanak", "kdsubanak", "bulan", "tahun", "nogaji", "kdjns", "nip", "nmpeg",
        "kdduduk", "kdgol", "npwp", "nmrek", "nm_bank", "rekening", "kdbankspan", "nmbankspan",
        "kdpos", "kdnegara", "kdkppn", "tipesup", "gjpokok", "tjistri", "tjanak", "tjupns",
        "tjstruk", "tjfungs", "tjdaerah", "tjpencil", "tjlain", "tjkompen", "pembul", "tjberas",
        "tjpph", "potpfkbul", "potpfk2", "GajiKotor", "potpfk10", "potpph", "potswrum",
        "potkelbtj", "potlain", "pottabrum", "bersih", "sandi", "kdkawin", "Status",
        "kdjab", "thngj", "kdgapok", "bpjs", "bpjs2"
    ]
    HEADERS_BPMP = [
        "Masa Pajak", "Tahun Pajak", "Status Pegawai", "Posisi", "NPWP/NIK/TIN",
        "Nomor Passport", "Kode Objek Pajak", "Penghasilan Kotor", "Tarif", "ID TKU",
        "Tgl Pemotongan", "TER A", "TER B", "TER C"
    ]
    HEADERS_MASTER = [
        "No", "PNS/PPPK", "Nama", "NIK", "ID PENERIMA TKU", "KDGOL", "KODE OBJEK PAJAK",
        "KDKAWIN", "STATUS", "NIP", "nmrek", "nm_bank", "rekening", "kdbankspan",
        "nmbankspan", "kdpos", "ID TKU", "AKTIF/TIDAK", "Keterangan"
    ]
    
    # Mapping KDKAWIN ke STATUS
    KDKAWIN_MAP = {
        "1000": "TK/0", "1001": "TK/1", "1002": "TK/2" , "1100": "K/0", 
        "1101": "K/1", "1102": "K/2"
    }
    
    # ===== PERUBAHAN PENTING: FUNGSI UNTUK MEMPERTAHANKAN FORMAT ANGKA ASLI =====
    def format_nilai_asli(nilai):
        """Mempertahankan format nilai asli tanpa .00 atau notasi ilmiah"""
        if pd.isna(nilai):
            return ''
        
        # Jika nilai sudah string, kembalikan langsung
        if isinstance(nilai, str):
            return nilai.strip()
        
        # Jika nilai float atau integer
        if isinstance(nilai, (int, float)):
            # Cek apakah nilai integer sebenarnya (tanpa desimal)
            if isinstance(nilai, float) and nilai.is_integer():
                return str(int(nilai))
            else:
                # Untuk float dengan desimal, tampilkan tanpa trailing zeros
                nilai_str = str(nilai)
                if '.' in nilai_str:
                    # Hapus trailing zeros setelah titik desimal
                    nilai_str = nilai_str.rstrip('0').rstrip('.')
                return nilai_str
        
        # Untuk tipe data lainnya
        return str(nilai)
    
    def clean_numeric_series(series):
        """Membersihkan series numerik agar tidak ada .00 yang tidak perlu"""
        return series.apply(format_nilai_asli)
    
    def format_angka_panjang(nilai):
        """Format khusus untuk angka panjang seperti NIP, rekening, dll."""
        if pd.isna(nilai):
            return ''
        
        nilai_str = str(nilai)
        
        # Hapus notasi ilmiah (e+)
        if 'e+' in nilai_str.lower():
            try:
                # Coba konversi ke integer tanpa notasi ilmiah
                nilai_float = float(nilai_str)
                # Format tanpa notasi ilmiah dan tanpa desimal jika integer
                if nilai_float.is_integer():
                    return str(int(nilai_float))
                else:
                    # Format dengan semua digit
                    return format(nilai_float, 'f').rstrip('0').rstrip('.')
            except:
                return nilai_str
        
        # Hapus .0, .00, .000, dll
        if '.' in nilai_str:
            parts = nilai_str.split('.')
            # Jika bagian desimal hanya berisi 0, hapus bagian desimal
            if len(parts) == 2 and all(c == '0' for c in parts[1]):
                return parts[0]
        
        return nilai_str
    # ===== END PERUBAHAN =====
    
    # ===== FUNGSI VALIDASI DUPLIKASI =====
    def check_duplicates(df, column_name, file_name):
        """Cek duplikasi di kolom tertentu dan return dataframe duplikat"""
        if df is None or column_name not in df.columns:
            return None, []
        
        # Ambil series dan bersihkan - gunakan format asli
        series = df[column_name].apply(format_nilai_asli)
        
        # Identifikasi duplikat (abaikan nilai kosong atau 'nan')
        mask = (series != 'nan') & (series != '') & (series.notna())
        series_filtered = series[mask]
        
        duplicates = series_filtered[series_filtered.duplicated(keep=False)]
        
        if len(duplicates) == 0:
            return None, []
        
        # Buat dataframe duplikat
        dup_indices = []
        for value in duplicates.unique():
            indices = df.index[series == value].tolist()
            dup_indices.extend(indices)
        
        df_duplicates = df.iloc[dup_indices].copy()
        df_duplicates = df_duplicates.reset_index(drop=True)
        df_duplicates.insert(0, 'Baris_Asli', [i+2 for i in dup_indices])  # +2 untuk header dan index 0-based
        df_duplicates.insert(1, 'Nilai_Duplikat', series.iloc[dup_indices].values)
        
        return df_duplicates, duplicates.unique().tolist()
    
    def highlight_duplicates(df, column_name, file_name):
        """Highlight baris duplikat di preview dataframe"""
        if df is None or column_name not in df.columns:
            return df
        
        series = df[column_name].apply(format_nilai_asli)
        mask = (series != 'nan') & (series != '') & (series.notna())
        duplicates_mask = series[mask].duplicated(keep=False)
        
        # Buat styling function
        def highlight_duplicate_rows(row):
            idx = row.name
            if idx in df.index[duplicates_mask]:
                return ['background-color: #FF6B6B'] * len(row)
            return [''] * len(row)
        
        return df.style.apply(highlight_duplicate_rows, axis=1)
    
    def konversi_kode_objek(kdgol):
        """Konversi kdgol ke kode objek pajak"""
        if pd.isna(kdgol):
            return "-"
        kdgol = format_nilai_asli(kdgol)
        if kdgol.startswith("3"):
            return "21-402-02"
        elif kdgol.startswith("4"):
            return "21-402-03"
        elif kdgol.startswith("2"):
            return "21-402-04"
        elif kdgol.startswith("1"):
            return "21-402-04"
        else:
            return "-"
    
    def konversi_status(kdkawin):
        """Konversi kdkawin ke status"""
        if pd.isna(kdkawin):
            return "-"
        kdkawin = format_nilai_asli(kdkawin)
        return KDKAWIN_MAP.get(kdkawin, "-")
    
    def read_excel_flexible(uploaded_file, expected_headers, label):
        """Baca Excel dengan pencarian header fleksibel dan pertahankan format asli"""
        if uploaded_file is None:
            return None
        
        try:
            uploaded_file.seek(0)
            wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
            sheet = wb.active
            
            total_rows = sheet.max_row
            total_columns = sheet.max_column
            expected_set = set(h.lower().strip() for h in expected_headers)
            
            st.write(f"📊 {label}: {total_rows} baris, {total_columns} kolom terdeteksi")
            
            # Cari baris header
            best_row = -1
            max_matches = 0
            best_row_values = []
            
            for row_idx in range(1, min(21, total_rows + 1)):
                row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[row_idx]]
                row_values_lower = [v.lower() for v in row_values]
                row_set = set(v for v in row_values_lower if v)
                matches = len(expected_set.intersection(row_set))
                if matches > max_matches:
                    max_matches = matches
                    best_row = row_idx
                    best_row_values = row_values
            
            if best_row > 0 and max_matches >= len(expected_headers) // 2:
                # Handle kolom kosong di kiri
                first_non_empty_col = next((i for i, val in enumerate(best_row_values) if val), 0)
                
                # Ekstrak data mulai dari baris setelah header
                data = []
                for row in sheet.iter_rows(min_row=best_row + 1, max_row=total_rows,
                                          min_col=first_non_empty_col + 1, values_only=True):
                    # ===== PERUBAHAN: GUNAKAN FORMAT ASLI UNTUK SETIAP SEL =====
                    formatted_row = []
                    for cell in row:
                        if isinstance(cell, (int, float)):
                            # Pertahankan format asli angka
                            formatted_row.append(format_nilai_asli(cell))
                        else:
                            formatted_row.append(cell)
                    data.append(formatted_row)
                    # ===== END PERUBAHAN =====
                
                # Ambil header yang valid (mulai dari kolom pertama yang tidak kosong)
                actual_headers = best_row_values[first_non_empty_col:]
                
                # Handle kolom duplikat dan kosong
                seen = {}
                unique_headers = []
                for i, col in enumerate(actual_headers):
                    if not col or col.strip() == "":
                        col = f"Unnamed_{i}"
                    
                    # Handle duplikat
                    original_col = col
                    counter = 1
                    while col in seen:
                        col = f"{original_col}_{counter}"
                        counter += 1
                    
                    seen[col] = True
                    unique_headers.append(col)
                
                # Batasi jumlah kolom sesuai data
                if len(data) > 0:
                    max_data_cols = len(data[0])
                    unique_headers = unique_headers[:max_data_cols]
                
                df = pd.DataFrame(data, columns=unique_headers)
                
                # Drop kolom yang sepenuhnya kosong
                df = df.dropna(axis=1, how='all')
                
                # Drop baris yang sepenuhnya kosong
                df = df.dropna(how='all').reset_index(drop=True)
                
                st.success(f"✅ {label} berhasil dibaca! Header di baris {best_row}, {len(df)} baris data")
                
            else:
                # Fallback: baca seluruh sheet
                data = []
                for row in sheet.iter_rows(min_row=1, max_row=total_rows, values_only=True):
                    # ===== PERUBAHAN: GUNAKAN FORMAT ASLI UNTUK SETIAP SEL =====
                    formatted_row = []
                    for cell in row:
                        if isinstance(cell, (int, float)):
                            formatted_row.append(format_nilai_asli(cell))
                        else:
                            formatted_row.append(cell)
                    data.append(formatted_row)
                    # ===== END PERUBAHAN =====
                
                if len(data) < 2:
                    st.error(f"❌ {label}: File tidak memiliki cukup data")
                    return None
                
                # Handle header duplikat
                headers = data[0]
                seen = {}
                unique_headers = []
                for i, col in enumerate(headers):
                    col_str = str(col).strip() if col else f"Unnamed_{i}"
                    
                    original_col = col_str
                    counter = 1
                    while col_str in seen:
                        col_str = f"{original_col}_{counter}"
                        counter += 1
                    
                    seen[col_str] = True
                    unique_headers.append(col_str)
                
                df = pd.DataFrame(data[1:], columns=unique_headers)
                df = df.dropna(axis=1, how='all')
                df = df.dropna(how='all').reset_index(drop=True)
                
                st.warning(f"⚠️ {label}: Header tidak sepenuhnya cocok, menggunakan baris pertama")
            
            # ===== PERUBAHAN: BERSIHKAN KOLOM NUMERIK PENTING =====
            # Daftar kolom yang perlu diformat khusus (angka panjang)
            numeric_cols_keywords = ['nip', 'npwp', 'nik', 'rekening', 'nogaji', 'id']
            
            for col in df.columns:
                col_lower = str(col).lower()
                if any(keyword in col_lower for keyword in numeric_cols_keywords):
                    df[col] = df[col].apply(format_angka_panjang)
            # ===== END PERUBAHAN =====
            
            # Tampilkan kolom yang ditemukan
            st.info(f"Kolom ditemukan: {', '.join(df.columns.tolist()[:10])}{'...' if len(df.columns) > 10 else ''}")
            
            return df
        
        except Exception as e:
            st.error(f"❌ Error membaca {label}: {e}")
            import traceback
            st.code(traceback.format_exc())
            return None
    
    def fuzzy_match_row(nama, nip, df_master, threshold=80):
        """Cari baris yang cocok menggunakan fuzzy matching"""
        if df_master is None or df_master.empty:
            return None
        
        best_match_idx = None
        best_score = 0
        
        # Cari kolom Nama dan NIP dengan pencarian fleksibel
        nama_col = None
        nip_col = None
        
        for col in df_master.columns:
            col_upper = str(col).upper()
            if 'NAMA' in col_upper and nama_col is None:
                nama_col = col
            if 'NIP' in col.upper() and nip_col is None:
                nip_col = col
        
        if not nama_col and not nip_col:
            return None
        
        for idx, row in df_master.iterrows():
            # Match berdasarkan Nama dan NIP
            nama_master = format_nilai_asli(row.get(nama_col, '')) if nama_col else ''
            nip_master = format_nilai_asli(row.get(nip_col, '')) if nip_col else ''
            
            nama_score = fuzz.ratio(str(nama).lower(), nama_master.lower()) if nama and nama_master else 0
            nip_score = fuzz.ratio(str(nip).lower(), nip_master.lower()) if nip and nip_master else 0
            
            # Gabungan score (prioritas NIP lebih tinggi)
            if nip_col and nama_col:
                combined_score = (nip_score * 0.7) + (nama_score * 0.3)
            elif nip_col:
                combined_score = nip_score
            elif nama_col:
                combined_score = nama_score
            else:
                combined_score = 0
            
            if combined_score > best_score and combined_score >= threshold:
                best_score = combined_score
                best_match_idx = idx
        
        return best_match_idx
    
    def process_data(df_mentah, df_bpmp, df_master_existing=None):
        """Proses data dari file mentah dan BPMP ke format master"""
        
        if df_mentah is None or df_bpmp is None:
            st.warning("⚠️ Pastikan file Data Mentah dan BPMP sudah di-upload!")
            return None
        
        # Normalisasi kolom - handle duplikat
        df_mentah.columns = [str(col).strip() for col in df_mentah.columns]
        df_bpmp.columns = [str(col).strip() for col in df_bpmp.columns]
        
        # Debug: tampilkan kolom yang tersedia
        st.write("**Kolom Data Mentah:**", df_mentah.columns.tolist()[:15])
        st.write("**Kolom Data BPMP:**", df_bpmp.columns.tolist())
        
        # Buat DataFrame hasil dengan merge berdasarkan fuzzy matching
        hasil = []
        
        for idx_mentah, row_mentah in df_mentah.iterrows():
            # ===== PERUBAHAN: GUNAKAN FORMAT ASLI UNTUK DATA =====
            # Ambil data dari file mentah dengan penanganan error
            nip = format_nilai_asli(row_mentah.get('nip', '')) if 'nip' in df_mentah.columns else ''
            nama = format_nilai_asli(row_mentah.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else ''
            npwp_mentah = format_nilai_asli(row_mentah.get('npwp', '')) if 'npwp' in df_mentah.columns else ''
            
            # Cari data BPMP yang cocok menggunakan fuzzy matching
            matched_bpmp = None
            best_match_score = 0
            
            for idx_bpmp, row_bpmp in df_bpmp.iterrows():
                # Cek apakah kolom NPWP/NIK/TIN ada
                nik_col = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col = col
                        break
                
                if nik_col:
                    nik_bpmp = format_nilai_asli(row_bpmp.get(nik_col, ''))
                    
                    # Matching berdasarkan NPWP/NIK
                    if npwp_mentah and nik_bpmp and npwp_mentah != '' and nik_bpmp != '':
                        score = fuzz.ratio(npwp_mentah, nik_bpmp)
                        if score > best_match_score and score >= 80:
                            best_match_score = score
                            matched_bpmp = row_bpmp
            
            # Jika tidak ada match berdasarkan NPWP, coba match berdasarkan urutan baris
            if matched_bpmp is None and idx_mentah < len(df_bpmp):
                matched_bpmp = df_bpmp.iloc[idx_mentah]
            
            # Ambil data dari matched BPMP
            posisi = ''
            nik_bpmp = npwp_mentah # default ke NPWP dari mentah
            
            if matched_bpmp is not None:
                # Cari kolom Posisi
                posisi_col = None
                for col in df_bpmp.columns:
                    if 'POSISI' in col.upper():
                        posisi_col = col
                        break
                if posisi_col:
                    posisi = format_nilai_asli(matched_bpmp.get(posisi_col, ''))
                
                # Cari kolom NIK
                nik_col = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col = col
                        break
                if nik_col:
                    nik_from_bpmp = format_nilai_asli(matched_bpmp.get(nik_col, ''))
                    if nik_from_bpmp:
                        nik_bpmp = nik_from_bpmp
            # ===== END PERUBAHAN =====
            
            # Ambil data lainnya dari mentah
            kdgol = format_nilai_asli(row_mentah.get('kdgol', '')) if 'kdgol' in df_mentah.columns else ''
            kdkawin = format_nilai_asli(row_mentah.get('kdkawin', '')) if 'kdkawin' in df_mentah.columns else ''
            
            # Filter PNS/PPPK
            pns_pppk = ''
            if pd.notna(posisi) and posisi:
                posisi_str = str(posisi).upper().strip()
                if 'PNS' in posisi_str:
                    pns_pppk = 'PNS'
                # Untuk PPPK biarkan kosong sesuai instruksi
            
            # ===== PERUBAHAN: FORMAT ID PENERIMA TKU DAN ID TKU DENGAN FORMAT ASLI =====
            # Format ID PENERIMA TKU: ambil NIK lalu tambahkan "000000" di akhir
            id_penerima_tku = f"{nik_bpmp}000000" if nik_bpmp and nik_bpmp.strip() != '' else ''
            
            # Format ID TKU: nilai default "0001658723701000000000" untuk semua baris
            id_tku = "0001658723701000000000"
            # ===== END PERUBAHAN =====
            
            row_data = {
                'No': idx_mentah + 1,
                'PNS/PPPK': pns_pppk,
                'Nama': nama,
                'NIK': nik_bpmp,
                'ID PENERIMA TKU': id_penerima_tku,  # Menggunakan format baru
                'KDGOL': kdgol,
                'KODE OBJEK PAJAK': konversi_kode_objek(kdgol),
                'KDKAWIN': kdkawin,
                'STATUS': konversi_status(kdkawin),
                'NIP': nip,
                'nmrek': format_nilai_asli(row_mentah.get('nmrek', '')),
                'nm_bank': format_nilai_asli(row_mentah.get('nm_bank', '')),
                'rekening': format_angka_panjang(row_mentah.get('rekening', '')),
                'kdbankspan': format_nilai_asli(row_mentah.get('kdbankspan', '')),
                'nmbankspan': format_nilai_asli(row_mentah.get('nmbankspan', '')),
                'kdpos': format_nilai_asli(row_mentah.get('kdpos', '')),
                'ID TKU': id_tku,  # Menggunakan nilai default
                'AKTIF/TIDAK': 'AKTIF',
                'Keterangan': ''
            }
            
            hasil.append(row_data)
        
        df_hasil = pd.DataFrame(hasil)
        
        # ===== PERUBAHAN: BERSIHKAN KOLOM NUMERIK DI HASIL =====
        # Pastikan kolom numerik penting diformat dengan benar
        numeric_cols = ['NIP', 'NIK', 'ID PENERIMA TKU', 'rekening']
        for col in numeric_cols:
            if col in df_hasil.columns:
                df_hasil[col] = df_hasil[col].apply(format_angka_panjang)
        # ===== END PERUBAHAN =====
        
        # Merge dengan master existing jika ada
        if df_master_existing is not None and not df_master_existing.empty:
            # Normalisasi kolom master existing
            df_master_existing.columns = [str(col).strip() for col in df_master_existing.columns]
            
            st.write("**Kolom Master Existing:**", df_master_existing.columns.tolist())
            
            # Tandai data yang sudah ada
            for idx, row in df_hasil.iterrows():
                match_idx = fuzzy_match_row(row['Nama'], row['NIP'], df_master_existing)
                
                if match_idx is not None:
                    # Data sudah ada - tandai KUNING
                    df_hasil.at[idx, 'Status_Color'] = 'KUNING'
                    
                    # Cari kolom Keterangan dengan fleksibel
                    ket_col = None
                    for col in df_master_existing.columns:
                        if 'KETERANGAN' in str(col).upper():
                            ket_col = col
                            break
                    
                    # Update keterangan jika ada
                    if ket_col:
                        existing_ket = df_master_existing.at[match_idx, ket_col]
                        if pd.notna(existing_ket) and existing_ket:
                            df_hasil.at[idx, 'Keterangan'] = format_nilai_asli(existing_ket)
                else:
                    # Data baru
                    df_hasil.at[idx, 'Status_Color'] = 'HIJAU'
            
            # Cek data lama yang tidak ada di bulan ini (MERAH)
            for idx, row_old in df_master_existing.iterrows():
                # Cari kolom Nama dan NIP di master existing
                nama_col = None
                nip_col = None
                
                for col in df_master_existing.columns:
                    col_upper = str(col).upper()
                    if 'NAMA' in col_upper and nama_col is None:
                        nama_col = col
                    if 'NIP' in col_upper and nip_col is None:
                        nip_col = col
                
                if not nama_col or not nip_col:
                    continue
                
                nama_old = format_nilai_asli(row_old.get(nama_col, ''))
                nip_old = format_nilai_asli(row_old.get(nip_col, ''))
                
                match_idx = fuzzy_match_row(nama_old, nip_old, df_hasil)
                
                if match_idx is None:
                    # Data tidak ada di bulan ini - tambahkan dengan status TIDAK AKTIF
                    row_old_dict = row_old.to_dict()
                    
                    # ===== PERUBAHAN: UNTUK DATA LAMA YANG TIDAK AKTIF =====
                    # Format ID PENERIMA TKU dari NIK master lama
                    nik_master = format_nilai_asli(row_old_dict.get('NIK', ''))
                    id_penerima_tku_old = f"{nik_master}000000" if nik_master and nik_master.strip() != '' else ''
                    
                    # ID TKU tetap menggunakan nilai default
                    id_tku_old = "0001658723701000000000"
                    # ===== END PERUBAHAN =====
                    
                    # Map kolom dari master existing ke format hasil
                    new_row = {
                        'No': len(df_hasil) + 1,
                        'PNS/PPPK': format_nilai_asli(row_old_dict.get('PNS/PPPK', '')),
                        'Nama': nama_old,
                        'NIK': format_nilai_asli(row_old_dict.get('NIK', '')),
                        'ID PENERIMA TKU': id_penerima_tku_old,  # Menggunakan format baru
                        'KDGOL': format_nilai_asli(row_old_dict.get('KDGOL', '')),
                        'KODE OBJEK PAJAK': format_nilai_asli(row_old_dict.get('KODE OBJEK PAJAK', '')),
                        'KDKAWIN': format_nilai_asli(row_old_dict.get('KDKAWIN', '')),
                        'STATUS': format_nilai_asli(row_old_dict.get('STATUS', '')),
                        'NIP': nip_old,
                        'nmrek': format_nilai_asli(row_old_dict.get('nmrek', '')),
                        'nm_bank': format_nilai_asli(row_old_dict.get('nm_bank', '')),
                        'rekening': format_angka_panjang(row_old_dict.get('rekening', '')),
                        'kdbankspan': format_nilai_asli(row_old_dict.get('kdbankspan', '')),
                        'nmbankspan': format_nilai_asli(row_old_dict.get('nmbankspan', '')),
                        'kdpos': format_nilai_asli(row_old_dict.get('kdpos', '')),
                        'ID TKU': id_tku_old,  # Menggunakan nilai default
                        'AKTIF/TIDAK': 'TIDAK',
                        'Keterangan': format_nilai_asli(row_old_dict.get('Keterangan', '')),
                        'Status_Color': 'MERAH'
                    }
                    
                    df_hasil = pd.concat([df_hasil, pd.DataFrame([new_row])], ignore_index=True)
        else:
            # Semua data baru
            df_hasil['Status_Color'] = 'HIJAU'
        
        return df_hasil
    
    def create_excel_with_colors(df):
        """Buat Excel dengan warna berdasarkan status"""
        output = BytesIO()
        
        # Hapus kolom helper
        df_export = df.drop(columns=['Status_Color'], errors='ignore')
        
        # ===== PERUBAHAN: FORMAT KOLOM NUMERIK SEBELUM EXPORT =====
        # Format kolom numerik untuk export Excel
        for col in df_export.columns:
            if col in ['NIP', 'NIK', 'ID PENERIMA TKU', 'rekening', 'ID TKU']:
                df_export[col] = df_export[col].apply(lambda x: format_angka_panjang(x) if pd.notna(x) else '')
        # ===== END PERUBAHAN =====
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Master Data')
            
            workbook = writer.book
            worksheet = writer.sheets['Master Data']
            
            # Define fills
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            # Apply colors
            for idx, row in df.iterrows():
                excel_row = idx + 2 # +2 karena header di row 1
                color = row.get('Status_Color', '')
                
                if color == 'KUNING':
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = yellow_fill
                elif color == 'MERAH':
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = red_fill
                elif color == 'ORANGE':
                    # Orange untuk nama yang ada tapi data berbeda
                    worksheet.cell(row=excel_row, column=3).fill = orange_fill # Kolom Nama
                elif color == 'HIJAU':
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=excel_row, column=col).fill = green_fill
        
        output.seek(0)
        return output
    
    # ===== UI UNTUK FITUR MASTER DATA =====
    st.title("🔍 CROSCHECK DATA PNS")
    
    # ===== PANDUAN PENGGUNAAN =====
    with st.expander("📘 **PANDUAN PENGGUNAAN & INFORMASI SISTEM**", expanded=False):
        st.markdown("""
        ### 🎯 **TUJUAN SISTEM**
        Sistem ini dirancang untuk melakukan **Croscheck Data PNS** dengan fitur:
        1. **Validasi Data** - Memastikan konsistensi data antara file Data Mentah, BPMP, dan Master
        2. **Tracking Perubahan** - Melacak perubahan data pegawai bulanan
        3. **Deteksi Duplikasi** - Mengidentifikasi NIP dan NPWP/NIK yang duplikat
        4. **Pembaruan Master Data** - Menghasilkan file master data terbaru
        
        ### 📁 **FORMAT FILE YANG DIDUKUNG**
        - **Hanya format Excel**: `.xls` dan `.xlsx`
        - **Jumlah File**: 3 file (2 wajib, 1 opsional)
        
        ### 🔄 **ALUR KERJA SISTEM**
        1. **Upload 3 File**:
           - **File 1**: Data Mentah (wajib) - Data bulanan dari sistem penggajian
           - **File 2**: Data BPMP (wajib) - Data perpajakan
           - **File 3**: Master Existing (opsional) - Data master sebelumnya untuk tracking
        
        2. **Validasi Otomatis**:
           - Sistem akan mengecek duplikasi NIP dan NPWP/NIK
           - Tampilkan peringatan jika ditemukan duplikasi
        
        3. **Proses Data**:
           - Klik tombol "Proses Data" untuk memulai croscheck
           - Sistem akan menggabungkan data dari 3 file
        
        4. **Analisis Hasil**:
           - Lihat hasil di 5 tab berbeda
           - Download hasil sesuai kebutuhan
        
        ### ⚠️ **ATURAN PENTING**
        1. **NIP HARUS UNIK** - Tidak boleh ada duplikasi NIP di file mana pun
        2. **NPWP/NIK HARUS UNIK** - Tidak boleh ada duplikasi NPWP/NIK di file mana pun
        3. **Format Header** - Header file harus sesuai dengan template
        4. **Data Kosong** - NIP dan NPWP tidak boleh kosong
        
        ### 🎨 **KETERANGAN WARNA**
        - 🟢 **Hijau**: Data baru/belum ada di master lama
        - 🟡 **Kuning**: Data sudah ada (tidak berubah)
        - 🔴 **Merah**: Data tidak aktif (ada di master lama, tidak di bulan ini)
        - 🟠 **Orange**: Nama ada tapi data berbeda
        
        ### 📊 **FITUR TAB ANALISIS**
        1. **Hasil Master Data Baru** - Data master hasil croscheck
        2. **Perbandingan Master Lama vs Baru** - Analisis perubahan
        3. **Validasi Data Mentah vs BPMP** - Cek kesesuaian data
        4. **Validasi Data Mentah vs Master** - Validasi berdasarkan master
        5. **Analisis Detail Perubahan** - Statistik dan trend
        
        ### ❗ **CATATAN TEKNIS**
        - Sistem menggunakan **fuzzy matching** untuk pencocokan data
        - **ID PENERIMA TKU**: Format = NIK + "000000"
        - **ID TKU**: Nilai tetap "0001658723701000000000" untuk semua
        - **Status Kawin**: Otomatis dikonversi dari kode (1000 → TK/0, 1100 → K/0, dll)
        - **Kode Objek Pajak**: Otomatis dikonversi dari KDGOL
        
        ### 🚨 **VALIDASI DUPLIKASI**
        Sistem akan **OTOMATIS** mengecek:
        - **NIP duplikat** di Data Mentah, Data Master, dan Hasil
        - **NPWP/NIK duplikat** di semua file
        - Jika ditemukan duplikasi, **tombol download akan dinonaktifkan** sampai diperbaiki
        """)
    
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.subheader("📁 File 1: Data Mentah")
        uploaded_mentah = st.file_uploader("Upload Data Mentah", type=["xls", "xlsx"], key="mentah")
    with col2:
        st.subheader("📁 File 2: Data BPMP")
        uploaded_bpmp = st.file_uploader("Upload Data BPMP", type=["xls", "xlsx"], key="bpmp")
    with col3:
        st.subheader("📁 File 3: Master Existing (Opsional)")
        uploaded_master = st.file_uploader("Upload Master Lama", type=["xls", "xlsx"], key="master")
    
    st.markdown("---")
    
    # Baca file
    df_mentah = read_excel_flexible(uploaded_mentah, HEADERS_MENTAH, "Data Mentah")
    df_bpmp = read_excel_flexible(uploaded_bpmp, HEADERS_BPMP, "Data BPMP")
    df_master_existing = read_excel_flexible(uploaded_master, HEADERS_MASTER, "Master Existing")
    
    # ===== VALIDASI DUPLIKASI UNTUK SEMUA FILE =====
    st.subheader("🔍 Validasi Duplikasi NIP dan NPWP/NIK")
    
    # Inisialisasi status duplikasi
    duplicate_status = {
        'mentah_nip': False,
        'mentah_npwp': False,
        'master_nip': False,
        'master_nik': False,
        'bpmp_nik': False,
        'hasil_nip': False,
        'hasil_nik': False
    }
    
    # Cek duplikasi di Data Mentah
    if df_mentah is not None:
        col_check1, col_check2 = st.columns(2)
        
        with col_check1:
            # Cek NIP duplikat di Data Mentah
            if 'nip' in df_mentah.columns:
                df_nip_dup_mentah, dup_nip_values = check_duplicates(df_mentah, 'nip', 'Data Mentah')
                if df_nip_dup_mentah is not None:
                    duplicate_status['mentah_nip'] = True
                    st.error(f"❌ **DITEMUKAN {len(dup_nip_values)} NIP DUPLIKAT DI DATA MENTAH**")
                    with st.expander("🔍 Lihat Detail NIP Duplikat di Data Mentah"):
                        st.dataframe(df_nip_dup_mentah[['Baris_Asli', 'Nilai_Duplikat', 'nmpeg']].head(20))
                else:
                    st.success("✅ NIP di Data Mentah: TIDAK ADA DUPLIKASI")
        
        with col_check2:
            # Cek NPWP duplikat di Data Mentah
            if 'npwp' in df_mentah.columns:
                df_npwp_dup_mentah, dup_npwp_values = check_duplicates(df_mentah, 'npwp', 'Data Mentah')
                if df_npwp_dup_mentah is not None:
                    duplicate_status['mentah_npwp'] = True
                    st.error(f"❌ **DITEMUKAN {len(dup_npwp_values)} NPWP DUPLIKAT DI DATA MENTAH**")
                    with st.expander("🔍 Lihat Detail NPWP Duplikat di Data Mentah"):
                        st.dataframe(df_npwp_dup_mentah[['Baris_Asli', 'Nilai_Duplikat', 'nmpeg']].head(20))
                else:
                    st.success("✅ NPWP di Data Mentah: TIDAK ADA DUPLIKASI")
    
    # Cek duplikasi di Master Existing
    if df_master_existing is not None:
        col_check3, col_check4 = st.columns(2)
        
        with col_check3:
            # Cek NIP duplikat di Master Existing
            # Cari kolom NIP dengan pencarian fleksibel
            nip_col_master = None
            for col in df_master_existing.columns:
                if 'NIP' in str(col).upper():
                    nip_col_master = col
                    break
            
            if nip_col_master:
                df_nip_dup_master, dup_nip_master_values = check_duplicates(df_master_existing, nip_col_master, 'Master Existing')
                if df_nip_dup_master is not None:
                    duplicate_status['master_nip'] = True
                    st.error(f"❌ **DITEMUKAN {len(dup_nip_master_values)} NIP DUPLIKAT DI MASTER EXISTING**")
                    with st.expander("🔍 Lihat Detail NIP Duplikat di Master Existing"):
                        st.dataframe(df_nip_dup_master[[f'Baris_Asli', 'Nilai_Duplikat', 'Nama' if 'Nama' in df_master_existing.columns else df_master_existing.columns[2]]].head(20))
                else:
                    st.success("✅ NIP di Master Existing: TIDAK ADA DUPLIKASI")
        
        with col_check4:
            # Cek NIK duplikat di Master Existing
            # Cari kolom NIK dengan pencarian fleksibel
            nik_col_master = None
            for col in df_master_existing.columns:
                if 'NIK' in str(col).upper() and 'PENERIMA' not in str(col).upper():
                    nik_col_master = col
                    break
            
            if nik_col_master:
                df_nik_dup_master, dup_nik_master_values = check_duplicates(df_master_existing, nik_col_master, 'Master Existing')
                if df_nik_dup_master is not None:
                    duplicate_status['master_nik'] = True
                    st.error(f"❌ **DITEMUKAN {len(dup_nik_master_values)} NIK DUPLIKAT DI MASTER EXISTING**")
                    with st.expander("🔍 Lihat Detail NIK Duplikat di Master Existing"):
                        st.dataframe(df_nik_dup_master[[f'Baris_Asli', 'Nilai_Duplikat', 'Nama' if 'Nama' in df_master_existing.columns else df_master_existing.columns[2]]].head(20))
                else:
                    st.success("✅ NIK di Master Existing: TIDAK ADA DUPLIKASI")
    
    # Cek duplikasi di Data BPMP
    if df_bpmp is not None:
        # Cari kolom NPWP/NIK/TIN di BPMP
        nik_col_bpmp = None
        for col in df_bpmp.columns:
            if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                nik_col_bpmp = col
                break
        
        if nik_col_bpmp:
            df_nik_dup_bpmp, dup_nik_bpmp_values = check_duplicates(df_bpmp, nik_col_bpmp, 'Data BPMP')
            if df_nik_dup_bpmp is not None:
                duplicate_status['bpmp_nik'] = True
                st.error(f"❌ **DITEMUKAN {len(dup_nik_bpmp_values)} NPWP/NIK DUPLIKAT DI DATA BPMP**")
                with st.expander("🔍 Lihat Detail NPWP/NIK Duplikat di Data BPMP"):
                    # Cari kolom nama di BPMP
                    nama_col_bpmp = None
                    for col in df_bpmp.columns:
                        if 'NAMA' in col.upper() or 'PEGAWAI' in col.upper():
                            nama_col_bpmp = col
                            break
                    
                    if nama_col_bpmp:
                        st.dataframe(df_nik_dup_bpmp[['Baris_Asli', 'Nilai_Duplikat', nama_col_bpmp]].head(20))
                    else:
                        st.dataframe(df_nik_dup_bpmp[['Baris_Asli', 'Nilai_Duplikat']].head(20))
            else:
                st.success("✅ NPWP/NIK di Data BPMP: TIDAK ADA DUPLIKASI")
    
    # Simpan status duplikasi ke session state
    st.session_state['duplicate_status'] = duplicate_status
    
    # Tampilkan status keseluruhan
    has_duplicates = any(duplicate_status.values())
    
    if has_duplicates:
        st.error("""
        ⚠️ **PERINGATAN: DITEMUKAN DUPLIKASI DATA!**
        
        **Sistem tidak dapat melanjutkan proses sebelum duplikasi diperbaiki:**
        1. Periksa file yang memiliki duplikasi
        2. Perbaiki data duplikat di file sumber
        3. Upload ulang file yang sudah diperbaiki
        4. Pastikan NIP dan NPWP/NIK UNIK di semua file
        
        **Tombol download akan dinonaktifkan sampai semua duplikasi diperbaiki.**
        """)
    else:
        st.success("✅ **SEMUA FILE VALID: Tidak ditemukan duplikasi NIP dan NPWP/NIK**")
    
    st.markdown("---")
    
    # Tampilkan preview data yang di-upload
    if df_mentah is not None:
        with st.expander("👀 Preview Data Mentah (dengan highlight duplikasi)"):
            # Tampilkan dengan highlight jika ada duplikasi
            if 'nip' in df_mentah.columns:
                styled_df = highlight_duplicates(df_mentah.head(500), 'nip', 'Data Mentah')
                st.dataframe(styled_df)
            else:
                st.dataframe(df_mentah.head(500))
    
    if df_bpmp is not None:
        with st.expander("👀 Preview Data BPMP"):
            st.dataframe(df_bpmp.head(500))
    
    if df_master_existing is not None:
        with st.expander("👀 Preview Master Existing"):
            st.dataframe(df_master_existing.head(500))
    
    # Proses data
    if st.button("🔄 Proses Data", type="primary", disabled=has_duplicates):
        if has_duplicates:
            st.error("⚠️ Tidak dapat memproses data karena terdapat duplikasi. Perbaiki terlebih dahulu.")
        else:
            with st.spinner("Memproses data..."):
                df_hasil = process_data(df_mentah, df_bpmp, df_master_existing)
                
                if df_hasil is not None:
                    # Cek duplikasi di hasil
                    df_nip_dup_hasil, dup_nip_hasil = check_duplicates(df_hasil, 'NIP', 'Hasil')
                    df_nik_dup_hasil, dup_nik_hasil = check_duplicates(df_hasil, 'NIK', 'Hasil')
                    
                    if df_nip_dup_hasil is not None:
                        duplicate_status['hasil_nip'] = True
                        st.error(f"❌ Ditemukan {len(dup_nip_hasil)} NIP duplikat di Hasil")
                        with st.expander("🔍 Lihat Detail NIP Duplikat di Hasil"):
                            st.dataframe(df_nip_dup_hasil[['Baris_Asli', 'Nilai_Duplikat', 'Nama']].head(20))
                    
                    if df_nik_dup_hasil is not None:
                        duplicate_status['hasil_nik'] = True
                        st.error(f"❌ Ditemukan {len(dup_nik_hasil)} NIK duplikat di Hasil")
                        with st.expander("🔍 Lihat Detail NIK Duplikat di Hasil"):
                            st.dataframe(df_nik_dup_hasil[['Baris_Asli', 'Nilai_Duplikat', 'Nama']].head(20))
                    
                    # Update status duplikasi
                    st.session_state['duplicate_status'] = duplicate_status
                    has_duplicates = any(duplicate_status.values())
                    
                    if not has_duplicates:
                        st.session_state['df_hasil'] = df_hasil
                        st.session_state['df_master_existing'] = df_master_existing
                        st.session_state['df_mentah'] = df_mentah
                        st.session_state['df_bpmp'] = df_bpmp
                        st.success("✅ Data berhasil diproses!")
                    else:
                        st.error("❌ Proses data gagal karena menghasilkan data duplikat. Periksa file sumber.")
    
    # Tampilkan hasil
    if 'df_hasil' in st.session_state:
        st.markdown("---")
        
        # Cek apakah ada duplikasi di hasil sebelum menampilkan tab
        current_has_duplicates = st.session_state.get('duplicate_status', {}).get('hasil_nip', False) or \
                                st.session_state.get('duplicate_status', {}).get('hasil_nik', False)
        
        # Tabs untuk membandingkan
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📊 Hasil Master Data Baru",
            "📋 Perbandingan dengan Master Lama",
            "🔄 Validasi Data Mentah vs BPMP",
            "⚖️ Validasi Data Mentah vs Master",
            "📈 Analisis Perubahan"
        ])
        
        with tab1:
            st.subheader("📋 Hasil Master Data Baru")
            
            df_display = st.session_state['df_hasil'].copy()
            
            # Cek duplikasi di tab ini juga
            if current_has_duplicates:
                st.error("""
                ⚠️ **PERINGATAN: DITEMUKAN DUPLIKASI DI DATA HASIL!**
                
                Data hasil mengandung NIP atau NIK yang duplikat. 
                **Tombol download dinonaktifkan** sampai duplikasi diperbaiki.
                
                **Langkah perbaikan:**
                1. Periksa Data Mentah dan Data BPMP sumber
                2. Pastikan NIP dan NPWP/NIK unik
                3. Proses ulang data setelah diperbaiki
                """)
            
            # Fungsi untuk styling
            def highlight_rows(row):
                color = row.get('Status_Color', '')
                if color == 'KUNING':
                    return ['background-color: #FFFF00'] * len(row)
                elif color == 'MERAH':
                    return ['background-color: #FF6B6B'] * len(row)
                elif color == 'ORANGE':
                    return ['background-color: #FFA500' if i == 2 else '' for i in range(len(row))]
                elif color == 'HIJAU':
                    return ['background-color: #90EE90'] * len(row)
                return [''] * len(row)
            
            # Tampilkan dengan styling
            df_show = df_display.drop(columns=['Status_Color'], errors='ignore')
            st.dataframe(df_show.style.apply(highlight_rows, axis=1), height=400)
            
            # Legend
            st.markdown("""
            **Keterangan Warna:**
            - 🟢 **Hijau**: Data baru (belum ada di master lama)
            - 🟡 **Kuning**: Data sudah ada di master lama (tidak berubah)
            - 🟠 **Orange**: Nama ada tapi data berubah
            - 🔴 **Merah**: Pegawai tidak aktif (ada di master lama tapi tidak ada di bulan ini)
            """)
            
            # Download button khusus untuk tab 1
            st.markdown("---")
            st.subheader("📥 Download Data Master Baru")
            
            # Pilihan format download
            download_format = st.radio(
                "Pilih format download:",
                ["Excel dengan warna", "Excel tanpa warna", "CSV"],
                horizontal=True,
                key="tab1_download"
            )
            
            # Nonaktifkan tombol jika ada duplikasi
            download_disabled = current_has_duplicates
            
            if download_format == "Excel dengan warna":
                excel_file = create_excel_with_colors(df_display)
                file_name = "master_data_pegawai.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif download_format == "Excel tanpa warna":
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_show.to_excel(writer, index=False, sheet_name='Master Data')
                output.seek(0)
                excel_file = output
                file_name = "master_data_pegawai_tanpa_warna.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:  # CSV
                csv_data = df_show.to_csv(index=False)
                excel_file = BytesIO(csv_data.encode())
                file_name = "master_data_pegawai.csv"
                mime_type = "text/csv"
            
            st.download_button(
                label=f"📥 Download {download_format}" + (" ⚠️ (Dinonaktifkan - Ada Duplikasi)" if download_disabled else ""),
                data=excel_file,
                file_name=file_name,
                mime=mime_type,
                key="download_master_baru",
                disabled=download_disabled
            )
            
            if download_disabled:
                st.warning("⚠️ Tombol download dinonaktifkan karena terdapat duplikasi data. Perbaiki duplikasi terlebih dahulu.")
            
            # Statistik
            st.markdown("---")
            st.subheader("📊 Statistik Master Data Baru")
            
            col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
            
            with col_stat1:
                total_aktif = len(df_display[df_display['AKTIF/TIDAK'] == 'AKTIF'])
                st.metric("👥 Total Aktif", total_aktif)
            
            with col_stat2:
                total_tidak_aktif = len(df_display[df_display['AKTIF/TIDAK'] == 'TIDAK'])
                st.metric("❌ Tidak Aktif", total_tidak_aktif)
            
            with col_stat3:
                total_baru = len(df_display[df_display['Status_Color'] == 'HIJAU'])
                st.metric("🆕 Data Baru", total_baru)
            
            with col_stat4:
                total_duplikasi = len(df_display[df_display['Status_Color'] == 'KUNING'])
                st.metric("⚠️ Data Tidak Berubah", total_duplikasi)
        
        with tab2:
            st.subheader("🔍 Perbandingan Master Lama vs Master Baru")
            
            if 'df_master_existing' in st.session_state and st.session_state['df_master_existing'] is not None:
                df_old = st.session_state['df_master_existing']
                df_new = st.session_state['df_hasil'].drop(columns=['Status_Color'], errors='ignore')
                
                # Cek duplikasi di data perbandingan
                df_nip_dup_new, _ = check_duplicates(df_new, 'NIP', 'Master Baru')
                df_nik_dup_new, _ = check_duplicates(df_new, 'NIK', 'Master Baru')
                
                if df_nip_dup_new is not None or df_nik_dup_new is not None:
                    st.error("⚠️ **DITEMUKAN DUPLIKASI DI DATA PERBANDINGAN!** Tombol download dinonaktifkan.")
                    comparison_has_duplicates = True
                else:
                    comparison_has_duplicates = False
                
                # Opsi filter
                st.markdown("### 🎯 Filter Perbandingan")
                col_filter1, col_filter2 = st.columns(2)
                
                with col_filter1:
                    show_option = st.radio(
                        "Tampilkan:",
                        ["Semua Data", "Hanya yang Berbeda", "Hanya yang Sama"],
                        horizontal=True,
                        key="tab2_filter"
                    )
                
                with col_filter2:
                    highlight_option = st.checkbox("Highlight Perbedaan per Kolom", value=True, key="tab2_highlight")
                
                st.markdown("---")
                
                # ===== PERUBAHAN: FUNGSI COMPARE ROWS DENGAN IGNORE KETERANGAN =====
                def compare_rows(row_old, row_new):
                    """Bandingkan dua row dan return dict perbedaan - IGNORE KETERANGAN"""
                    differences = {}
                    
                    for col in HEADERS_MASTER:
                        if col == 'No' or col == 'Status_Color' or col == 'Keterangan':
                            continue
                        
                        val_old = format_nilai_asli(row_old.get(col, '')) if row_old is not None else ''
                        val_new = format_nilai_asli(row_new.get(col, '')) if row_new is not None else ''
                        
                        if val_old != val_new:
                            differences[col] = {
                                'old': val_old,
                                'new': val_new
                            }
                    
                    return differences
                # ===== END PERUBAHAN =====
                
                # Buat dataframe perbandingan
                comparison_data = []
                
                for idx_new, row_new in df_new.iterrows():
                    nama_new = format_nilai_asli(row_new.get('Nama', ''))
                    nip_new = format_nilai_asli(row_new.get('NIP', ''))
                    
                    # Cari matching row di master lama
                    match_idx = fuzzy_match_row(nama_new, nip_new, df_old)
                    
                    if match_idx is not None:
                        row_old = df_old.iloc[match_idx]
                        differences = compare_rows(row_old, row_new)
                        
                        comparison_row = {
                            'Nama': nama_new,
                            'NIP': nip_new,
                            'Status': 'SAMA' if not differences else 'BERBEDA',
                            'Jumlah Perbedaan': len(differences),
                            'Kolom Berbeda': ', '.join(differences.keys()) if differences else '-',
                            'row_new': row_new,
                            'row_old': row_old,
                            'differences': differences
                        }
                    else:
                        comparison_row = {
                            'Nama': nama_new,
                            'NIP': nip_new,
                            'Status': 'BARU',
                            'Jumlah Perbedaan': 0,
                            'Kolom Berbeda': 'Data Baru (tidak ada di master lama)',
                            'row_new': row_new,
                            'row_old': None,
                            'differences': {}
                        }
                    
                    comparison_data.append(comparison_row)
                
                # Tambahkan data yang hilang (ada di master lama tapi tidak di master baru)
                for idx_old, row_old in df_old.iterrows():
                    nama_old = format_nilai_asli(row_old.get('Nama', ''))
                    nip_old = format_nilai_asli(row_old.get('NIP', ''))
                    
                    # Cek apakah kolom ada
                    nama_col = None
                    nip_col = None
                    for col in df_old.columns:
                        col_upper = str(col).upper()
                        if 'NAMA' in col_upper and nama_col is None:
                            nama_col = col
                        if 'NIP' in col_upper and nip_col is None:
                            nip_col = col
                    
                    if nama_col and nip_col:
                        nama_old = format_nilai_asli(row_old.get(nama_col, ''))
                        nip_old = format_nilai_asli(row_old.get(nip_col, ''))
                        
                        match_idx = fuzzy_match_row(nama_old, nip_old, df_new)
                        
                        if match_idx is None:
                            comparison_row = {
                                'Nama': nama_old,
                                'NIP': nip_old,
                                'Status': 'HILANG',
                                'Jumlah Perbedaan': 0,
                                'Kolom Berbeda': 'Tidak ada di master baru',
                                'row_new': None,
                                'row_old': row_old,
                                'differences': {}
                            }
                            comparison_data.append(comparison_row)
                
                df_comparison = pd.DataFrame(comparison_data)
                
                # Filter berdasarkan pilihan
                if show_option == "Hanya yang Berbeda":
                    df_comparison = df_comparison[df_comparison['Status'].isin(['BERBEDA', 'BARU', 'HILANG'])]
                elif show_option == "Hanya yang Sama":
                    df_comparison = df_comparison[df_comparison['Status'] == 'SAMA']
                
                # Tampilkan ringkasan
                st.markdown("### 📊 Ringkasan Perbandingan")
                col_summary1, col_summary2, col_summary3, col_summary4 = st.columns(4)
                
                with col_summary1:
                    total_sama = len([x for x in comparison_data if x['Status'] == 'SAMA'])
                    st.metric("✅ Data Sama", total_sama)
                
                with col_summary2:
                    total_berbeda = len([x for x in comparison_data if x['Status'] == 'BERBEDA'])
                    st.metric("⚠️ Data Berbeda", total_berbeda)
                
                with col_summary3:
                    total_baru = len([x for x in comparison_data if x['Status'] == 'BARU'])
                    st.metric("🆕 Data Baru", total_baru)
                
                with col_summary4:
                    total_hilang = len([x for x in comparison_data if x['Status'] == 'HILANG'])
                    st.metric("❌ Data Hilang", total_hilang)
                
                st.markdown("---")
                
                # Tampilkan tabel perbandingan ringkas
                st.markdown("### 📋 Tabel Ringkasan Perbandingan")
                
                def color_status(val):
                    if val == 'SAMA':
                        return 'background-color: #90EE90'
                    elif val == 'BERBEDA':
                        return 'background-color: #FFA500'
                    elif val == 'BARU':
                        return 'background-color: #87CEEB'
                    elif val == 'HILANG':
                        return 'background-color: #FF6B6B'
                    return ''
                
                df_summary = df_comparison[['Nama', 'NIP', 'Status', 'Jumlah Perbedaan', 'Kolom Berbeda']].copy()
                st.dataframe(
                    df_summary.style.applymap(color_status, subset=['Status']),
                    height=400
                )
                
                # Download button khusus untuk tab 2
                st.markdown("---")
                st.subheader("📥 Download Data Perbandingan")
                
                # Buat dataframe untuk download
                df_comparison_download = df_comparison[['Nama', 'NIP', 'Status', 'Jumlah Perbedaan', 'Kolom Berbeda']].copy()
                
                # Tambahkan detail perbedaan jika ada
                if highlight_option and 'row_old' in df_comparison.columns and 'row_new' in df_comparison.columns:
                    # Buat dataframe detail
                    detail_data = []
                    for idx, row in df_comparison.iterrows():
                        if row['Status'] == 'BERBEDA' and row['differences']:
                            for col, diff in row['differences'].items():
                                detail_data.append({
                                    'Nama': row['Nama'],
                                    'NIP': row['NIP'],
                                    'Kolom': col,
                                    'Nilai Master Lama': diff['old'],
                                    'Nilai Master Baru': diff['new']
                                })
                    
                    if detail_data:
                        df_detail = pd.DataFrame(detail_data)
                        
                        # Gabungkan dengan summary
                        output_comparison = BytesIO()
                        with pd.ExcelWriter(output_comparison, engine='openpyxl') as writer:
                            df_comparison_download.to_excel(writer, index=False, sheet_name='Ringkasan Perbandingan')
                            df_detail.to_excel(writer, index=False, sheet_name='Detail Perbedaan')
                        
                        file_name = "perbandingan_master_lama_baru_detil.xlsx"
                    else:
                        output_comparison = BytesIO()
                        with pd.ExcelWriter(output_comparison, engine='openpyxl') as writer:
                            df_comparison_download.to_excel(writer, index=False, sheet_name='Ringkasan Perbandingan')
                        
                        file_name = "perbandingan_master_lama_baru.xlsx"
                else:
                    output_comparison = BytesIO()
                    with pd.ExcelWriter(output_comparison, engine='openpyxl') as writer:
                        df_comparison_download.to_excel(writer, index=False, sheet_name='Ringkasan Perbandingan')
                    
                    file_name = "perbandingan_master_lama_baru.xlsx"
                
                output_comparison.seek(0)
                
                st.download_button(
                    label="📥 Download Hasil Perbandingan (Excel)" + (" ⚠️ (Dinonaktifkan - Ada Duplikasi)" if comparison_has_duplicates else ""),
                    data=output_comparison,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_perbandingan",
                    disabled=comparison_has_duplicates
                )
                
                if comparison_has_duplicates:
                    st.warning("⚠️ Tombol download dinonaktifkan karena terdapat duplikasi data di hasil perbandingan.")
                
                st.markdown("---")
                
                # Detail perbandingan per pegawai
                if highlight_option and not df_comparison.empty:
                    st.markdown("### 🔍 Detail Perbandingan per Pegawai")
                    
                    # Pilih pegawai untuk detail
                    pegawai_options = [f"{row['Nama']} ({row['NIP']}) - {row['Status']}"
                                      for _, row in df_comparison.iterrows()]
                    
                    if pegawai_options:
                        selected_pegawai = st.selectbox(
                            "Pilih pegawai untuk melihat detail perbandingan:",
                            pegawai_options,
                            key="tab2_select_pegawai"
                        )
                        
                        selected_idx = pegawai_options.index(selected_pegawai)
                        selected_row = comparison_data[selected_idx]
                        
                        st.markdown(f"#### 👤 {selected_row['Nama']} (NIP: {selected_row['NIP']})")
                        st.markdown(f"**Status:** `{selected_row['Status']}`")
                        
                        if selected_row['Status'] == 'BERBEDA':
                            st.warning(f"⚠️ Ditemukan **{selected_row['Jumlah Perbedaan']}** perbedaan")
                            
                            # Buat tabel perbandingan detail
                            detail_data = []
                            
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan']:  # Ignore Keterangan
                                    continue
                                
                                val_old = format_nilai_asli(selected_row['row_old'].get(col, '')) if selected_row['row_old'] is not None else '-'
                                val_new = format_nilai_asli(selected_row['row_new'].get(col, '')) if selected_row['row_new'] is not None else '-'
                                
                                is_different = (col in selected_row['differences'])
                                
                                detail_data.append({
                                    'Kolom': col,
                                    'Master Lama': val_old,
                                    'Master Baru': val_new,
                                    'Status': '❌ BERBEDA' if is_different else '✅ SAMA'
                                })
                            
                            df_detail = pd.DataFrame(detail_data)
                            
                            # Styling untuk highlight perbedaan
                            def highlight_diff(row):
                                if row['Status'] == '❌ BERBEDA':
                                    return ['background-color: #FFE4E1'] * len(row)
                                return [''] * len(row)
                            
                            st.dataframe(
                                df_detail.style.apply(highlight_diff, axis=1),
                                height=500
                            )
                        
                        elif selected_row['Status'] == 'SAMA':
                            st.success("✅ Semua data sama dengan master lama")
                            
                            # Tampilkan data
                            detail_data = []
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan']:  # Ignore Keterangan
                                    continue
                                val = format_nilai_asli(selected_row['row_new'].get(col, ''))
                                detail_data.append({
                                    'Kolom': col,
                                    'Nilai': val
                                })
                            
                            st.dataframe(pd.DataFrame(detail_data), height=400)
                        
                        elif selected_row['Status'] == 'BARU':
                            st.info("🆕 Data baru, tidak ada di master lama")
                            
                            detail_data = []
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan']:  # Ignore Keterangan
                                    continue
                                val = format_nilai_asli(selected_row['row_new'].get(col, ''))
                                detail_data.append({
                                    'Kolom': col,
                                    'Nilai': val
                                })
                            
                            st.dataframe(pd.DataFrame(detail_data), height=400)
                        
                        elif selected_row['Status'] == 'HILANG':
                            st.error("❌ Data tidak ada di master baru (tidak aktif)")
                            
                            detail_data = []
                            for col in HEADERS_MASTER:
                                if col in ['No', 'Status_Color', 'Keterangan']:  # Ignore Keterangan
                                    continue
                                val = format_nilai_asli(selected_row['row_old'].get(col, ''))
                                detail_data.append({
                                    'Kolom': col,
                                    'Nilai': val
                                })
                            
                            st.dataframe(pd.DataFrame(detail_data), height=400)
            
            else:
                st.info("ℹ️ Tidak ada Master Lama yang di-upload untuk dibandingkan")
        
        with tab3:
            st.subheader("🔄 Validasi Data Mentah vs Data BPMP")
            st.markdown("**Validasi kesesuaian antara Data Mentah (Primary) dengan Data BPMP**")
            
            # Ambil dataframe dari session state atau dari file yang diupload
            if 'df_mentah' in st.session_state and 'df_bpmp' in st.session_state:
                df_mentah = st.session_state['df_mentah']
                df_bpmp = st.session_state['df_bpmp']
                
                # Cek duplikasi di data untuk validasi
                df_nip_dup_mentah_valid, _ = check_duplicates(df_mentah, 'nip', 'Data Mentah (Validasi)')
                df_npwp_dup_mentah_valid, _ = check_duplicates(df_mentah, 'npwp', 'Data Mentah (Validasi)')
                
                # Cari kolom NPWP/NIK/TIN di BPMP
                nik_col_bpmp = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col_bpmp = col
                        break
                
                if nik_col_bpmp:
                    df_nik_dup_bpmp_valid, _ = check_duplicates(df_bpmp, nik_col_bpmp, 'Data BPMP (Validasi)')
                else:
                    df_nik_dup_bpmp_valid = None
                
                validation_has_duplicates = (df_nip_dup_mentah_valid is not None or 
                                           df_npwp_dup_mentah_valid is not None or 
                                           df_nik_dup_bpmp_valid is not None)
                
                if validation_has_duplicates:
                    st.error("""
                    ⚠️ **DITEMUKAN DUPLIKASI DI DATA VALIDASI!**
                    
                    **Tombol download dinonaktifkan** karena terdapat duplikasi di:
                    - Data Mentah (NIP atau NPWP duplikat)
                    - Data BPMP (NPWP/NIK duplikat)
                    
                    Perbaiki duplikasi terlebih dahulu sebelum mendownload hasil validasi.
                    """)
                
                # Tampilkan jumlah baris
                st.markdown("### 📊 Informasi Jumlah Data")
                col_info1, col_info2, col_info3 = st.columns(3)
                
                with col_info1:
                    st.metric("📄 Total Baris Data Mentah", len(df_mentah))
                
                with col_info2:
                    st.metric("📄 Total Baris Data BPMP", len(df_bpmp))
                
                with col_info3:
                    selisih = len(df_mentah) - len(df_bpmp)
                    st.metric("⚖️ Selisih", selisih,
                             delta=f"Data Mentah {'lebih banyak' if selisih > 0 else 'lebih sedikit' if selisih < 0 else 'sama'}")
                
                if selisih != 0:
                    st.warning(f"⚠️ Jumlah data tidak sama! Data Mentah: {len(df_mentah)}, Data BPMP: {len(df_bpmp)}")
                else:
                    st.success("✅ Jumlah data sama antara Data Mentah dan BPMP")
                
                st.markdown("---")
                
                # ===== PERUBAHAN: CARI KOLOM UNTUK PERBANDINGAN TAMBAHAN =====
                # Cari kolom di Data Mentah untuk perbandingan
                bulan_col_mentah = None
                tahun_col_mentah = None
                gaji_kotor_col_mentah = None
                kdkawin_col_mentah = None
                
                for col in df_mentah.columns:
                    col_lower = str(col).lower()
                    if 'bulan' in col_lower and bulan_col_mentah is None:
                        bulan_col_mentah = col
                    elif 'tahun' in col_lower and tahun_col_mentah is None:
                        tahun_col_mentah = col
                    elif 'gajikotor' in col_lower.replace(' ', '') and gaji_kotor_col_mentah is None:
                        gaji_kotor_col_mentah = col
                    elif 'kdkawin' in col_lower and kdkawin_col_mentah is None:
                        kdkawin_col_mentah = col
                
                # Cari kolom di Data BPMP untuk perbandingan
                masa_pajak_col_bpmp = None
                tahun_pajak_col_bpmp = None
                penghasilan_kotor_col_bpmp = None
                status_col_bpmp = None
                
                for col in df_bpmp.columns:
                    col_lower = str(col).lower()
                    if 'masa' in col_lower and 'pajak' in col_lower and masa_pajak_col_bpmp is None:
                        masa_pajak_col_bpmp = col
                    elif 'tahun' in col_lower and 'pajak' in col_lower and tahun_pajak_col_bpmp is None:
                        tahun_pajak_col_bpmp = col
                    elif 'penghasilan' in col_lower and 'kotor' in col_lower and penghasilan_kotor_col_bpmp is None:
                        penghasilan_kotor_col_bpmp = col
                    elif 'status' in col_lower and 'pegawai' not in col_lower and status_col_bpmp is None:
                        status_col_bpmp = col
                # ===== END PERUBAHAN =====
                
                # Proses validasi
                st.markdown("### 🔍 Hasil Validasi NIP, NPWP, dan Data Lainnya")
                
                validation_data = []
                
                # Cari kolom NPWP/NIK/TIN di BPMP
                nik_col_bpmp = None
                for col in df_bpmp.columns:
                    if 'NPWP' in col.upper() or 'NIK' in col.upper() or 'TIN' in col.upper():
                        nik_col_bpmp = col
                        break
                
                # ===== PERUBAHAN PENTING: BUAT DICTIONARY UNTUK MAPPING BPMP BERDASARKAN NIK =====
                # Buat mapping dari NIK ke baris BPMP untuk pencarian yang lebih cepat
                bpmp_mapping = {}
                if nik_col_bpmp:
                    for idx_bpmp, row_bpmp in df_bpmp.iterrows():
                        nik_bpmp = format_nilai_asli(row_bpmp.get(nik_col_bpmp, ''))
                        if nik_bpmp:
                            # Simpan row dan data tambahan
                            bpmp_data = {
                                'row': row_bpmp,
                                'masa_pajak': format_nilai_asli(row_bpmp.get(masa_pajak_col_bpmp, '')) if masa_pajak_col_bpmp else '',
                                'tahun_pajak': format_nilai_asli(row_bpmp.get(tahun_pajak_col_bpmp, '')) if tahun_pajak_col_bpmp else '',
                                'penghasilan_kotor': format_nilai_asli(row_bpmp.get(penghasilan_kotor_col_bpmp, '')) if penghasilan_kotor_col_bpmp else '',
                                'status_bpmp': format_nilai_asli(row_bpmp.get(status_col_bpmp, '')) if status_col_bpmp else ''
                            }
                            bpmp_mapping[nik_bpmp] = bpmp_data
                # ===== END PERUBAHAN =====
                
                for idx_mentah, row_mentah in df_mentah.iterrows():
                    nip_mentah = format_nilai_asli(row_mentah.get('nip', '')) if 'nip' in df_mentah.columns else ''
                    npwp_mentah = format_nilai_asli(row_mentah.get('npwp', '')) if 'npwp' in df_mentah.columns else ''
                    nama_mentah = format_nilai_asli(row_mentah.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else ''
                    
                    # ===== PERUBAHAN: AMBIL DATA TAMBAHAN DARI MENTAH =====
                    bulan_mentah = format_nilai_asli(row_mentah.get(bulan_col_mentah, '')) if bulan_col_mentah else ''
                    tahun_mentah = format_nilai_asli(row_mentah.get(tahun_col_mentah, '')) if tahun_col_mentah else ''
                    gaji_kotor_mentah = format_nilai_asli(row_mentah.get(gaji_kotor_col_mentah, '')) if gaji_kotor_col_mentah else ''
                    kdkawin_mentah = format_nilai_asli(row_mentah.get(kdkawin_col_mentah, '')) if kdkawin_col_mentah else ''
                    status_kawin_mentah = konversi_status(kdkawin_mentah) if kdkawin_col_mentah else ''
                    # ===== END PERUBAHAN =====
                    
                    # Cari di BPMP berdasarkan NPWP menggunakan mapping
                    found_in_bpmp = False
                    npwp_bpmp = '-'
                    match_score = 0
                    rekomendasi = ''
                    
                    # ===== PERUBAHAN: INISIALISASI DATA TAMBAHAN BPMP =====
                    masa_pajak_bpmp = ''
                    tahun_pajak_bpmp = ''
                    penghasilan_kotor_bpmp = ''
                    status_bpmp_value = ''
                    status_bulan = 'TIDAK ADA DATA'
                    status_tahun = 'TIDAK ADA DATA'
                    status_gaji = 'TIDAK ADA DATA'
                    status_kawin = 'TIDAK ADA DATA'
                    # ===== END PERUBAHAN =====
                    
                    if nik_col_bpmp and npwp_mentah and npwp_mentah in bpmp_mapping:
                        # Exact match ditemukan
                        found_in_bpmp = True
                        npwp_bpmp = npwp_mentah
                        match_score = 100
                        
                        bpmp_data = bpmp_mapping[npwp_mentah]
                        masa_pajak_bpmp = bpmp_data['masa_pajak']
                        tahun_pajak_bpmp = bpmp_data['tahun_pajak']
                        penghasilan_kotor_bpmp = bpmp_data['penghasilan_kotor']
                        status_bpmp_value = bpmp_data['status_bpmp']
                        
                        # ===== PERUBAHAN: BANDINGKAN DATA TAMBAHAN =====
                        # Bandingkan data tambahan
                        if bulan_mentah and masa_pajak_bpmp:
                            try:
                                # Coba konversi ke angka untuk perbandingan
                                bulan_mentah_num = int(bulan_mentah) if bulan_mentah.isdigit() else 0
                                masa_pajak_num = int(masa_pajak_bpmp) if masa_pajak_bpmp.isdigit() else 0
                                status_bulan = 'SESUAI' if bulan_mentah_num == masa_pajak_num else 'TIDAK SESUAI'
                            except:
                                status_bulan = 'TIDAK SESUAI' if bulan_mentah != masa_pajak_bpmp else 'SESUAI'
                        
                        if tahun_mentah and tahun_pajak_bpmp:
                            try:
                                tahun_mentah_num = int(tahun_mentah) if tahun_mentah.isdigit() else 0
                                tahun_pajak_num = int(tahun_pajak_bpmp) if tahun_pajak_bpmp.isdigit() else 0
                                status_tahun = 'SESUAI' if tahun_mentah_num == tahun_pajak_num else 'TIDAK SESUAI'
                            except:
                                status_tahun = 'TIDAK SESUAI' if tahun_mentah != tahun_pajak_bpmp else 'SESUAI'
                        
                        if gaji_kotor_mentah and penghasilan_kotor_bpmp:
                            try:
                                # Bersihkan format angka
                                gaji_mentah_clean = gaji_kotor_mentah.replace('.', '').replace(',', '.')
                                gaji_bpmp_clean = penghasilan_kotor_bpmp.replace('.', '').replace(',', '.')
                                
                                gaji_mentah_num = float(gaji_mentah_clean)
                                gaji_bpmp_num = float(gaji_bpmp_clean)
                                
                                # Toleransi 1 untuk perbedaan pembulatan
                                status_gaji = 'SESUAI' if abs(gaji_mentah_num - gaji_bpmp_num) <= 1 else 'TIDAK SESUAI'
                            except:
                                status_gaji = 'TIDAK SESUAI' if gaji_kotor_mentah != penghasilan_kotor_bpmp else 'SESUAI'
                        
                        # ===== TAMBAHAN: PERBANDINGAN STATUS KAWIN =====
                        if status_kawin_mentah and status_bpmp_value:
                            # Normalisasi nilai untuk perbandingan
                            status_kawin_mentah_clean = status_kawin_mentah.strip().upper()
                            status_bpmp_clean = status_bpmp_value.strip().upper()
                            
                            # Mapping untuk perbandingan fleksibel
                            status_mapping = {
                                'TK/0': ['TK', 'TK/0', 'TK 0', 'TIDAK KAWIN'],
                                'TK/1': ['TK/1', 'TK 1', 'TIDAK KAWIN 1'],
                                'TK/2': ['TK/2', 'TK 2', 'TIDAK KAWIN 2'],
                                'K/0': ['K', 'K/0', 'K 0', 'KAWIN', 'KAWIN 0'],
                                'K/1': ['K/1', 'K 1', 'KAWIN 1'],
                                'K/2': ['K/2', 'K 2', 'KAWIN 2']
                            }
                            
                            # Cek apakah status cocok
                            match_found = False
                            if status_kawin_mentah_clean == status_bpmp_clean:
                                match_found = True
                            else:
                                # Cek mapping fleksibel
                                for key, values in status_mapping.items():
                                    if status_kawin_mentah_clean == key and status_bpmp_clean in values:
                                        match_found = True
                                        break
                                    elif status_bpmp_clean == key and status_kawin_mentah_clean in values:
                                        match_found = True
                                        break
                            
                            status_kawin = 'SESUAI' if match_found else 'TIDAK SESUAI'
                        elif status_kawin_mentah and not status_bpmp_value:
                            status_kawin = 'TIDAK ADA DATA BPMP'
                        elif not status_kawin_mentah and status_bpmp_value:
                            status_kawin = 'TIDAK ADA DATA MENTAH'
                        # ===== END TAMBAHAN =====
                        # ===== END PERUBAHAN =====
                    
                    elif nik_col_bpmp and npwp_mentah:
                        # Coba fuzzy matching jika exact match tidak ditemukan
                        best_score = 0
                        best_nik = None
                        
                        for nik in bpmp_mapping.keys():
                            score = fuzz.ratio(npwp_mentah, nik)
                            if score > best_score and score >= 80:
                                best_score = score
                                best_nik = nik
                        
                        if best_nik:
                            found_in_bpmp = True
                            npwp_bpmp = best_nik
                            match_score = best_score
                            
                            bpmp_data = bpmp_mapping[best_nik]
                            masa_pajak_bpmp = bpmp_data['masa_pajak']
                            tahun_pajak_bpmp = bpmp_data['tahun_pajak']
                            penghasilan_kotor_bpmp = bpmp_data['penghasilan_kotor']
                            status_bpmp_value = bpmp_data['status_bpmp']
                            
                            # Bandingkan data tambahan untuk fuzzy match juga
                            if bulan_mentah and masa_pajak_bpmp:
                                try:
                                    bulan_mentah_num = int(bulan_mentah) if bulan_mentah.isdigit() else 0
                                    masa_pajak_num = int(masa_pajak_bpmp) if masa_pajak_bpmp.isdigit() else 0
                                    status_bulan = 'SESUAI' if bulan_mentah_num == masa_pajak_num else 'TIDAK SESUAI'
                                except:
                                    status_bulan = 'TIDAK SESUAI' if bulan_mentah != masa_pajak_bpmp else 'SESUAI'
                            
                            if tahun_mentah and tahun_pajak_bpmp:
                                try:
                                    tahun_mentah_num = int(tahun_mentah) if tahun_mentah.isdigit() else 0
                                    tahun_pajak_num = int(tahun_pajak_bpmp) if tahun_pajak_bpmp.isdigit() else 0
                                    status_tahun = 'SESUAI' if tahun_mentah_num == tahun_pajak_num else 'TIDAK SESUAI'
                                except:
                                    status_tahun = 'TIDAK SESUAI' if tahun_mentah != tahun_pajak_bpmp else 'SESUAI'
                            
                            if gaji_kotor_mentah and penghasilan_kotor_bpmp:
                                try:
                                    gaji_mentah_clean = gaji_kotor_mentah.replace('.', '').replace(',', '.')
                                    gaji_bpmp_clean = penghasilan_kotor_bpmp.replace('.', '').replace(',', '.')
                                    
                                    gaji_mentah_num = float(gaji_mentah_clean)
                                    gaji_bpmp_num = float(gaji_bpmp_clean)
                                    
                                    status_gaji = 'SESUAI' if abs(gaji_mentah_num - gaji_bpmp_num) <= 1 else 'TIDAK SESUAI'
                                except:
                                    status_gaji = 'TIDAK SESUAI' if gaji_kotor_mentah != penghasilan_kotor_bpmp else 'SESUAI'
                            
                            # ===== TAMBAHAN: PERBANDINGAN STATUS KAWIN UNTUK FUZZY MATCH =====
                            if status_kawin_mentah and status_bpmp_value:
                                # Normalisasi nilai untuk perbandingan
                                status_kawin_mentah_clean = status_kawin_mentah.strip().upper()
                                status_bpmp_clean = status_bpmp_value.strip().upper()
                                
                                # Mapping untuk perbandingan fleksibel
                                status_mapping = {
                                    'TK/0': ['TK', 'TK/0', 'TK 0', 'TIDAK KAWIN'],
                                    'TK/1': ['TK/1', 'TK 1', 'TIDAK KAWIN 1'],
                                    'TK/2': ['TK/2', 'TK 2', 'TIDAK KAWIN 2'],
                                    'K/0': ['K', 'K/0', 'K 0', 'KAWIN', 'KAWIN 0'],
                                    'K/1': ['K/1', 'K 1', 'KAWIN 1'],
                                    'K/2': ['K/2', 'K 2', 'KAWIN 2']
                                }
                                
                                # Cek apakah status cocok
                                match_found = False
                                if status_kawin_mentah_clean == status_bpmp_clean:
                                    match_found = True
                                else:
                                    # Cek mapping fleksibel
                                    for key, values in status_mapping.items():
                                        if status_kawin_mentah_clean == key and status_bpmp_clean in values:
                                            match_found = True
                                            break
                                        elif status_bpmp_clean == key and status_kawin_mentah_clean in values:
                                            match_found = True
                                            break
                                
                                status_kawin = 'SESUAI' if match_found else 'TIDAK SESUAI'
                            elif status_kawin_mentah and not status_bpmp_value:
                                status_kawin = 'TIDAK ADA DATA BPMP'
                            elif not status_kawin_mentah and status_bpmp_value:
                                status_kawin = 'TIDAK ADA DATA MENTAH'
                            # ===== END TAMBAHAN =====
                    
                    # Tentukan status utama
                    if not nip_mentah and not npwp_mentah:
                        status_utama = 'DATA KOSONG'
                        rekomendasi = 'Lengkapi NIP dan NPWP di Data Mentah'
                    elif not nip_mentah:
                        status_utama = 'NIP KOSONG'
                        rekomendasi = 'Lengkapi NIP di Data Mentah'
                    elif not npwp_mentah:
                        status_utama = 'NPWP KOSONG'
                        rekomendasi = 'Lengkapi NPWP di Data Mentah'
                    elif found_in_bpmp:
                        status_utama = 'VALID'
                        rekomendasi = f'Data lengkap dan cocok (Match: {match_score}%)'
                    else:
                        status_utama = 'TIDAK ADA DI BPMP'
                        rekomendasi = f'Tambahkan pegawai "{nama_mentah}" dengan NPWP {npwp_mentah} ke Data BPMP'
                    
                    validation_data.append({
                        'No': idx_mentah + 1,
                        'Nama': nama_mentah,
                        'NIP (Data Mentah)': nip_mentah if nip_mentah else '❌ KOSONG',
                        'NPWP (Data Mentah)': npwp_mentah if npwp_mentah else '❌ KOSONG',
                        'NPWP (Data BPMP)': npwp_bpmp if found_in_bpmp else '❌ TIDAK ADA',
                        # ===== PERUBAHAN: TAMBAHAN DATA PERBANDINGAN =====
                        'Bulan (Mentah)': bulan_mentah if bulan_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Masa Pajak (BPMP)': masa_pajak_bpmp if found_in_bpmp and masa_pajak_col_bpmp else '❌ TIDAK ADA',
                        'Status Bulan': status_bulan,
                        'Tahun (Mentah)': tahun_mentah if tahun_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Tahun Pajak (BPMP)': tahun_pajak_bpmp if found_in_bpmp and tahun_pajak_col_bpmp else '❌ TIDAK ADA',
                        'Status Tahun': status_tahun,
                        'GajiKotor (Mentah)': gaji_kotor_mentah if gaji_kotor_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Penghasilan Kotor (BPMP)': penghasilan_kotor_bpmp if found_in_bpmp and penghasilan_kotor_col_bpmp else '❌ TIDAK ADA',
                        'Status Gaji Kotor': status_gaji,
                        # ===== TAMBAHAN: PERBANDINGAN STATUS KAWIN =====
                        'KDKAWIN (Mentah)': kdkawin_mentah if kdkawin_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Status Kawin (Mentah)': status_kawin_mentah if kdkawin_col_mentah else '❌ TIDAK ADA KOLOM',
                        'Status (BPMP)': status_bpmp_value if found_in_bpmp and status_col_bpmp else '❌ TIDAK ADA',
                        'Status Perbandingan Kawin': status_kawin,
                        # ===== END TAMBAHAN =====
                        'Status': status_utama,
                        'Rekomendasi': rekomendasi
                    })
                
                df_validation = pd.DataFrame(validation_data)
                
                # Simpan ke session state untuk download
                st.session_state['df_validation_bpmp'] = df_validation
                
                # Statistik validasi utama
                st.markdown("### 📊 Ringkasan Validasi Utama")
                col_val1, col_val2, col_val3, col_val4 = st.columns(4)
                
                with col_val1:
                    total_valid = len(df_validation[df_validation['Status'] == 'VALID'])
                    st.metric("✅ Data Valid", total_valid)
                
                with col_val2:
                    total_tidak_ada = len(df_validation[df_validation['Status'] == 'TIDAK ADA DI BPMP'])
                    st.metric("❌ Tidak Ada di BPMP", total_tidak_ada)
                
                with col_val3:
                    total_kosong = len(df_validation[df_validation['Status'].str.contains('KOSONG')])
                    st.metric("⚠️ Data Kosong", total_kosong)
                
                with col_val4:
                    persentase_valid = round((total_valid / len(df_validation) * 100), 1) if len(df_validation) > 0 else 0
                    st.metric("📈 Persentase Valid", f"{persentase_valid}%")
                
                # ===== PERUBAHAN: STATISTIK VALIDASI TAMBAHAN =====
                st.markdown("### 📊 Ringkasan Validasi Data Tambahan")
                col_val5, col_val6, col_val7, col_val8 = st.columns(4)
                
                with col_val5:
                    total_bulan_sesuai = len(df_validation[df_validation['Status Bulan'] == 'SESUAI'])
                    total_bulan_data = len(df_validation[df_validation['Status Bulan'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_bulan = round((total_bulan_sesuai / total_bulan_data * 100), 1) if total_bulan_data > 0 else 0
                    st.metric("📅 Bulan Sesuai", f"{total_bulan_sesuai}/{total_bulan_data}", f"{persentase_bulan}%")
                
                with col_val6:
                    total_tahun_sesuai = len(df_validation[df_validation['Status Tahun'] == 'SESUAI'])
                    total_tahun_data = len(df_validation[df_validation['Status Tahun'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_tahun = round((total_tahun_sesuai / total_tahun_data * 100), 1) if total_tahun_data > 0 else 0
                    st.metric("📅 Tahun Sesuai", f"{total_tahun_sesuai}/{total_tahun_data}", f"{persentase_tahun}%")
                
                with col_val7:
                    total_gaji_sesuai = len(df_validation[df_validation['Status Gaji Kotor'] == 'SESUAI'])
                    total_gaji_data = len(df_validation[df_validation['Status Gaji Kotor'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_gaji = round((total_gaji_sesuai / total_gaji_data * 100), 1) if total_gaji_data > 0 else 0
                    st.metric("💰 Gaji Sesuai", f"{total_gaji_sesuai}/{total_gaji_data}", f"{persentase_gaji}%")
                
                with col_val8:
                    total_kawin_sesuai = len(df_validation[df_validation['Status Perbandingan Kawin'] == 'SESUAI'])
                    total_kawin_data = len(df_validation[df_validation['Status Perbandingan Kawin'].isin(['SESUAI', 'TIDAK SESUAI'])])
                    persentase_kawin = round((total_kawin_sesuai / total_kawin_data * 100), 1) if total_kawin_data > 0 else 0
                    st.metric("💍 Status Kawin Sesuai", f"{total_kawin_sesuai}/{total_kawin_data}", f"{persentase_kawin}%")
                # ===== END PERUBAHAN =====
                
                st.markdown("---")
                
                # Filter tampilan
                col_filter_val1, col_filter_val2 = st.columns(2)
                
                with col_filter_val1:
                    filter_status = st.selectbox(
                        "Filter Status Utama:",
                        ["Semua", "VALID", "TIDAK ADA DI BPMP", "NIP KOSONG", "NPWP KOSONG", "DATA KOSONG"],
                        key="tab3_filter"
                    )
                
                with col_filter_val2:
                    show_rekomendasi = st.checkbox("Tampilkan Kolom Rekomendasi", value=True, key="tab3_rekomendasi")
                
                # Apply filter
                if filter_status != "Semua":
                    df_validation_display = df_validation[df_validation['Status'] == filter_status]
                else:
                    df_validation_display = df_validation
                
                # ===== PERUBAHAN PENTING: SELALU TAMPILKAN SEMUA DATA PERBANDINGAN =====
                # Tentukan kolom yang akan ditampilkan - SELALU sertakan semua kolom perbandingan
                kolom_tampil = [
                    'No', 'Nama', 
                    'NIP (Data Mentah)', 'NPWP (Data Mentah)', 'NPWP (Data BPMP)',
                    'Bulan (Mentah)', 'Masa Pajak (BPMP)', 'Status Bulan',
                    'Tahun (Mentah)', 'Tahun Pajak (BPMP)', 'Status Tahun',
                    'GajiKotor (Mentah)', 'Penghasilan Kotor (BPMP)', 'Status Gaji Kotor',
                    # ===== TAMBAHAN: KOLOM STATUS KAWIN =====
                    'KDKAWIN (Mentah)', 'Status Kawin (Mentah)', 'Status (BPMP)', 'Status Perbandingan Kawin',
                    # ===== END TAMBAHAN =====
                    'Status'
                ]
                
                if show_rekomendasi:
                    kolom_tampil.append('Rekomendasi')
                
                df_validation_display = df_validation_display[kolom_tampil]
                # ===== END PERUBAHAN PENTING =====
                
                st.markdown("### 📋 Tabel Validasi Lengkap")
                st.info("✅ **Semua data perbandingan (Bulan, Tahun, Gaji, Status Kawin) ditampilkan untuk memudahkan validasi**")
                
                # Styling function
                def color_validation_status(row):
                    colors = [''] * len(row)
                    status = row['Status'] if 'Status' in row.index else ''
                    
                    if status == 'VALID':
                        colors = ['background-color: #90EE90'] * len(row)
                    elif status == 'TIDAK ADA DI BPMP':
                        colors = ['background-color: #FF6B6B'] * len(row)
                    elif 'KOSONG' in status:
                        colors = ['background-color: #FFD700'] * len(row)
                    
                    # Warna untuk status tambahan yang tidak sesuai
                    for i, col_name in enumerate(row.index):
                        if 'Status Bulan' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                        elif 'Status Tahun' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                        elif 'Status Gaji Kotor' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                        elif 'Status Perbandingan Kawin' in col_name and row[col_name] == 'TIDAK SESUAI':
                            colors[i] = 'background-color: #FFA07A'
                    
                    return colors
                
                # Tampilkan tabel dengan scroll horizontal agar semua kolom terlihat
                st.dataframe(
                    df_validation_display.style.apply(color_validation_status, axis=1),
                    height=500,
                    use_container_width=True
                )
                
                # Download button khusus untuk tab 3
                st.markdown("---")
                st.subheader("📥 Download Hasil Validasi Data Mentah vs BPMP")
                
                # Buat Excel untuk download
                output_validation = BytesIO()
                with pd.ExcelWriter(output_validation, engine='openpyxl') as writer:
                    # Sheet 1: Data lengkap
                    df_validation.to_excel(writer, index=False, sheet_name='Validasi Lengkap')
                    
                    # Sheet 2: Data yang perlu perbaikan
                    df_perlu_perbaikan = df_validation[
                        (df_validation['Status'] == 'TIDAK ADA DI BPMP') |
                        (df_validation['Status'].str.contains('KOSONG'))
                    ]
                    df_perlu_perbaikan.to_excel(writer, index=False, sheet_name='Perlu Perbaikan')
                    
                    # Sheet 3: Data yang tidak sesuai (bulan, tahun, gaji, status kawin)
                    df_tidak_sesuai = df_validation[
                        (df_validation['Status Bulan'] == 'TIDAK SESUAI') |
                        (df_validation['Status Tahun'] == 'TIDAK SESUAI') |
                        (df_validation['Status Gaji Kotor'] == 'TIDAK SESUAI') |
                        (df_validation['Status Perbandingan Kawin'] == 'TIDAK SESUAI')
                    ]
                    df_tidak_sesuai.to_excel(writer, index=False, sheet_name='Data Tidak Sesuai')
                    
                    workbook = writer.book
                    
                    # Warna untuk sheet Perlu Perbaikan
                    if not df_perlu_perbaikan.empty:
                        worksheet = writer.sheets['Perlu Perbaikan']
                        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        yellow_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        
                        for idx, row in df_perlu_perbaikan.iterrows():
                            excel_row = list(df_perlu_perbaikan.index).index(idx) + 2
                            
                            if row['Status'] == 'TIDAK ADA DI BPMP':
                                for col in range(1, len(df_perlu_perbaikan.columns) + 1):
                                    worksheet.cell(row=excel_row, column=col).fill = red_fill
                            elif 'KOSONG' in row['Status']:
                                for col in range(1, len(df_perlu_perbaikan.columns) + 1):
                                    worksheet.cell(row=excel_row, column=col).fill = yellow_fill
                    
                    # Warna untuk sheet Data Tidak Sesuai
                    if not df_tidak_sesuai.empty:
                        worksheet = writer.sheets['Data Tidak Sesuai']
                        orange_fill = PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')
                        
                        for idx, row in df_tidak_sesuai.iterrows():
                            excel_row = list(df_tidak_sesuai.index).index(idx) + 2
                            
                            # Highlight kolom yang tidak sesuai
                            for col_idx, col_name in enumerate(df_tidak_sesuai.columns, start=1):
                                if 'Status Bulan' in col_name and row[col_name] == 'TIDAK SESUAI':
                                    worksheet.cell(row=excel_row, column=col_idx).fill = orange_fill
                                elif 'Status Tahun' in col_name and row[col_name] == 'TIDAK SESUAI':
                                    worksheet.cell(row=excel_row, column=col_idx).fill = orange_fill
                                elif 'Status Gaji Kotor' in col_name and row[col_name] == 'TIDAK SESUAI':
                                    worksheet.cell(row=excel_row, column=col_idx).fill = orange_fill
                                elif 'Status Perbandingan Kawin' in col_name and row[col_name] == 'TIDAK SESUAI':
                                    worksheet.cell(row=excel_row, column=col_idx).fill = orange_fill
                
                output_validation.seek(0)
                
                st.download_button(
                    label="📥 Download Hasil Validasi (Excel)" + (" ⚠️ (Dinonaktifkan - Ada Duplikasi)" if validation_has_duplicates else ""),
                    data=output_validation,
                    file_name="validasi_mentah_vs_bpmp.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_validasi_bpmp",
                    disabled=validation_has_duplicates
                )
                
                if validation_has_duplicates:
                    st.warning("⚠️ Tombol download dinonaktifkan karena terdapat duplikasi data di Data Mentah atau Data BPMP.")
                
                # Legend
                st.markdown("""
                **Keterangan Warna:**
                - 🟢 **Hijau**: Data VALID (NIP dan NPWP ada di kedua file)
                - 🔴 **Merah**: Data TIDAK ADA DI BPMP (ada di Data Mentah tapi tidak ada di BPMP)
                - 🟡 **Kuning**: Data KOSONG (NIP atau NPWP kosong di Data Mentah)
                - 🟠 **Salmon**: Data tambahan TIDAK SESUAI (bulan, tahun, gaji, atau status kawin tidak cocok)
                
                **Kolom Perbandingan:**
                - **Bulan (Mentah)** vs **Masa Pajak (BPMP)**
                - **Tahun (Mentah)** vs **Tahun Pajak (BPMP)**
                - **GajiKotor (Mentah)** vs **Penghasilan Kotor (BPMP)**
                - **Status Kawin (Mentah)** vs **Status (BPMP)**
                """)
                
                st.markdown("---")
                
                # Download rekomendasi perbaikan
                if total_tidak_ada > 0 or total_kosong > 0:
                    st.markdown("### 📥 Download Data yang Perlu Diperbaiki")
                    
                    df_perlu_perbaikan = df_validation[
                        (df_validation['Status'] == 'TIDAK ADA DI BPMP') |
                        (df_validation['Status'].str.contains('KOSONG'))
                    ]
                    
                    if not df_perlu_perbaikan.empty:
                        st.info(f"💡 Total **{len(df_perlu_perbaikan)}** data yang perlu diperbaiki")
                else:
                    st.success("🎉 Semua data valid! Tidak ada yang perlu diperbaiki.")
            
            else:
                st.warning("⚠️ Pastikan Data Mentah dan Data BPMP sudah di-upload untuk melakukan validasi")
        
        with tab4:
            st.subheader("⚖️ Validasi Data Mentah vs Data Master")
            st.markdown("**Data Master sebagai PRIMARY (Data Benar & Akurat) - Validasi Data Mentah**")
            
            if 'df_mentah' in st.session_state and 'df_master_existing' in st.session_state and st.session_state['df_master_existing'] is not None:
                df_mentah = st.session_state['df_mentah']
                df_master = st.session_state['df_master_existing']
                
                # Cek duplikasi di data untuk validasi
                df_nip_dup_mentah_master, _ = check_duplicates(df_mentah, 'nip', 'Data Mentah (Validasi Master)')
                df_npwp_dup_mentah_master, _ = check_duplicates(df_mentah, 'npwp', 'Data Mentah (Validasi Master)')
                
                # Cari kolom NIP dan NIK di Master
                nip_col_master = None
                nik_col_master = None
                for col in df_master.columns:
                    if 'NIP' in str(col).upper():
                        nip_col_master = col
                    if 'NIK' in str(col).upper() and 'PENERIMA' not in str(col).upper():
                        nik_col_master = col
                
                df_nip_dup_master_valid, _ = check_duplicates(df_master, nip_col_master, 'Master (Validasi)') if nip_col_master else (None, [])
                df_nik_dup_master_valid, _ = check_duplicates(df_master, nik_col_master, 'Master (Validasi)') if nik_col_master else (None, [])
                
                master_validation_has_duplicates = (df_nip_dup_mentah_master is not None or 
                                                  df_npwp_dup_mentah_master is not None or
                                                  df_nip_dup_master_valid is not None or
                                                  df_nik_dup_master_valid is not None)
                
                if master_validation_has_duplicates:
                    st.error("""
                    ⚠️ **DITEMUKAN DUPLIKASI DI DATA VALIDASI MASTER!**
                    
                    **Tombol download dinonaktifkan** karena terdapat duplikasi di:
                    - Data Mentah (NIP atau NPWP duplikat)
                    - Data Master (NIP atau NIK duplikat)
                    
                    Perbaiki duplikasi terlebih dahulu sebelum mendownload hasil validasi.
                    """)
                
                # VALIDASI PRIMARY KEY: Cek NIP duplikat di Data Mentah
                st.markdown("### 🔑 Validasi Primary Key (NIP)")
                
                if 'nip' in df_mentah.columns:
                    nip_series = df_mentah['nip'].apply(format_nilai_asli)
                    nip_duplicates = nip_series[nip_series.duplicated(keep=False)]
                    
                    if len(nip_duplicates) > 0:
                        st.error(f"❌ **PERINGATAN: Ditemukan {len(nip_duplicates)} NIP yang duplikat di Data Mentah!**")
                        st.warning("⚠️ NIP adalah Primary Key dan harus UNIQUE. Data dengan NIP duplikat tidak dapat diproses dengan benar.")
                        
                        # Tampilkan NIP yang duplikat
                        duplicate_nips = nip_duplicates.unique()
                        duplicate_data = []
                        
                        for dup_nip in duplicate_nips:
                            if dup_nip and dup_nip != 'nan':
                                dup_rows = df_mentah[nip_series == dup_nip]
                                for idx, row in dup_rows.iterrows():
                                    duplicate_data.append({
                                        'Baris': idx + 2, # +2 karena header dan index mulai dari 0
                                        'NIP (DUPLIKAT)': dup_nip,
                                        'Nama': format_nilai_asli(row.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else '-',
                                        'KDGOL': format_nilai_asli(row.get('kdgol', '')) if 'kdgol' in df_mentah.columns else '-',
                                        'Jumlah Duplikat': len(dup_rows)
                                    })
                        
                        df_duplicates = pd.DataFrame(duplicate_data)
                        
                        st.markdown("#### 📋 Daftar NIP yang Duplikat:")
                        st.dataframe(
                            df_duplicates.style.applymap(lambda x: 'background-color: #FF6B6B'),
                            height=300
                        )
                        
                        # Download duplikat
                        output_duplicates = BytesIO()
                        with pd.ExcelWriter(output_duplicates, engine='openpyxl') as writer:
                            df_duplicates.to_excel(writer, index=False, sheet_name='NIP Duplikat')
                            
                            workbook = writer.book
                            worksheet = writer.sheets['NIP Duplikat']
                            
                            red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                            
                            for row_idx in range(2, len(df_duplicates) + 2):
                                for col_idx in range(1, len(df_duplicates.columns) + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill
                        
                        output_duplicates.seek(0)
                        
                        st.download_button(
                            label="📥 Download Daftar NIP Duplikat (Excel)",
                            data=output_duplicates,
                            file_name="nip_duplikat_data_mentah.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_duplikat"
                        )
                        
                        st.error("⛔ **PERBAIKI DUPLIKASI NIP TERLEBIH DAHULU** sebelum melanjutkan validasi dengan Master!")
                        st.markdown("---")
                        
                    else:
                        st.success("✅ Validasi Primary Key berhasil: Tidak ada NIP yang duplikat di Data Mentah")
                        st.markdown("---")
                else:
                    st.warning("⚠️ Kolom 'nip' tidak ditemukan di Data Mentah")
                    st.markdown("---")
                
                # Mapping kolom yang akan dibandingkan
                # Data Master = Data Mentah
                comparison_mapping = {
                    'Nama': 'nmpeg',
                    'NIP': 'nip',
                    'NIK': 'npwp',
                    'KDGOL': 'kdgol',
                    'KDKAWIN': 'kdkawin'
                }
                
                st.markdown("### 📊 Informasi Jumlah Data")
                col_info1, col_info2, col_info3 = st.columns(3)
                
                with col_info1:
                    st.metric("📄 Total Baris Data Master", len(df_master))
                
                with col_info2:
                    st.metric("📄 Total Baris Data Mentah", len(df_mentah))
                
                with col_info3:
                    selisih = len(df_master) - len(df_mentah)
                    st.metric("⚖️ Selisih", selisih,
                             delta=f"Master {'lebih banyak' if selisih > 0 else 'lebih sedikit' if selisih < 0 else 'sama'}")
                
                st.markdown("---")
                
                # Cari kolom di master dengan pencarian fleksibel
                master_cols = {}
                for master_col in comparison_mapping.keys():
                    for col in df_master.columns:
                        if master_col.upper() in str(col).upper():
                            master_cols[master_col] = col
                            break
                
                st.markdown("### 🔍 Hasil Validasi Data Mentah vs Master")
                st.info(f"**Mapping Kolom:** Nama={comparison_mapping['Nama']}, NIP={comparison_mapping['NIP']}, NIK={comparison_mapping['NIK']}, KDGOL={comparison_mapping['KDGOL']}, KDKAWIN={comparison_mapping['KDKAWIN']}")
                
                validation_master_data = []
                
                for idx_mentah, row_mentah in df_mentah.iterrows():
                    # Ambil data dari mentah
                    nip_mentah = format_nilai_asli(row_mentah.get('nip', '')) if 'nip' in df_mentah.columns else ''
                    nmpeg_mentah = format_nilai_asli(row_mentah.get('nmpeg', '')) if 'nmpeg' in df_mentah.columns else ''
                    npwp_mentah = format_nilai_asli(row_mentah.get('npwp', '')) if 'npwp' in df_mentah.columns else ''
                    kdgol_mentah = format_nilai_asli(row_mentah.get('kdgol', '')) if 'kdgol' in df_mentah.columns else ''
                    kdkawin_mentah = format_nilai_asli(row_mentah.get('kdkawin', '')) if 'kdkawin' in df_mentah.columns else ''
                    
                    # PRIMARY KEY: Cari matching di master berdasarkan NIP EXACT MATCH
                    match_found = False
                    nama_master = '-'
                    nip_master = '-'
                    nik_master = '-'
                    kdgol_master = '-'
                    kdkawin_master = '-'
                    
                    errors = []
                    
                    # Cari NIP di Master (Primary Key - harus exact match)
                    if 'NIP' in master_cols and nip_mentah:
                        nip_col_master = master_cols['NIP']
                        
                        for idx_master, row_master in df_master.iterrows():
                            nip_master_check = format_nilai_asli(row_master.get(nip_col_master, ''))
                            
                            # PRIMARY KEY: Exact match untuk NIP (tidak pakai fuzzy)
                            if nip_master_check and nip_mentah == nip_master_check:
                                match_found = True
                                nip_master = nip_master_check
                                
                                # Ambil data dari master untuk field lainnya
                                if 'Nama' in master_cols:
                                    nama_master = format_nilai_asli(row_master.get(master_cols['Nama'], ''))
                                if 'NIK' in master_cols:
                                    nik_master = format_nilai_asli(row_master.get(master_cols['NIK'], ''))
                                if 'KDGOL' in master_cols:
                                    kdgol_master = format_nilai_asli(row_master.get(master_cols['KDGOL'], ''))
                                if 'KDKAWIN' in master_cols:
                                    kdkawin_master = format_nilai_asli(row_master.get(master_cols['KDKAWIN'], ''))
                                
                                # Bandingkan HANYA field selain NIP (NIP sudah match sebagai Primary Key)
                                if nama_master and nmpeg_mentah:
                                    if fuzz.ratio(nmpeg_mentah.lower(), nama_master.lower()) < 90:
                                        errors.append('Nama')
                                
                                if nik_master and npwp_mentah:
                                    # Bandingkan NIK master dengan NPWP mentah
                                    if fuzz.ratio(nik_master, npwp_mentah) < 90:
                                        errors.append('NIK/NPWP')
                                
                                if kdgol_master and kdgol_mentah:
                                    if kdgol_master != kdgol_mentah:
                                        errors.append('KDGOL')
                                
                                if kdkawin_master and kdkawin_mentah:
                                    if kdkawin_master != kdkawin_mentah:
                                        errors.append('KDKAWIN')
                                
                                break
                    
                    # Tentukan status berdasarkan Primary Key (NIP)
                    if not nip_mentah or nip_mentah == '':
                        status = 'NIP KOSONG'
                        rekomendasi = 'NIP kosong di Data Mentah - tidak dapat diproses'
                    elif not match_found:
                        # NIP tidak ditemukan di Master
                        status = 'MASTER BELUM LENGKAP'
                        rekomendasi = f'NIP {nip_mentah} tidak ada di Master. Lengkapi Data Master terlebih dahulu.'
                    elif errors:
                        # NIP match, tapi field lain berbeda
                        status = 'TIDAK SESUAI'
                        # Rekomendasi HANYA untuk field yang berbeda (bukan NIP)
                        rekomendasi = f'Perbaiki kolom: {", ".join(errors)} di Data Mentah agar sesuai dengan Master'
                    else:
                        status = 'SESUAI'
                        rekomendasi = 'Semua data sesuai dengan Master'
                    
                    validation_master_data.append({
                        'No': idx_mentah + 1,
                        'NIP (Mentah)': nip_mentah if nip_mentah else '❌ KOSONG',
                        'NIP (Master)': nip_master if match_found else '❌ TIDAK ADA',
                        'Nama (Mentah)': nmpeg_mentah,
                        'Nama (Master)': nama_master,
                        'NPWP (Mentah)': npwp_mentah,
                        'NIK (Master)': nik_master,
                        'KDGOL (Mentah)': kdgol_mentah,
                        'KDGOL (Master)': kdgol_master,
                        'KDKAWIN (Mentah)': kdkawin_mentah,
                        'KDKAWIN (Master)': kdkawin_master,
                        'Status': status,
                        'Kolom Bermasalah': ', '.join(errors) if errors else ('-' if match_found else 'NIP tidak ada di Master'),
                        'Rekomendasi': rekomendasi
                    })
                
                df_validation_master = pd.DataFrame(validation_master_data)
                
                # Simpan ke session state untuk download
                st.session_state['df_validation_master'] = df_validation_master
                
                # Statistik
                st.markdown("### 📊 Ringkasan Validasi")
                col_val1, col_val2, col_val3, col_val4 = st.columns(4)
                
                with col_val1:
                    total_sesuai = len(df_validation_master[df_validation_master['Status'] == 'SESUAI'])
                    st.metric("✅ Data Sesuai", total_sesuai)
                
                with col_val2:
                    total_tidak_sesuai = len(df_validation_master[df_validation_master['Status'] == 'TIDAK SESUAI'])
                    st.metric("⚠️ Tidak Sesuai", total_tidak_sesuai)
                
                with col_val3:
                    total_master_belum = len(df_validation_master[df_validation_master['Status'] == 'MASTER BELUM LENGKAP'])
                    st.metric("❌ Master Belum Lengkap", total_master_belum)
                
                with col_val4:
                    total_nip_kosong = len(df_validation_master[df_validation_master['Status'] == 'NIP KOSONG'])
                    st.metric("⚠️ NIP Kosong", total_nip_kosong)
                
                st.markdown("---")
                
                # Filter
                col_filter1, col_filter2 = st.columns(2)
                
                with col_filter1:
                    filter_status_master = st.selectbox(
                        "Filter Status:",
                        ["Semua", "SESUAI", "TIDAK SESUAI", "MASTER BELUM LENGKAP", "NIP KOSONG"],
                        key="filter_master"
                    )
                
                with col_filter2:
                    show_rekomendasi_master = st.checkbox("Tampilkan Rekomendasi", value=True, key="show_rek_master")
                
                # Apply filter
                if filter_status_master != "Semua":
                    df_validation_master_display = df_validation_master[df_validation_master['Status'] == filter_status_master]
                else:
                    df_validation_master_display = df_validation_master
                
                if not show_rekomendasi_master:
                    df_validation_master_display = df_validation_master_display.drop(columns=['Rekomendasi'])
                
                st.markdown("### 📋 Tabel Validasi")
                
                # Styling dengan highlight per kolom
                def highlight_master_differences(row):
                    colors = [''] * len(row)
                    status = row['Status']
                    
                    if status == 'SESUAI':
                        return ['background-color: #90EE90'] * len(row)
                    elif status == 'MASTER BELUM LENGKAP':
                        return ['background-color: #FFD700'] * len(row)
                    elif status == 'NIP KOSONG':
                        return ['background-color: #FFA500'] * len(row)
                    elif status == 'TIDAK SESUAI':
                        # Highlight merah HANYA kolom Data Mentah yang bermasalah (bukan NIP)
                        kolom_bermasalah = str(row.get('Kolom Bermasalah', '')).split(', ')
                        
                        for i, col_name in enumerate(row.index):
                            # NIP tidak di-highlight karena Primary Key (tidak bisa salah)
                            if 'Nama' in kolom_bermasalah and '(Mentah)' in col_name and 'Nama' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif 'NIK/NPWP' in kolom_bermasalah and '(Mentah)' in col_name and 'NPWP' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif 'KDGOL' in kolom_bermasalah and '(Mentah)' in col_name and 'KDGOL' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif 'KDKAWIN' in kolom_bermasalah and '(Mentah)' in col_name and 'KDKAWIN' in col_name:
                                colors[i] = 'background-color: #FF6B6B'
                            elif col_name in ['Status', 'Kolom Bermasalah', 'Rekomendasi']:
                                colors[i] = 'background-color: #FFA500'
                    
                    return colors
                
                st.dataframe(
                    df_validation_master_display.style.apply(highlight_master_differences, axis=1),
                    height=500
                )
                
                # Download button khusus untuk tab 4
                st.markdown("---")
                st.subheader("📥 Download Hasil Validasi Data Mentah vs Master")
                
                # Buat Excel untuk download
                output_validation_master = BytesIO()
                with pd.ExcelWriter(output_validation_master, engine='openpyxl') as writer:
                    # Sheet 1: Data lengkap
                    df_validation_master.to_excel(writer, index=False, sheet_name='Validasi Lengkap')
                    
                    # Sheet 2: Data yang perlu diperbaiki
                    df_perlu_perbaikan_master = df_validation_master[
                        (df_validation_master['Status'] == 'TIDAK SESUAI') |
                        (df_validation_master['Status'] == 'MASTER BELUM LENGKAP') |
                        (df_validation_master['Status'] == 'NIP KOSONG')
                    ]
                    
                    if not df_perlu_perbaikan_master.empty:
                        df_perlu_perbaikan_master.to_excel(writer, index=False, sheet_name='Perlu Perbaikan')
                        
                        workbook = writer.book
                        worksheet = writer.sheets['Perlu Perbaikan']
                        
                        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        yellow_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                        
                        for idx, row in df_perlu_perbaikan_master.iterrows():
                            excel_row = list(df_perlu_perbaikan_master.index).index(idx) + 2
                            
                            if row['Status'] == 'TIDAK SESUAI':
                                # Highlight kolom Data Mentah yang bermasalah (bukan NIP)
                                kolom_bermasalah = str(row['Kolom Bermasalah']).split(', ')
                                
                                for col_idx, col_name in enumerate(df_perlu_perbaikan_master.columns, start=1):
                                    if any(kb in col_name for kb in kolom_bermasalah) and '(Mentah)' in col_name and 'NIP' not in col_name:
                                        worksheet.cell(row=excel_row, column=col_idx).fill = red_fill
                            
                            elif row['Status'] == 'MASTER BELUM LENGKAP':
                                for col in range(1, len(df_perlu_perbaikan_master.columns) + 1):
                                    worksheet.cell(row=excel_row, column=col).fill = yellow_fill
                            
                            elif row['Status'] == 'NIP KOSONG':
                                for col in range(1, len(df_perlu_perbaikan_master.columns) + 1):
                                    worksheet.cell(row=excel_row, column=col).fill = orange_fill
                
                output_validation_master.seek(0)
                
                st.download_button(
                    label="📥 Download Hasil Validasi (Excel)" + (" ⚠️ (Dinonaktifkan - Ada Duplikasi)" if master_validation_has_duplicates else ""),
                    data=output_validation_master,
                    file_name="validasi_mentah_vs_master.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_validasi_master",
                    disabled=master_validation_has_duplicates
                )
                
                if master_validation_has_duplicates:
                    st.warning("⚠️ Tombol download dinonaktifkan karena terdapat duplikasi data di Data Mentah atau Data Master.")
                
                # Legend
                st.markdown("""
                **Keterangan Warna:**
                - 🟢 **Hijau**: Data SESUAI dengan Master (NIP match dan semua field cocok)
                - 🔴 **Merah**: Highlight pada kolom Data Mentah yang TIDAK SESUAI (NIP match tapi field lain berbeda)
                - 🟠 **Orange**: NIP KOSONG di Data Mentah / Kolom informasi
                - 🟡 **Kuning**: MASTER BELUM LENGKAP (NIP tidak ditemukan di Master)
                
                **Note:** NIP adalah Primary Key - jika NIP tidak match, tidak ada rekomendasi perbaikan karena data tidak dapat dicocokkan.
                """)
                
                st.markdown("---")
                
                # Pisahkan info
                if 'df_perlu_perbaikan_master' in locals() and not df_perlu_perbaikan_master.empty:
                    total_master_belum_lengkap = len(df_perlu_perbaikan_master[df_perlu_perbaikan_master['Status'] == 'MASTER BELUM LENGKAP'])
                    total_tidak_sesuai_perbaikan = len(df_perlu_perbaikan_master[df_perlu_perbaikan_master['Status'] == 'TIDAK SESUAI'])
                    total_nip_kosong_perbaikan = len(df_perlu_perbaikan_master[df_perlu_perbaikan_master['Status'] == 'NIP KOSONG'])
                    
                    st.info(f"""
                    💡 Total **{len(df_perlu_perbaikan_master)}** data yang perlu ditindaklanjuti:
                    - 🟡 **{total_master_belum_lengkap}** NIP tidak ditemukan di Master (Master belum lengkap)
                    - 🔴 **{total_tidak_sesuai_perbaikan}** field (selain NIP) perlu diperbaiki di Data Mentah
                    - 🟠 **{total_nip_kosong_perbaikan}** NIP kosong di Data Mentah
                    
                    **Catatan:** NIP adalah Primary Key - tidak ada rekomendasi perbaikan untuk NIP yang tidak match.
                    """)
                else:
                    st.success("🎉 Semua data valid! Data Mentah sesuai dengan Master dan Master sudah lengkap.")
            
            else:
                st.warning("⚠️ Pastikan Data Mentah dan Master Existing sudah di-upload untuk melakukan validasi")
        
        with tab5:
            st.subheader("📈 Analisis Detail Perubahan")
            
            df_display = st.session_state['df_hasil'].copy()
            
            # Cek duplikasi di data analisis
            df_nip_dup_analisis, _ = check_duplicates(df_display, 'NIP', 'Analisis')
            df_nik_dup_analisis, _ = check_duplicates(df_display, 'NIK', 'Analisis')
            
            analisis_has_duplicates = (df_nip_dup_analisis is not None or df_nik_dup_analisis is not None)
            
            if analisis_has_duplicates:
                st.error("⚠️ **DITEMUKAN DUPLIKASI DI DATA ANALISIS!** Tombol download dinonaktifkan.")
            
            if 'df_master_existing' in st.session_state and st.session_state['df_master_existing'] is not None:
                
                # Pegawai yang tidak aktif
                df_tidak_aktif = df_display[df_display['AKTIF/TIDAK'] == 'TIDAK']
                if not df_tidak_aktif.empty:
                    st.markdown("### ❌ Pegawai Tidak Aktif Bulan Ini")
                    st.dataframe(df_tidak_aktif[['No', 'Nama', 'NIP', 'KDGOL', 'Keterangan']].style.apply(
                        lambda x: ['background-color: #FF6B6B'] * len(x), axis=1
                    ))
                    
                    # Statistik tidak aktif
                    col_tidak1, col_tidak2 = st.columns(2)
                    with col_tidak1:
                        st.metric("Total Tidak Aktif", len(df_tidak_aktif))
                    with col_tidak2:
                        persen_tidak = round((len(df_tidak_aktif) / len(df_display) * 100), 1)
                        st.metric("Persentase Tidak Aktif", f"{persen_tidak}%")
                else:
                    st.success("✅ Tidak ada pegawai yang tidak aktif")
                
                st.markdown("---")
                
                # Pegawai baru
                df_baru = df_display[df_display['Status_Color'] == 'HIJAU']
                if not df_baru.empty:
                    st.markdown("### 🆕 Pegawai Baru Bulan Ini")
                    st.dataframe(df_baru[['No', 'Nama', 'NIP', 'KDGOL', 'PNS/PPPK']].style.apply(
                        lambda x: ['background-color: #90EE90'] * len(x), axis=1
                    ))
                    
                    # Statistik baru
                    col_baru1, col_baru2 = st.columns(2)
                    with col_baru1:
                        st.metric("Total Baru", len(df_baru))
                    with col_baru2:
                        persen_baru = round((len(df_baru) / len(df_display) * 100), 1)
                        st.metric("Persentase Baru", f"{persen_baru}%")
                else:
                    st.info("ℹ️ Tidak ada pegawai baru")
                
                st.markdown("---")
                
                # Pegawai dengan data berubah
                df_berubah = df_display[df_display['Status_Color'] == 'KUNING']
                if not df_berubah.empty:
                    st.markdown("### 🔄 Pegawai dengan Data Berubah")
                    st.dataframe(df_berubah[['No', 'Nama', 'NIP', 'KDGOL', 'Keterangan']].style.apply(
                        lambda x: ['background-color: #FFFF00'] * len(x), axis=1
                    ))
                    
                    # Statistik berubah
                    col_berubah1, col_berubah2 = st.columns(2)
                    with col_berubah1:
                        st.metric("Total Berubah", len(df_berubah))
                    with col_berubah2:
                        persen_berubah = round((len(df_berubah) / len(df_display) * 100), 1)
                        st.metric("Persentase Berubah", f"{persen_berubah}%")
                else:
                    st.info("ℹ️ Tidak ada pegawai dengan data berubah")
                
                st.markdown("---")
                
                # Distribusi status
                st.markdown("### 📊 Distribusi Status Pegawai")
                col_chart1, col_chart2 = st.columns(2)
                
                with col_chart1:
                    status_counts = df_display['AKTIF/TIDAK'].value_counts()
                    if not status_counts.empty:
                        st.bar_chart(status_counts)
                        st.caption("Perbandingan Aktif vs Tidak Aktif")
                    else:
                        st.info("Tidak ada data distribusi status")
                
                with col_chart2:
                    if 'Status_Color' in df_display.columns:
                        color_counts = df_display['Status_Color'].value_counts()
                        if not color_counts.empty:
                            color_mapping = {
                                'HIJAU': 'Baru',
                                'KUNING': 'Tidak Berubah',
                                'MERAH': 'Tidak Aktif',
                                'ORANGE': 'Berubah'
                            }
                            color_counts.index = [color_mapping.get(c, c) for c in color_counts.index]
                            st.bar_chart(color_counts)
                            st.caption("Distribusi Perubahan Data")
                        else:
                            st.info("Tidak ada data perubahan")
                
                st.markdown("---")
                
                # Download button khusus untuk tab 5
                st.subheader("📥 Download Analisis Perubahan")
                
                # Buat Excel untuk analisis
                output_analisis = BytesIO()
                with pd.ExcelWriter(output_analisis, engine='openpyxl') as writer:
                    # Sheet 1: Ringkasan Analisis
                    summary_data = {
                        'Kategori': ['Total Data', 'Data Baru', 'Data Tidak Aktif', 'Data Berubah', 'Data Tidak Berubah'],
                        'Jumlah': [
                            len(df_display),
                            len(df_baru) if 'df_baru' in locals() else 0,
                            len(df_tidak_aktif) if 'df_tidak_aktif' in locals() else 0,
                            len(df_berubah) if 'df_berubah' in locals() else 0,
                            len(df_display[df_display['Status_Color'] == 'KUNING']) if 'Status_Color' in df_display.columns else 0
                        ],
                        'Persentase': [
                            '100%',
                            f"{round((len(df_baru) / len(df_display) * 100), 1)}%" if 'df_baru' in locals() else '0%',
                            f"{round((len(df_tidak_aktif) / len(df_display) * 100), 1)}%" if 'df_tidak_aktif' in locals() else '0%',
                            f"{round((len(df_berubah) / len(df_display) * 100), 1)}%" if 'df_berubah' in locals() else '0%',
                            f"{round((len(df_display[df_display['Status_Color'] == 'KUNING']) / len(df_display) * 100), 1)}%" if 'Status_Color' in df_display.columns else '0%'
                        ]
                    }
                    df_summary_analisis = pd.DataFrame(summary_data)
                    df_summary_analisis.to_excel(writer, index=False, sheet_name='Ringkasan Analisis')
                    
                    # Sheet 2: Data Tidak Aktif
                    if 'df_tidak_aktif' in locals() and not df_tidak_aktif.empty:
                        df_tidak_aktif_export = df_tidak_aktif.drop(columns=['Status_Color'], errors='ignore')
                        df_tidak_aktif_export.to_excel(writer, index=False, sheet_name='Pegawai Tidak Aktif')
                    
                    # Sheet 3: Data Baru
                    if 'df_baru' in locals() and not df_baru.empty:
                        df_baru_export = df_baru.drop(columns=['Status_Color'], errors='ignore')
                        df_baru_export.to_excel(writer, index=False, sheet_name='Pegawai Baru')
                    
                    # Sheet 4: Data Berubah
                    if 'df_berubah' in locals() and not df_berubah.empty:
                        df_berubah_export = df_berubah.drop(columns=['Status_Color'], errors='ignore')
                        df_berubah_export.to_excel(writer, index=False, sheet_name='Pegawai Berubah')
                    
                    # Sheet 5: Data Lengkap
                    df_display_export = df_display.drop(columns=['Status_Color'], errors='ignore')
                    df_display_export.to_excel(writer, index=False, sheet_name='Data Lengkap')
                
                output_analisis.seek(0)
                
                st.download_button(
                    label="📥 Download Analisis Perubahan (Excel)" + (" ⚠️ (Dinonaktifkan - Ada Duplikasi)" if analisis_has_duplicates else ""),
                    data=output_analisis,
                    file_name="analisis_perubahan_pegawai.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_analisis",
                    disabled=analisis_has_duplicates
                )
                
                if analisis_has_duplicates:
                    st.warning("⚠️ Tombol download dinonaktifkan karena terdapat duplikasi data di hasil analisis.")
                
                # Ringkasan statistik
                st.markdown("### 📈 Ringkasan Statistik")
                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                
                with col_stat1:
                    st.metric("Total Pegawai", len(df_display))
                
                with col_stat2:
                    st.metric("Pegawai Baru", len(df_baru) if 'df_baru' in locals() else 0)
                
                with col_stat3:
                    st.metric("Tidak Aktif", len(df_tidak_aktif) if 'df_tidak_aktif' in locals() else 0)
                
                with col_stat4:
                    st.metric("Data Berubah", len(df_berubah) if 'df_berubah' in locals() else 0)
            
            else:
                st.info("ℹ️ Upload Master Lama untuk melihat analisis perubahan detail")
                
                # Tampilkan statistik dasar saja
                st.markdown("### 📊 Statistik Dasar")
                col_stat1, col_stat2 = st.columns(2)
                
                with col_stat1:
                    st.metric("Total Pegawai", len(df_display))
                
                with col_stat2:
                    total_aktif = len(df_display[df_display['AKTIF/TIDAK'] == 'AKTIF'])
                    st.metric("Pegawai Aktif", total_aktif)
                
                # Download button untuk data dasar
                st.markdown("---")
                st.subheader("📥 Download Data Dasar")
                
                output_dasar = BytesIO()
                df_display_export = df_display.drop(columns=['Status_Color'], errors='ignore')
                
                with pd.ExcelWriter(output_dasar, engine='openpyxl') as writer:
                    df_display_export.to_excel(writer, index=False, sheet_name='Data Pegawai')
                
                output_dasar.seek(0)
                
                st.download_button(
                    label="📥 Download Data Pegawai (Excel)" + (" ⚠️ (Dinonaktifkan - Ada Duplikasi)" if analisis_has_duplicates else ""),
                    data=output_dasar,
                    file_name="data_pegawai_dasar.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_dasar",
                    disabled=analisis_has_duplicates
                )

# Untuk running langsung file ini (testing)
if __name__ == "__main__":
    show()