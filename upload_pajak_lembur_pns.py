import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def check_duplicate_nips(df_mentah):
    """Cek NIP duplikat di data mentah dan return baris yang duplikat"""
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    duplicates = df_mentah[df_mentah.duplicated('nip_clean', keep=False)]
    return duplicates

def check_new_data(df_mentah, df_master):
    """Cek NIP yang ada di data mentah tapi tidak ada di data master"""
    # Bersihkan NIP
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    df_master['NIP_clean'] = df_master['NIP'].astype(str).str.strip()
    
    # Cari NIP yang tidak ada di master
    master_nips = set(df_master['NIP_clean'].tolist())
    mentah_nips = set(df_mentah['nip_clean'].tolist())
    
    new_nips = mentah_nips - master_nips
    new_data = df_mentah[df_mentah['nip_clean'].isin(new_nips)]
    
    return new_data

def show():
    # Tombol kembali
    if st.button("← Kembali ke Dashboard PNS"):
        st.session_state.current_page = 'dashboard_pns'
        st.rerun()
    
    st.title("⏰ Upload Pajak Lembur PNS")
    st.markdown("---")
    
    # ========== PANDUAN PENGGUNAAN ==========
    with st.expander("📚 Panduan Penggunaan - Baca Sebelum Mulai", expanded=False):
        st.markdown("""
        ### **🎯 Cara Menggunakan Fitur Ini:**
        
        1. **Siapkan 2 File Excel**:
           - **Data Mentah**: File transaksi lembur harian/bulanan PNS
           - **Data Master**: Database pegawai PNS
        
        2. **Format Data Mentah (WAJIB)**:
           ```
           Kolom yang HARUS ADA:
           - nip          : Nomor Induk Pegawai
           - kotor        : Nilai penghasilan kotor lembur
           - pajak        : Nilai PPh yang sudah dipotong
           - bln          : Bulan transaksi
           - thn          : Tahun transaksi
           
           Kolom khusus lembur:
           - jamlemburharikerja : Jam lembur hari kerja
           - jamlemburharilibur : Jam lembur hari libur
           
           Kolom tambahan yang diperlukan:
           - nmpeg        : Nama pegawai
           - npwp         : NPWP pegawai
           - kdgol        : Kode golongan
           - dan kolom lain sesuai template
           ```
           
        3. **Format Data Master (WAJIB)**:
           ```
           Kolom yang HARUS ADA:
           - NIP                : Nomor Induk Pegawai
           - NIK                : Digunakan sebagai NPWP di hasil akhir
           - STATUS             : Status PTKP (K/TK/HB/etc.)
           
           Kolom YANG DICARI:
           - ID PENERIMA TKU    : ID penerima penghasilan
           - KODE OBJEK PAJAK   : Kode objek pajak (untuk lembur)
           
           Kolom tambahan (jika ada):
           - PNS/PPPK       : Jenis pegawai
           - Nama           : Nama lengkap
           - KDKAWIN        : Kode status kawin
           - nmrek          : Nama rekening
           - rekening       : Nomor rekening
           ```
        
        4. **PROSES OTOMATIS**:
           - Sistem akan melakukan **inner join** berdasarkan NIP
           - Hanya data dengan NIP yang match di kedua file yang diproses
           - Data tanpa match di Data Master akan ditampilkan sebagai warning
        
        5. **PERHATIAN KHUSUS**:
           ```
           ID TKU Pemotong sekarang menggunakan angka DEFAULT
           Warna hijau muda (data sistem), bukan manual (orange)
           
           Kolom manual (orange) yang perlu diisi:
           - Nomor Dok. Referensi     → dari Nomor SP2D
           - Tanggal Dok. Referensi   → dari Tanggal SP2D
           - Tanggal Pemotongan       → dari Tanggal Invoice
           ```
        
        6. **VALIDASI DATA**:
           - NIP di Data Mentah harus ada di Data Master
           - **NIP duplikat** di data mentah akan ditandai merah dan harus diperbaiki
           - **Data baru** (NIP tidak ada di master) akan ditandai hijau dan harus ditambahkan ke master
           - Sistem hanya memproses data yang memiliki kecocokan NIP
           - Data tidak match akan ditampilkan di bagian akhir
        """)
        
        # Template untuk download
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 Template Data Mentah (Minimal)",
                data=pd.DataFrame({
                    'kdsatker': ['123456'],
                    'bln': ['1'],
                    'thn': ['2024'],
                    'tgl': ['2024-01-15'],
                    'nogaji': ['001'],
                    'nip': ['123456789012345678'],
                    'nmpeg': ['Nama Contoh'],
                    'kdgol': ['IV/a'],
                    'npwp': ['123456789012345'],
                    'kdbankspan': ['014'],
                    'nmbankspan': ['BCA'],
                    'norek': ['1234567890'],
                    'nmrek': ['Nama Rekening'],
                    'nmcabbank': ['Cabang Pusat'],
                    'jamlemburharikerja': ['10'],
                    'jamlemburharilibur': ['5'],
                    'kotor': ['1000000'],
                    'pajak': ['50000'],
                    'bersih': ['950000']
                }).to_csv(index=False).encode('utf-8'),
                file_name="template_data_mentah_lembur_pns.csv",
                mime="text/csv"
            )
        
        with col2:
            st.download_button(
                label="📥 Template Data Master (Minimal)",
                data=pd.DataFrame({
                    'NIP': ['123456789012345678', '987654321098765432'],
                    'NIK': ['123456789012345', '987654321098765'],
                    'ID PENERIMA TKU': ['ID001', 'ID002'],
                    'STATUS': ['K', 'TK'],
                    'KODE OBJEK PAJAK': ['21-200-01', '21-200-01'],  # Kode khusus lembur
                    'Nama': ['Nama Contoh 1', 'Nama Contoh 2'],
                    'PNS/PPPK': ['PNS', 'PNS']
                }).to_csv(index=False).encode('utf-8'),
                file_name="template_data_master_lembur_pns.csv",
                mime="text/csv"
            )
    
    st.markdown("---")
    
    # ========== INFORMASI PENTING ==========
    st.info("""
    **📊 INFORMASI PENTING:**
    
    **1. PROSES JOIN DATA:**
    - Sistem menggunakan **INNER JOIN** berdasarkan NIP
    - **NIP duplikat** di data mentah akan ditandai merah dan harus diperbaiki
    - **Data baru** (NIP tidak ada di master) akan ditandai hijau dan harus ditambahkan ke master
    - Data tanpa match akan ditampilkan terpisah untuk tindak lanjut
    
    **2. PERUBAHAN PENTING:**
    ```
    ID TKU Pemotong sekarang menggunakan ANGKA DEFAULT
    Nilai: "0001658723701000000000"
    
    Kolom yang perlu diisi MANUAL (warna orange):
    - Nomor Dok. Referensi     → dari Nomor SP2D
    - Tanggal Dok. Referensi   → dari Tanggal SP2D
    - Tanggal Pemotongan       → dari Tanggal Invoice
    ```
    
    **3. KOLOM HASIL BP 21 YANG DIHASILKAN:**
    ```
    1. Masa Pajak               → dari kolom 'bln' (bulan)
    2. Tahun Pajak              → dari kolom 'thn' (tahun)
    3. NPWP                     → dari kolom 'NIK' di Data Master
    4. ID TKU Penerima          → dari 'ID PENERIMA TKU' di Data Master
    5. Status PTKP              → dari kolom 'STATUS' di Data Master
    6. Fasilitas                → default "DTP"
    7. Kode Objek Pajak         → dari Data Master (khusus lembur)
    8. Penghasilan              → dari kolom 'kotor' di Data Mentah
    9. Deemed                   → default "100"
    10. Tarif                   → otomatis dihitung (pajak/kotor)×100
    11. Jenis Dok. Referensi    → default "CommercialInvoice"
    12. Nomor Dok. Referensi    → manual (orange)
    13. Tanggal Dok. Referensi  → manual (orange)
    14. ID TKU Pemotong         → DEFAULT: "0001658723701000000000"
    15. Tanggal Pemotongan      → manual (orange)
    ```
    """)
    
    # ========== INFORMASI FORMAT WARNA ==========
    st.warning("""
    **🎨 INFORMASI FORMAT WARNA DI HASIL DOWNLOAD:**
    
    **1. Data Hijau Muda (#E6FFE6):**
    - Data hasil sistem (sudah diproses otomatis)
    - Bisa langsung dicopy/duplikat
    - Tidak perlu perubahan
    
    **2. Data Oranye (#FFD580):**
    - Data yang harus diisi **MANUAL** dari DAFTAR SP2D SATKER
    - **HANYA 3 KOLOM yang perlu diisi manual sekarang:**
      - **Nomor Dok. Referensi** → Ambil dari **Nomor SP2D**
      - **Tanggal Dok. Referensi** → Ambil dari **Tanggal SP2D**
      - **Tanggal Pemotongan** → Ambil dari **Tanggal Invoice**
    
    **3. ID TKU Pemotong sekarang DEFAULT:**
    - Menggunakan nilai default: **"0001658723701000000000"**
    - Warna hijau muda (sudah terisi otomatis)
    - Tidak perlu diisi manual lagi
    """)
    
    # ========== FORM UPLOAD ==========
    st.subheader("📤 Upload File Excel")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### **📄 Data Mentah (Transaksi Lembur)**")
        st.caption("File Excel berisi transaksi lembur harian/bulanan")
        
        with st.expander("ℹ️ Detail Kolom Data Mentah", expanded=False):
            st.markdown("""
            **Kolom WAJIB ada:**
            - `nip` : Nomor Induk Pegawai (kunci join, TIDAK BOLEH DUPLIKAT)
            - `kotor` : Nilai penghasilan kotor lembur
            - `pajak` : Nilai PPh (untuk menghitung tarif)
            - `bln` : Bulan (1-12, tanpa leading zero)
            - `thn` : Tahun (4 digit)
            
            **Kolom khusus lembur:**
            - `jamlemburharikerja` : Jam lembur hari kerja
            - `jamlemburharilibur` : Jam lembur hari libur
            
            **Kolom PENTING:**
            - `nmpeg` : Nama pegawai
            - `npwp` : NPWP (jika ada)
            - `kdgol` : Kode golongan
            - `norek` : Nomor rekening
            - `nmrek` : Nama pemilik rekening
            """)
        
        uploaded_file_raw = st.file_uploader(
            "**Pilih file Data Mentah**",
            type=['xlsx', 'xls'],
            key="raw_data_lembur_pns",
            help="Upload file transaksi lembur PNS. Pastikan minimal ada kolom nip, kotor, pajak, bln, thn"
        )
    
    with col2:
        st.markdown("#### **📋 Data Master (Referensi)**")
        st.caption("Database referensi pegawai PNS")
        
        with st.expander("ℹ️ Detail Kolom Data Master", expanded=False):
            st.markdown("""
            **Kolom WAJIB ada:**
            - `NIP` : Nomor Induk Pegawai (kunci join)
            - `NIK` : Akan digunakan sebagai NPWP di output
            - `STATUS` : Status PTKP (K/TK/HB/etc.)
            
            **Kolom YANG DICARI:**
            - `ID PENERIMA TKU` : ID penerima penghasilan
            - `KODE OBJEK PAJAK` : Kode objek pajak (khusus lembur)
            
            **Kolom tambahan (jika ada):**
            - `PNS/PPPK` : Jenis pegawai
            - `Nama` : Nama lengkap
            - `KDKAWIN` : Kode status kawin
            - `nmrek` : Nama rekening
            - `rekening` : Nomor rekening
            """)
        
        uploaded_file_master = st.file_uploader(
            "**Pilih file Data Master**",
            type=['xlsx', 'xls'],
            key="master_data_lembur_pns",
            help="Upload database master pegawai PNS"
        )
    
    st.markdown("---")
    
    # ========== PROSES DATA ==========
    if uploaded_file_raw is not None and uploaded_file_master is not None:
        try:
            # Baca kedua file
            df_raw = pd.read_excel(uploaded_file_raw)
            df_master = pd.read_excel(uploaded_file_master)
            
            # BERSIHKAN NAMA KOLOM (hapus spasi di awal/akhir)
            df_raw.columns = df_raw.columns.str.strip()
            df_master.columns = df_master.columns.str.strip()
            
            st.success(f"✅ Data mentah berhasil diupload: {uploaded_file_raw.name}")
            st.success(f"✅ Data master berhasil diupload: {uploaded_file_master.name}")
            
            # ========== VALIDASI AWAL ==========
            st.subheader("🔍 Validasi Data dan Deteksi Data Baru")
            
            # Validasi kolom wajib di Data Mentah
            required_raw = ['nip', 'kotor', 'pajak', 'bln', 'thn']
            missing_raw = [col for col in required_raw if col not in df_raw.columns]
            
            if missing_raw:
                st.error(f"❌ **ERROR**: Kolom berikut tidak ditemukan di Data Mentah:")
                for col in missing_raw:
                    st.write(f"   - **{col}**")
                st.warning("💡 **Solusi**: Pastikan semua kolom wajib ada dan penulisannya benar")
                st.stop()
            
            # Validasi kolom wajib di Data Master
            required_master = ['NIP', 'NIK', 'STATUS']
            missing_master = [col for col in required_master if col not in df_master.columns]
            
            if missing_master:
                st.error(f"❌ **ERROR**: Kolom berikut tidak ditemukan di Data Master:")
                for col in missing_master:
                    st.write(f"   - **{col}**")
                st.warning("💡 **Solusi**: Pastikan NIP, NIK, dan STATUS ada di Data Master")
                st.stop()
            
            # ========== CEK NIP DUPLIKAT DI DATA MENTAH ==========
            duplicates = check_duplicate_nips(df_raw)
            if not duplicates.empty:
                st.session_state.duplicate_nips_df = duplicates
                st.error(f"❌ Ditemukan {len(duplicates)} NIP duplikat di Data Mentah!")
                
                # Tampilkan baris duplikat dengan warna merah
                st.warning("**Baris dengan NIP duplikat (ditandai merah):**")
                
                # Buat DataFrame dengan styling untuk duplikat
                def highlight_duplicates(df_original, duplicates_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris duplikat
                    mask = styled_df['nip_clean'].isin(duplicates_df['nip_clean'].unique())
                    
                    # Buat list warna
                    colors = ['background-color: #ffcccc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_duplicates = highlight_duplicates(df_raw.head(50), duplicates)
                st.dataframe(styled_duplicates, use_container_width=True)
                
                if len(df_raw) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_raw)} baris")
                
                st.error("**PERBAIKI NIP DUPLIKAT SEBELUM MELANJUTKAN!**")
                st.stop()
            
            # ========== CEK DATA BARU (NIP DI MENTAH TAPI TIDAK DI MASTER) ==========
            new_data = check_new_data(df_raw, df_master)
            
            if not new_data.empty:
                st.session_state.new_data_df = new_data
                st.warning(f"⚠️ Ditemukan {len(new_data)} data baru di Data Mentah yang tidak ada di Data Master!")
                
                # Tampilkan data baru dengan warna hijau
                st.info("**Data baru (ditandai hijau - perlu ditambahkan ke Data Master):**")
                
                def highlight_new_data(df_original, new_data_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris data baru
                    mask = styled_df['nip_clean'].isin(new_data_df['nip_clean'].unique())
                    
                    # Buat list warna
                    colors = ['background-color: #ccffcc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_new_data = highlight_new_data(df_raw.head(50), new_data)
                st.dataframe(styled_new_data, use_container_width=True)
                
                if len(df_raw) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_raw)} baris")
                
                # Tampilkan detail data baru yang perlu ditambahkan
                st.markdown("---")
                st.error("**DATA BARU HARUS DITAMBAHKAN KE DATA MASTER SEBELUM MELANJUTKAN!**")
                
                # Tampilkan tabel detail data baru
                st.info("**Detail Data Baru yang Harus Ditambahkan ke Data Master:**")
                
                # Ekstrak informasi penting dari data baru
                new_data_details = new_data[['nip_clean', 'nmpeg', 'npwp']].copy()
                new_data_details = new_data_details.drop_duplicates('nip_clean')
                new_data_details.columns = ['NIP', 'Nama', 'NPWP']
                
                st.dataframe(new_data_details, use_container_width=True)
                
                # Tambahkan tombol untuk menuju ke halaman croscheck_pns
                if st.button("➕ Tambahkan Data Baru ke Master (Croscheck PNS)", type="primary", use_container_width=True):
                    st.session_state.current_page = 'croscheck_pns'
                    st.session_state.selected_menu = 'croscheck'
                    st.rerun()
                
                st.stop()  # Hentikan proses sampai data baru ditangani
            
            # Jika tidak ada data baru, lanjutkan
            st.success("✅ Tidak ditemukan data baru. Semua NIP di Data Mentah ada di Data Master.")
            
            # Preview data yang diupload
            col1, col2 = st.columns(2)
            with col1:
                with st.expander(f"📄 Data Mentah ({len(df_raw)} baris)"):
                    st.write(f"**Kolom yang terdeteksi ({len(df_raw.columns)}):**")
                    cols_per_row = 3
                    cols = list(df_raw.columns)
                    for i in range(0, len(cols), cols_per_row):
                        row_cols = cols[i:i+cols_per_row]
                        col_text = " | ".join([f"`{col}`" for col in row_cols])
                        st.write(f"  {col_text}")
                    st.dataframe(df_raw.head(3))
            
            with col2:
                with st.expander(f"📄 Data Master ({len(df_master)} baris)"):
                    st.write(f"**Kolom yang terdeteksi ({len(df_master.columns)}):**")
                    cols = list(df_master.columns)
                    for i in range(0, len(cols), cols_per_row):
                        row_cols = cols[i:i+cols_per_row]
                        col_text = " | ".join([f"`{col}`" for col in row_cols])
                        st.write(f"  {col_text}")
                    st.dataframe(df_master.head(3))
            
            st.markdown("---")
            
            # ========== TOMBOL PROSES ==========
            if st.button("🔄 **PROSES DATA & GENERATE BP 21**", 
                        use_container_width=True, 
                        type="primary",
                        help="Klik untuk memproses data dan menghasilkan file BP 21 untuk lembur"):
                
                with st.spinner("🔄 Sedang memproses data lembur..."):
                    try:
                        # Membersihkan dan mempersiapkan data
                        df_raw_clean = df_raw.copy()
                        df_master_clean = df_master.copy()
                        
                        # Pastikan NIP dalam format string untuk join
                        df_raw_clean['nip'] = df_raw_clean['nip'].astype(str).str.strip()
                        df_master_clean['NIP'] = df_master_clean['NIP'].astype(str).str.strip()
                        
                        # Ambil NIP unik dari data mentah sebagai primary key
                        nip_list = df_raw_clean['nip'].unique()
                        
                        # Filter data master hanya untuk NIP yang ada di data mentah
                        df_master_filtered = df_master_clean[df_master_clean['NIP'].isin(nip_list)].copy()
                        
                        # Informasi proses join
                        st.info(f"🔍 **Proses Matching Data:**")
                        st.info(f"   • NIP unik di Data Mentah: {len(nip_list)}")
                        st.info(f"   • NIP ditemukan di Data Master: {len(df_master_filtered)}")
                        
                        # Join data berdasarkan NIP (inner join untuk hanya ambil yang match)
                        df_merged = pd.merge(
                            df_raw_clean,
                            df_master_filtered,
                            left_on='nip',
                            right_on='NIP',
                            how='inner'  # Hanya ambil yang match
                        )
                        
                        # ========== BUAT DATA HASIL BP 21 ==========
                        df_result = pd.DataFrame()
                        
                        # 1. Masa Pajak: ambil dari bln (tanpa leading zero)
                        df_result['Masa Pajak'] = df_merged['bln'].astype(str).str.lstrip('0')
                        
                        # 2. Tahun Pajak: ambil dari thn
                        df_result['Tahun Pajak'] = df_merged['thn'].astype(str)
                        
                        # 3. NPWP: ambil dari NIK (data master) - hilangkan .0
                        df_result['NPWP'] = df_merged['NIK'].apply(lambda x: str(x).replace('.0', '') if pd.notna(x) else '')
                        
                        # 4. ID TKU Penerima Penghasilan: cari kolom yang cocok
                        id_tku_col = None
                        for col in df_merged.columns:
                            if 'ID PENERIMA TKU' in col.upper() or 'ID_PENERIMA_TKU' in col.upper():
                                id_tku_col = col
                                break
                        
                        if id_tku_col:
                            df_result['ID TKU Penerima Penghasilan'] = df_merged[id_tku_col].astype(str)
                            st.success(f"✅ Kolom ID TKU Penerima ditemukan: {id_tku_col}")
                        else:
                            df_result['ID TKU Penerima Penghasilan'] = ''
                            st.warning("⚠️ Kolom 'ID PENERIMA TKU' tidak ditemukan, diisi dengan nilai kosong")
                        
                        # 5. Status PTKP: ambil dari STATUS
                        df_result['Status PTKP'] = df_merged['STATUS'].astype(str)
                        
                        # 6. Fasilitas: default "DTP"
                        df_result['Fasilitas'] = 'DTP'
                        
                        # 7. Kode Objek Pajak: cari kolom yang cocok
                        kode_pajak_col = None
                        for col in df_merged.columns:
                            if 'KODE OBJEK PAJAK' in col.upper() or 'KODE_OBJEK_PAJAK' in col.upper():
                                kode_pajak_col = col
                                break
                        
                        if kode_pajak_col:
                            df_result['Kode Objek Pajak'] = df_merged[kode_pajak_col].astype(str)
                            st.success(f"✅ Kolom Kode Objek Pajak ditemukan: {kode_pajak_col}")
                        else:
                            df_result['Kode Objek Pajak'] = ''
                            st.warning("⚠️ Kolom 'KODE OBJEK PAJAK' tidak ditemukan, diisi dengan nilai kosong")
                        
                        # 8. Penghasilan: ambil dari kotor
                        df_result['Penghasilan'] = df_merged['kotor'].astype(float)
                        
                        # 9. Deemed: default "100"
                        df_result['Deemed'] = '100'
                        
                        # 10. Tarif: hitung otomatis (pajak / kotor) × 100
                        # Hindari pembagian dengan nol
                        df_result['Tarif'] = df_merged.apply(
                            lambda row: round((row['pajak'] / row['kotor'] * 100), 2) if row['kotor'] != 0 else 0,
                            axis=1
                        )
                        
                        st.info("📊 **Mode Perhitungan**: Tarif dihitung otomatis = (pajak / kotor) × 100")
                        
                        # 11. Jenis Dok. Referensi: default "CommercialInvoice"
                        df_result['Jenis Dok. Referensi'] = 'CommercialInvoice'
                        
                        # 12. Nomor Dok. Referensi: kosong (diisi manual) - WARNA ORANYE
                        df_result['Nomor Dok. Referensi'] = ''
                        
                        # 13. Tanggal Dok. Referensi: kosong (diisi manual) - WARNA ORANYE
                        df_result['Tanggal Dok. Referensi'] = ''
                        
                        # 14. ID TKU Pemotong: DEFAULT "0001658723701000000000"
                        df_result['ID TKU Pemotong'] = '0001658723701000000000'
                        st.success("✅ ID TKU Pemotong diisi dengan nilai default: 0001658723701000000000")
                        
                        # 15. Tanggal Pemotongan: kosong (diisi manual) - WARNA ORANYE
                        df_result['Tanggal Pemotongan'] = ''
                        
                        # Hitung statistik
                        processed_count = len(df_result)
                        master_matched = processed_count  # Semua data adalah yang match karena inner join
                        not_matched = len(df_raw) - master_matched
                        
                        st.success(f"✅ **Data berhasil diproses!** Total {processed_count} baris data BP 21 untuk lembur")
                        
                        # ========== TAMPILKAN HASIL ==========
                        st.markdown("---")
                        st.subheader("📊 Ringkasan Hasil Proses")
                        
                        # Metrik ringkasan
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Data Mentah", len(df_raw))
                        with col2:
                            st.metric("NIP Unik", len(nip_list))
                        with col3:
                            st.metric("✅ Match & Diproses", master_matched)
                        with col4:
                            st.metric("❌ Tidak Match", not_matched)
                        
                        # Informasi perhitungan
                        st.info(f"""
                        **📈 Statistik Hasil:**
                        - **Total Penghasilan Lembur:** Rp {df_result['Penghasilan'].sum():,.0f}
                        - **Rata-rata Tarif:** {df_result['Tarif'].mean():.2f}%
                        - **Rentang Tarif:** {df_result['Tarif'].min():.2f}% - {df_result['Tarif'].max():.2f}%
                        - **Total Pajak Dipotong:** Rp {df_merged['pajak'].sum():,.0f}
                        """)
                        
                        # Preview hasil dengan warna indikator
                        st.subheader("📄 Preview Hasil (Warna hanya ilustrasi)")
                        
                        # Buat DataFrame untuk preview dengan indikasi kolom
                        preview_df = df_result.head(10).copy()
                        
                        # Tentukan kolom yang harus berwarna oranye (HANYA 3 KOLOM SEKARANG)
                        orange_columns_preview = ['Nomor Dok. Referensi', 'Tanggal Dok. Referensi', 'Tanggal Pemotongan']
                        
                        # Tampilkan dataframe dengan styling sederhana
                        st.dataframe(preview_df)
                        
                        # Tampilkan legenda warna di bawah
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            <div style="background-color: #e6ffe6; padding: 10px; border-radius: 5px; border: 1px solid #ccc; margin-top: 10px;">
                            <strong>🟢 Hijau Muda:</strong> Data hasil sistem (siap copy)
                            </div>
                            """, unsafe_allow_html=True)
                        with col2:
                            st.markdown("""
                            <div style="background-color: #fff0e6; padding: 10px; border-radius: 5px; border: 1px solid #ccc; margin-top: 10px;">
                            <strong>🟠 Oranye:</strong> Isi manual dari DAFTAR SP2D SATKER
                            </div>
                            """, unsafe_allow_html=True)
                        
                        # Tampilkan informasi kolom warna
                        st.info(f"""
                        **Kolom dengan warna Hijau Muda (Data Sistem):** {', '.join([col for col in preview_df.columns if col not in orange_columns_preview])}
                        
                        **Kolom dengan warna Oranye (Data Manual):** {', '.join(orange_columns_preview)}
                        
                        **📝 CATATAN:** ID TKU Pemotong sekarang menggunakan nilai default: 0001658723701000000000
                        """)
                        
                        # Informasi pengisian manual
                        with st.expander("📝 PETUNJUK PENGISIAN MANUAL", expanded=False):
                            st.markdown("""
                            ### **Kolom yang perlu diisi MANUAL (Warna Oranye):**
                            
                            1. **Nomor Dok. Referensi** → Nomor SP2D
                               - Contoh: `SP2D-001/BPK/2024`
                            
                            2. **Tanggal Dok. Referensi** → Tanggal SP2D
                               - Format: DD/MM/YYYY
                               - Contoh: `15/01/2024`
                            
                            3. **Tanggal Pemotongan** → Tanggal Invoice
                               - Format: DD/MM/YYYY
                               - Contoh: `10/01/2024`
                            
                            **Sumber data:** DAFTAR SP2D SATKER (download dari sistem keuangan)
                            
                            ### **Kolom yang sudah OTOMATIS (Hijau Muda):**
                            
                            4. **ID TKU Pemotong** → Menggunakan nilai default: 0001658723701000000000
                               - Tidak perlu diisi manual
                               - Warna hijau muda di file hasil
                            
                            **Catatan khusus lembur:**
                            - Pastikan Kode Objek Pajak sesuai untuk transaksi lembur
                            - ID TKU Pemotong menggunakan nilai default
                            """)
                        
                        # ========== DATA TIDAK MATCH ==========
                        if not_matched > 0:
                            st.markdown("---")
                            st.warning(f"⚠️ **PERHATIAN**: Ditemukan {not_matched} data yang tidak memiliki match di Data Master")
                            
                            # Cari NIP yang tidak match
                            matched_nips = df_master_filtered['NIP'].unique()
                            no_match_data = df_raw_clean[~df_raw_clean['nip'].isin(matched_nips)].copy()
                            
                            # Tampilkan informasi detail
                            st.write(f"**NIP yang tidak ditemukan di Data Master:**")
                            no_match_display = no_match_data[['nip', 'nmpeg']].drop_duplicates('nip')
                            no_match_display.columns = ['NIP', 'Nama Pegawai']
                            st.dataframe(no_match_display, use_container_width=True, height=200)
                            
                            # Download data tidak match
                            csv_no_match = no_match_display.to_csv(index=False).encode('utf-8')
                            st.download_button(
                                label="📥 Download Data Tidak Match",
                                data=csv_no_match,
                                file_name="data_tidak_match_lembur_pns.csv",
                                mime="text/csv",
                                help="Download daftar NIP yang tidak ditemukan di Data Master"
                            )
                        
                        # ========== SIMPAN KE SESSION STATE ==========
                        st.session_state.hasil_pajak_lembur_pns = df_result
                        st.session_state.jumlah_data_lembur_pns = processed_count
                        st.session_state.tidak_match_lembur_pns = not_matched
                        
                        st.balloons()
                        st.success("🎉 **Data BP 21 untuk lembur siap untuk didownload dengan format warna!**")
                        
                    except Exception as e:
                        st.error(f"❌ **ERROR**: Terjadi kesalahan saat memproses data lembur")
                        st.error(f"Detail error: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
            
            # ========== TOMBOL DOWNLOAD DENGAN FORMAT WARNA ==========
            if 'hasil_pajak_lembur_pns' in st.session_state:
                st.markdown("---")
                st.subheader("💾 Download Hasil dengan Format Warna")
                
                hasil_final = st.session_state.hasil_pajak_lembur_pns
                jumlah_data = st.session_state.get('jumlah_data_lembur_pns', 0)
                tidak_match = st.session_state.get('tidak_match_lembur_pns', 0)
                
                # Info file
                st.info(f"📄 **Nama file:** BP21_Pajak_Lembur_PNS_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                st.info(f"📊 **Jumlah data:** {jumlah_data} baris")
                
                if tidak_match > 0:
                    st.warning(f"⚠️ **Data tidak match:** {tidak_match} baris (tidak termasuk dalam file)")
                
                # Buat file Excel dengan format warna menggunakan openpyxl
                output = BytesIO()
                
                # Buat workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "BP21_Pajak_Lembur_PNS"
                
                # Tulis header
                headers = list(hasil_final.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                
                # Tulis data
                for row_num, row_data in enumerate(hasil_final.values, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Tentukan kolom untuk warna
                # Kolom yang diwarnai oranye (harus diisi manual) - HANYA 3 KOLOM SEKARANG
                orange_columns = ['Nomor Dok. Referensi', 'Tanggal Dok. Referensi', 'Tanggal Pemotongan']
                
                # Cari indeks kolom orange
                orange_indices = []
                for col_name in orange_columns:
                    if col_name in headers:
                        orange_indices.append(headers.index(col_name) + 1)  # +1 karena openpyxl mulai dari 1
                
                # Terapkan warna untuk semua sel data (baris 2 ke atas)
                for row in ws.iter_rows(min_row=2, max_row=len(hasil_final) + 1, 
                                       min_col=1, max_col=len(headers)):
                    for cell in row:
                        # Jika kolom ini termasuk orange columns, beri warna orange
                        if cell.column in orange_indices:
                            cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                        else:
                            # Untuk kolom lainnya, beri warna hijau muda
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Simpan workbook ke BytesIO
                wb.save(output)
                output.seek(0)
                
                # Tombol download dengan format warna
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="📥 **Download Hasil (Excel dengan Format Warna)**",
                        data=output,
                        file_name=f"BP21_Pajak_Lembur_PNS_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        help="File Excel dengan format warna: Hijau untuk data sistem, Oranye untuk data manual"
                    )
                
                # Informasi tambahan
                with st.expander("ℹ️ Informasi Format File yang Didownload", expanded=False):
                    st.markdown(f"""
                    ### **🎨 Format Warna di File Excel:**
                    
                    **1. Header (Baris 1):**
                    - Warna: **Hijau Tua** (#C6E0B4)
                    - Font: **Bold**
                    
                    **2. Data Hasil Sistem (Siap Copy):**
                    - Warna: **Hijau Muda** (#E6FFE6)
                    - Data ini sudah diproses oleh sistem
                    - Bisa langsung digunakan/dicopy
                    - Total: **{len(headers) - 3} kolom** dari **{len(headers)} kolom**
                    
                    **3. Data yang Harus Diisi Manual:**
                    - Warna: **Oranye** (#FFD580)
                    - Kolom yang perlu diisi manual:
                      1. **Nomor Dok. Referensi** → dari Nomor SP2D
                      2. **Tanggal Dok. Referensi** → dari Tanggal SP2D
                      3. **Tanggal Pemotongan** → dari Tanggal Invoice
                    
                    **4. PERUBAHAN PENTING:**
                    - **ID TKU Pemotong** sekarang menggunakan nilai default: 0001658723701000000000
                    - Warna **Hijau Muda** (data sistem)
                    - Tidak perlu diisi manual
                    
                    ### **📋 Kolom dengan Warna Hijau Muda (Data Sistem):**
                    {', '.join([f'**{col}**' for col in headers if col not in orange_columns])}
                    
                    ### **📋 Kolom dengan Warna Oranye (Data Manual):**
                    {', '.join([f'**{col}**' for col in orange_columns if col in headers])}
                    
                    ### **Sumber Data Manual:**
                    - **DAFTAR SP2D SATKER** (download dari sistem keuangan)
                    - **Khusus lembur**: ID TKU Pemotong menggunakan nilai default 0001658723701000000000
                    """)
                
                # Tombol reset
                if st.button("🔄 Reset Proses", use_container_width=True):
                    if 'hasil_pajak_lembur_pns' in st.session_state:
                        del st.session_state.hasil_pajak_lembur_pns
                    if 'new_data_df' in st.session_state:
                        del st.session_state.new_data_df
                    if 'duplicate_nips_df' in st.session_state:
                        del st.session_state.duplicate_nips_df
                    st.rerun()
        
        except Exception as e:
            st.error(f"❌ **ERROR**: Terjadi kesalahan saat membaca file")
            st.error(f"Detail: {str(e)}")
            st.info("💡 **Tips**: Pastikan file Excel tidak corrupt. Coba buka file di Excel terlebih dahulu.")
    
    elif uploaded_file_raw is not None:
        st.warning("⚠️ **Langkah 2**: Silakan upload juga **Data Master** untuk melanjutkan")
    elif uploaded_file_master is not None:
        st.warning("⚠️ **Langkah 1**: Silakan upload **Data Mentah** terlebih dahulu")
    
    st.markdown("---")
    st.caption("""
    🔧 **Dukungan Teknis**: 
    - **NIP duplikat**: Pastikan tidak ada NIP yang sama di Data Mentah
    - **Data baru**: NIP yang tidak ada di Data Master akan ditandai hijau dan harus ditambahkan melalui Croscheck PNS
    - Pastikan format NIP konsisten di kedua file
    - Untuk lembur, pastikan Kode Objek Pajak sesuai dengan ketentuan lembur
    - Kolom 'pajak' di Data Mentah adalah nilai PPh yang sudah dipotong
    - **ID TKU Pemotong** sekarang menggunakan nilai default: 0001658723701000000000
    - Kolom manual (orange) yang perlu diisi: Nomor SP2D, Tanggal SP2D, dan Tanggal Invoice
    - Jika ada error, coba download template dan sesuaikan data Anda dengan format yang diberikan
    """)