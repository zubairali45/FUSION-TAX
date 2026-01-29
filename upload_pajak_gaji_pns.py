import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Header definitions
HEADERS_MENTAH = [
    "kdsatker", "kdanak", "kdsubanak", "bulan", "tahun", "nogaji", "kdjns", "nip", "nmpeg",
    "kdduduk", "kdgol", "npwp", "nmrek", "nm_bank", "rekening", "kdbankspan", "nmbankspan",
    "kdpos", "kdnegara", "kdkppn", "tipesup", "gjpokok", "tjistri", "tjanak", "tjupns",
    "tjstruk", "tjfungs", "tjdaerah", "tjpencil", "tjlain", "tjkompen", "pembul", "tjberas",
    "tjpph", "potpfkbul", "potpfk2", "GajiKotor", "potpfk10", "potpph", "potswrum",
    "potkelbtj", "potlain", "pottabrum", "bersih", "sandi", "kdkawin",
    "kdjab", "thngj", "kdgapok", "bpjs", "bpjs2"
]

# DIUBAH: Urutan header sesuai permintaan
HEADERS_BPMP = [
    "Masa Pajak", "Tahun Pajak", "Status Pegawai", "NPWP/NIK/TIN",
    "Nomor Passport", "Status", "Posisi", "Sertifikat/Fasilitas", "Kode Objek Pajak", 
    "Penghasilan Kotor", "Tarif", "ID TKU", "Tgl Pemotongan", "TER A", "TER B", "TER C"
]

HEADERS_MASTER = [
    "No", "PNS/PPPK", "Nama", "NIK", "ID PENERIMA TKU", "KDGOL", "KODE OBJEK PAJAK",
    "KDKAWIN", "STATUS", "NIP", "nmrek", "nm_bank", "rekening", "kdbankspan",
    "nmbankspan", "kdpos", "AKTIF/TIDAK", "Keterangan"
]

# Kolom yang wajib ada (tidak termasuk yang opsional)
REQUIRED_MASTER = ["NIP", "NIK", "STATUS"]

# Kolom untuk menghitung gaji jika tidak ada kolom gajikotor
GAJI_COMPONENTS = [
    "gjpokok", "tjistri", "tjanak", "tjupns",
    "tjstruk", "tjfungs", "tjdaerah", "tjpencil", "tjlain", "tjkompen", 
    "pembul", "tjberas", "tjpph", "potpfkbul", "potpfk2"
]

# Kolom wajib untuk data mentah
REQUIRED_MENTAH = ["nip", "bulan", "tahun"]

def validate_headers(df, expected_headers, file_type):
    """Validasi header file yang diupload"""
    df_headers = df.columns.tolist()
    
    # Untuk data master, cek kolom wajib saja
    if file_type == "Data Master":
        missing = [h for h in REQUIRED_MASTER if h not in df_headers]
        if missing:
            st.error(f"❌ Header wajib yang hilang di file {file_type}: {', '.join(missing)}")
            return False
            
        return True
    
    # Untuk data mentah, cek kolom wajib dan beri peringatan jika GajiKotor tidak ada
    elif file_type == "Data Mentah":
        # Cek kolom wajib
        missing_required = [h for h in REQUIRED_MENTAH if h not in df_headers]
        if missing_required:
            st.error(f"❌ Header wajib yang hilang di file {file_type}: {', '.join(missing_required)}")
            return False
        
        # Cek apakah GajiKotor ada
        if 'GajiKotor' not in df_headers and 'gajikotor' not in df_headers:
            st.warning("⚠️ Kolom 'GajiKotor' tidak ditemukan. Sistem akan menghitung otomatis dari komponen gaji.")
            
            # Cek apakah komponen gaji yang diperlukan ada
            missing_components = []
            for component in GAJI_COMPONENTS:
                if component not in df_headers:
                    missing_components.append(component)
            
            if missing_components:
                st.warning(f"⚠️ Beberapa komponen gaji tidak ditemukan: {', '.join(missing_components)}")
                st.info("Sistem akan menganggap nilai 0 untuk komponen yang tidak ditemukan.")
        
        # Cek header tambahan
        extra = [h for h in df_headers if h not in expected_headers]
        if extra:
            st.warning(f"⚠️ Header tambahan di file {file_type} (akan diabaikan): {', '.join(extra)}")
        
        return True
    
    # Untuk data lain, validasi seperti biasa
    else:
        missing = [h for h in expected_headers if h not in df_headers]
        extra = [h for h in df_headers if h not in expected_headers]
        
        if missing:
            st.error(f"❌ Header yang hilang di file {file_type}: {', '.join(missing)}")
            return False
        if extra:
            st.warning(f"⚠️ Header tambahan di file {file_type} (akan diabaikan): {', '.join(extra)}")
        
        return True

def check_duplicate_nips(df_mentah):
    """Cek NIP duplikat di data mentah dan return baris yang duplikat"""
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    duplicates = df_mentah[df_mentah.duplicated('nip_clean', keep=False)]
    return duplicates

def check_new_data(df_mentah, df_master):
    """Cek NIP yang ada di data mentah tapi tidak ada di data master"""
    # Bersihkan NIP
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    df_master['NIP_clean'] = df_master['NIP'].astype(str).str.strip()
    
    # Cari NIP yang tidak ada di master
    master_nips = set(df_master['NIP_clean'].tolist())
    mentah_nips = set(df_mentah['nip_clean'].tolist())
    
    new_nips = mentah_nips - master_nips
    new_data = df_mentah[df_mentah['nip_clean'].isin(new_nips)]
    
    return new_data

def calculate_gaji_kotor(row):
    """Hitung gaji kotor jika kolom gajikotor tidak ada"""
    total = 0
    missing_columns = []
    
    for col in GAJI_COMPONENTS:
        if col in row and pd.notna(row[col]):
            try:
                total += float(row[col])
            except (ValueError, TypeError):
                missing_columns.append(f"{col} (nilai tidak valid)")
        else:
            missing_columns.append(col)
    
    if missing_columns:
        st.warning(f"⚠️ Kolom {', '.join(missing_columns)} tidak ditemukan atau nilainya tidak valid untuk perhitungan gaji. Dianggap 0.")
    
    return total

def process_data_to_bpmp(df_mentah, df_master):
    """Proses data mentah dan master menjadi format BPMP"""
    try:
        # List untuk menyimpan hasil
        hasil_bpmp = []
        
        # Counter untuk tracking
        total_mentah = len(df_mentah)
        berhasil = 0
        gagal = 0
        
        # Progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Informasi perhitungan gaji
        gunakan_perhitungan_sistem = False
        if 'GajiKotor' not in df_mentah.columns and 'gajikotor' not in df_mentah.columns:
            gunakan_perhitungan_sistem = True
            st.info("ℹ️ Menggunakan perhitungan sistem untuk Penghasilan Kotor")
        
        # Loop setiap baris di data mentah
        for idx, row_mentah in df_mentah.iterrows():
            nip_mentah = str(row_mentah['nip']).strip()
            
            # Cari NIP di data master
            df_master['NIP'] = df_master['NIP'].astype(str).str.strip()
            match_master = df_master[df_master['NIP'] == nip_mentah]
            
            if not match_master.empty:
                row_master = match_master.iloc[0]
                
                # Tentukan gaji kotor
                gaji_kotor = 0
                if 'gajikotor' in row_mentah and pd.notna(row_mentah['gajikotor']):
                    gaji_kotor = row_mentah['gajikotor']
                elif 'GajiKotor' in row_mentah and pd.notna(row_mentah['GajiKotor']):
                    gaji_kotor = row_mentah['GajiKotor']
                else:
                    # Hitung otomatis dari komponen gaji
                    gaji_kotor = calculate_gaji_kotor(row_mentah)
                
                # DIUBAH: Buat dictionary untuk baris BPMP sesuai urutan baru
                bpmp_row = {
                    "Masa Pajak": row_mentah['bulan'],
                    "Tahun Pajak": row_mentah['tahun'],
                    "Status Pegawai": "Resident",
                    "NPWP/NIK/TIN": row_master['NIK'],
                    "Nomor Passport": "",
                    "Status": row_master['STATUS'],
                    "Posisi": "pns",
                    "Sertifikat/Fasilitas": "DTP",
                    "Kode Objek Pajak": "21-100-01",
                    "Penghasilan Kotor": gaji_kotor,
                    "Tarif": "",
                    "ID TKU": "0001658723701000000000",
                    "Tgl Pemotongan": "",
                    "TER A": "",
                    "TER B": "",
                    "TER C": ""
                }
                
                hasil_bpmp.append(bpmp_row)
                berhasil += 1
            else:
                gagal += 1
                if gagal <= 10:
                    st.warning(f"⚠️ NIP {nip_mentah} tidak ditemukan di data master (baris {idx+1})")
            
            # Update progress
            progress = (idx + 1) / total_mentah
            progress_bar.progress(progress)
            status_text.text(f"Memproses: {idx+1}/{total_mentah} | Berhasil: {berhasil} | Gagal: {gagal}")
        
        if gagal > 10:
            st.warning(f"⚠️ ... dan {gagal - 10} NIP lainnya tidak ditemukan")
        
        progress_bar.empty()
        status_text.empty()
        
        # Convert ke DataFrame dengan urutan kolom yang benar
        if hasil_bpmp:
            # DIUBAH: Pastikan DataFrame dibuat dengan urutan HEADERS_BPMP
            df_hasil = pd.DataFrame(hasil_bpmp, columns=HEADERS_BPMP)
            
            # Tampilkan informasi tentang perhitungan gaji
            if gunakan_perhitungan_sistem:
                st.success(f"✅ Penghasilan Kotor dihitung otomatis dari {len(GAJI_COMPONENTS)} komponen gaji")
            
            return df_hasil, berhasil, gagal
        else:
            return None, 0, gagal
            
    except Exception as e:
        st.error(f"❌ Error saat memproses data: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None, 0, 0

def convert_df_to_excel(df):
    """Convert DataFrame ke Excel dengan styling warna sesuai permintaan"""
    output = BytesIO()
    
    # Buat workbook dan worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data BPMP'
    
    # Definisikan warna-warna
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Merah untuk header
    red_font = Font(color='FFFFFF', bold=True)  # Font putih untuk header merah
    bold_font = Font(bold=True)  # Font tebal untuk header lainnya
    green_fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')  # Hijau muda untuk isi baris
    
    # Kolom yang harus berwarna merah di header
    red_header_columns = ["Tarif", "TER A", "TER B", "TER C"]
    
    # Kolom yang isi barisnya harus hijau muda (kolom 1-10)
    green_data_columns = [
        "Masa Pajak", "Tahun Pajak", "Status Pegawai", "NPWP/NIK/TIN",
        "Nomor Passport", "Status", "Posisi", "Sertifikat/Fasilitas", 
        "Kode Objek Pajak", "Penghasilan Kotor", "ID TKU"
    ]
    
    # Kolom yang isi barisnya harus merah (kolom 11-16)
    red_data_columns = ["Tarif", "TER A", "TER B", "TER C"]
    
    # Tulis header
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        
        # Beri warna merah untuk kolom tertentu
        if header in red_header_columns:
            cell.fill = red_fill
            cell.font = red_font
        else:
            cell.font = bold_font
    
    # Tulis data dengan pewarnaan
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Dapatkan nama kolom berdasarkan indeks
            column_name = df.columns[col_idx-1]
            
            # Beri warna hijau muda untuk isi baris kolom 1-10
            if column_name in green_data_columns:
                cell.fill = green_fill
            # Beri warna merah untuk isi baris kolom tertentu
            elif column_name in red_data_columns:
                cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')  # Merah muda untuk isi
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Simpan ke BytesIO
    wb.save(output)
    output.seek(0)
    return output

def create_template_mentah():
    """Membuat template Excel untuk data mentah"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Data Mentah'
    
    # Data contoh untuk template
    data = {
        'nip': ['123456789012345678'],
        'bulan': [1],
        'tahun': [2024],
        'gjpokok': [5000000],
        'tjistri': [300000],
        'tjanak': [200000],
        'tjupns': [100000],
        'tjstruk': [50000],
        'tjfungs': [40000],
        'tjdaerah': [30000],
        'tjpencil': [20000],
        'tjlain': [10000],
        'tjkompen': [5000],
        'pembul': [3000],
        'tjberas': [25000],
        'tjpph': [15000],
        'potpfkbul': [2000],
        'potpfk2': [1000],
        'kdsatker': ['123456'],
        'nogaji': ['001'],
        'nmpeg': ['Nama Contoh'],
        'kdgol': ['IV/a'],
        'npwp': ['123456789012345'],
        'kdkawin': ['K']
    }
    
    # Tulis header
    for col_idx, header in enumerate(data.keys(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Tulis data
    for col_idx, (key, values) in enumerate(data.items(), 1):
        ws.cell(row=2, column=col_idx, value=values[0])
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def create_template_master():
    """Membuat template Excel untuk data master"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Data Master'
    
    # Data contoh untuk template
    data = {
        'NIP': ['123456789012345678', '987654321098765432'],
        'NIK': ['123456789012345', '987654321098765'],
        'STATUS': ['K', 'TK'],
        'KDKAWIN': ['K', 'TK'],
        'Nama': ['Nama Contoh 1', 'Nama Contoh 2'],
        'PNS/PPPK': ['PNS', 'PNS']
    }
    
    # Tulis header
    for col_idx, header in enumerate(data.keys(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Tulis data (2 baris contoh)
    for row_idx in range(2):
        for col_idx, (key, values) in enumerate(data.items(), 1):
            ws.cell(row=row_idx+2, column=col_idx, value=values[row_idx])
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def show():
    # Header dengan tombol kembali
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("← Kembali ke Dashboard PNS"):
            st.session_state.current_page = 'dashboard_pns'
            st.rerun()
    
    st.title("💰 Upload Pajak Gaji PNS")
    st.markdown("---")
    
    # ========== PANDUAN PENGGUNAAN ==========
    with st.expander("📚 Panduan Penggunaan - Baca Sebelum Mulai", expanded=False):
        st.markdown("""
        ### **🎯 Cara Menggunakan Fitur Ini:**
        
        1. **Siapkan 2 File Excel**:
           - **Data Mentah Gaji**: Hasil download sistem penggajian per bulan
           - **Data Master**: Database pegawai PNS
        
        2. **Format Data Mentah (WAJIB)**:
           ```
           Kolom yang HARUS ADA (minimal):
           - nip         : Nomor Induk Pegawai (tidak boleh duplikat)
           - bulan       : Bulan transaksi (1-12)
           - tahun       : Tahun transaksi
           
           Kolom Gaji Kotor (PILIH SALAH SATU):
           - OPTIONAL: gajikotor atau GajiKotor → Total penghasilan kotor (manual/opsional)
           - JIKA TIDAK ADA: Sistem hitung otomatis dari 15 komponen gaji:
               gjpokok, tjistri, tjanak, tjupns, tjstruk, tjfungs, tjdaerah,
               tjpencil, tjlain, tjkompen, pembul, tjberas, tjpph, potpfkbul, potpfk2
           ```
           
        3. **Format Data Master (WAJIB)**:
           ```
           Kolom yang HARUS ADA:
           - NIP               : Nomor Induk Pegawai (untuk matching)
           - NIK               : Digunakan sebagai NPWP/NIK/TIN di output
           - STATUS            : Status PTKP (K/TK/HB/etc.)
           
           Kolom PENTING:
           - KDKAWIN          : Kode status kawin (untuk dibandingkan dengan kdkawin di data mentah)
           
           Kolom tambahan (jika ada):
           - PNS/PPPK       : Jenis pegawai
           - Nama           : Nama lengkap
           - KODE OBJEK PAJAK : Kode objek pajak
           ```
        
        4. **PROSES OTOMATIS**:
           - Sistem akan melakukan **inner join** berdasarkan NIP
           - **NIP duplikat** di data mentah akan ditandai merah
           - **Data baru** (NIP tidak ada di master) akan ditandai hijau
           - **Perbedaan kdkawin** akan ditandai kuning
           - **ID TKU akan otomatis diisi dengan nilai default**: 0001658723701000000000
           - **Penghasilan Kotor**: Jika tidak ada kolom gajikotor, sistem hitung otomatis
        
        5. **FORMAT OUTPUT (BPMP)**:
           ```
           Hasil akan dikonversi ke 16 kolom format BPMP:
           1. Masa Pajak → dari kolom 'bulan'
           2. Tahun Pajak → dari kolom 'tahun'
           3. Status Pegawai → default "Resident"
           4. NPWP/NIK/TIN → dari kolom 'NIK' di Data Master
           5. Nomor Passport → kosong
           6. Status → dari kolom 'STATUS' di Data Master
           7. Posisi → default "pns"
           8. Sertifikat/Fasilitas → default "DTP"
           9. Kode Objek Pajak → default "21-100-01"
           10. Penghasilan Kotor → dari 'gajikotor' atau hasil perhitungan sistem
           11. Tarif → kosong (berisi rumus di aplikasi BPMP)
           12. ID TKU → default "0001658723701000000000" (untuk semua data)
           13. Tgl Pemotongan → kosong
           14-16. TER A, TER B, TER C → kosong (berisi rumus di aplikasi BPMP)
           ```
        
        6. **🎨 KODE WARNA DI HASIL DOWNLOAD**:
           **HEADER:**
           - **Merah dengan font putih**: Tarif, TER A, TER B, TER C
           - **Hitam tebal**: Header lainnya
           
           **ISI DATA:**
           - **Hijau Muda (Soft Green)**: Kolom 1-10 + ID TKU (data hasil sistem)
           - **Merah Muda (Light Red)**: Kolom 11-16 (berisi rumus di aplikasi BPMP)
           
           **⚠️ PERHATIAN KHUSUS:**
           - Kolom berwarna merah/muda mengandung **RUMUS/FORMULA** di aplikasi BPMP
           - **JANGAN DISALIN** kolom tersebut saat menyalin ke aplikasi BPMP
           - Hanya salin kolom yang berwarna hijau muda dan kolom tanpa warna
        """)
        
        # Template untuk download - hanya format Excel (xlsx)
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 Template Data Mentah (Excel)",
                data=create_template_mentah(),
                file_name="template_data_mentah_gaji_pns.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            st.download_button(
                label="📥 Template Data Master (Excel)",
                data=create_template_master(),
                file_name="template_data_master_gaji_pns.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.markdown("---")
    
    # ========== INFORMASI PENTING ==========
    st.info("""
    **📊 INFORMASI PENTING:**
    
    **1. PROSES MATCHING DATA:**
    - Sistem menggunakan **inner join** berdasarkan NIP
    - **NIP duplikat** di data mentah akan ditandai merah dan harus diperbaiki
    - **Data baru** (NIP tidak ada di master) akan ditandai hijau dan perlu ditambahkan ke master
    - **Perbedaan kdkawin** antara mentah dan master akan ditandai kuning
    - Proses matching bersifat case-sensitive dan sensitive terhadap spasi
    
    **2. PERHITUNGAN GAJI KOTOR:**
    - **Prioritas 1**: Jika kolom **gajikotor** (lowercase) ada, gunakan nilai tersebut
    - **Prioritas 2**: Jika kolom **GajiKotor** (original) ada, gunakan nilai tersebut
    - **Prioritas 3**: Jika tidak ada kolom gajikotor, sistem hitung otomatis dari 15 komponen gaji:
      ```
      Penghasilan Kotor = SUM(
          gjpokok, tjistri, tjanak, tjupns, tjstruk, tjfungs, 
          tjdaerah, tjpencil, tjlain, tjkompen, pembul, tjberas, 
          tjpph, potpfkbul, potpfk2
      )
      ```
    - Komponen gaji yang hilang akan dianggap 0 dan muncul warning
    
    **3. KOLOM DEFAULT YANG AKAN DIHASILKAN:**
    ```
    Status Pegawai    : "Resident" (untuk WNI)
    Posisi           : "pns" (identitas sebagai PNS)
    Sertifikat       : "DTP" (fasilitas yang digunakan)
    Kode Objek Pajak : "21-100-01" (kode standar)
    ID TKU           : "0001658723701000000000" (nilai default untuk semua data)
    ```
    
    **4. DETEKSI PERUBAHAN:**
    - **kdkawin** di data mentah vs **KDKAWIN** di master
    - Jika berbeda: baris di data mentah akan ditandai kuning
    - Ini menunjukkan ada perubahan status yang perlu diupdate di master
    """)
    
    # ========== PERINGATAN KHUSUS TENTANG WARNA ==========
    st.warning("""
    **🎯 PANDUAN PENYALINAN KE APLIKASI BPMP:**
    
    **SALIN kolom ini saja (Hijau Muda & Tanpa Warna):**
    1. Masa Pajak
    2. Tahun Pajak
    3. Status Pegawai
    4. NPWP/NIK/TIN
    5. Nomor Passport
    6. Status
    7. Posisi
    8. Sertifikat/Fasilitas
    9. Kode Objek Pajak
    10. Penghasilan Kotor
    12. ID TKU (selalu berisi: 0001658723701000000000)
    13. Tgl Pemotongan (jika ada)
    
    **JANGAN SALIN kolom ini (Merah Muda):**
    11. Tarif (berisi rumus)
    14. TER A (berisi rumus)
    15. TER B (berisi rumus)
    16. TER C (berisi rumus)
    
    **Alasan:** Kolom merah muda akan dihitung otomatis oleh aplikasi BPMP menggunakan rumus/formula internal.
    """)
    
    # Initialize session state
    if 'df_mentah' not in st.session_state:
        st.session_state.df_mentah = None
    if 'df_bpmp' not in st.session_state:
        st.session_state.df_bpmp = None
    if 'df_master' not in st.session_state:
        st.session_state.df_master = None
    if 'df_hasil' not in st.session_state:
        st.session_state.df_hasil = None
    if 'new_data_df' not in st.session_state:
        st.session_state.new_data_df = None
    if 'duplicate_nips_df' not in st.session_state:
        st.session_state.duplicate_nips_df = None
    
    # ===== BAGIAN 1: UPLOAD DATA MENTAH =====
    st.subheader("📤 1. Upload Data Mentah Gaji")
    st.markdown("File Excel yang berisi data gaji mentah dari sistem penggajian")
    
    with st.expander("ℹ️ Detail Kolom Data Mentah", expanded=False):
        st.markdown("""
        **Kolom WAJIB ada (minimal):**
        - `nip` : Nomor Induk Pegawai (kunci join, TIDAK BOLEH DUPLIKAT)
        - `bulan` : Bulan transaksi (1-12)
        - `tahun` : Tahun transaksi (4 digit)
        
        **Kolom Gaji Kotor (OPSIONAL - PILIH SALAH SATU):**
        - `gajikotor` atau `GajiKotor` : Total penghasilan kotor (manual)
        **ATAU** semua komponen gaji untuk dihitung sistem:
        - `gjpokok`, `tjistri`, `tjanak`, `tjupns`, `tjstruk`, `tjfungs`, `tjdaerah`,
          `tjpencil`, `tjlain`, `tjkompen`, `pembul`, `tjberas`, `tjpph`, `potpfkbul`, `potpfk2`
        
        **Kolom PENTING lainnya:**
        - `kdsatker` : Kode satker
        - `nogaji` : Nomor gaji
        - `nmpeg` : Nama pegawai
        - `kdgol` : Kode golongan
        - `npwp` : NPWP (jika ada)
        - `kdkawin` : Kode status kawin (untuk dibandingkan dengan master)
        
        **Total kolom yang direkomendasikan:** 51 kolom (sesuai sistem penggajian)
        """)
    
    uploaded_mentah = st.file_uploader(
        "**Pilih file Data Mentah Gaji**",
        type=['xlsx', 'xls'],
        key="mentah_uploader",
        help="Upload file Excel (xlsx atau xls) hasil download sistem penggajian"
    )
    
    if uploaded_mentah:
        try:
            # Hanya membaca file Excel (xlsx, xls)
            df = pd.read_excel(uploaded_mentah)
            
            if validate_headers(df, HEADERS_MENTAH, "Data Mentah"):
                # Cek duplikat NIP
                duplicates = check_duplicate_nips(df)
                if not duplicates.empty:
                    st.session_state.duplicate_nips_df = duplicates
                    st.error(f"❌ Ditemukan {len(duplicates)} NIP duplikat di Data Mentah!")
                    
                    # Tampilkan baris duplikat dengan warna merah
                    st.warning("**Baris dengan NIP duplikat (ditandai merah):**")
                    
                    # Buat DataFrame dengan styling untuk duplikat
                    def highlight_duplicates(df_original, duplicates_df):
                        # Buat salinan untuk styling
                        styled_df = df_original.copy()
                        
                        # Tandai baris duplikat
                        mask = styled_df['nip'].astype(str).str.strip().isin(
                            duplicates_df['nip_clean'].unique()
                        )
                        
                        # Buat list warna
                        colors = ['background-color: #ffcccc' if mask.iloc[i] else '' 
                                 for i in range(len(styled_df))]
                        
                        return styled_df.style.apply(lambda x: colors, axis=0)
                    
                    styled_duplicates = highlight_duplicates(df.head(50), duplicates)
                    st.dataframe(styled_duplicates, use_container_width=True)
                    
                    if len(df) > 50:
                        st.info(f"Menampilkan 50 baris pertama dari total {len(df)} baris")
                    
                    st.error("**PERBAIKI NIP DUPLIKAT SEBELUM MELANJUTKAN!**")
                    st.session_state.df_mentah = None
                else:
                    st.session_state.df_mentah = df
                    st.success(f"✅ File '{uploaded_mentah.name}' berhasil diupload!")
                    st.info(f"📊 Total data: {len(df)} baris")
                    
                    # Cek dan tampilkan informasi tentang gajikotor
                    if 'GajiKotor' not in df.columns and 'gajikotor' not in df.columns:
                        st.info("ℹ️ **Kolom GajiKotor tidak ditemukan.** Sistem akan menghitung otomatis dari komponen gaji.")
                        
                        # Tampilkan contoh komponen gaji yang ada
                        komponen_ada = [col for col in GAJI_COMPONENTS if col in df.columns]
                        if komponen_ada:
                            st.success(f"✅ Ditemukan {len(komponen_ada)} komponen gaji untuk perhitungan otomatis")
                        else:
                            st.warning("⚠️ Tidak ditemukan komponen gaji untuk perhitungan. Hasil Penghasilan Kotor akan 0.")
                    
                    with st.expander("👁️ Preview Data Mentah"):
                        # Tampilkan semua data
                        st.dataframe(df, use_container_width=True)
                        st.caption(f"Menampilkan semua {len(df)} baris data")
            else:
                st.session_state.df_mentah = None
                
        except Exception as e:
            st.error(f"❌ Error membaca file: {str(e)}")
            st.session_state.df_mentah = None
    
    st.markdown("---")
    
    # ===== BAGIAN 2: UPLOAD DATA MASTER =====
    st.subheader("📋 2. Upload Data Master PNS")
    st.markdown("File Excel yang berisi data master pegawai (NIP, NIK, Status, dll)")
    
    with st.expander("ℹ️ Detail Kolom Data Master", expanded=False):
        st.markdown("""
        **Kolom WAJIB ada:**
        - `NIP` : Nomor Induk Pegawai (kunci join)
        - `NIK` : Akan digunakan sebagai NPWP/NIK/TIN di output
        - `STATUS` : Status PTKP (K/TK/HB/etc.)
        
        **Kolom PENTING:**
        - `KDKAWIN` : Kode status kawin (untuk dibandingkan dengan kdkawin di data mentah)
        
        **Kolom tambahan (jika ada):**
        - `PNS/PPPK` : Jenis pegawai
        - `Nama` : Nama lengkap
        - `KODE OBJEK PAJAK` : Kode objek pajak
        """)
    
    uploaded_master = st.file_uploader(
        "**Pilih file Data Master**",
        type=['xlsx', 'xls'],
        key="master_uploader",
        help="Upload database master pegawai PNS dalam format Excel (xlsx atau xls)"
    )
    
    if uploaded_master:
        try:
            # Hanya membaca file Excel (xlsx, xls)
            df = pd.read_excel(uploaded_master)
            
            if validate_headers(df, HEADERS_MASTER, "Data Master"):
                st.session_state.df_master = df
                st.success(f"✅ File '{uploaded_master.name}' berhasil diupload!")
                st.info(f"📊 Total pegawai: {len(df)} orang")
                
                with st.expander("👁️ Preview Data Master"):
                    # Tampilkan semua data
                    st.dataframe(df, use_container_width=True)
                    st.caption(f"Menampilkan semua {len(df)} baris data")
            else:
                st.session_state.df_master = None
                
        except Exception as e:
            st.error(f"❌ Error membaca file: {str(e)}")
            st.session_state.df_master = None
    
    st.markdown("---")
    
    # ===== BAGIAN 3: DETEKSI DATA BARU DAN PERBEDAAN =====
    if st.session_state.df_mentah is not None and st.session_state.df_master is not None:
        st.subheader("🔍 3. Deteksi Data Baru dan Perbedaan")
        
        # Cek data baru (NIP di mentah tapi tidak di master)
        new_data = check_new_data(st.session_state.df_mentah, st.session_state.df_master)
        
        if not new_data.empty:
            st.session_state.new_data_df = new_data
            st.warning(f"⚠️ Ditemukan {len(new_data)} data baru di Data Mentah yang tidak ada di Data Master!")
            
            # Tampilkan data baru dengan warna hijau
            st.info("**Data baru (ditandai hijau - perlu ditambahkan ke Data Master):**")
            
            def highlight_new_data(df_original, new_data_df):
                # Buat salinan untuk styling
                styled_df = df_original.copy()
                
                # Tandai baris data baru
                mask = styled_df['nip'].astype(str).str.strip().isin(
                    new_data_df['nip_clean'].unique()
                )
                
                # Buat list warna
                colors = ['background-color: #ccffcc' if mask.iloc[i] else '' 
                         for i in range(len(styled_df))]
                
                return styled_df.style.apply(lambda x: colors, axis=0)
            
            styled_new_data = highlight_new_data(st.session_state.df_mentah.head(50), new_data)
            st.dataframe(styled_new_data, use_container_width=True)
            
            if len(st.session_state.df_mentah) > 50:
                st.info(f"Menampilkan 50 baris pertama dari total {len(st.session_state.df_mentah)} baris")
            
            # TOMBOL UNTUK MENUJU KE HALAMAN CROSSCHECK PNS
            st.markdown("---")
            st.error("**DATA BARU HARUS DITAMBAHKAN KE DATA MASTER SEBELUM MELANJUTKAN!**")
            
            # Tambahkan tombol untuk menuju ke halaman croscheck_pns
            if st.button("➕ Tambahkan Data Baru ke Master (Croscheck PNS)", type="primary", use_container_width=True):
                st.session_state.current_page = 'croscheck_pns'
                st.session_state.selected_menu = 'croscheck'
                st.rerun()
            
            st.stop()  # Hentikan proses sampai data baru ditangani
        
        # Cek perbedaan kdkawin antara mentah dan master
        st.session_state.df_mentah['nip_clean'] = st.session_state.df_mentah['nip'].astype(str).str.strip()
        st.session_state.df_master['NIP_clean'] = st.session_state.df_master['NIP'].astype(str).str.strip()
        
        # Gabungkan untuk membandingkan kdkawin
        merged = pd.merge(
            st.session_state.df_mentah[['nip_clean', 'kdkawin']],
            st.session_state.df_master[['NIP_clean', 'KDKAWIN']],
            left_on='nip_clean',
            right_on='NIP_clean',
            how='inner'
        )
        
        # Cari perbedaan
        if 'kdkawin' in merged.columns and 'KDKAWIN' in merged.columns:
            merged['kdkawin'] = merged['kdkawin'].astype(str).str.strip()
            merged['KDKAWIN'] = merged['KDKAWIN'].astype(str).str.strip()
            differences = merged[merged['kdkawin'] != merged['KDKAWIN']]
            
            if not differences.empty:
                st.warning(f"⚠️ Ditemukan {len(differences)} perbedaan kdkawin antara Data Mentah dan Data Master!")
                
                # Tampilkan perbedaan
                st.info("**Perbedaan kdkawin (ditandai kuning - perlu diperiksa):**")
                
                def highlight_kdkawin_differences(df_mentah_original, differences_df):
                    # Buat salinan untuk styling
                    styled_df = df_mentah_original.copy()
                    
                    # Tandai baris dengan perbedaan kdkawin
                    mask = styled_df['nip_clean'].isin(differences_df['nip_clean'])
                    
                    # Buat list warna (hanya kolom kdkawin yang kuning)
                    colors = []
                    for i in range(len(styled_df)):
                        if mask.iloc[i]:
                            row_colors = [''] * len(styled_df.columns)
                            if 'kdkawin' in styled_df.columns:
                                idx = list(styled_df.columns).index('kdkawin')
                                row_colors[idx] = 'background-color: #ffffcc'
                            colors.append(row_colors)
                        else:
                            colors.append([''] * len(styled_df.columns))
                    
                    return styled_df.style.apply(lambda x: colors[df_mentah_original.index.get_loc(x.name)] 
                                                if x.name in df_mentah_original.index else [''] * len(x), axis=1)
                
                styled_differences = highlight_kdkawin_differences(
                    st.session_state.df_mentah.head(50), 
                    differences
                )
                st.dataframe(styled_differences, use_container_width=True)
                
                # Tampilkan tabel perbandingan
                st.info("**Detail Perbandingan kdkawin:**")
                comparison_df = differences[['nip_clean', 'kdkawin', 'KDKAWIN']].copy()
                comparison_df.columns = ['NIP', 'kdkawin (Data Mentah)', 'KDKAWIN (Data Master)']
                st.dataframe(comparison_df, use_container_width=True)
                
                st.warning("**Data Mentah menggunakan nilai kdkawin yang baru!**")
        
        # Jika tidak ada data baru, lanjutkan
        st.success("✅ Tidak ditemukan data baru. Semua NIP di Data Mentah ada di Data Master.")
        
    # ===== BAGIAN 4: PROSES DATA =====
    st.subheader("🔄 4. Proses Data ke Format BPMP")
    
    # Cek apakah semua file sudah diupload
    if st.session_state.df_mentah is None:
        st.warning("⚠️ **Langkah 1**: Silakan upload Data Mentah terlebih dahulu")
    elif st.session_state.df_master is None:
        st.warning("⚠️ **Langkah 2**: Silakan upload Data Master terlebih dahulu")
    elif st.session_state.new_data_df is not None and not st.session_state.new_data_df.empty:
        st.warning("⚠️ **Langkah 3**: Ada data baru yang harus ditambahkan ke Data Master terlebih dahulu")
    else:
        st.success("✅ Semua file sudah siap!")
        
        # Tampilkan statistik
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Data Mentah", f"{len(st.session_state.df_mentah)} baris")
            st.caption("Data gaji yang akan diproses")
        with col2:
            st.metric("Data Master", f"{len(st.session_state.df_master)} pegawai")
            st.caption("Database referensi pegawai")
        
        st.markdown("---")
        
        # Informasi sebelum proses
        st.info("""
        **🔍 INFORMASI PROSES:**
        - Sistem akan melakukan **matching berdasarkan NIP**
        - **Penghasilan Kotor** akan diambil dari:
          1. Kolom `gajikotor` jika ada
          2. Jika tidak, dihitung otomatis dari 15 komponen gaji
        - Hanya data dengan NIP yang match di kedua file yang akan diproses
        - **ID TKU akan otomatis diisi dengan: 0001658723701000000000** (untuk semua data)
        - Progress bar akan menunjukkan status pemrosesan
        """)
        
        if st.button("🚀 **PROSES DATA KE FORMAT BPMP**", type="primary", use_container_width=True):
            with st.spinner("⏳ Memproses data..."):
                df_hasil, berhasil, gagal = process_data_to_bpmp(
                    st.session_state.df_mentah,
                    st.session_state.df_master
                )
                
                if df_hasil is not None:
                    st.session_state.df_hasil = df_hasil
                    
                    st.success("✅ Proses selesai!")
                    
                    # Tampilkan statistik hasil
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("✅ Berhasil Diproses", berhasil, 
                                 help="Data dengan NIP yang ditemukan di Data Master")
                    with col2:
                        st.metric("❌ Tidak Match", gagal,
                                 help="Data dengan NIP yang tidak ditemukan di Data Master")
                    with col3:
                        st.metric("📊 Total Output", len(df_hasil),
                                 help="Total baris yang akan dihasilkan")
                    
                    st.markdown("---")
                    
                    # Preview hasil
                    st.subheader("📄 Preview Hasil BPMP")
                    st.info(f"**Format kolom hasil ({len(df_hasil.columns)} kolom):**")
                    
                    # Tampilkan preview dataframe
                    preview_df = df_hasil.head(10).copy()
                    st.dataframe(preview_df)
                    
                    # Informasi warna di preview
                    st.info("""
                    **🎨 LEGENDA WARNA (di File Excel Hasil):**
                    
                    **Header:**
                    - **Merah dengan font putih**: Tarif, TER A, TER B, TER C
                    - **Hitam tebal**: Semua header lainnya
                    
                    **Isi Data:**
                    - **Hijau Muda (#C6EFCE)**: Kolom 1-10 + ID TKU (data hasil sistem)
                    - **Merah Muda (#FF9999)**: Tarif, TER A, TER B, TER C (berisi rumus)
                    
                    **⚠️ Ingat: Jangan salin kolom merah muda ke aplikasi BPMP!**
                    """)
                else:
                    st.error("❌ Tidak ada data yang berhasil diproses!")
    
    # ===== BAGIAN 5: DOWNLOAD HASIL =====
    if st.session_state.df_hasil is not None:
        st.markdown("---")
        st.subheader("💾 Download Hasil dengan Format Warna")
        
        df_hasil = st.session_state.df_hasil
        jumlah_data = len(df_hasil)
        
        # Info file
        st.info(f"📄 **Nama file:** Data_BPMP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        st.info(f"📊 **Jumlah data:** {jumlah_data} baris")
        
        # Peringatan khusus tentang kolom merah
        st.warning("""
        **⚠️ PERHATIAN KHUSUS - BACA SEBELUM MENYALIN KE APLIKASI BPMP:**
        
        **Kolom yang JANGAN DISALIN (berisi rumus di aplikasi BPMP):**
        1. **Tarif** (kolom 11)
        2. **TER A** (kolom 14)
        3. **TER B** (kolom 15)
        4. **TER C** (kolom 16)
        
        **Kolom yang HARUS DISALIN:**
        - **Kolom 1-10** (Hijau Muda)
        - **Kolom 12-13** (ID TKU dan Tgl Pemotongan - tanpa warna/hijau muda)
        
        **Cara menyalin yang benar:**
        1. Buka file hasil di Excel
        2. **SALIN** hanya kolom 1-10 + kolom 12-13
        3. **LEWATI** kolom 11 dan 14-16 (kolom merah muda)
        4. Paste ke aplikasi BPMP
        """)
        
        # Generate Excel file
        excel_file = convert_df_to_excel(df_hasil)
        filename = f"Data_BPMP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Tombol download
        st.download_button(
            label="📥 **DOWNLOAD FILE EXCEL BPMP**",
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            help="Download file Excel dengan format warna siap untuk aplikasi BPMP"
        )
        
        # Informasi tambahan
        with st.expander("ℹ️ Informasi Detail Format File", expanded=False):
            st.markdown(f"""
            ### **🎨 DETAIL FORMAT WARNA DI FILE EXCEL:**
            
            **HEADER (Baris 1):**
            - **Merah dengan font putih:** Tarif, TER A, TER B, TER C
            - **Hitam tebal:** Semua header lainnya
            
            **ISI DATA (Baris 2 ke atas):**
            - **Hijau Muda (#C6EFCE):** 
              {', '.join([f'**{col}**' for col in df_hasil.columns if col not in ['Tarif', 'TER A', 'TER B', 'TER C', 'Tgl Pemotongan']])}
            - **Merah Muda (#FF9999):** Tarif, TER A, TER B, TER C
            - **Tanpa warna:** Tgl Pemotongan
            
            ### **📋 MAPPING DATA:**
            1. **Masa Pajak** → `bulan` dari Data Mentah
            2. **Tahun Pajak** → `tahun` dari Data Mentah
            3. **Status Pegawai** → "Resident" (default untuk WNI)
            4. **NPWP/NIK/TIN** → `NIK` dari Data Master
            5. **Nomor Passport** → kosong (untuk WNI)
            6. **Status** → `STATUS` dari Data Master
            7. **Posisi** → "pns" (default)
            8. **Sertifikat/Fasilitas** → "DTP" (default)
            9. **Kode Objek Pajak** → "21-100-01" (default)
            10. **Penghasilan Kotor** → dari `gajikotor` atau hasil perhitungan sistem
            11. **Tarif** → kosong (rumus di aplikasi BPMP)
            12. **ID TKU** → **default "0001658723701000000000"** (untuk semua data)
            13. **Tgl Pemotongan** → kosong (diisi manual)
            14-16. **TER A, B, C** → kosong (rumus di aplikasi BPMP)
            
            ### **📊 STATISTIK DATA:**
            - **Total baris:** {jumlah_data}
            - **Total penghasilan kotor:** Rp {df_hasil['Penghasilan Kotor'].sum():,.0f}
            - **Rata-rata penghasilan:** Rp {df_hasil['Penghasilan Kotor'].mean():,.0f}
            - **Status unik:** {df_hasil['Status'].nunique()} jenis
            - **ID TKU:** Sama untuk semua data (0001658723701000000000)
            """)
        
        # Tombol reset
        if st.button("🔄 **Reset Proses & Upload File Baru**", use_container_width=True):
            if 'df_mentah' in st.session_state:
                del st.session_state.df_mentah
            if 'df_master' in st.session_state:
                del st.session_state.df_master
            if 'df_hasil' in st.session_state:
                del st.session_state.df_hasil
            if 'new_data_df' in st.session_state:
                del st.session_state.new_data_df
            if 'duplicate_nips_df' in st.session_state:
                del st.session_state.duplicate_nips_df
            st.rerun()
    
    # Footer
    st.markdown("---")
    st.caption("""
    🔧 **Dukungan Teknis**: 
    - **NIP duplikat**: Pastikan tidak ada NIP yang sama di Data Mentah
    - **Data baru**: NIP yang tidak ada di Data Master akan ditandai hijau dan harus ditambahkan
    - **Perbedaan kdkawin**: Baris dengan perbedaan akan ditandai kuning
    - **Perhitungan gaji**: Sistem otomatis hitung jika kolom `gajikotor` tidak ada
    - **ID TKU**: Tidak diambil dari Data Master, menggunakan nilai default untuk semua data
    - **Format file**: Hanya mendukung format Excel (.xlsx, .xls)
    - Untuk error: Coba download template dan sesuaikan data dengan format yang diberikan
    """)
    

if __name__ == "__main__":
    show()