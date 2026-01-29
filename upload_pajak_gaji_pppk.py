import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Header definitions untuk PPPK
HEADERS_MENTAH_PPPK = [
    "kdsatker", "kdanak", "kdsubanak", "bulan", "tahun", "nogaji", "kdjns", "nip", "nmpeg",
    "kdduduk", "kdgol", "npwp", "nmrek", "nm_bank", "rekening", "kdbankspan", "nmbankspan",
    "kdpos", "kdnegara", "kdkppn", "tipesup", "gjpokok", "tjistri", "tjanak", "tjupns",
    "tjstruk", "tjfungs", "tjdaerah", "tjpencil", "tjlain", "tjkompen", "pembul", "tjberas",
    "tjpph", "potpfkbul", "potpfk2", "GajiKotor", "potpfk10", "potpph", "potswrum",
    "potkelbtj", "potlain", "pottabrum", "bersih", "sandi", "kdkawin", "kdjab",
    "thngj", "kdgapok", "bpjs", "bpjs2"
]

# DIUBAH: Urutan header sesuai permintaan
HEADERS_BPMP = [
    "Masa Pajak", "Tahun Pajak", "Status Pegawai", "NPWP/NIK/TIN",
    "Nomor Passport", "Status", "Posisi", "Sertifikat/Fasilitas", "Kode Objek Pajak", 
    "Penghasilan Kotor", "Tarif", "ID TKU", "Tgl Pemotongan", "TER A", "TER B", "TER C"
]

HEADERS_MASTER = [
    "No", "PNS/PPPK", "Nama", "NIK", "ID PENERIMA TKU", "KDGOL", "KODE OBJEK PAJAK",
    "KDKAWIN", "STATUS", "NIP", "nmrek", "nm_bank", "rekening", "kdbankspan",
    "nmbankspan", "kdpos", "AKTIF/TIDAK", "Keterangan"
]

# Kolom yang wajib ada (tidak termasuk yang opsional)
REQUIRED_MASTER = ["NIP", "NIK", "STATUS"]

# Kolom untuk menghitung gaji jika tidak ada kolom gajikotor
GAJI_COMPONENTS = [
    "gjpokok", "tjistri", "tjanak", "tjupns",
    "tjstruk", "tjfungs", "tjdaerah", "tjpencil", "tjlain", "tjkompen", 
    "pembul", "tjberas", "tjpph", "potpfkbul", "potpfk2"
]

# Kolom wajib untuk data mentah
REQUIRED_MENTAH = ["nip", "bulan", "tahun"]

def validate_headers(df, expected_headers, file_type):
    """Validasi header file yang diupload"""
    df_headers = df.columns.tolist()
    
    # Untuk data master, cek kolom wajib saja
    if file_type == "Data Master":
        missing = [h for h in REQUIRED_MASTER if h not in df_headers]
        if missing:
            st.error(f"❌ Header wajib yang hilang di file {file_type}: {', '.join(missing)}")
            return False
            
        # Cek apakah ada ID TKU atau ID PENERIMA TKU
        if "ID TKU" not in df_headers and "ID PENERIMA TKU" not in df_headers:
            st.error(f"❌ File {file_type} harus memiliki kolom 'ID TKU' atau 'ID PENERIMA TKU'")
            return False
            
        return True
    
    # Untuk data mentah, cek kolom wajib dan beri peringatan jika GajiKotor tidak ada
    elif file_type == "Data Mentah":
        # Cek kolom wajib
        missing_required = [h for h in REQUIRED_MENTAH if h not in df_headers]
        if missing_required:
            st.error(f"❌ Header wajib yang hilang di file {file_type}: {', '.join(missing_required)}")
            return False
        
        # Cek apakah GajiKotor ada
        if 'GajiKotor' not in df_headers and 'gajikotor' not in df_headers:
            st.warning("⚠️ Kolom 'GajiKotor' tidak ditemukan. Sistem akan menghitung otomatis dari komponen gaji.")
            
            # Cek apakah komponen gaji yang diperlukan ada
            missing_components = []
            for component in GAJI_COMPONENTS:
                if component not in df_headers:
                    missing_components.append(component)
            
            if missing_components:
                st.warning(f"⚠️ Beberapa komponen gaji tidak ditemukan: {', '.join(missing_components)}")
                st.info("Sistem akan menganggap nilai 0 untuk komponen yang tidak ditemukan.")
        
        # Cek header tambahan
        extra = [h for h in df_headers if h not in expected_headers]
        if extra:
            st.warning(f"⚠️ Header tambahan di file {file_type} (akan diabaikan): {', '.join(extra)}")
        
        return True
    
    # Untuk data lain, validasi seperti biasa
    else:
        missing = [h for h in expected_headers if h not in df_headers]
        extra = [h for h in df_headers if h not in expected_headers]
        
        if missing:
            st.error(f"❌ Header yang hilang di file {file_type}: {', '.join(missing)}")
            return False
        if extra:
            st.warning(f"⚠️ Header tambahan di file {file_type} (akan diabaikan): {', '.join(extra)}")
        
        return True

def check_duplicate_nips(df_mentah):
    """Cek NIP duplikat di data mentah dan return baris yang duplikat"""
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    duplicates = df_mentah[df_mentah.duplicated('nip_clean', keep=False)]
    return duplicates

def check_new_data(df_mentah, df_master):
    """Cek NIP yang ada di data mentah tapi tidak ada di data master"""
    # Bersihkan NIP
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    df_master['NIP_clean'] = df_master['NIP'].astype(str).str.strip()
    
    # Cari NIP yang tidak ada di master
    master_nips = set(df_master['NIP_clean'].tolist())
    mentah_nips = set(df_mentah['nip_clean'].tolist())
    
    new_nips = mentah_nips - master_nips
    new_data = df_mentah[df_mentah['nip_clean'].isin(new_nips)]
    
    return new_data

def calculate_gaji_kotor(row):
    """Hitung gaji kotor jika kolom gajikotor tidak ada"""
    total = 0
    missing_columns = []
    
    for col in GAJI_COMPONENTS:
        if col in row and pd.notna(row[col]):
            try:
                total += float(row[col])
            except (ValueError, TypeError):
                missing_columns.append(f"{col} (nilai tidak valid)")
        else:
            missing_columns.append(col)
    
    if missing_columns:
        st.warning(f"⚠️ Kolom {', '.join(missing_columns)} tidak ditemukan atau nilainya tidak valid untuk perhitungan gaji. Dianggap 0.")
    
    return total

def process_data_to_bpmp(df_mentah, df_master):
    """Proses data mentah dan master menjadi format BPMP"""
    try:
        # Deteksi nama kolom ID TKU yang digunakan
        id_tku_col = None
        if "ID TKU" in df_master.columns:
            id_tku_col = "ID TKU"
        elif "ID PENERIMA TKU" in df_master.columns:
            id_tku_col = "ID PENERIMA TKU"
        else:
            st.error("❌ Kolom ID TKU tidak ditemukan di data master")
            return None, 0, 0
        
        st.info(f"ℹ️ Menggunakan kolom: **{id_tku_col}** dari data master")
        
        # List untuk menyimpan hasil
        hasil_bpmp = []
        
        # Counter untuk tracking
        total_mentah = len(df_mentah)
        berhasil = 0
        gagal = 0
        
        # Progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Informasi perhitungan gaji
        gunakan_perhitungan_sistem = False
        if 'GajiKotor' not in df_mentah.columns and 'gajikotor' not in df_mentah.columns:
            gunakan_perhitungan_sistem = True
            st.info("ℹ️ Menggunakan perhitungan sistem untuk Penghasilan Kotor")
        
        # Loop setiap baris di data mentah
        for idx, row_mentah in df_mentah.iterrows():
            nip_mentah = str(row_mentah['nip']).strip()
            
            # Cari NIP di data master
            df_master['NIP'] = df_master['NIP'].astype(str).str.strip()
            match_master = df_master[df_master['NIP'] == nip_mentah]
            
            if not match_master.empty:
                row_master = match_master.iloc[0]
                
                # Tentukan gaji kotor
                gaji_kotor = 0
                if 'gajikotor' in row_mentah and pd.notna(row_mentah['gajikotor']):
                    gaji_kotor = row_mentah['gajikotor']
                elif 'GajiKotor' in row_mentah and pd.notna(row_mentah['GajiKotor']):
                    gaji_kotor = row_mentah['GajiKotor']
                else:
                    # Hitung otomatis dari komponen gaji
                    gaji_kotor = calculate_gaji_kotor(row_mentah)
                
                # DIUBAH: Buat dictionary untuk baris BPMP sesuai urutan baru
                bpmp_row = {
                    "Masa Pajak": row_mentah['bulan'],
                    "Tahun Pajak": row_mentah['tahun'],
                    "Status Pegawai": "Resident",
                    "NPWP/NIK/TIN": row_master['NIK'],
                    "Nomor Passport": "",
                    "Status": row_master['STATUS'],
                    "Posisi": "PNS",  # Huruf besar untuk PPPK
                    "Sertifikat/Fasilitas": "DTP",
                    "Kode Objek Pajak": "21-100-01",
                    "Penghasilan Kotor": gaji_kotor,
                    "Tarif": "",
                    "ID TKU": row_master[id_tku_col],
                    "Tgl Pemotongan": "",
                    "TER A": "",
                    "TER B": "",
                    "TER C": ""
                }
                
                hasil_bpmp.append(bpmp_row)
                berhasil += 1
            else:
                gagal += 1
                if gagal <= 10:
                    st.warning(f"⚠️ NIP {nip_mentah} tidak ditemukan di data master (baris {idx+1})")
            
            # Update progress
            progress = (idx + 1) / total_mentah
            progress_bar.progress(progress)
            status_text.text(f"Memproses: {idx+1}/{total_mentah} | Berhasil: {berhasil} | Gagal: {gagal}")
        
        if gagal > 10:
            st.warning(f"⚠️ ... dan {gagal - 10} NIP lainnya tidak ditemukan")
        
        progress_bar.empty()
        status_text.empty()
        
        # Convert ke DataFrame dengan urutan kolom yang benar
        if hasil_bpmp:
            # DIUBAH: Pastikan DataFrame dibuat dengan urutan HEADERS_BPMP
            df_hasil = pd.DataFrame(hasil_bpmp, columns=HEADERS_BPMP)
            
            # Tampilkan informasi tentang perhitungan gaji
            if gunakan_perhitungan_sistem:
                st.success(f"✅ Penghasilan Kotor dihitung otomatis dari {len(GAJI_COMPONENTS)} komponen gaji")
            
            return df_hasil, berhasil, gagal
        else:
            return None, 0, gagal
            
    except Exception as e:
        st.error(f"❌ Error saat memproses data: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None, 0, 0

def convert_df_to_excel(df):
    """Convert DataFrame ke Excel dengan styling warna sesuai permintaan"""
    output = BytesIO()
    
    # Buat workbook dan worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data BPMP'
    
    # Definisikan warna-warna
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Merah untuk header
    red_font = Font(color='FFFFFF', bold=True)  # Font putih untuk header merah
    bold_font = Font(bold=True)  # Font tebal untuk header lainnya
    green_fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')  # Hijau muda untuk isi baris
    
    # Kolom yang harus berwarna merah di header
    red_header_columns = ["Tarif", "TER A", "TER B", "TER C"]
    
    # Kolom yang isi barisnya harus hijau muda (kolom 1-10)
    green_data_columns = [
        "Masa Pajak", "Tahun Pajak", "Status Pegawai", "NPWP/NIK/TIN",
        "Nomor Passport", "Status", "Posisi", "Sertifikat/Fasilitas", 
        "Kode Objek Pajak", "Penghasilan Kotor", "ID TKU"
    ]
    
    # Kolom yang isi barisnya harus merah (kolom 11-16)
    red_data_columns = ["Tarif", "TER A", "TER B", "TER C"]
    
    # Tulis header
    for col_idx, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        
        # Beri warna merah untuk kolom tertentu
        if header in red_header_columns:
            cell.fill = red_fill
            cell.font = red_font
        else:
            cell.font = bold_font
    
    # Tulis data dengan pewarnaan
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Dapatkan nama kolom berdasarkan indeks
            column_name = df.columns[col_idx-1]
            
            # Beri warna hijau muda untuk isi baris kolom 1-10
            if column_name in green_data_columns:
                cell.fill = green_fill
            # Beri warna merah untuk isi baris kolom tertentu
            elif column_name in red_data_columns:
                cell.fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')  # Merah muda untuk isi
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Simpan ke BytesIO
    wb.save(output)
    output.seek(0)
    return output

def create_template_mentah():
    """Membuat template Excel untuk data mentah PPPK"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Data Mentah PPPK'
    
    # Data contoh untuk template
    data = {
        'nip': ['123456789012345678'],
        'bulan': [1],
        'tahun': [2024],
        'GajiKotor': [8000000],  # Opsional: bisa dihitung otomatis
        'gjpokok': [5000000],
        'tjistri': [300000],
        'tjanak': [200000],
        'tjupns': [100000],
        'tjstruk': [50000],
        'tjfungs': [40000],
        'tjdaerah': [30000],
        'tjpencil': [20000],
        'tjlain': [10000],
        'tjkompen': [5000],
        'pembul': [3000],
        'tjberas': [25000],
        'tjpph': [15000],
        'potpfkbul': [2000],
        'potpfk2': [1000],
        'kdsatker': ['123456'],
        'nogaji': ['001'],
        'nmpeg': ['Nama Contoh PPPK'],
        'kdgol': ['IV/a'],
        'npwp': ['123456789012345'],
        'kdkawin': ['K']
    }
    
    # Tulis header
    for col_idx, header in enumerate(data.keys(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Tulis data
    for col_idx, (key, values) in enumerate(data.items(), 1):
        ws.cell(row=2, column=col_idx, value=values[0])
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def create_template_master():
    """Membuat template Excel untuk data master PPPK"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Data Master PPPK'
    
    # Data contoh untuk template
    data = {
        'NIP': ['123456789012345678', '987654321098765432'],
        'NIK': ['123456789012345', '987654321098765'],
        'STATUS': ['K', 'TK'],
        'ID TKU': ['TKU1234567890123456789', 'TKU9876543210987654321'],
        'ID PENERIMA TKU': ['TKU1234567890123456789', 'TKU9876543210987654321'],
        'KDKAWIN': ['K', 'TK'],
        'Nama': ['Nama Contoh PPPK 1', 'Nama Contoh PPPK 2'],
        'PNS/PPPK': ['PPPK', 'PPPK']
    }
    
    # Tulis header
    for col_idx, header in enumerate(data.keys(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Tulis data (2 baris contoh)
    for row_idx in range(2):
        for col_idx, (key, values) in enumerate(data.items(), 1):
            ws.cell(row=row_idx+2, column=col_idx, value=values[row_idx])
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def show():
    # Header dengan tombol kembali
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("← Kembali ke Dashboard PPPK"):
            st.session_state.current_page = 'dashboard_pppk'
            st.rerun()
    
    st.title("💰 Upload Pajak Gaji PPPK")
    st.markdown("---")
    
    # ========== PANDUAN PENGGUNAAN ==========
    with st.expander("📚 Panduan Penggunaan - Baca Sebelum Mulai", expanded=False):
        st.markdown("""
        ### **🎯 Cara Menggunakan Fitur Ini:**
        
        1. **Siapkan 2 File Excel**:
           - **Data Mentah Gaji PPPK**: Hasil download sistem penggajian per bulan
           - **Data Master PPPK**: Database pegawai PPPK
        
        2. **Format Data Mentah PPPK (WAJIB)**:
           ```
           Kolom yang HARUS ADA (minimal):
           - nip         : Nomor Induk Pegawai PPPK (tidak boleh duplikat)
           - bulan       : Bulan transaksi (1-12)
           - tahun       : Tahun transaksi
           
           Kolom Gaji Kotor (PILIH SALAH SATU):
           - OPTIONAL: GajiKotor atau gajikotor → Total penghasilan kotor (manual/opsional)
           - JIKA TIDAK ADA: Sistem hitung otomatis dari 15 komponen gaji:
               gjpokok, tjistri, tjanak, tjupns, tjstruk, tjfungs, tjdaerah,
               tjpencil, tjlain, tjkompen, pembul, tjberas, tjpph, potpfkbul, potpfk2
           ```
           
        3. **Format Data Master PPPK (WAJIB)**:
           ```
           Kolom yang HARUS ADA:
           - NIP               : Nomor Induk Pegawai PPPK (untuk matching)
           - NIK               : Digunakan sebagai NPWP/NIK/TIN di output
           - STATUS            : Status PTKP (K/TK/HB/etc.)
           
           Kolom YANG HARUS ADA (salah satu):
           - ID TKU           : ID penerima penghasilan
           - ID PENERIMA TKU  : ID penerima penghasilan (alternatif)
           
           Kolom PENTING:
           - KDKAWIN          : Kode status kawin (untuk dibandingkan dengan kdkawin di data mentah)
           
           Kolom tambahan (jika ada):
           - PNS/PPPK       : Jenis pegawai (isi dengan "PPPK")
           - Nama           : Nama lengkap
           - KODE OBJEK PAJAK : Kode objek pajak
           ```
        
        4. **PROSES OTOMATIS**:
           - Sistem akan melakukan **inner join** berdasarkan NIP
           - **NIP duplikat** di data mentah akan ditandai merah
           - **Data baru** (NIP tidak ada di master) akan ditandai hijau
           - **Perbedaan kdkawin** akan ditandai kuning
           - **ID TKU akan diambil dari Data Master** (tidak ada nilai default)
           - **Penghasilan Kotor**: Jika tidak ada kolom GajiKotor, sistem hitung otomatis
           - **Posisi**: Akan diisi dengan "PNS" (huruf besar untuk PPPK)
        
        5. **FORMAT OUTPUT (BPMP)**:
           ```
           Hasil akan dikonversi ke 16 kolom format BPMP:
           1. Masa Pajak → dari kolom 'bulan'
           2. Tahun Pajak → dari kolom 'tahun'
           3. Status Pegawai → default "Resident"
           4. NPWP/NIK/TIN → dari kolom 'NIK' di Data Master
           5. Nomor Passport → kosong
           6. Status → dari kolom 'STATUS' di Data Master
           7. Posisi → default "PNS" (huruf besar untuk PPPK)
           8. Sertifikat/Fasilitas → default "DTP"
           9. Kode Objek Pajak → default "21-100-01"
           10. Penghasilan Kotor → dari 'GajiKotor' atau hasil perhitungan sistem
           11. Tarif → kosong (berisi rumus di aplikasi BPMP)
           12. ID TKU → dari Data Master
           13. Tgl Pemotongan → kosong
           14-16. TER A, TER B, TER C → kosong (berisi rumus di aplikasi BPMP)
           ```
        
        6. **🎨 KODE WARNA DI HASIL DOWNLOAD**:
           **HEADER:**
           - **Merah dengan font putih**: Tarif, TER A, TER B, TER C
           - **Hitam tebal**: Header lainnya
           
           **ISI DATA:**
           - **Hijau Muda (Soft Green)**: Kolom 1-10 + ID TKU (data hasil sistem)
           - **Merah Muda (Light Red)**: Kolom 11-16 (berisi rumus di aplikasi BPMP)
           
           **⚠️ PERHATIAN KHUSUS:**
           - Kolom berwarna merah/muda mengandung **RUMUS/FORMULA** di aplikasi BPMP
           - **JANGAN DISALIN** kolom tersebut saat menyalin ke aplikasi BPMP
           - Hanya salin kolom yang berwarna hijau muda dan kolom tanpa warna
        """)
        
        # Template untuk download - hanya format Excel (xlsx)
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 Template Data Mentah PPPK (Excel)",
                data=create_template_mentah(),
                file_name="template_data_mentah_gaji_pppk.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            st.download_button(
                label="📥 Template Data Master PPPK (Excel)",
                data=create_template_master(),
                file_name="template_data_master_gaji_pppk.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.markdown("---")
    
    # ========== INFORMASI PENTING ==========
    st.info("""
    **📊 INFORMASI PENTING UNTUK PPPK:**
    
    **1. PERBEDAAN DENGAN PNS:**
    - **Posisi**: "PNS" (huruf besar) - beda dengan PNS yang menggunakan "pns" (huruf kecil)
    - **ID TKU**: Diambil dari Data Master (tidak ada nilai default seperti PNS)
    - **Format NIP PPPK**: 18 digit (sama dengan PNS)
    - **Kode Objek Pajak**: Sama "21-100-01"
    - **Status Pegawai**: Sama "Resident" (untuk WNI)
    
    **2. PROSES MATCHING DATA:**
    - Sistem menggunakan **inner join** berdasarkan NIP PPPK
    - **NIP duplikat** di data mentah akan ditandai merah dan harus diperbaiki
    - **Data baru** (NIP tidak ada di master) akan ditandai hijau dan perlu ditambahkan ke master
    - **Perbedaan kdkawin** antara mentah dan master akan ditandai kuning
    - Proses matching bersifat case-sensitive dan sensitive terhadap spasi
    
    **3. PERHITUNGAN GAJI KOTOR:**
    - **Prioritas 1**: Jika kolom **gajikotor** (lowercase) ada, gunakan nilai tersebut
    - **Prioritas 2**: Jika kolom **GajiKotor** (original) ada, gunakan nilai tersebut
    - **Prioritas 3**: Jika tidak ada kolom GajiKotor, sistem hitung otomatis dari 15 komponen gaji:
      ```
      Penghasilan Kotor = SUM(
          gjpokok, tjistri, tjanak, tjupns, tjstruk, tjfungs, 
          tjdaerah, tjpencil, tjlain, tjkompen, pembul, tjberas, 
          tjpph, potpfkbul, potpfk2
      )
      ```
    - Komponen gaji yang hilang akan dianggap 0 dan muncul warning
    
    **4. KOLOM DEFAULT YANG AKAN DIHASILKAN:**
    ```
    Status Pegawai    : "Resident" (untuk WNI)
    Posisi           : "PNS" (huruf besar untuk PPPK)
    Sertifikat       : "DTP" (fasilitas yang digunakan)
    Kode Objek Pajak : "21-100-01" (kode standar)
    ID TKU           : Diambil dari Data Master (WAJIB ADA)
    ```
    
    **5. DETEKSI PERUBAHAN:**
    - **kdkawin** di data mentah vs **KDKAWIN** di master
    - Jika berbeda: baris di data mentah akan ditandai kuning
    - Ini menunjukkan ada perubahan status yang perlu diupdate di master
    """)
    
    # ========== PERINGATAN KHUSUS TENTANG WARNA ==========
    st.warning("""
    **🎯 PANDUAN PENYALINAN KE APLIKASI BPMP:**
    
    **SALIN kolom ini saja (Hijau Muda & Tanpa Warna):**
    1. Masa Pajak
    2. Tahun Pajak
    3. Status Pegawai
    4. NPWP/NIK/TIN
    5. Nomor Passport
    6. Status
    7. Posisi ← **"PNS" (huruf besar) untuk PPPK**
    8. Sertifikat/Fasilitas
    9. Kode Objek Pajak
    10. Penghasilan Kotor
    12. ID TKU
    13. Tgl Pemotongan (jika ada)
    
    **JANGAN SALIN kolom ini (Merah Muda):**
    11. Tarif (berisi rumus)
    14. TER A (berisi rumus)
    15. TER B (berisi rumus)
    16. TER C (berisi rumus)
    
    **Alasan:** Kolom merah muda akan dihitung otomatis oleh aplikasi BPMP menggunakan rumus/formula internal.
    """)
    
    # Initialize session state
    if 'df_mentah_pppk' not in st.session_state:
        st.session_state.df_mentah_pppk = None
    if 'df_bpmp_pppk' not in st.session_state:
        st.session_state.df_bpmp_pppk = None
    if 'df_master_pppk' not in st.session_state:
        st.session_state.df_master_pppk = None
    if 'df_hasil_pppk' not in st.session_state:
        st.session_state.df_hasil_pppk = None
    if 'new_data_df_pppk' not in st.session_state:
        st.session_state.new_data_df_pppk = None
    if 'duplicate_nips_df_pppk' not in st.session_state:
        st.session_state.duplicate_nips_df_pppk = None
    
    # ===== BAGIAN 1: UPLOAD DATA MENTAH PPPK =====
    st.subheader("📤 1. Upload Data Mentah Gaji PPPK")
    st.markdown("File Excel yang berisi data gaji mentah dari sistem penggajian PPPK")
    
    with st.expander("ℹ️ Detail Kolom Data Mentah PPPK", expanded=False):
        st.markdown("""
        **Kolom WAJIB ada (minimal):**
        - `nip` : Nomor Induk Pegawai PPPK (kunci join, TIDAK BOLEH DUPLIKAT)
        - `bulan` : Bulan transaksi (1-12)
        - `tahun` : Tahun transaksi (4 digit)
        
        **Kolom Gaji Kotor (OPSIONAL - PILIH SALAH SATU):**
        - `GajiKotor` atau `gajikotor` : Total penghasilan kotor (manual)
        **ATAU** semua komponen gaji untuk dihitung sistem:
        - `gjpokok`, `tjistri`, `tjanak`, `tjupns`, `tjstruk`, `tjfungs`, `tjdaerah`,
          `tjpencil`, `tjlain`, `tjkompen`, `pembul`, `tjberas`, `tjpph`, `potpfkbul`, `potpfk2`
        
        **Kolom PENTING lainnya:**
        - `kdsatker` : Kode satker
        - `nogaji` : Nomor gaji
        - `nmpeg` : Nama pegawai PPPK
        - `kdgol` : Kode golongan
        - `npwp` : NPWP (jika ada)
        - `kdkawin` : Kode status kawin (untuk dibandingkan dengan master)
        
        **Total kolom yang direkomendasikan:** 50 kolom (sesuai sistem penggajian PPPK)
        """)
    
    uploaded_mentah = st.file_uploader(
        "**Pilih file Data Mentah Gaji PPPK**",
        type=['xlsx', 'xls'],
        key="mentah_pppk_uploader",
        help="Upload file Excel (xlsx atau xls) hasil download sistem penggajian PPPK"
    )
    
    if uploaded_mentah:
        try:
            # Hanya membaca file Excel (xlsx, xls)
            df = pd.read_excel(uploaded_mentah)
            
            if validate_headers(df, HEADERS_MENTAH_PPPK, "Data Mentah"):
                # Cek duplikat NIP
                duplicates = check_duplicate_nips(df)
                if not duplicates.empty:
                    st.session_state.duplicate_nips_df_pppk = duplicates
                    st.error(f"❌ Ditemukan {len(duplicates)} NIP duplikat di Data Mentah PPPK!")
                    
                    # Tampilkan baris duplikat dengan warna merah
                    st.warning("**Baris dengan NIP duplikat (ditandai merah):**")
                    
                    # Buat DataFrame dengan styling untuk duplikat
                    def highlight_duplicates(df_original, duplicates_df):
                        # Buat salinan untuk styling
                        styled_df = df_original.copy()
                        
                        # Tandai baris duplikat
                        mask = styled_df['nip'].astype(str).str.strip().isin(
                            duplicates_df['nip_clean'].unique()
                        )
                        
                        # Buat list warna
                        colors = ['background-color: #ffcccc' if mask.iloc[i] else '' 
                                 for i in range(len(styled_df))]
                        
                        return styled_df.style.apply(lambda x: colors, axis=0)
                    
                    styled_duplicates = highlight_duplicates(df.head(50), duplicates)
                    st.dataframe(styled_duplicates, use_container_width=True)
                    
                    if len(df) > 50:
                        st.info(f"Menampilkan 50 baris pertama dari total {len(df)} baris")
                    
                    st.error("**PERBAIKI NIP DUPLIKAT SEBELUM MELANJUTKAN!**")
                    st.session_state.df_mentah_pppk = None
                else:
                    st.session_state.df_mentah_pppk = df
                    st.success(f"✅ File '{uploaded_mentah.name}' berhasil diupload!")
                    st.info(f"📊 Total data: {len(df)} baris")
                    
                    # Cek dan tampilkan informasi tentang GajiKotor
                    if 'GajiKotor' not in df.columns and 'gajikotor' not in df.columns:
                        st.info("ℹ️ **Kolom GajiKotor tidak ditemukan.** Sistem akan menghitung otomatis dari komponen gaji.")
                        
                        # Tampilkan contoh komponen gaji yang ada
                        komponen_ada = [col for col in GAJI_COMPONENTS if col in df.columns]
                        if komponen_ada:
                            st.success(f"✅ Ditemukan {len(komponen_ada)} komponen gaji untuk perhitungan otomatis")
                        else:
                            st.warning("⚠️ Tidak ditemukan komponen gaji untuk perhitungan. Hasil Penghasilan Kotor akan 0.")
                    
                    with st.expander("👁️ Preview Data Mentah PPPK"):
                        # Tampilkan semua data
                        st.dataframe(df, use_container_width=True)
                        st.caption(f"Menampilkan semua {len(df)} baris data")
            else:
                st.session_state.df_mentah_pppk = None
                
        except Exception as e:
            st.error(f"❌ Error membaca file: {str(e)}")
            st.session_state.df_mentah_pppk = None
    
    st.markdown("---")
    
    # ===== BAGIAN 2: UPLOAD DATA MASTER PPPK =====
    st.subheader("📋 2. Upload Data Master PPPK")
    st.markdown("File Excel yang berisi data master pegawai PPPK (NIP, NIK, Status, ID TKU, dll)")
    
    with st.expander("ℹ️ Detail Kolom Data Master PPPK", expanded=False):
        st.markdown("""
        **Kolom WAJIB ada:**
        - `NIP` : Nomor Induk Pegawai PPPK (kunci join)
        - `NIK` : Akan digunakan sebagai NPWP/NIK/TIN di output
        - `STATUS` : Status PTKP (K/TK/HB/etc.)
        
        **Kolom YANG HARUS ADA (salah satu):**
        - `ID TKU` : ID pemotong/penerima TKU
        - `ID PENERIMA TKU` : ID penerima TKU (alternatif)
        
        **Kolom PENTING:**
        - `KDKAWIN` : Kode status kawin (untuk dibandingkan dengan kdkawin di data mentah)
        
        **Kolom tambahan (jika ada):**
        - `PNS/PPPK` : Jenis pegawai (harus berisi "PPPK")
        - `Nama` : Nama lengkap
        - `KODE OBJEK PAJAK` : Kode objek pajak
        """)
    
    uploaded_master = st.file_uploader(
        "**Pilih file Data Master PPPK**",
        type=['xlsx', 'xls'],
        key="master_pppk_uploader",
        help="Upload database master pegawai PPPK dalam format Excel (xlsx atau xls)"
    )
    
    if uploaded_master:
        try:
            # Hanya membaca file Excel (xlsx, xls)
            df = pd.read_excel(uploaded_master)
            
            if validate_headers(df, HEADERS_MASTER, "Data Master"):
                st.session_state.df_master_pppk = df
                st.success(f"✅ File '{uploaded_master.name}' berhasil diupload!")
                st.info(f"📊 Total pegawai PPPK: {len(df)} orang")
                
                with st.expander("👁️ Preview Data Master PPPK"):
                    # Tampilkan semua data
                    st.dataframe(df, use_container_width=True)
                    st.caption(f"Menampilkan semua {len(df)} baris data")
            else:
                st.session_state.df_master_pppk = None
                
        except Exception as e:
            st.error(f"❌ Error membaca file: {str(e)}")
            st.session_state.df_master_pppk = None
    
    st.markdown("---")
    
    # ===== BAGIAN 3: DETEKSI DATA BARU DAN PERBEDAAN =====
    if st.session_state.df_mentah_pppk is not None and st.session_state.df_master_pppk is not None:
        st.subheader("🔍 3. Deteksi Data Baru dan Perbedaan")
        
        # Cek data baru (NIP di mentah tapi tidak di master)
        new_data = check_new_data(st.session_state.df_mentah_pppk, st.session_state.df_master_pppk)
        
        if not new_data.empty:
            st.session_state.new_data_df_pppk = new_data
            st.warning(f"⚠️ Ditemukan {len(new_data)} data baru di Data Mentah yang tidak ada di Data Master!")
            
            # Tampilkan data baru dengan warna hijau
            st.info("**Data baru (ditandai hijau - perlu ditambahkan ke Data Master):**")
            
            def highlight_new_data(df_original, new_data_df):
                # Buat salinan untuk styling
                styled_df = df_original.copy()
                
                # Tandai baris data baru
                mask = styled_df['nip'].astype(str).str.strip().isin(
                    new_data_df['nip_clean'].unique()
                )
                
                # Buat list warna
                colors = ['background-color: #ccffcc' if mask.iloc[i] else '' 
                         for i in range(len(styled_df))]
                
                return styled_df.style.apply(lambda x: colors, axis=0)
            
            styled_new_data = highlight_new_data(st.session_state.df_mentah_pppk.head(50), new_data)
            st.dataframe(styled_new_data, use_container_width=True)
            
            if len(st.session_state.df_mentah_pppk) > 50:
                st.info(f"Menampilkan 50 baris pertama dari total {len(st.session_state.df_mentah_pppk)} baris")
            
            # TOMBOL UNTUK MENUJU KE HALAMAN CROSSCHECK PPPK
            st.markdown("---")
            st.error("**DATA BARU HARUS DITAMBAHKAN KE DATA MASTER SEBELUM MELANJUTKAN!**")
            
            # Tambahkan tombol untuk menuju ke halaman croscheck_pppk
            if st.button("➕ Tambahkan Data Baru ke Master (Croscheck PPPK)", type="primary", use_container_width=True):
                st.session_state.current_page = 'croscheck_pppk'
                st.session_state.selected_menu = 'croscheck_pppk'
                st.rerun()
            
            st.stop()  # Hentikan proses sampai data baru ditangani
        
        # Cek perbedaan kdkawin antara mentah dan master
        st.session_state.df_mentah_pppk['nip_clean'] = st.session_state.df_mentah_pppk['nip'].astype(str).str.strip()
        st.session_state.df_master_pppk['NIP_clean'] = st.session_state.df_master_pppk['NIP'].astype(str).str.strip()
        
        # Gabungkan untuk membandingkan kdkawin
        merged = pd.merge(
            st.session_state.df_mentah_pppk[['nip_clean', 'kdkawin']],
            st.session_state.df_master_pppk[['NIP_clean', 'KDKAWIN']],
            left_on='nip_clean',
            right_on='NIP_clean',
            how='inner'
        )
        
        # Cari perbedaan
        if 'kdkawin' in merged.columns and 'KDKAWIN' in merged.columns:
            merged['kdkawin'] = merged['kdkawin'].astype(str).str.strip()
            merged['KDKAWIN'] = merged['KDKAWIN'].astype(str).str.strip()
            differences = merged[merged['kdkawin'] != merged['KDKAWIN']]
            
            if not differences.empty:
                st.warning(f"⚠️ Ditemukan {len(differences)} perbedaan kdkawin antara Data Mentah dan Data Master!")
                
                # Tampilkan perbedaan
                st.info("**Perbedaan kdkawin (ditandai kuning - perlu diperiksa):**")
                
                def highlight_kdkawin_differences(df_mentah_original, differences_df):
                    # Buat salinan untuk styling
                    styled_df = df_mentah_original.copy()
                    
                    # Tandai baris dengan perbedaan kdkawin
                    mask = styled_df['nip_clean'].isin(differences_df['nip_clean'])
                    
                    # Buat list warna (hanya kolom kdkawin yang kuning)
                    colors = []
                    for i in range(len(styled_df)):
                        if mask.iloc[i]:
                            row_colors = [''] * len(styled_df.columns)
                            if 'kdkawin' in styled_df.columns:
                                idx = list(styled_df.columns).index('kdkawin')
                                row_colors[idx] = 'background-color: #ffffcc'
                            colors.append(row_colors)
                        else:
                            colors.append([''] * len(styled_df.columns))
                    
                    return styled_df.style.apply(lambda x: colors[df_mentah_original.index.get_loc(x.name)] 
                                                if x.name in df_mentah_original.index else [''] * len(x), axis=1)
                
                styled_differences = highlight_kdkawin_differences(
                    st.session_state.df_mentah_pppk.head(50), 
                    differences
                )
                st.dataframe(styled_differences, use_container_width=True)
                
                # Tampilkan tabel perbandingan
                st.info("**Detail Perbandingan kdkawin:**")
                comparison_df = differences[['nip_clean', 'kdkawin', 'KDKAWIN']].copy()
                comparison_df.columns = ['NIP', 'kdkawin (Data Mentah)', 'KDKAWIN (Data Master)']
                st.dataframe(comparison_df, use_container_width=True)
                
                st.warning("**Data Mentah menggunakan nilai kdkawin yang baru!**")
        
        # Jika tidak ada data baru, lanjutkan
        st.success("✅ Tidak ditemukan data baru. Semua NIP di Data Mentah ada di Data Master.")
        
    # ===== BAGIAN 4: PROSES DATA =====
    st.subheader("🔄 4. Proses Data ke Format BPMP")
    
    # Cek apakah semua file sudah diupload
    if st.session_state.df_mentah_pppk is None:
        st.warning("⚠️ **Langkah 1**: Silakan upload Data Mentah PPPK terlebih dahulu")
    elif st.session_state.df_master_pppk is None:
        st.warning("⚠️ **Langkah 2**: Silakan upload Data Master PPPK terlebih dahulu")
    elif st.session_state.new_data_df_pppk is not None and not st.session_state.new_data_df_pppk.empty:
        st.warning("⚠️ **Langkah 3**: Ada data baru yang harus ditambahkan ke Data Master terlebih dahulu")
    else:
        st.success("✅ Semua file sudah siap!")
        
        # Tampilkan statistik
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Data Mentah PPPK", f"{len(st.session_state.df_mentah_pppk)} baris")
            st.caption("Data gaji PPPK yang akan diproses")
        with col2:
            st.metric("Data Master PPPK", f"{len(st.session_state.df_master_pppk)} pegawai")
            st.caption("Database referensi pegawai PPPK")
        
        st.markdown("---")
        
        # Informasi sebelum proses
        st.info("""
        **🔍 INFORMASI PROSES KHUSUS PPPK:**
        - Sistem akan melakukan **matching berdasarkan NIP PPPK**
        - **Posisi** akan otomatis diisi: **"PNS"** (huruf besar)
        - **ID TKU** akan diambil dari Data Master (kolom 'ID TKU' atau 'ID PENERIMA TKU')
        - **Penghasilan Kotor** akan diambil dari:
          1. Kolom `GajiKotor` jika ada
          2. Jika tidak, dihitung otomatis dari 15 komponen gaji
        - Hanya data dengan NIP yang match di kedua file yang akan diproses
        - Progress bar akan menunjukkan status pemrosesan
        """)
        
        if st.button("🚀 **PROSES DATA PPPK KE FORMAT BPMP**", type="primary", use_container_width=True):
            with st.spinner("⏳ Memproses data PPPK..."):
                df_hasil, berhasil, gagal = process_data_to_bpmp(
                    st.session_state.df_mentah_pppk,
                    st.session_state.df_master_pppk
                )
                
                if df_hasil is not None:
                    st.session_state.df_hasil_pppk = df_hasil
                    
                    st.success("✅ Proses selesai!")
                    
                    # Tampilkan statistik hasil
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("✅ Berhasil Diproses", berhasil, 
                                 help="Data PPPK dengan NIP yang ditemukan di Data Master")
                    with col2:
                        st.metric("❌ Tidak Match", gagal,
                                 help="Data PPPK dengan NIP yang tidak ditemukan di Data Master")
                    with col3:
                        st.metric("📊 Total Output", len(df_hasil),
                                 help="Total baris PPPK yang akan dihasilkan")
                    
                    st.markdown("---")
                    
                    # Preview hasil
                    st.subheader("📄 Preview Hasil BPMP untuk PPPK")
                    st.info(f"**Format kolom hasil ({len(df_hasil.columns)} kolom):**")
                    
                    # Tampilkan preview dataframe
                    preview_df = df_hasil.head(10).copy()
                    st.dataframe(preview_df)
                    
                    # Verifikasi posisi PPPK
                    if 'Posisi' in preview_df.columns:
                        posisi_values = preview_df['Posisi'].unique()
                        if len(posisi_values) == 1 and posisi_values[0] == "PNS":
                            st.success(f"✅ Posisi PPPK sudah benar: **{posisi_values[0]}** (huruf besar)")
                        else:
                            st.warning(f"⚠️ Posisi PPPK: {posisi_values} - Harusnya 'PNS' (huruf besar)")
                    
                    # Verifikasi ID TKU
                    if 'ID TKU' in preview_df.columns:
                        id_tku_unique = preview_df['ID TKU'].nunique()
                        st.info(f"ℹ️ ID TKU unik: {id_tku_unique} jenis")
                    
                    # Informasi warna di preview
                    st.info("""
                    **🎨 LEGENDA WARNA (di File Excel Hasil):**
                    
                    **Header:**
                    - **Merah dengan font putih**: Tarif, TER A, TER B, TER C
                    - **Hitam tebal**: Semua header lainnya
                    
                    **Isi Data:**
                    - **Hijau Muda (#C6EFCE)**: Kolom 1-10 + ID TKU (data hasil sistem)
                    - **Merah Muda (#FF9999)**: Tarif, TER A, TER B, TER C (berisi rumus)
                    
                    **⚠️ Ingat: Jangan salin kolom merah muda ke aplikasi BPMP!**
                    """)
                else:
                    st.error("❌ Tidak ada data PPPK yang berhasil diproses!")
    
    # ===== BAGIAN 5: DOWNLOAD HASIL =====
    if st.session_state.df_hasil_pppk is not None:
        st.markdown("---")
        st.subheader("💾 Download Hasil dengan Format Warna")
        
        df_hasil = st.session_state.df_hasil_pppk
        jumlah_data = len(df_hasil)
        
        # Info file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Data_BPMP_PPPK_{timestamp}.xlsx"
        
        st.info(f"📄 **Nama file:** {filename}")
        st.info(f"📊 **Jumlah data PPPK:** {jumlah_data} baris")
        
        # Statistik tambahan
        col1, col2, col3 = st.columns(3)
        with col1:
            total_gaji = df_hasil['Penghasilan Kotor'].sum()
            st.metric("Total Penghasilan", f"Rp {total_gaji:,.0f}")
        with col2:
            avg_gaji = df_hasil['Penghasilan Kotor'].mean()
            st.metric("Rata-rata Penghasilan", f"Rp {avg_gaji:,.0f}")
        with col3:
            status_count = df_hasil['Status'].nunique()
            st.metric("Jenis Status", status_count)
        
        # Peringatan khusus tentang kolom merah
        st.warning("""
        **⚠️ PERHATIAN KHUSUS - BACA SEBELUM MENYALIN KE APLIKASI BPMP:**
        
        **Kolom yang JANGAN DISALIN (berisi rumus di aplikasi BPMP):**
        1. **Tarif** (kolom 11)
        2. **TER A** (kolom 14)
        3. **TER B** (kolom 15)
        4. **TER C** (kolom 16)
        
        **Kolom yang HARUS DISALIN:**
        - **Kolom 1-10** (Hijau Muda)
        - **Kolom 12-13** (ID TKU dan Tgl Pemotongan - tanpa warna/hijau muda)
        
        **Khusus PPPK:**
        - Pastikan kolom **Posisi** berisi **"PNS"** (huruf besar)
        - **ID TKU** diambil dari Data Master (tidak ada nilai default)
        
        **Cara menyalin yang benar:**
        1. Buka file hasil di Excel
        2. **SALIN** hanya kolom 1-10 + kolom 12-13
        3. **LEWATI** kolom 11 dan 14-16 (kolom merah muda)
        4. Paste ke aplikasi BPMP
        5. Verifikasi kolom Posisi = "PNS" (huruf besar)
        """)
        
        # Generate Excel file
        excel_file = convert_df_to_excel(df_hasil)
        
        # Tombol download
        st.download_button(
            label="📥 **DOWNLOAD FILE EXCEL BPMP PPPK**",
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            help="Download file Excel dengan format warna siap untuk aplikasi BPMP"
        )
        
        # Informasi tambahan
        with st.expander("ℹ️ Informasi Detail Format File PPPK", expanded=False):
            st.markdown(f"""
            ### **🎨 DETAIL FORMAT WARNA DI FILE EXCEL:**
            
            **HEADER (Baris 1):**
            - **Merah dengan font putih:** Tarif, TER A, TER B, TER C
            - **Hitam tebal:** Semua header lainnya
            
            **ISI DATA (Baris 2 ke atas):**
            - **Hijau Muda (#C6EFCE):** 
              {', '.join([f'**{col}**' for col in df_hasil.columns if col not in ['Tarif', 'TER A', 'TER B', 'TER C', 'Tgl Pemotongan']])}
            - **Merah Muda (#FF9999):** Tarif, TER A, TER B, TER C
            - **Tanpa warna:** Tgl Pemotongan
            
            ### **📋 MAPPING DATA KHUSUS PPPK:**
            1. **Masa Pajak** → `bulan` dari Data Mentah
            2. **Tahun Pajak** → `tahun` dari Data Mentah
            3. **Status Pegawai** → "Resident" (default untuk WNI)
            4. **NPWP/NIK/TIN** → `NIK` dari Data Master
            5. **Nomor Passport** → kosong (untuk WNI)
            6. **Status** → `STATUS` dari Data Master
            7. **Posisi** → **"PNS"** (huruf besar - khusus PPPK)
            8. **Sertifikat/Fasilitas** → "DTP" (default)
            9. **Kode Objek Pajak** → "21-100-01" (default)
            10. **Penghasilan Kotor** → dari `GajiKotor` atau hasil perhitungan sistem
            11. **Tarif** → kosong (rumus di aplikasi BPMP)
            12. **ID TKU** → `ID TKU` atau `ID PENERIMA TKU` dari Data Master
            13. **Tgl Pemotongan** → kosong (diisi manual)
            14-16. **TER A, B, C** → kosong (rumus di aplikasi BPMP)
            
            ### **📊 STATISTIK DATA PPPK:**
            - **Total baris PPPK:** {jumlah_data}
            - **Total penghasilan kotor:** Rp {df_hasil['Penghasilan Kotor'].sum():,.0f}
            - **Rata-rata penghasilan PPPK:** Rp {df_hasil['Penghasilan Kotor'].mean():,.0f}
            - **Status unik:** {df_hasil['Status'].nunique()} jenis
            - **Posisi:** {df_hasil['Posisi'].iloc[0] if len(df_hasil) > 0 else 'N/A'} (harus "PNS")
            - **ID TKU unik:** {df_hasil['ID TKU'].nunique()} jenis
            """)
        
        # Tombol reset
        if st.button("🔄 **Reset Proses & Upload File Baru**", use_container_width=True):
            if 'df_mentah_pppk' in st.session_state:
                del st.session_state.df_mentah_pppk
            if 'df_master_pppk' in st.session_state:
                del st.session_state.df_master_pppk
            if 'df_hasil_pppk' in st.session_state:
                del st.session_state.df_hasil_pppk
            if 'new_data_df_pppk' in st.session_state:
                del st.session_state.new_data_df_pppk
            if 'duplicate_nips_df_pppk' in st.session_state:
                del st.session_state.duplicate_nips_df_pppk
            st.rerun()
    
    # Footer
    st.markdown("---")
    st.caption("""
    🔧 **Dukungan Teknis untuk PPPK**: 
    - **NIP duplikat**: Pastikan tidak ada NIP yang sama di Data Mentah PPPK
    - **Data baru**: NIP yang tidak ada di Data Master akan ditandai hijau dan harus ditambahkan
    - **Perbedaan kdkawin**: Baris dengan perbedaan akan ditandai kuning
    - **ID TKU**: Wajib ada di Data Master (kolom 'ID TKU' atau 'ID PENERIMA TKU')
    - **Perhitungan gaji**: Sistem otomatis hitung jika kolom `GajiKotor` tidak ada
    - **Posisi**: Otomatis diisi "PNS" (huruf besar) untuk PPPK
    - **Format file**: Hanya mendukung format Excel (.xlsx, .xls)
    - Untuk error: Coba download template dan sesuaikan data dengan format yang diberikan
    """)
    

if __name__ == "__main__":
    show()