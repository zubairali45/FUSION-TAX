import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def find_column_by_keywords(df, keywords_list):
    """Mencari kolom berdasarkan daftar kata kunci (case insensitive)"""
    for col in df.columns:
        col_lower = str(col).lower()
        for keyword in keywords_list:
            if keyword in col_lower:
                return col
    return None

def check_duplicate_nips(df_mentah):
    """Cek NIP duplikat di data mentah dan return baris yang duplikat"""
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    duplicates = df_mentah[df_mentah.duplicated('nip_clean', keep=False)]
    return duplicates

def check_new_data(df_mentah, df_master):
    """Cek NIP yang ada di data mentah tapi tidak ada di data master"""
    # Bersihkan NIP
    df_mentah['nip_clean'] = df_mentah['nip'].astype(str).str.strip()
    df_master['NIP_clean'] = df_master['NIP'].astype(str).str.strip()
    
    # Cari NIP yang tidak ada di master
    master_nips = set(df_master['NIP_clean'].tolist())
    mentah_nips = set(df_mentah['nip_clean'].tolist())
    
    new_nips = mentah_nips - master_nips
    new_data = df_mentah[df_mentah['nip_clean'].isin(new_nips)]
    
    return new_data

def create_template_mentah():
    """Membuat template Excel untuk data mentah pajak makan"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Data Mentah'
    
    # Data contoh untuk template
    data = {
        'kdsatker': ['123456'],
        'bln': ['1'],
        'thn': ['2024'],
        'tgl': ['2024-01-15'],
        'nogaji': ['001'],
        'nip': ['123456789012345678'],
        'nmpeg': ['Nama Contoh'],
        'kdgol': ['IV/a'],
        'npwp': ['123456789012345'],
        'kdbankspan': ['014'],
        'nmbankspan': ['BCA'],
        'norek': ['1234567890'],
        'nmrek': ['Nama Rekening'],
        'nmcabbank': ['Cabang Pusat'],
        'jmlhari': ['20'],
        'kotor': ['1000000'],
        'bersih': ['950000']
    }
    
    # Tulis header
    for col_idx, header in enumerate(data.keys(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Tulis data
    for col_idx, (key, values) in enumerate(data.items(), 1):
        ws.cell(row=2, column=col_idx, value=values[0])
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def create_template_master():
    """Membuat template Excel untuk data master pajak makan"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template Data Master'
    
    # Data contoh untuk template
    data = {
        'NIP': ['123456789012345678', '987654321098765432'],
        'NIK': ['123456789012345', '987654321098765'],
        'ID PENERIMA TKU': ['ID001', 'ID002'],
        'STATUS': ['K', 'TK'],
        'KODE OBJEK PAJAK': ['21-402-02', '21-402-03'],
        'Nama': ['Nama Contoh 1', 'Nama Contoh 2'],
        'PNS/PPPK': ['PNS', 'PNS']
    }
    
    # Tulis header
    for col_idx, header in enumerate(data.keys(), 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Tulis data (2 baris contoh)
    for row_idx in range(2):
        for col_idx, (key, values) in enumerate(data.items(), 1):
            ws.cell(row=row_idx+2, column=col_idx, value=values[row_idx])
    
    # Sesuaikan lebar kolom
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

def show():
    # Tombol kembali
    if st.button("← Kembali ke Dashboard PNS"):
        st.session_state.current_page = 'dashboard_pns'
        st.rerun()
    
    st.title("🍽️ Upload Pajak Makan PNS")
    st.markdown("---")
    
    # ========== PANDUAN PENGGUNAAN ==========
    with st.expander("📚 Panduan Penggunaan - Baca Sebelum Mulai", expanded=False):
        st.markdown("""
        ### **🎯 Cara Menggunakan Fitur Ini:**
        
        1. **Siapkan 2 File Excel**:
           - **Data Mentah**: File transaksi harian/bulanan PNS
           - **Data Master**: Database pegawai PNS
        
        2. **Format Data Mentah (WAJIB)**:
           ```
           Kolom yang HARUS ADA:
           - nip          : Nomor Induk Pegawai
           - kotor        : Nilai penghasilan kotor
           - bln          : Bulan transaksi
           - thn          : Tahun transaksi
           
           Kolom tambahan yang diperlukan:
           - nmpeg        : Nama pegawai
           - npwp         : NPWP pegawai
           - kdgol        : Kode golongan
           - dan kolom lain sesuai template
           ```
           
        3. **Format Data Master (WAJIB)**:
           ```
           Kolom yang HARUS ADA:
           - NIP                : Nomor Induk Pegawai
           - NIK                : Digunakan sebagai NPWP di hasil akhir
           - STATUS             : Status PTKP (K/TK/HB/etc.)
           - KODE OBJEK PAJAK   : Kode objek pajak (21-402-02, 21-402-03, 21-402-04)
           
           Kolom YANG DICARI:
           - ID PENERIMA TKU    : ID penerima penghasilan
           
           Kolom tambahan (jika ada):
           - PNS/PPPK       : Jenis pegawai
           - Nama           : Nama lengkap
           - KDKAWIN        : Kode status kawin
           - nmrek          : Nama rekening
           - rekening       : Nomor rekening
           ```
        
        4. **PROSES OTOMATIS**:
           - Sistem akan melakukan **inner join** berdasarkan NIP
           - **Deteksi data baru**: NIP yang ada di data mentah tapi tidak ada di data master akan ditandai
           - **Data baru harus ditambahkan ke Data Master** melalui halaman Croscheck sebelum diproses
           - Data tanpa match di Data Master akan ditampilkan sebagai warning
        
        5. **PERHITUNGAN TARIF OTOMATIS**:
           ```
           Logika perhitungan tarif berdasarkan KODE OBJEK PAJAK:
           - 21-402-02 → Tarif = 5%
           - 21-402-03 → Tarif = 15%
           - 21-402-04 → Tarif = 0%
           
           Sistem TIDAK menggunakan kolom 'pph' atau 'potongan' lagi
           Tarif diambil langsung dari KODE OBJEK PAJAK di Data Master
           ```
        
        6. **VALIDASI DATA**:
           - NIP di Data Mentah harus ada di Data Master
           - **NIP duplikat** di data mentah akan ditandai merah
           - **Data baru** akan ditandai hijau dan harus ditambahkan ke master
           - Sistem hanya memproses data yang memiliki kecocokan NIP
           - Data tidak match akan ditampilkan di bagian akhir
        """)
        
        # Template untuk download - hanya format Excel (xlsx)
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 Template Data Mentah (Excel)",
                data=create_template_mentah(),
                file_name="template_data_mentah_pajak_makan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            st.download_button(
                label="📥 Template Data Master (Excel)",
                data=create_template_master(),
                file_name="template_data_master_pajak_makan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.markdown("---")
    
    # ========== INFORMASI PENTING ==========
    st.info("""
    **📊 INFORMASI PENTING:**
    
    **1. PROSES JOIN DATA:**
    - Sistem menggunakan **INNER JOIN** berdasarkan NIP
    - **NIP duplikat** di data mentah akan ditandai merah dan harus diperbaiki
    - **Data baru** (NIP tidak ada di master) akan ditandai hijau dan harus ditambahkan ke master
    - Data tanpa match akan ditampilkan terpisah untuk tindak lanjut
    
    **2. KOLOM HASIL BP 21 YANG DIHASILKAN:**
    ```
    1. Masa Pajak               → dari kolom 'bln' (bulan)
    2. Tahun Pajak              → dari kolom 'thn' (tahun)
    3. NPWP                     → dari kolom 'NIK' di Data Master
    4. ID TKU Penerima          → dari 'ID PENERIMA TKU' di Data Master
    5. Status PTKP              → dari kolom 'STATUS' di Data Master
    6. Fasilitas                → default "DTP"
    7. Kode Objek Pajak         → dari Data Master
    8. Penghasilan              → dari kolom 'kotor' di Data Mentah
    9. Deemed                   → default "100"
    10. Tarif                   → otomatis dari KODE OBJEK PAJAK
    11. Jenis Dok. Referensi    → default "CommercialInvoice"
    12. Nomor Dok. Referensi    → manual (dari SP2D)
    13. Tanggal Dok. Referensi  → manual (dari SP2D, format: bulan/tanggal/tahun)
    14. ID TKU Pemotong         → DEFAULT: "0001658723701000000000"
    15. Tanggal Pemotongan      → manual (dari Invoice, format: bulan/tanggal/tahun)
    ```
    
    **3. LOGIKA TARIF BARU:**
    - **21-402-02** → Tarif 5%
    - **21-402-03** → Tarif 15%
    - **21-402-04** → Tarif 0%
    - Sistem TIDAK menggunakan kolom 'pph' atau 'potongan' lagi
    
    **4. FORMAT TANGGAL BARU:**
    - **Tanggal Dok. Referensi** → Format: bulan/tanggal/tahun (contoh: 8/4/2025)
    - **Tanggal Pemotongan** → Format: bulan/tanggal/tahun (contoh: 8/4/2025)
    
    **5. ID TKU PEMOTONG BARU:**
    - **Nilai Default**: "0001658723701000000000"
    - Semua data akan menggunakan ID TKU Pemotong yang sama
    - Tidak perlu mengambil dari Data Master lagi
    
    **⚠️ CATATAN PENTING:**
    - Pastikan NIP di kedua file memiliki format yang sama
    - Sistem akan menghapus spasi di awal/akhir NIP
    - Data Master harus mencakup semua NIP yang ada di Data Mentah
    - Pastikan KODE OBJEK PAJAK di Data Master sesuai format (21-402-02, 21-402-03, 21-402-04)
    - **Data baru harus ditambahkan ke Data Master sebelum diproses!**
    """)
    
    # ========== INFORMASI FORMAT WARNA ==========
    st.warning("""
    **🎨 INFORMASI FORMAT WARNA DI HASIL DOWNLOAD:**
    
    **1. Data Hijau Muda (#E6FFE6):**
    - Data hasil sistem (sudah diproses otomatis)
    - Bisa langsung dicopy/duplikat
    - Tidak perlu perubahan
    
    **2. Data Oranye (#FFD580):**
    - Data yang harus diisi **MANUAL** dari DAFTAR SP2D SATKER
    - Kolom yang perlu diisi:
      - **Nomor Dok. Referensi** → Ambil dari **Nomor SP2D**
      - **Tanggal Dok. Referensi** → Ambil dari **Tanggal SP2D** (format: bulan/tanggal/tahun)
      - **Tanggal Pemotongan** → Ambil dari **Tanggal Invoice** (format: bulan/tanggal/tahun)
    """)
    
    # ========== FORM UPLOAD ==========
    st.subheader("📤 Upload File Excel")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### **📄 Data Mentah (Transaksi)**")
        st.caption("File Excel berisi transaksi harian/bulanan")
        
        with st.expander("ℹ️ Detail Kolom Data Mentah", expanded=False):
            st.markdown("""
            **Kolom WAJIB ada:**
            - `nip` : Nomor Induk Pegawai (kunci join, TIDAK BOLEH DUPLIKAT)
            - `kotor` : Nilai penghasilan kotor
            - `bln` : Bulan (1-12, tanpa leading zero)
            - `thn` : Tahun (4 digit)
            
            **Kolom PENTING untuk tanggal:**
            - `tgl` atau kolom tanggal lainnya untuk referensi
            - Pastikan format tanggal konsisten
            
            **Kolom PENTING lainnya:**
            - `nmpeg` : Nama pegawai
            - `npwp` : NPWP (jika ada)
            - `kdgol` : Kode golongan
            - `norek` : Nomor rekening
            - `nmrek` : Nama pemilik rekening
            """)
        
        uploaded_file_raw = st.file_uploader(
            "**Pilih file Data Mentah**",
            type=['xlsx', 'xls'],
            key="raw_data_pns",
            help="Upload file transaksi PNS dalam format Excel (xlsx atau xls). Pastikan minimal ada kolom nip, kotor, bln, thn"
        )
    
    with col2:
        st.markdown("#### **📋 Data Master (Referensi)**")
        st.caption("Database referensi pegawai PNS")
        
        with st.expander("ℹ️ Detail Kolom Data Master", expanded=False):
            st.markdown("""
            **Kolom WAJIB ada:**
            - `NIP` : Nomor Induk Pegawai (kunci join)
            - `NIK` : Akan digunakan sebagai NPWP di output
            - `STATUS` : Status PTKP (K/TK/HB/etc.)
            - `KODE OBJEK PAJAK` : Kode objek pajak (21-402-02, 21-402-03, 21-402-04)
            
            **Kolom YANG DICARI:**
            - `ID PENERIMA TKU` : ID penerima penghasilan
            
            **Kolom tambahan (jika ada):**
            - `PNS/PPPK` : Jenis pegawai
            - `Nama` : Nama lengkap
            - `KDKAWIN` : Kode status kawin
            - `nmrek` : Nama rekening
            - `rekening` : Nomor rekening
            """)
        
        uploaded_file_master = st.file_uploader(
            "**Pilih file Data Master**",
            type=['xlsx', 'xls'],
            key="master_data_pns",
            help="Upload database master pegawai PNS dalam format Excel (xlsx atau xls)"
        )
    
    st.markdown("---")
    
    # ========== PROSES DATA ==========
    if uploaded_file_raw is not None and uploaded_file_master is not None:
        try:
            # Baca kedua file
            df_raw = pd.read_excel(uploaded_file_raw)
            df_master = pd.read_excel(uploaded_file_master)
            
            # BERSIHKAN NAMA KOLOM (hapus spasi di awal/akhir)
            df_raw.columns = df_raw.columns.str.strip()
            df_master.columns = df_master.columns.str.strip()
            
            st.success(f"✅ Data mentah berhasil diupload: {uploaded_file_raw.name}")
            st.success(f"✅ Data master berhasil diupload: {uploaded_file_master.name}")
            
            # ========== VALIDASI AWAL DAN DETEKSI DATA ==========
            st.subheader("🔍 Validasi Data dan Deteksi Data Baru")
            
            # Validasi kolom wajib di Data Mentah
            required_raw = ['nip', 'kotor', 'bln', 'thn']
            missing_raw = [col for col in required_raw if col not in df_raw.columns]
            
            if missing_raw:
                st.error(f"❌ **ERROR**: Kolom berikut tidak ditemukan di Data Mentah:")
                for col in missing_raw:
                    st.write(f"   - **{col}**")
                st.warning("💡 **Solusi**: Pastikan semua kolom wajib ada dan penulisannya benar")
                st.stop()
            
            # Validasi kolom wajib di Data Master
            required_master = ['NIP', 'NIK', 'STATUS']
            missing_master = [col for col in required_master if col not in df_master.columns]
            
            if missing_master:
                st.error(f"❌ **ERROR**: Kolom berikut tidak ditemukan di Data Master:")
                for col in missing_master:
                    st.write(f"   - **{col}**")
                st.warning("💡 **Solusi**: Pastikan NIP, NIK, dan STATUS ada di Data Master")
                st.stop()
            
            # Cari kolom KODE OBJEK PAJAK di Data Master
            kode_pajak_keywords = ['kode objek pajak', 'kode_objek_pajak', 'objek pajak']
            kode_pajak_col = find_column_by_keywords(df_master, kode_pajak_keywords)
            
            if not kode_pajak_col:
                st.error("❌ **ERROR**: Kolom 'KODE OBJEK PAJAK' tidak ditemukan di Data Master")
                st.warning("💡 **Solusi**: Pastikan Data Master memiliki kolom 'KODE OBJEK PAJAK' dengan format: 21-402-02, 21-402-03, 21-402-04")
                st.stop()
            
            # ========== CEK NIP DUPLIKAT DI DATA MENTAH ==========
            duplicates = check_duplicate_nips(df_raw)
            if not duplicates.empty:
                st.session_state.duplicate_nips_df = duplicates
                st.error(f"❌ Ditemukan {len(duplicates)} NIP duplikat di Data Mentah!")
                
                # Tampilkan baris duplikat dengan warna merah
                st.warning("**Baris dengan NIP duplikat (ditandai merah):**")
                
                # Buat DataFrame dengan styling untuk duplikat
                def highlight_duplicates(df_original, duplicates_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris duplikat
                    mask = styled_df['nip'].astype(str).str.strip().isin(
                        duplicates_df['nip_clean'].unique()
                    )
                    
                    # Buat list warna
                    colors = ['background-color: #ffcccc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_duplicates = highlight_duplicates(df_raw.head(50), duplicates)
                st.dataframe(styled_duplicates, use_container_width=True)
                
                if len(df_raw) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_raw)} baris")
                
                st.error("**PERBAIKI NIP DUPLIKAT SEBELUM MELANJUTKAN!**")
                st.stop()
            
            # ========== CEK DATA BARU (NIP DI MENTAH TAPI TIDAK DI MASTER) ==========
            new_data = check_new_data(df_raw, df_master)
            
            if not new_data.empty:
                st.session_state.new_data_df = new_data
                st.warning(f"⚠️ Ditemukan {len(new_data)} data baru di Data Mentah yang tidak ada di Data Master!")
                
                # Tampilkan data baru dengan warna hijau
                st.info("**Data baru (ditandai hijau - perlu ditambahkan ke Data Master):**")
                
                def highlight_new_data(df_original, new_data_df):
                    # Buat salinan untuk styling
                    styled_df = df_original.copy()
                    
                    # Tandai baris data baru
                    mask = styled_df['nip'].astype(str).str.strip().isin(
                        new_data_df['nip_clean'].unique()
                    )
                    
                    # Buat list warna
                    colors = ['background-color: #ccffcc' if mask.iloc[i] else '' 
                             for i in range(len(styled_df))]
                    
                    return styled_df.style.apply(lambda x: colors, axis=0)
                
                styled_new_data = highlight_new_data(df_raw.head(50), new_data)
                st.dataframe(styled_new_data, use_container_width=True)
                
                if len(df_raw) > 50:
                    st.info(f"Menampilkan 50 baris pertama dari total {len(df_raw)} baris")
                
                # Tampilkan detail data baru yang perlu ditambahkan
                st.markdown("---")
                st.error("**DATA BARU HARUS DITAMBAHKAN KE DATA MASTER SEBELUM MELANJUTKAN!**")
                
                # Tampilkan tabel detail data baru
                st.info("**Detail Data Baru yang Harus Ditambahkan ke Data Master:**")
                
                # Ekstrak informasi penting dari data baru
                new_data_details = new_data[['nip', 'nmpeg', 'npwp', 'kdgol']].copy()
                new_data_details = new_data_details.drop_duplicates('nip')
                new_data_details.columns = ['NIP', 'Nama', 'NPWP', 'Kode Golongan']
                
                st.dataframe(new_data_details, use_container_width=True)
                
                # Tambahkan tombol untuk menuju ke halaman croscheck_pns
                if st.button("➕ Tambahkan Data Baru ke Master (Croscheck PNS)", type="primary", use_container_width=True):
                    st.session_state.current_page = 'croscheck_pns'
                    st.session_state.selected_menu = 'croscheck'
                    st.rerun()
                
                st.stop()  # Hentikan proses sampai data baru ditangani
            
            # Jika tidak ada data baru, lanjutkan
            st.success("✅ Tidak ditemukan data baru. Semua NIP di Data Mentah ada di Data Master.")
            
            # Preview data yang diupload
            col1, col2 = st.columns(2)
            with col1:
                with st.expander(f"📄 Data Mentah ({len(df_raw)} baris)"):
                    st.write(f"**Kolom yang terdeteksi ({len(df_raw.columns)}):**")
                    cols_per_row = 3
                    cols = list(df_raw.columns)
                    for i in range(0, len(cols), cols_per_row):
                        row_cols = cols[i:i+cols_per_row]
                        col_text = " | ".join([f"`{col}`" for col in row_cols])
                        st.write(f"  {col_text}")
                    st.dataframe(df_raw.head(3))
            
            with col2:
                with st.expander(f"📄 Data Master ({len(df_master)} baris)"):
                    st.write(f"**Kolom yang terdeteksi ({len(df_master.columns)}):**")
                    cols = list(df_master.columns)
                    for i in range(0, len(cols), cols_per_row):
                        row_cols = cols[i:i+cols_per_row]
                        col_text = " | ".join([f"`{col}`" for col in row_cols])
                        st.write(f"  {col_text}")
                    st.dataframe(df_master.head(3))
            
            st.markdown("---")
            
            # ========== TOMBOL PROSES ==========
            if st.button("🔄 **PROSES DATA & GENERATE BP 21**", 
                        use_container_width=True, 
                        type="primary",
                        help="Klik untuk memproses data dan menghasilkan file BP 21"):
                
                with st.spinner("🔄 Sedang memproses data..."):
                    try:
                        # Membersihkan dan mempersiapkan data
                        df_raw_clean = df_raw.copy()
                        df_master_clean = df_master.copy()
                        
                        # Pastikan NIP dalam format string untuk join
                        df_raw_clean['nip'] = df_raw_clean['nip'].astype(str).str.strip()
                        df_master_clean['NIP'] = df_master_clean['NIP'].astype(str).str.strip()
                        
                        # Ambil NIP unik dari data mentah sebagai primary key
                        nip_list = df_raw_clean['nip'].unique()
                        
                        # Filter data master hanya untuk NIP yang ada di data mentah
                        df_master_filtered = df_master_clean[df_master_clean['NIP'].isin(nip_list)].copy()
                        
                        # Informasi proses join
                        st.info(f"🔍 **Proses Matching Data:**")
                        st.info(f"   • NIP unik di Data Mentah: {len(nip_list)}")
                        st.info(f"   • NIP ditemukan di Data Master: {len(df_master_filtered)}")
                        
                        # Join data berdasarkan NIP (inner join untuk hanya ambil yang match)
                        df_merged = pd.merge(
                            df_raw_clean,
                            df_master_filtered,
                            left_on='nip',
                            right_on='NIP',
                            how='inner'  # Hanya ambil yang match
                        )
                        
                        # ========== BUAT DATA HASIL BP 21 ==========
                        df_result = pd.DataFrame()
                        
                        # 1. Masa Pajak: ambil dari bln (tanpa leading zero)
                        df_result['Masa Pajak'] = df_merged['bln'].astype(str).str.lstrip('0')
                        
                        # 2. Tahun Pajak: ambil dari thn
                        df_result['Tahun Pajak'] = df_merged['thn'].astype(str)
                        
                        # 3. NPWP: ambil dari NIK (data master) - hilangkan .0
                        df_result['NPWP'] = df_merged['NIK'].apply(lambda x: str(x).replace('.0', '') if pd.notna(x) else '')
                        
                        # 4. ID TKU Penerima Penghasilan: cari kolom yang cocok
                        id_tku_keywords = ['id penerima tku', 'id_penerima_tku', 'id tku', 'id penerima', 'tku']
                        id_tku_col = find_column_by_keywords(df_merged, id_tku_keywords)
                        
                        if id_tku_col:
                            df_result['ID TKU Penerima Penghasilan'] = df_merged[id_tku_col].astype(str)
                            st.success(f"✅ Kolom ID TKU ditemukan: {id_tku_col}")
                        else:
                            df_result['ID TKU Penerima Penghasilan'] = ''
                            st.warning("⚠️ Kolom 'ID PENERIMA TKU' tidak ditemukan, diisi dengan nilai kosong")
                        
                        # 5. Status PTKP: ambil dari STATUS
                        df_result['Status PTKP'] = df_merged['STATUS'].astype(str)
                        
                        # 6. Fasilitas: default "DTP"
                        df_result['Fasilitas'] = 'DTP'
                        
                        # 7. Kode Objek Pajak: dari Data Master
                        df_result['Kode Objek Pajak'] = df_merged[kode_pajak_col].astype(str)
                        st.success(f"✅ Kolom Kode Objek Pajak ditemukan: {kode_pajak_col}")
                        
                        # 8. Penghasilan: ambil dari kotor
                        df_result['Penghasilan'] = df_merged['kotor'].astype(float)
                        
                        # 9. Deemed: default "100"
                        df_result['Deemed'] = '100'
                        
                        # 10. Tarif: LOGIKA BARU BERDASARKAN KODE OBJEK PAJAK
                        def get_tarif_from_kode(kode):
                            """Mendapatkan tarif berdasarkan kode objek pajak"""
                            kode = str(kode).strip()
                            if kode == '21-402-02':
                                return 5.0
                            elif kode == '21-402-03':
                                return 15.0
                            elif kode == '21-402-04':
                                return 0.0
                            else:
                                # Jika kode tidak dikenali, default ke 0
                                return 0.0
                        
                        # Terapkan fungsi ke setiap baris
                        df_result['Tarif'] = df_merged[kode_pajak_col].apply(get_tarif_from_kode)
                        
                        # Hitung statistik tarif
                        tarif_counts = df_result['Tarif'].value_counts()
                        tarif_info = []
                        for tarif, count in tarif_counts.items():
                            kode_mapping = {
                                5.0: '21-402-02',
                                15.0: '21-402-03',
                                0.0: '21-402-04'
                            }
                            kode = kode_mapping.get(tarif, f'Tidak dikenali (tarif {tarif})')
                            tarif_info.append(f"  • Tarif {tarif:.0f}% ({kode}): {count} baris")
                        
                        st.info("📊 **Mode Perhitungan BARU**: Tarif diambil dari KODE OBJEK PAJAK")
                        for info in tarif_info:
                            st.info(info)
                        
                        # 11. Jenis Dok. Referensi: default "CommercialInvoice"
                        df_result['Jenis Dok. Referensi'] = 'CommercialInvoice'
                        
                        # 12. Nomor Dok. Referensi: Cari kolom dengan berbagai variasi nama
                        nomor_ref_keywords = [
                            'nomor dok. referensi', 
                            'nomor dok referensi',
                            'nomor referensi',
                            'nomor dokumen',
                            'no dok referensi',
                            'nomor sp2d',
                            'no sp2d',
                            'sp2d',
                            'nomor dok',
                            'referensi'
                        ]
                        nomor_ref_col = find_column_by_keywords(df_merged, nomor_ref_keywords)
                        
                        if nomor_ref_col and not df_merged[nomor_ref_col].isna().all():
                            # Debug info
                            st.info(f"🔍 Ditemukan kolom '{nomor_ref_col}' untuk Nomor Dok. Referensi")
                            st.info(f"📋 Sample data dari kolom ini: {df_merged[nomor_ref_col].head(3).tolist()}")
                            
                            df_result['Nomor Dok. Referensi'] = df_merged[nomor_ref_col].astype(str)
                            st.success(f"✅ Kolom Nomor Dok. Referensi ditemukan: {nomor_ref_col}")
                        else:
                            df_result['Nomor Dok. Referensi'] = ''
                            st.warning("⚠️ Kolom Nomor Dok. Referensi tidak ditemukan atau kosong, diisi dengan nilai kosong (warna oranye)")
                        
                        # 13. Tanggal Dok. Referensi: Cari kolom dengan berbagai variasi nama
                        tanggal_ref_keywords = [
                            'tanggal dok. referensi',
                            'tanggal dok referensi',
                            'tanggal referensi',
                            'tanggal dokumen',
                            'tgl dok referensi',
                            'tanggal sp2d',
                            'tgl sp2d',
                            'tanggal dok',
                            'tgl referensi'
                        ]
                        tanggal_ref_col = find_column_by_keywords(df_merged, tanggal_ref_keywords)
                        
                        if tanggal_ref_col and not df_merged[tanggal_ref_col].isna().all():
                            # Debug info
                            st.info(f"🔍 Ditemukan kolom '{tanggal_ref_col}' untuk Tanggal Dok. Referensi")
                            st.info(f"📋 Sample data dari kolom ini: {df_merged[tanggal_ref_col].head(3).tolist()}")
                            
                            # Konversi ke format bulan/tanggal/tahun (8/4/2025)
                            try:
                                # Coba parse tanggal dengan berbagai format
                                dates = pd.to_datetime(df_merged[tanggal_ref_col], errors='coerce')
                                # Format ke bulan/tanggal/tahun (tanpa leading zero)
                                df_result['Tanggal Dok. Referensi'] = dates.apply(
                                    lambda x: f"{x.month}/{x.day}/{x.year}" if pd.notna(x) else ''
                                )
                                st.success(f"✅ Kolom Tanggal Dok. Referensi ditemukan: {tanggal_ref_col}")
                                st.info(f"📅 Format tanggal: bulan/tanggal/tahun (contoh: 8/4/2025)")
                            except Exception as e:
                                st.warning(f"⚠️ Gagal mengonversi format tanggal di kolom {tanggal_ref_col}. Menggunakan format asli: {str(e)}")
                                df_result['Tanggal Dok. Referensi'] = df_merged[tanggal_ref_col].astype(str)
                        else:
                            df_result['Tanggal Dok. Referensi'] = ''
                            st.warning("⚠️ Kolom Tanggal Dok. Referensi tidak ditemukan atau kosong, diisi dengan nilai kosong (warna oranye)")
                        
                        # 14. ID TKU Pemotong: DEFAULT "0001658723701000000000"
                        df_result['ID TKU Pemotong'] = '0001658723701000000000'
                        st.success("✅ ID TKU Pemotong diatur default: 0001658723701000000000")
                        
                        # 15. Tanggal Pemotongan: Cari kolom dengan berbagai variasi nama
                        tanggal_potong_keywords = [
                            'tanggal pemotongan',
                            'tgl pemotongan',
                            'tanggal potong',
                            'tgl potong',
                            'tanggal invoice',
                            'tgl invoice',
                            'invoice date',
                            'tanggal transaksi',
                            'tgl transaksi'
                        ]
                        tanggal_pemotongan_col = find_column_by_keywords(df_merged, tanggal_potong_keywords)
                        
                        if tanggal_pemotongan_col and not df_merged[tanggal_pemotongan_col].isna().all():
                            # Debug info
                            st.info(f"🔍 Ditemukan kolom '{tanggal_pemotongan_col}' untuk Tanggal Pemotongan")
                            st.info(f"📋 Sample data dari kolom ini: {df_merged[tanggal_pemotongan_col].head(3).tolist()}")
                            
                            # Konversi ke format bulan/tanggal/tahun (8/4/2025)
                            try:
                                # Coba parse tanggal dengan berbagai format
                                dates = pd.to_datetime(df_merged[tanggal_pemotongan_col], errors='coerce')
                                # Format ke bulan/tanggal/tahun (tanpa leading zero)
                                df_result['Tanggal Pemotongan'] = dates.apply(
                                    lambda x: f"{x.month}/{x.day}/{x.year}" if pd.notna(x) else ''
                                )
                                st.success(f"✅ Kolom Tanggal Pemotongan ditemukan: {tanggal_pemotongan_col}")
                                st.info(f"📅 Format tanggal: bulan/tanggal/tahun (contoh: 8/4/2025)")
                            except Exception as e:
                                st.warning(f"⚠️ Gagal mengonversi format tanggal di kolom {tanggal_pemotongan_col}. Menggunakan format asli: {str(e)}")
                                df_result['Tanggal Pemotongan'] = df_merged[tanggal_pemotongan_col].astype(str)
                        else:
                            df_result['Tanggal Pemotongan'] = ''
                            st.warning("⚠️ Kolom Tanggal Pemotongan tidak ditemukan atau kosong, diisi dengan nilai kosong (warna oranye)")
                        
                        # Tampilkan informasi kolom yang ditemukan
                        st.info("📊 **Deteksi Kolom Otomatis:**")
                        st.info(f"  • Nomor Dok. Referensi: {nomor_ref_col if nomor_ref_col else 'Tidak ditemukan'}")
                        st.info(f"  • Tanggal Dok. Referensi: {tanggal_ref_col if tanggal_ref_col else 'Tidak ditemukan'}")
                        st.info(f"  • Tanggal Pemotongan: {tanggal_pemotongan_col if tanggal_pemotongan_col else 'Tidak ditemukan'}")
                        
                        # Hitung statistik
                        processed_count = len(df_result)
                        master_matched = processed_count  # Semua data adalah yang match karena inner join
                        not_matched = len(df_raw) - master_matched
                        
                        st.success(f"✅ **Data berhasil diproses!** Total {processed_count} baris data BP 21")
                        
                        # ========== TAMPILKAN HASIL ==========
                        st.markdown("---")
                        st.subheader("📊 Ringkasan Hasil Proses")
                        
                        # Metrik ringkasan
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Data Mentah", len(df_raw))
                        with col2:
                            st.metric("NIP Unik", len(nip_list))
                        with col3:
                            st.metric("✅ Match & Diproses", master_matched)
                        with col4:
                            st.metric("❌ Tidak Match", not_matched)
                        
                        # Informasi perhitungan
                        st.info(f"""
                        **📈 Statistik Hasil:**
                        - **Total Penghasilan:** Rp {df_result['Penghasilan'].sum():,.0f}
                        - **Rata-rata Tarif:** {df_result['Tarif'].mean():.2f}%
                        - **Distribusi Kode Objek Pajak:**
                        """)
                        
                        for tarif, count in tarif_counts.items():
                            kode_mapping = {
                                5.0: '21-402-02',
                                15.0: '21-402-03',
                                0.0: '21-402-04'
                            }
                            kode = kode_mapping.get(tarif, f'Tidak dikenali')
                            st.info(f"  • {kode} (Tarif {tarif:.0f}%): {count} baris")
                        
                        # Informasi format tanggal
                        st.info(f"""
                        **📅 Format Tanggal yang Digunakan:**
                        - **Tanggal Dok. Referensi:** bulan/tanggal/tahun (contoh: 8/4/2025)
                        - **Tanggal Pemotongan:** bulan/tanggal/tahun (contoh: 8/4/2025)
                        
                        **🔑 ID TKU Pemotong:**
                        - **Nilai Default:** 0001658723701000000000
                        - Semua baris menggunakan ID TKU Pemotong yang sama
                        """)
                        
                        # Preview hasil dengan warna indikator
                        st.subheader("📄 Preview Hasil (Warna hanya ilustrasi)")

                        # Buat DataFrame untuk preview dengan indikasi kolom
                        preview_df = df_result.head(10).copy()

                        # Tentukan kolom yang harus berwarna oranye
                        orange_columns_preview = ['Nomor Dok. Referensi', 'Tanggal Dok. Referensi', 'Tanggal Pemotongan']

                        # Tampilkan dataframe dengan styling sederhana
                        st.dataframe(preview_df)

                        # Tampilkan legenda warna di bawah
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("""
                            <div style="background-color: #e6ffe6; padding: 10px; border-radius: 5px; border: 1px solid #ccc; margin-top: 10px;">
                            <strong>🟢 Hijau Muda:</strong> Data hasil sistem (siap copy)
                            </div>
                            """, unsafe_allow_html=True)
                        with col2:
                            st.markdown("""
                            <div style="background-color: #fff0e6; padding: 10px; border-radius: 5px; border: 1px solid #ccc; margin-top: 10px;">
                            <strong>🟠 Oranye:</strong> Isi manual dari DAFTAR SP2D SATKER
                            </div>
                            """, unsafe_allow_html=True)

                        # Tampilkan informasi kolom warna
                        st.info(f"""
                        **Kolom dengan warna Hijau Muda (Data Sistem):** {', '.join([col for col in preview_df.columns if col not in orange_columns_preview])}

                        **Kolom dengan warna Oranye (Data Manual):** {', '.join(orange_columns_preview)}
                        """)
                        
                        # Informasi pengisian manual
                        with st.expander("📝 PETUNJUK PENGISIAN MANUAL", expanded=False):
                            st.markdown("""
                            ### **Kolom yang perlu diisi MANUAL (Warna Oranye):**
                            
                            1. **Nomor Dok. Referensi** → Nomor SP2D
                               - Contoh: `SP2D-001/BPK/2024`
                            
                            2. **Tanggal Dok. Referensi** → Tanggal SP2D
                               - Format: **bulan/tanggal/tahun** (contoh: 8/4/2025)
                               - Contoh: `8/4/2025` untuk 4 Agustus 2025
                            
                            3. **Tanggal Pemotongan** → Tanggal Invoice
                               - Format: **bulan/tanggal/tahun** (contoh: 8/4/2025)
                               - Contoh: `8/4/2025` untuk 4 Agustus 2025
                            
                            **Sumber data:** DAFTAR SP2D SATKER (download dari sistem keuangan)
                            
                            ### **Kolom ID TKU Pemotong:**
                            - **Nilai tetap:** 0001658723701000000000
                            - Sudah diisi otomatis oleh sistem
                            - Tidak perlu perubahan
                            """)
                        
                        # ========== DATA TIDAK MATCH ==========
                        if not_matched > 0:
                            st.markdown("---")
                            st.warning(f"⚠️ **PERHATIAN**: Ditemukan {not_matched} data yang tidak memiliki match di Data Master")
                            
                            # Cari NIP yang tidak match
                            matched_nips = df_master_filtered['NIP'].unique()
                            no_match_data = df_raw_clean[~df_raw_clean['nip'].isin(matched_nips)].copy()
                            
                            # Tampilkan informasi detail
                            st.write(f"**NIP yang tidak ditemukan di Data Master:**")
                            no_match_display = no_match_data[['nip', 'nmpeg']].drop_duplicates('nip')
                            no_match_display.columns = ['NIP', 'Nama Pegawai']
                            st.dataframe(no_match_display, use_container_width=True, height=200)
                            
                            # Download data tidak match
                            csv_no_match = no_match_display.to_csv(index=False).encode('utf-8')
                            st.download_button(
                                label="📥 Download Data Tidak Match",
                                data=csv_no_match,
                                file_name="data_tidak_match_pajak_makan.csv",
                                mime="text/csv",
                                help="Download daftar NIP yang tidak ditemukan di Data Master"
                            )
                        
                        # ========== SIMPAN KE SESSION STATE ==========
                        st.session_state.hasil_pajak_makan_pns = df_result
                        st.session_state.jumlah_data_pns = processed_count
                        st.session_state.tidak_match_pns = not_matched
                        
                        st.balloons()
                        st.success("🎉 **Data BP 21 siap untuk didownload dengan format warna!**")
                        
                    except Exception as e:
                        st.error(f"❌ **ERROR**: Terjadi kesalahan saat memproses data")
                        st.error(f"Detail error: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
            
            # ========== TOMBOL DOWNLOAD DENGAN FORMAT WARNA ==========
            if 'hasil_pajak_makan_pns' in st.session_state:
                st.markdown("---")
                st.subheader("💾 Download Hasil dengan Format Warna")
                
                hasil_final = st.session_state.hasil_pajak_makan_pns
                jumlah_data = st.session_state.get('jumlah_data_pns', 0)
                tidak_match = st.session_state.get('tidak_match_pns', 0)
                
                # Info file
                st.info(f"📄 **Nama file:** BP21_Pajak_Makan_PNS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                st.info(f"📊 **Jumlah data:** {jumlah_data} baris")
                
                if tidak_match > 0:
                    st.warning(f"⚠️ **Data tidak match:** {tidak_match} baris (tidak termasuk dalam file)")
                
                # Buat file Excel dengan format warna menggunakan openpyxl
                output = BytesIO()
                
                # Buat workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "BP21_Pajak_Makan_PNS"
                
                # Tulis header
                headers = list(hasil_final.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                
                # Tulis data
                for row_num, row_data in enumerate(hasil_final.values, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Tentukan kolom untuk warna
                # Kolom yang diwarnai oranye (harus diisi manual)
                orange_columns = ['Nomor Dok. Referensi', 'Tanggal Dok. Referensi', 'Tanggal Pemotongan']
                
                # Cari indeks kolom orange
                orange_indices = []
                for col_name in orange_columns:
                    if col_name in headers:
                        orange_indices.append(headers.index(col_name) + 1)  # +1 karena openpyxl mulai dari 1
                
                # Terapkan warna untuk semua sel data (baris 2 ke atas)
                for row in ws.iter_rows(min_row=2, max_row=len(hasil_final) + 1, 
                                       min_col=1, max_col=len(headers)):
                    for cell in row:
                        # Jika kolom ini termasuk orange columns, cek apakah kosong
                        if cell.column in orange_indices:
                            # Jika sel kosong atau hanya berisi spasi, beri warna oranye
                            if cell.value is None or str(cell.value).strip() == '':
                                cell.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                            else:
                                # Jika sudah ada isi, beri warna hijau muda
                                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        else:
                            # Untuk kolom lainnya, beri warna hijau muda
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Simpan workbook ke BytesIO
                wb.save(output)
                output.seek(0)
                
                # Tombol download dengan format warna
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="📥 **Download Hasil (Excel dengan Format Warna)**",
                        data=output,
                        file_name=f"BP21_Pajak_Makan_PNS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        help="File Excel dengan format warna: Hijau untuk data sistem, Oranye untuk data manual"
                    )
                
                # Informasi tambahan
                with st.expander("ℹ️ Informasi Format File yang Didownload", expanded=False):
                    st.markdown(f"""
                    ### **🎨 Format Warna di File Excel:**
                    
                    **1. Header (Baris 1):**
                    - Warna: **Hijau Tua** (#C6E0B4)
                    - Font: **Bold**
                    
                    **2. Data Hasil Sistem (Siap Copy):**
                    - Warna: **Hijau Muda** (#E6FFE6)
                    - Data ini sudah diproses oleh sistem
                    - Bisa langsung digunakan/dicopy
                    - Total: **{len(headers) - 3} kolom** dari **{len(headers)} kolom**
                    
                    **3. Data yang Harus Diisi Manual:**
                    - Warna: **Oranye** (#FFD580)
                    - Kolom yang kosong akan berwarna oranye
                    - Kolom yang sudah terisi akan berwarna hijau muda
                    - Kolom yang perlu diisi manual:
                      1. **Nomor Dok. Referensi** → dari Nomor SP2D
                      2. **Tanggal Dok. Referensi** → dari Tanggal SP2D (format: bulan/tanggal/tahun)
                      3. **Tanggal Pemotongan** → dari Tanggal Invoice (format: bulan/tanggal/tahun)
                    
                    ### **📋 Kolom dengan Warna Hijau Muda (Data Sistem):**
                    {', '.join([f'**{col}**' for col in headers if col not in orange_columns])}
                    
                    ### **📋 Kolom dengan Warna Oranye (Data Manual):**
                    {', '.join([f'**{col}**' for col in orange_columns if col in headers])}
                    
                    ### **Format Tanggal:**
                    - **Contoh:** 8/4/2025 (untuk 4 Agustus 2025)
                    - **Bulan:** tanpa leading zero (1-12)
                    - **Tanggal:** tanpa leading zero (1-31)
                    - **Tahun:** 4 digit (2025)
                    
                    ### **Kolom ID TKU Pemotong:**
                    - **Nilai tetap:** 0001658723701000000000
                    - Warna hijau muda (sudah diisi otomatis)
                    - Tidak perlu perubahan
                    
                    ### **Sumber Data Manual:**
                    - **DAFTAR SP2D SATKER** (download dari sistem keuangan)
                    - Isi semua baris dengan data yang sama untuk setiap SP2D
                    """)
                
                # Tombol reset
                if st.button("🔄 Reset Proses", use_container_width=True):
                    if 'hasil_pajak_makan_pns' in st.session_state:
                        del st.session_state.hasil_pajak_makan_pns
                    if 'new_data_df' in st.session_state:
                        del st.session_state.new_data_df
                    if 'duplicate_nips_df' in st.session_state:
                        del st.session_state.duplicate_nips_df
                    st.rerun()
        
        except Exception as e:
            st.error(f"❌ **ERROR**: Terjadi kesalahan saat membaca file")
            st.error(f"Detail: {str(e)}")
            st.info("💡 **Tips**: Pastikan file Excel tidak corrupt. Coba buka file di Excel terlebih dahulu.")
    
    elif uploaded_file_raw is not None:
        st.warning("⚠️ **Langkah 2**: Silakan upload juga **Data Master** untuk melanjutkan")
    elif uploaded_file_master is not None:
        st.warning("⚠️ **Langkah 1**: Silakan upload **Data Mentah** terlebih dahulu")
    
    st.markdown("---")
    st.caption("""
    🔧 **Dukungan Teknis**: 
    - **NIP duplikat**: Pastikan tidak ada NIP yang sama di Data Mentah
    - **Data baru**: NIP yang tidak ada di Data Master akan ditandai hijau dan harus ditambahkan
    - **Format file**: Hanya mendukung format Excel (.xlsx, .xls)
    - Untuk perhitungan tarif: Sistem menggunakan KODE OBJEK PAJAK dari Data Master
    - Mapping tarif: 21-402-02=5%, 21-402-03=15%, 21-402-04=0%
    - Sistem TIDAK menggunakan kolom 'pph' atau 'potongan' lagi
    - Format tanggal: bulan/tanggal/tahun (contoh: 8/4/2025)
    - ID TKU Pemotong: nilai default tetap "0001658723701000000000"
    - Untuk data manual (kolom referensi): Download DAFTAR SP2D SATKER dari sistem keuangan
    - Jika ada error, coba download template dan sesuaikan data Anda dengan format yang diberikan
    """)

if __name__ == "__main__":
    show()